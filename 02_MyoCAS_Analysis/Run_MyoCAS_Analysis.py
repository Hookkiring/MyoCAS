"""
MyoCAS Displacement Analysis and A/IA/N Classification

This script performs the main MyoCAS analysis workflow for time-lapse
microscopy videos of contracting myotubes.

Main functions:
    1. Extract frames from input videos.
    2. Identify candidate contraction/relaxation frames based on pixel-intensity changes.
    3. Quantify optical flow-derived contractile displacement.
    4. Extract displacement-derived temporal and scalar features.
    5. Classify each region as inactive, active, or noise using a trained classifier.
    6. Optionally perform myotube segmentation and area quantification.
    7. Save analysis results, plots, videos, and Excel summary files.

Classes:
    INACTIVE: inactive region
    ACTIVE  : active region
    NOISE   : noise region

Notes:
    - Replace all example paths with your own input, checkpoint, and output paths.
    - The trained classifier checkpoint should match the model architecture used in this script.
    - Pixel-to-micrometer calibration should be set according to the microscope setting.
    
KR:
    이 스크립트는 MyoCAS의 주요 분석 코드입니다.
    time-lapse 근관 영상을 입력으로 받아 프레임 추출, 최대 수축/이완 후보 프레임 탐색,
    optical flow 기반 수축 변위 계산, displacement feature 추출, A/IA/N 분류,
    선택적 segmentation 및 면적 계산을 수행합니다.

    INACTIVE는 비활성 영역, ACTIVE는 활성 영역, NOISE는 노이즈 영역을 의미합니다.
    사용 전 입력 영상 경로, 학습된 checkpoint 경로, pixel-to-micrometer 보정값을
    실험 조건에 맞게 수정해야 합니다.

"""

import cv2
import os
import numpy as np
import matplotlib.pyplot as plt
import shutil
import time
from openpyxl import Workbook, load_workbook
import natsort
import gc
from scipy.fft import fft, fftfreq
from matplotlib import colors
import plotly.graph_objects as go
import matplotlib.ticker as ticker
from matplotlib import rcParams
import pywt
import torch
from moviepy.editor import VideoFileClip, CompositeVideoClip, ImageClip, ColorClip
from moviepy.video.fx.all import resize
from torch.utils.data import Dataset, DataLoader
from torchvision.models.segmentation import deeplabv3_resnet50, DeepLabV3_ResNet50_Weights
from torchvision.models.segmentation import deeplabv3_resnet101, DeepLabV3_ResNet101_Weights
from tqdm import tqdm
import albumentations as A
from albumentations.pytorch import ToTensorV2
from torch.optim.lr_scheduler import ReduceLROnPlateau
import torch.nn as nn
import torch.nn.functional as F
import albumentations as A
from albumentations.pytorch import ToTensorV2
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm
import multiprocessing
import ast
import pandas as pd
from PIL import Image
from io import BytesIO
# =========================================================
# USER SETTINGS
# KR: 사용자가 주로 수정해야 하는 주요 설정값
# =========================================================

# Directory containing input time-lapse videos.
# KR: 분석할 time-lapse 영상들이 들어 있는 폴더 경로입니다.
VIDEO_FILE_PATH = r"path/to/input_videos"

# Trained checkpoint for A/IA/N classification.
# KR: active/inactive/noise 분류에 사용할 학습된 classifier checkpoint 경로입니다.
BEST_CLASSIFICATION_CHECKPOINT_FILE = r"weights/MyoCAS_A_IA_N_classification.pth"

# Trained checkpoint for myotube segmentation.
# KR: 근관 segmentation에 사용할 학습된 checkpoint 경로입니다.

BEST_SEGMENTATION_CHECKPOINT_FILE = r"weights/MyoCAS_Segmentation.pth"

# Half-size of the square ROI used for pixel-intensity-based motion search.
# After the pixel with the largest intensity change is detected, a square ROI
# is cropped around that coordinate using this value.
# For example, ANALYSIS_SQUARE_SIZE_REFERENCE = 9 generates an ROI of
# approximately 18 × 18 pixels around the detected motion center.
# This ROI is used to search candidate maximal contraction and relaxation frames.
# KR: 픽셀 강도 변화가 가장 큰 좌표를 찾은 뒤, 그 주변을 작은 사각형 ROI로 자를 때 사용하는 half-size입니다.
# KR: 예를 들어 값이 9이면 변화 중심 좌표 주변 약 18 × 18 pixel 영역을 사용합니다.
# KR: 이 ROI는 최대 수축/최대 이완 후보 프레임을 찾는 데 사용됩니다.
ANALYSIS_SQUARE_SIZE_REFERENCE = 9

# Pixel-to-micrometer calibration factor.
# This value converts pixel displacement measured by optical flow into micrometers.
# Set this value according to the microscope magnification and image resolution.
# KR: optical flow로 계산된 pixel 단위 변위를 micrometer 단위로 변환하는 보정값입니다.
# KR: 현미경 배율과 이미지 해상도에 맞게 설정해야 합니다.
PIXEL_UM = 0.78125

# Area represented by one pixel, calculated from PIXEL_UM.
# This is used to convert pixel counts from segmentation masks into physical area.
# KR: 1 pixel이 나타내는 실제 면적입니다. segmentation mask의 pixel 수를 실제 면적으로 변환할 때 사용합니다.
AREA_MOVED = round((PIXEL_UM * PIXEL_UM), 4)

# Initial analysis window in seconds.
# The number of frames used for the first motion search is calculated as:
# fps × NUM_FRAMES.
# This helps focus the initial search on early contraction-related motion and
# reduces the chance of selecting large noise or bubble movement.
# KR: 초기 motion search에 사용할 시간 범위입니다.
# KR: 실제 사용 프레임 수는 fps × NUM_FRAMES로 계산됩니다.
# KR: 큰 노이즈나 bubble 움직임보다 초기 수축 관련 움직임을 우선 찾기 위한 설정입니다.
NUM_FRAMES = 4

# Manual region selection mode.
# "on": manually select the analysis region.
# "off": automatically detect the analysis region based on pixel-intensity changes.
# KR: 분석 영역을 수동으로 선택할지 여부입니다.
# KR: "on"이면 수동 선택, "off"이면 픽셀 강도 변화 기반으로 자동 탐색합니다.
MANUAL_SELECTION = 'off'

# Manual coordinate selection mode.
# "on": use manually specified coordinates defined by X_COORDINATES and Y_COORDINATES.
# "off": automatically determine the coordinates from image intensity changes.
# KR: 사용자가 지정한 좌표를 사용할지 여부입니다.
# KR: "on"이면 X_COORDINATES와 Y_COORDINATES 값을 사용하고, "off"이면 자동으로 좌표를 탐색합니다.
MANUAL_COORDINATE_SELECTION = 'off'

# A/IA/N classification mode.
# "on": classify each analyzed region as active, inactive, or noise using the trained classifier.
# "off": skip classifier-based region classification.
# KR: 분석 영역을 active, inactive, noise로 분류할지 여부입니다.
# KR: "on"이면 학습된 classifier를 사용하고, "off"이면 분류 과정을 건너뜁니다.
CLASSIFICATION_MODE = 'on'

# Frame removal mode.
# "on": remove or reorganize frame folders after analysis according to the pipeline logic.
# "off": keep extracted frames after analysis.
# KR: 분석 후 추출된 frame 폴더를 정리할지 여부입니다.
# KR: "on"이면 pipeline logic에 따라 frame 폴더를 이동 또는 삭제하고, "off"이면 그대로 유지합니다.
FRAME_REMOVE = 'on'

# Inactive/noise frame removal mode.
# "on": remove frame folders associated with inactive or noise-classified regions.
# "off": keep inactive/noise frame folders.
# KR: inactive 또는 noise로 분류된 영역의 frame 폴더를 제거할지 여부입니다.
# KR: 원본 분석 결과를 보존하려면 "off"로 두는 것이 안전합니다.
INACTIVE_NOISE_FRAME_REMOVE = 'off'

# Reverse contraction direction mode.
# "on": reverse the sign or direction of displacement when the contraction direction is inverted.
# "off": keep the displacement direction as calculated.
# KR: 수축 방향이 반대로 계산된 경우 displacement 방향을 반전할지 여부입니다.
# KR: 일반적으로는 "off"로 두고, 방향 보정이 필요한 경우에만 "on"으로 설정합니다.
REVERSE_CONTRACTION = 'off'

# Analysis video export mode.
# "on": generate output videos showing the analysis result.
# "off": skip video generation.
# Note: video generation can increase processing time and file size.
# KR: 분석 결과를 표시한 output video를 생성할지 여부입니다.
# KR: "on"이면 처리 시간이 길어지고 파일 크기가 커질 수 있습니다.
VIDEO_MODE = 'on'

# Optional displacement double-check mode.
# In the current implementation, predicted_class2 is initially set to the same
# prediction as predicted_class. This switch is provided as a placeholder for
# users who want to add their own manual or rule-based double-check logic.
# Keep this option "off" unless a custom predicted_class2 routine is implemented.
# KR: displacement 결과를 추가로 double-check하기 위한 선택적 placeholder입니다.
# KR: 현재 기본 구현에서는 predicted_class2가 predicted_class와 같은 값으로 초기화되므로,
# KR: 사용자가 별도의 manual rule 또는 custom checking logic을 추가하지 않으면 실질적인 차이는 없습니다.
# KR: 별도 double-check 루틴을 직접 추가하지 않는다면 "off"로 두는 것이 좋습니다.
DOUBLE_CHECK_DISPLACEMENT = 'off'

# Frame extraction skip mode.
# "on": reuse existing extracted frames if they are present and pass validation.
# "off": always extract frames again from the input video.
# KR: 이미 추출된 frame이 존재하고 validation을 통과하면 재사용할지 여부입니다.
# KR: "on"이면 중복 frame extraction을 줄여 분석 속도를 높일 수 있습니다.
FRAME_EXTRACTION_SKIP = 'on'

# GPU acceleration mode.
# "on": use CUDA-enabled GPU acceleration when available.
# "off": force CPU-based processing.
# If CUDA is not available, the script automatically falls back to CPU.
# KR: CUDA GPU를 사용할지 여부입니다.
# KR: "on"이어도 CUDA가 사용 불가능하면 자동으로 CPU를 사용합니다.
GPU_MODE = 'off'

# Extra figure export mode.
# "on": export additional displacement figure variants for manuscript preparation,
#       manual inspection, or figure layout adjustment.
# "off": export only the standard GitHub/public output figure.
# Note: keeping this option "off" is recommended for routine batch analysis
# because extra figures increase the number of output files.
#
# KR:
# 추가 figure variant를 저장할지 여부입니다.
# "on"이면 논문 figure 제작, 수동 확인, figure layout 조정을 위한
# 추가 Moved Area, Pixel, displacement plot들을 함께 저장합니다.
# "off"이면 GitHub/public release용 표준 figure 하나만 저장합니다.
# 일반적인 batch analysis에서는 output 파일 수가 늘어나므로 "off"를 권장합니다.
EXPORT_EXTRA_FIGURES = 'off'

# Number of output classes for A/IA/N classification.
# KR: A/IA/N 분류 모델의 출력 클래스 수입니다.
NB_CLASSES = 3

# Scalar features used for A/IA/N classification.
# KR: A/IA/N 분류 모델에 입력되는 scalar feature 목록입니다.
COLUMNS = [
    'MAX BAND ENERGY',
    'MAX Apeak',
    'MAX ABS DISP',
    'ABS NET CHANGE RATIO',
    'NUM SIGN CHANGES DIFF',
    'DIFF STD'
]

NB_INPUT = len(COLUMNS)

# Class labels used by the classifier.
# KR: classifier가 출력하는 class label입니다.
CLASS_NAMES = ['INACTIVE', 'ACTIVE', 'NOISE']

# Default coordinates used when manual coordinate selection is enabled.
# KR: manual coordinate selection이 켜져 있을 때 사용할 기본 좌표입니다.
X_COORDINATES, Y_COORDINATES = 22, 109
# =========================================================
# FILE, OUTPUT, AND VISUALIZATION SETTINGS
# KR: 파일명, 출력 폴더, 단위, 시각화 관련 설정값
# =========================================================

FRAME_PREFIX = 'frame_'  # Prefix for extracted frame image files
# KR: 추출된 frame 이미지 파일명 앞에 붙는 접두사입니다.

VIDEO_FILE_EXTENSION = ['mp4', 'mov', 'wmv', 'mkv', 'mpeg', 'flv', 'webm', 'avi']  # Supported video formats
# KR: 입력으로 허용할 video file 확장자 목록입니다.

FRAME_EXTENSION = 'png'  # Image format for extracted frames and saved figures
# KR: 추출 frame과 저장 figure의 이미지 확장자입니다.

VIDEO_CODEX = 'mp4v'  # Codec used for exported analysis videos
# KR: 분석 결과 video를 저장할 때 사용하는 codec입니다.

GRID_SPACING = 20  # Grid spacing for x/y axis visualization
# KR: x/y축 grid visualization에서 grid 간격을 설정합니다.

LINE_SPACING = 20  # Line spacing for vector or scale visualization
# KR: vector 또는 scale visualization에서 line 간격을 설정합니다.

# Units used in output files and figure labels
# KR: Excel output과 figure label에 표시되는 단위입니다.
LENGTH_UNIT = 'μm'
AREA_UNIT = 'μm²'
VELOCITY_UNIT = 'μm/s'
ACCELERATION_UNIT = 'μm/s²'

# Excel output file names for each classification category
# KR: classification 결과별 Excel output 파일명입니다.
STRONG_SAMPLE_EXCEL_NAME = 'ACTIVE_SAMPLE.xlsx'
WEAK_SAMPLE_EXCEL_NAME = 'INACTIVE_SAMPLE.xlsx'
WEAK_DOUBLECHECK_SAMPLE_EXCEL_NAME = 'INACTIVE_DOUBLE_SAMPLE.xlsx'
NOISE_SAMPLE_EXCEL_NAME = 'NOISE_SAMPLE.xlsx'
OPTICAL_NOISE_SAMPLE_EXCEL_NAME = 'OPTICAL_NOISE_SAMPLE.xlsx'
CHECK_SAMPLE_EXCEL_NAME = 'DOUBLE_CHECK.xlsx'

# Output folder names for each classification category
# KR: classification 결과별 output folder 이름입니다.
STRONG_SAMPLE_FOLDER_NAME = 'ACTIVE'
WEAK_SAMPLE_FOLDER_NAME = 'INACTIVE'
WEAK_DOUBLE_CHECK_SAMPLE_FOLDER_NAME = 'INACTIVE_DOUBLE'
NOISE_SAMPLE_FOLDER_NAME = 'NOISE'
OPTICAL_NOISE_SAMPLE_FOLDER_NAME = 'OPTICAL_NOISE'
CHECK_SAMPLE_FOLDER_NAME = 'DOUBLE_CHECK'

# Full output paths for Excel summary files
# KR: Excel summary file이 저장될 전체 경로입니다.
strong_sample_excel_path = os.path.join(VIDEO_FILE_PATH, STRONG_SAMPLE_EXCEL_NAME)
weak_sample_excel_path = os.path.join(VIDEO_FILE_PATH, WEAK_SAMPLE_EXCEL_NAME)
weak_double_sample_excel_path = os.path.join(VIDEO_FILE_PATH, WEAK_DOUBLECHECK_SAMPLE_EXCEL_NAME)
noise_sample_excel_path = os.path.join(VIDEO_FILE_PATH, NOISE_SAMPLE_EXCEL_NAME)
optical_noise_sample_excel_path = os.path.join(VIDEO_FILE_PATH, OPTICAL_NOISE_SAMPLE_EXCEL_NAME)
check_sample_excel_path = os.path.join(VIDEO_FILE_PATH, CHECK_SAMPLE_EXCEL_NAME)

# Reference contraction/stimulation frequency in Hz.
# This value is used as a fallback or reference frequency for FFT-based feature extraction
# when the dominant contraction frequency cannot be robustly determined from the signal.
# KR:
# 기준 수축/자극 주파수(Hz)입니다.
# 실제 FFT 분석에서는 displacement signal에서 dominant frequency를 자동으로 탐색하지만,
# 신호에서 주파수가 안정적으로 검출되지 않거나 기준 frequency window가 필요한 경우
# 이 값을 fallback/reference 값으로 사용합니다.
Hz = 1


# Default video frame rate used as a fallback value.
# The actual FPS is automatically read from each input video when possible.
# If FPS cannot be detected from the video metadata, this value is used instead.
# KR:
# video FPS를 자동으로 읽어오지 못했을 때 사용하는 fallback frame rate입니다.
# 실제 분석에서는 각 입력 video의 metadata에서 FPS를 자동으로 가져오지만,
# FPS 검출에 실패한 경우 이 값을 대신 사용합니다.
VIDEO_FPS = 30

# Default y-axis and FFT plot limits
# KR: displacement plot과 FFT plot의 기본 축 범위 설정입니다.
Y_LIM = 8
X_LIM_AMPLITUDE = 3
Y_LIM_AMPLITUDE = 500
X_TICKS = 5  # Tick interval for the x-axis
# KR: x축 tick 간격입니다.

rcParams['font.family'] = 'Arial'
plt.rcParams['figure.dpi'] = 300
use_gpu = (GPU_MODE.lower() == "on") and torch.cuda.is_available()
device = torch.device("cuda" if use_gpu else "cpu")

print(f"⚙ Compute Mode: {'GPU' if use_gpu else 'CPU'}")

print("\n" + "="*60)
print("MyoCAS analysis parameter check started.")
print("="*60)

print(f"ANALYSIS_SQUARE_SIZE_REFERENCE : {ANALYSIS_SQUARE_SIZE_REFERENCE} px")
print(f"PIXEL_UM                       : {PIXEL_UM} µm/px")
print(f"NUM_FRAMES                     : {NUM_FRAMES} frames")
print("-"*60)
print(f"MANUAL_SELECTION               : {MANUAL_SELECTION}")
print(f"CLASSIFICATION_MODE            : {CLASSIFICATION_MODE}")
print(f"FRAME_REMOVE                   : {FRAME_REMOVE}")
print(f"REVERSE_CONTRACTION            : {REVERSE_CONTRACTION}")
print(f"VIDEO_MODE                     : {VIDEO_MODE}")
print(f"INACTIVE_NOISE_FRAME_REMOVE    : {INACTIVE_NOISE_FRAME_REMOVE}")
print(f"DOUBLE_CHECK_DISPLACEMENT      : {DOUBLE_CHECK_DISPLACEMENT}")
print(f"GPU_MODE                       : {GPU_MODE}")

print("="*60)
print("Analysis started.")
print("="*60 + "\n")


# Create the base output folder if it does not exist
if not os.path.exists(VIDEO_FILE_PATH):
    os.makedirs(VIDEO_FILE_PATH)
    
    
# Check whether the base folder exists
if not os.path.exists(VIDEO_FILE_PATH):
    print("Base folder not found.")
    exit()

video_files = [
    f for f in natsort.natsorted(os.listdir(VIDEO_FILE_PATH))
    if any(f.lower().endswith(ext) for ext in VIDEO_FILE_EXTENSION)
]

print(f"Total number of videos: {len(video_files)}")

# =========================================================
# Get the list of input videos
# =========================================================
video_files = [
    f for f in natsort.natsorted(os.listdir(VIDEO_FILE_PATH))
    if any(f.lower().endswith(ext) for ext in VIDEO_FILE_EXTENSION)
]

print(f"\nTotal videos: {len(video_files)}")


# ==========================================================
# Folder / Video cleanup handler
# ==========================================================
def handle_folder_video(folder_path, video_path, destination_folder, remove_noise_flag):
    try:
        os.makedirs(destination_folder, exist_ok=True)

        if remove_noise_flag:
            # Remove the folder and move the corresponding video
            if os.path.exists(folder_path):
                shutil.rmtree(folder_path, ignore_errors=True)

            if os.path.exists(video_path):
                shutil.move(video_path, destination_folder)

        else:
            # Move both the folder and the corresponding video
            if os.path.exists(folder_path):
                shutil.move(folder_path, destination_folder)

            if os.path.exists(video_path):
                shutil.move(video_path, destination_folder)

    except Exception as e:
        print(f"[WARN] Failed to process folder/video: {e}")


class HybridLSTMClassifier(nn.Module):
    def __init__(
        self,
        seq_input_dim=2,
        scalar_input_dim=6,
        num_classes=3,
        lstm_hidden1=64,
        lstm_hidden2=32,
        dropout=0.3
    ):
        super().__init__()

        self.lstm1 = nn.LSTM(
            input_size=seq_input_dim,
            hidden_size=lstm_hidden1,
            batch_first=True,
            bidirectional=True
        )

        self.lstm2 = nn.LSTM(
            input_size=lstm_hidden1 * 2,
            hidden_size=lstm_hidden2,
            batch_first=True,
            bidirectional=True
        )

        self.dropout = nn.Dropout(dropout)

        # attention score
        self.attn_fc = nn.Linear(lstm_hidden2 * 2, 1)

        self.seq_fc = nn.Sequential(
            nn.Linear(lstm_hidden2 * 2, 64),
            nn.ReLU(),
            nn.LayerNorm(64),
            nn.Dropout(dropout),
            nn.Linear(64, 32),
            nn.ReLU()
        )

        self.scalar_fc = nn.Sequential(
            nn.Linear(scalar_input_dim, 32),
            nn.ReLU(),
            nn.LayerNorm(32),
            nn.Dropout(dropout),
            nn.Linear(32, 16),
            nn.ReLU()
        )

        self.classifier = nn.Sequential(
            nn.Linear(32 + 16, 128),
            nn.ReLU(),
            nn.LayerNorm(128),
            nn.Dropout(dropout),
            nn.Linear(128, 64),
            nn.ReLU(),
            nn.LayerNorm(64),
            nn.Dropout(dropout),
            nn.Linear(64, 32),
            nn.ReLU(),
            nn.Linear(32, num_classes)
        )

    def forward(self, seq_x, scalar_x, lengths):
        """
        seq_x   : [B, T, 2]
        scalar_x: [B, F]
        lengths : [B]
        """

        lengths_cpu = lengths.detach().cpu()

        packed = nn.utils.rnn.pack_padded_sequence(
            seq_x,
            lengths_cpu,
            batch_first=True,
            enforce_sorted=False
        )

        packed_out1, _ = self.lstm1(packed)
        packed_out2, _ = self.lstm2(packed_out1)

        lstm_out, _ = nn.utils.rnn.pad_packed_sequence(
            packed_out2,
            batch_first=True
        )  # [B, T_valid, hidden*2]

        lstm_out = self.dropout(lstm_out)

        # -----------------------------
        # Attention pooling with mask
        # -----------------------------
        B, T, C = lstm_out.shape

        mask = torch.arange(T, device=seq_x.device).unsqueeze(0) < lengths.unsqueeze(1)
        # mask shape: [B, T]

        attn_scores = self.attn_fc(lstm_out).squeeze(-1)   # [B, T]
        attn_scores = attn_scores.masked_fill(~mask, float('-inf'))
        attn_weights = torch.softmax(attn_scores, dim=1)   # [B, T]

        seq_feat = torch.sum(lstm_out * attn_weights.unsqueeze(-1), dim=1)  # [B, C]
        seq_feat = self.seq_fc(seq_feat)   # [B, 32]

        scalar_feat = self.scalar_fc(scalar_x)  # [B, 16]

        x = torch.cat([seq_feat, scalar_feat], dim=1)
        logits = self.classifier(x)

        return logits
    
    
def predict_single_sample_torch(model, displacement_list, scalar_values, device):
    if displacement_list is None or len(displacement_list) == 0:
        raise ValueError("displacement_list is empty.")

    seq_raw_np = np.array(displacement_list, dtype=np.float32)

    # -----------------------------
    # 1) Raw channel: preserve amplitude information with log compression
    # -----------------------------
    seq_raw_processed = np.sign(seq_raw_np) * np.log1p(np.abs(seq_raw_np))

    # -----------------------------
    # 2) Shape channel: normalize the waveform shape
    # -----------------------------
    max_abs = np.max(np.abs(seq_raw_np))
    if max_abs > 1e-8:
        seq_shape_processed = seq_raw_np / max_abs
    else:
        seq_shape_processed = seq_raw_np.copy()

    # -----------------------------
    # 3) Create a two-channel sequence
    # -----------------------------
    seq_2ch = np.stack(
        [seq_raw_processed, seq_shape_processed],
        axis=-1
    )  # [T, 2]

    valid_len = len(seq_2ch)

    x_seq = np.expand_dims(seq_2ch, axis=0).astype(np.float32)   # [1, T, 2]
    x_scalar = np.array(scalar_values, dtype=np.float32).reshape(1, -1)

    seq_tensor = torch.tensor(x_seq, dtype=torch.float32).to(device)
    scalar_tensor = torch.tensor(x_scalar, dtype=torch.float32).to(device)
    length_tensor = torch.tensor([valid_len], dtype=torch.long).to(device)

    with torch.no_grad():
        logits = model(seq_tensor, scalar_tensor, length_tensor)
        probs = torch.softmax(logits, dim=1)
        pred_class = torch.argmax(probs, dim=1).item()

    return pred_class, probs.squeeze(0).cpu().numpy()


# Save figure to the result folder
def Savefig(name, dpi, pad_inches):
    result_folder = os.path.join(folder_path, 'result')
    os.makedirs(result_folder, exist_ok=True)
    result_image_path = os.path.join(result_folder, f'{name}.{FRAME_EXTENSION}')

    # Save the plot
    plt.savefig(result_image_path, dpi=dpi, bbox_inches='tight', pad_inches=pad_inches)
    plt.close()

    print(f'{name} saved')
    

def Savefig_segmentation(pred_mask, name, dpi, pad_inches, segmentation_area=None):
    result_folder = os.path.join(folder_path, 'result')
    os.makedirs(result_folder, exist_ok=True)
    result_image_path = os.path.join(result_folder, f'{name}.{FRAME_EXTENSION}')

    plt.figure(figsize=(6, 6))
    plt.imshow(pred_mask, cmap='gray')
    plt.axis('off')

    # Display the segmentation area in the upper-right corner
    if segmentation_area is not None:
        plt.text(
            0.99, 0.01,  # Position in axes coordinates
            f'Area: {segmentation_area} {AREA_UNIT}', # Text to display
            color='red',
            fontsize=12,
            ha='right',  # Horizontal alignment
            va='bottom',  # Vertical alignment
            transform=plt.gca().transAxes  # Use axes-relative coordinates
        )

    plt.savefig(result_image_path, dpi=dpi, bbox_inches='tight', pad_inches=pad_inches)
    plt.close()

    print(f'{name} saved')
    
    
def extract_frames(video_file):

    video_path = os.path.join(VIDEO_FILE_PATH, video_file)
    folder_name = os.path.splitext(video_file)[0]
    output_dir = os.path.join(VIDEO_FILE_PATH, folder_name)

    # ---------------------------------------------------------
    # Check skip mode and validate existing extracted frames
    # ---------------------------------------------------------
    if FRAME_EXTRACTION_SKIP.lower() == 'on' and os.path.exists(output_dir):

        cap = cv2.VideoCapture(video_path, cv2.CAP_FFMPEG)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 2)

        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        cap.release()
        del cap 

        saved_frames = sum(
            1 for f in os.scandir(output_dir)
            if f.name.lower().endswith("." + FRAME_EXTENSION)
        )


        # Validate extracted frames when the frame count matches
        if total_frames > 0 and saved_frames == total_frames:

            print("Frame validation check...")

            try:
                sample_frames = sorted(os.listdir(output_dir))[:5]
                for sf in sample_frames:
                    img = cv2.imread(os.path.join(output_dir, sf))
                    if img is None:
                        raise Exception("Frame corrupted")
                    del img

                print("Frame validation passed.")
                return saved_frames

            except:
                print("Frame validation failed → re-extract")

        print("Frame mismatch → re-extract")

    # ---------------------------------------------------------
    # FRAME EXTRACTION
    # ---------------------------------------------------------
    os.makedirs(output_dir, exist_ok=True)

    cap = cv2.VideoCapture(video_path)
    frame_number = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        image_filename = os.path.join(
            output_dir,
            f"{FRAME_PREFIX}{frame_number:04d}.{FRAME_EXTENSION}"
        )

        cv2.imwrite(
        image_filename,
        frame,
        [cv2.IMWRITE_PNG_COMPRESSION, 1]  # Use low PNG compression for faster frame extraction
        )
        del frame
        if frame_number % 100 == 0:
            gc.collect()


        frame_number += 1

    cap.release()
    del cap  

    return frame_number


# ---------- HRNet backbone: simplified multi-scale implementation ----------
class BasicBlock(nn.Module):
    expansion = 1
    def __init__(self, inplanes, planes, stride=1, downsample=None):
        super().__init__()
        self.conv1 = nn.Conv2d(inplanes, planes, 3, stride, 1, bias=False)
        self.bn1 = nn.BatchNorm2d(planes)
        self.relu = nn.ReLU(inplace=True)
        self.conv2 = nn.Conv2d(planes, planes, 3, 1, 1, bias=False)
        self.bn2 = nn.BatchNorm2d(planes)
        self.downsample = downsample
        self.stride = stride

    def forward(self, x):
        identity = x
        out = self.relu(self.bn1(self.conv1(x)))
        out = self.bn2(self.conv2(out))
        if self.downsample is not None:
            identity = self.downsample(x)
        out += identity
        out = self.relu(out)
        return out

class CustomHRNet(nn.Module):
    def __init__(self, num_classes=2):
        super().__init__()
        # Stage 1
        self.stem = nn.Sequential(
            nn.Conv2d(3, 64, 3, 2, 1, bias=False),
            nn.BatchNorm2d(64),
            nn.ReLU(inplace=True),
            nn.Conv2d(64, 64, 3, 2, 1, bias=False),
            nn.BatchNorm2d(64),
            nn.ReLU(inplace=True)
        )
        self.layer1 = self._make_layer(BasicBlock, 64, 64, 4)
        # Stage 2: 2 branch
        self.stage2_branch1 = nn.Sequential(
            nn.Conv2d(64, 32, 3, 1, 1, bias=False),
            nn.BatchNorm2d(32),
            nn.ReLU(inplace=True)
        )
        self.stage2_branch2 = nn.Sequential(
            nn.Conv2d(64, 64, 3, 2, 1, bias=False),
            nn.BatchNorm2d(64),
            nn.ReLU(inplace=True)
        )
        # Stage 3: 3 branch
        self.stage3_branch1 = nn.Conv2d(32, 32, 3, 1, 1, bias=False)
        self.stage3_branch2 = nn.Conv2d(64, 64, 3, 1, 1, bias=False)
        self.stage3_branch3 = nn.Sequential(
            nn.Conv2d(64, 128, 3, 2, 1, bias=False),
            nn.BatchNorm2d(128),
            nn.ReLU(inplace=True)
        )
        # Feature fuse + upsample
        self.fuse_conv = nn.Conv2d(32 + 64 + 128, 256, 1)
        self.classifier = nn.Conv2d(256, num_classes, 1)
        
    def _make_layer(self, block, inplanes, planes, blocks, stride=1):
        layers = []
        downsample = None
        if stride != 1 or inplanes != planes * block.expansion:
            downsample = nn.Sequential(
                nn.Conv2d(inplanes, planes * block.expansion, 1, stride, bias=False),
                nn.BatchNorm2d(planes * block.expansion),
            )
        layers.append(block(inplanes, planes, stride, downsample))
        for _ in range(1, blocks):
            layers.append(block(planes * block.expansion, planes))
        return nn.Sequential(*layers)

    def forward(self, x):
        input_shape = x.shape[-2:]
        x = self.stem(x)     # [B, 64, 30, 30] for input 120x120
        x = self.layer1(x)   # [B, 64, 30, 30]
        # Stage 2 branches
        b1 = self.stage2_branch1(x)  # [B, 32, 30, 30]
        b2 = self.stage2_branch2(x)  # [B, 64, 15, 15]
        # Stage 3 branches
        b1_3 = self.stage3_branch1(b1)  # [B, 32, 30, 30]
        b2_3 = self.stage3_branch2(b2)  # [B, 64, 15, 15]
        b3_3 = self.stage3_branch3(b2)  # [B, 128, 8, 8]
        # Upsample all to same (b1_3) size
        b2_3_up = F.interpolate(b2_3, size=b1_3.shape[-2:], mode='bilinear', align_corners=False)
        b3_3_up = F.interpolate(b3_3, size=b1_3.shape[-2:], mode='bilinear', align_corners=False)
        # Concatenate
        feat = torch.cat([b1_3, b2_3_up, b3_3_up], dim=1)   # [B, 32+64+128, 30, 30]
        feat = self.fuse_conv(feat)   # [B, 256, 30, 30]
        feat = F.relu(feat)
        feat = F.interpolate(feat, size=input_shape, mode='bilinear', align_corners=False)
        out = self.classifier(feat)   # [B, num_classes, 120, 120]
        return out

# ---------- BatchNorm → GroupNorm ----------
def convert_batchnorm_to_groupnorm(module):
    for name, child in module.named_children():
        if isinstance(child, torch.nn.BatchNorm2d):
            setattr(module, name, torch.nn.GroupNorm(32, child.num_features))
        else:
            convert_batchnorm_to_groupnorm(child)

# ---------- HRNet OCR block ----------
class BasicBlock(nn.Module):
    expansion = 1
    def __init__(self, inplanes, planes, stride=1, downsample=None):
        super().__init__()
        self.conv1 = nn.Conv2d(inplanes, planes, kernel_size=3, stride=stride,
                            padding=1, bias=False)
        self.bn1 = nn.BatchNorm2d(planes)
        self.relu = nn.ReLU(inplace=True)

        self.conv2 = nn.Conv2d(planes, planes, kernel_size=3, stride=1,
                            padding=1, bias=False)
        self.bn2 = nn.BatchNorm2d(planes)

        self.downsample = downsample
        self.stride = stride

    def forward(self, x):
        identity = x
        out = self.relu(self.bn1(self.conv1(x)))
        out = self.bn2(self.conv2(out))
        if self.downsample is not None:
            identity = self.downsample(x)
        out += identity
        out = self.relu(out)
        return out

class HRNetOCR(nn.Module):
    def __init__(self, num_classes=2):
        super().__init__()
        self.conv1 = nn.Sequential(
            nn.Conv2d(3, 64, 3, stride=2, padding=1, bias=False),
            nn.BatchNorm2d(64),
            nn.ReLU(inplace=True)
        )
        self.layer1 = self._make_layer(BasicBlock, 64, 64, 4)
        self.stage2 = nn.Identity()
        self.stage3 = nn.Identity()
        self.stage4 = nn.Identity()
        self.final_layer = nn.Sequential(
            nn.Conv2d(64, 256, kernel_size=1, bias=False),
            nn.BatchNorm2d(256),
            nn.ReLU(inplace=True)
        )
        self.classifier = nn.Conv2d(256, num_classes, kernel_size=1)  

    def _make_layer(self, block, inplanes, planes, blocks, stride=1):
        layers = []
        downsample = None
        if stride != 1 or inplanes != planes * block.expansion:
            downsample = nn.Sequential(
                nn.Conv2d(inplanes, planes * block.expansion,
                        kernel_size=1, stride=stride, bias=False),
                nn.BatchNorm2d(planes * block.expansion),
            )
        layers.append(block(inplanes, planes, stride, downsample))
        for _ in range(1, blocks):
            layers.append(block(planes * block.expansion, planes))
        return nn.Sequential(*layers)

    def forward(self, x):
        x = self.conv1(x)
        x = self.layer1(x)
        x = self.stage2(x)
        x = self.stage3(x)
        x = self.stage4(x)
        x = self.final_layer(x)
        x = self.classifier(x)  
        return x


# ---------- DualNet: HRNet + DeepLabV3 ----------
class DualNet(nn.Module):
    def __init__(self, hrnet, deeplabv3, num_classes=2):
        super().__init__()
        self.hrnet = hrnet
        self.deeplabv3 = deeplabv3
        self.fuse = nn.Sequential(
            nn.Conv2d(4, 64, kernel_size=3, padding=1),  
            nn.GroupNorm(8, 64),                          
            nn.ReLU(inplace=True),
            nn.Dropout(0.2),                              
            nn.Conv2d(64, 2, kernel_size=1)              
        )

    def forward(self, x):
        input_shape = x.shape[-2:] 
        hrnet_out = self.hrnet(x)
        deeplab_out = self.deeplabv3(x)['out']
        if hrnet_out.shape[-2:] != deeplab_out.shape[-2:]:
            deeplab_out = F.interpolate(deeplab_out, size=hrnet_out.shape[-2:], mode='bilinear', align_corners=False)
        fusion = torch.cat([hrnet_out, deeplab_out], dim=1)
        out = self.fuse(fusion)
        out = F.interpolate(out, size=input_shape, mode='bilinear', align_corners=False)
        return out
    
# ---------- Loss ----------
class DiceIoUCELoss(nn.Module):
    def __init__(self, dice_weight=0.5, iou_weight=0.4, ce_weight=0.1, weight=None):
        super().__init__()
        self.dice_weight = dice_weight
        self.iou_weight = iou_weight
        self.ce_weight = ce_weight
        self.ce = nn.CrossEntropyLoss(weight=weight)

    def forward(self, inputs, targets):
        ce_loss = self.ce(inputs, targets)
        smooth = 1e-6
        inputs_soft = F.softmax(inputs, dim=1)[:, 1, :, :]
        targets_bin = (targets == 1).float()
        intersection = (inputs_soft * targets_bin).sum()
        dice_loss = 1 - ((2 * intersection + smooth) / (inputs_soft.sum() + targets_bin.sum() + smooth))
        union = inputs_soft.sum() + targets_bin.sum() - intersection
        iou_loss = 1 - ((intersection + smooth) / (union + smooth))
        total_loss = (self.dice_weight * dice_loss +
                    self.iou_weight * iou_loss +
                    self.ce_weight * ce_loss)
        return total_loss
    
# =========================================================
# MULTI EXTRACTION RUN
# =========================================================
print("\nParallel Frame Extraction START\n")


# Single-image segmentation inference using the preloaded model and device
def preprocess_and_segment(image_path):
    # ---------- Load image ----------
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    if img is None:
        raise ValueError(f"Failed to load image: {image_path}")

    # ---------- CLAHE + Gamma + Highpass + Smoothing ----------
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    clahe_img = clahe.apply(img)
    blur = cv2.GaussianBlur(clahe_img, (5, 5), 0)
    highpass = cv2.subtract(clahe_img, blur)
    enhanced = cv2.add(clahe_img, highpass)
    smoothed = cv2.GaussianBlur(enhanced, (5, 5), sigmaX=1.0)

    # ---------- Normalize and convert to tensor ----------
    transform = A.Compose([
        A.Normalize(mean=0.5, std=0.5),
        ToTensorV2()
    ])
    transformed = transform(image=smoothed)['image'].unsqueeze(0).repeat(1, 3, 1, 1)

    # ---------- Inference ----------
    with torch.no_grad():
        inputs = transformed.to(seg_device, non_blocking=True)
        outputs = seg_model(inputs)
        preds = torch.argmax(outputs, dim=1).squeeze().detach().cpu().numpy()

    # ---------- Post-process and save mask ----------
    pred_mask = (preds * 255).astype(np.uint8)

    white_pixel_count_segmentation = int(np.sum(pred_mask >= 128))
    segmentation_area = round(white_pixel_count_segmentation * AREA_MOVED)
    
    if EXPORT_EXTRA_FIGURES == 'on':
        Savefig_segmentation(
            pred_mask=pred_mask,
            name='Segmentation Result_Contration',
            dpi=200,
            pad_inches=0.1,
            segmentation_area=segmentation_area
        )

    # Optional GPU/CPU memory cleanup
    del inputs, outputs
    if seg_device.type == 'cuda':
        torch.cuda.empty_cache()

    
    return segmentation_area


def plot_fft(
    data_list, X_LIM_AMPLITUDE = X_LIM_AMPLITUDE, Y_LIM_AMPLITUDE = Y_LIM_AMPLITUDE,
    save_prefix='save_prefix',
):
    sampling_rate = fps

    # FFT
    fft_values = fft(data_list)
    frequencies = fftfreq(len(data_list), d=1/sampling_rate)

    positive_frequencies = frequencies[:len(frequencies)//2]
    positive_fft_values = np.abs(fft_values[:len(fft_values)//2])
    
    # Full frequency band above 0.5 Hz
    total_band = (positive_frequencies > 0.5)

    # Frequencies and FFT amplitudes within the selected band
    freqs_band = positive_frequencies[total_band]
    amps_band = positive_fft_values[total_band]

    # Find the frequency with the maximum amplitude
    max_idx = np.argmax(amps_band)  # Index of the maximum amplitude
    freq_at_max_amp = freqs_band[max_idx]  # Frequency at maximum amplitude
    freq_at_max_amp_number = round(freqs_band[max_idx],2)  
    amp_at_max_freq = amps_band[max_idx]    # Amplitude at the selected frequency
    if freq_at_max_amp < Hz-0.1:
        freq_at_max_amp = np.floor(Hz-0.1)
    else:
        freq_at_max_amp = round(freq_at_max_amp)
    half_width = 0.5  # ±0.5 Hz frequency window
    lower_bound = round(max(freq_at_max_amp - half_width, 0.5),2)
    upper_bound = round(freq_at_max_amp + half_width,2)

    target_band = (positive_frequencies >= lower_bound) & (positive_frequencies <= upper_bound)
    energy_peak_window = np.sum(positive_fft_values[target_band] ** 2)
    band_energy = round(energy_peak_window)  

    # Round and print FFT summary values
    freq_at_max_amp_round = round(freq_at_max_amp,2)
    amp_at_max_freq_round = round(amp_at_max_freq)
  
    print(f'Max amplitude {amp_at_max_freq_round} occurs at {freq_at_max_amp_number} Hz | Energy {band_energy} | Range {lower_bound} ~ {upper_bound}')
    

    # ---------- First plot ----------
    fig, ax = plt.subplots(figsize=(10, 10))
    ax.fill_between(positive_frequencies, 0, positive_fft_values, where=target_band, color=yellow_color, alpha=0.5)
    ax.plot(positive_frequencies, positive_fft_values, color=red_color, linewidth=4, zorder=10)
    ax.set_xlabel("Frequency (Hz)", fontweight='bold', fontsize=50)
    ax.set_ylabel("Amplitude", fontweight='bold', fontsize=50)
    ax.grid(False)
    ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=20)
    ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=20)
    ax.set_xlim(0, max(positive_frequencies)+1)
    ax.set_xticks(np.arange(0, max(positive_frequencies)+1, 1))
    ax.set_ylim(0, max(positive_fft_values) + 50)
    ax.set_yticks(np.arange(0, max(positive_fft_values) + 50, 100))
    for spine in ax.spines.values():
        spine.set_linewidth(4)
        spine.set_color('black')
    ax.spines['left'].set_position(('outward', 0))
    ax.spines['bottom'].set_position(('outward', 0))
    plt.xticks(fontweight='bold')
    plt.yticks(fontweight='bold')
    ax.text(0.95, 0.95, f'E [{lower_bound , upper_bound}]: {band_energy}', ha='right', va='top', transform=ax.transAxes, fontsize=20, color=yellow_color , weight='bold')
    ax.text(0.9, 0.9, f'Apeak: {amp_at_max_freq_round}', ha='right', va='top', transform=ax.transAxes, fontsize=20, color=blue_color , weight='bold')
    ax.axhline(y=amp_at_max_freq_round, color=blue_color, linestyle='--', linewidth=2)

    plt.tight_layout()
    Savefig(f'{save_prefix}', 300, 0.05)

    if EXPORT_EXTRA_FIGURES == 'on':
        # ---------- Second plot for manuscript-style visualization ----------
        fig, ax = plt.subplots(figsize=(10, 10))
        ax.fill_between(positive_frequencies, 0, positive_fft_values, where=target_band, color=yellow_color, alpha=0.5)
        ax.plot(positive_frequencies, positive_fft_values, color=red_color, linewidth=4, zorder=10)
        ax.grid(False)
        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=20)
        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=20)
        ax.set_xlim(0, X_LIM_AMPLITUDE)
        ax.set_xticks(np.arange(0, X_LIM_AMPLITUDE + 1, 1))
        ax.set_ylim(0, Y_LIM_AMPLITUDE)
        ax.set_yticks(np.arange(0, Y_LIM_AMPLITUDE + 50, 100))
        for spine in ax.spines.values():
            spine.set_linewidth(4)
            spine.set_color('black')
        ax.spines['left'].set_position(('outward', 0))
        ax.spines['bottom'].set_position(('outward', 0))
        plt.xticks(fontweight='bold')
        plt.yticks(fontweight='bold')
        plt.tight_layout()
        Savefig(f'{save_prefix} Paper', 300, 0.05)
        
        # ---------- Third plot for manuscript-style visualization ----------
        fig, ax = plt.subplots(figsize=(10, 10))
        ax.fill_between(positive_frequencies, 0, positive_fft_values, where=target_band, color=yellow_color, alpha=0.5)
        ax.plot(positive_frequencies, positive_fft_values, color=red_color, linewidth=4, zorder=10)
        ax.set_xlabel("Frequency (Hz)", fontweight='bold', fontsize=50)
        ax.set_ylabel("Amplitude", fontweight='bold', fontsize=50)
        ax.grid(False)
        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=20)
        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=20)
        ax.set_xlim(0, X_LIM_AMPLITUDE)
        ax.set_xticks(np.arange(0, X_LIM_AMPLITUDE + 1, 1))
        ax.set_ylim(0, Y_LIM_AMPLITUDE)
        ax.set_yticks(np.arange(0, Y_LIM_AMPLITUDE + 50, 100))
        for spine in ax.spines.values():
            spine.set_linewidth(4)
            spine.set_color('black')
        ax.spines['left'].set_position(('outward', 0))
        ax.spines['bottom'].set_position(('outward', 0))
        plt.xticks(fontweight='bold')
        plt.yticks(fontweight='bold')
        ax.text(0.95, 0.95, f'E [{lower_bound , upper_bound}]: {band_energy}', ha='right', va='top', transform=ax.transAxes, fontsize=20, color=yellow_color , weight='bold')
        ax.text(0.9, 0.9, f'Apeak: {amp_at_max_freq_round}', ha='right', va='top', transform=ax.transAxes, fontsize=20, color=blue_color , weight='bold')
        ax.axhline(y=amp_at_max_freq_round, color=blue_color, linestyle='--', linewidth=2)

        plt.tight_layout()
        Savefig(f'{save_prefix} paper2', 300, 0.05)
    
    return band_energy, amp_at_max_freq_round, freq_at_max_amp_round, freq_at_max_amp_number


    
# Automatically set the number of workers for CPU-based video decoding
cpu_count = multiprocessing.cpu_count()

max_workers = round(max(1, int(cpu_count * 2/3)))


with ThreadPoolExecutor(max_workers=max_workers) as executor:

    futures = [executor.submit(extract_frames, v) for v in video_files]

    for f in tqdm(futures, total=len(futures), desc="Frame Extraction"):
        f.result()

print("\nFrame extraction completed.\n")
print("\nMemory cleanup started.\n")

gc.collect()

try:
    cv2.destroyAllWindows()
except:
    pass

try:
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.ipc_collect()
except:
    pass

try:
    del futures
    del video_files
except:
    pass

try:
    if hasattr(os, "sync"):
        os.sync()
except:
    pass

gc.collect()

print("Memory Cleanup DONE\n")


# 예상 처리 시간을 저장할 리스트
# List for storing per-folder processing times
estimated_execution_times = []


#기준 이미지 초기화
# Initialize the reference image
reference_image = None

# 기존 엑셀 파일 열기
# Open an existing Excel file or create a new one
def load_or_create_workbook(filepath):
    try:
        wb = load_workbook(filename=filepath)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
    return wb, ws

good_wb, good_ws = load_or_create_workbook(strong_sample_excel_path)
weak_wb, weak_ws = load_or_create_workbook(weak_sample_excel_path)
weak_double_wb, weak_double_ws = load_or_create_workbook(weak_double_sample_excel_path)
noise_wb, noise_ws = load_or_create_workbook(noise_sample_excel_path)
check_wb, check_ws = load_or_create_workbook(check_sample_excel_path)
optical_noise_wb, optical_noise_ws = load_or_create_workbook(optical_noise_sample_excel_path)
    
# Add headers to Excel output files
# KR: Excel output 파일에 header를 추가합니다.
headers = (['Folder name', 'Maximum Intensity of Pixels', f'Myotube Area ({AREA_UNIT})',f'Maximum of Displacement ({LENGTH_UNIT})','MAX BAND ENERGY', 'MAX Apeak', 'Coordinate X', 'Coordinate Y', 'Hz', 'Displacement List', 'Max Abs Displacement', 'ABS Net Change Ratio', 'Sign Changes', 'Diff STD' ])

good_ws.append(headers)
weak_ws.append(headers)
noise_ws.append(headers)
check_ws.append(headers)

def safe_int(v, default=0):
    try:
        if v is None:
            return default
        v = float(v)
        if np.isnan(v):
            return default
        return int(round(v))
    except:
        return default

# Collect only folders in the input/output directory
# KR: 입력/출력 기본 폴더 안에서 실제 폴더 항목만 수집합니다.
output_folders = [f for f in natsort.natsorted(os.listdir(VIDEO_FILE_PATH)) if os.path.isdir(os.path.join(VIDEO_FILE_PATH, f))]

# Count the number of detected folders
# KR: 감지된 분석 대상 폴더 수를 계산합니다.
number_of_folders = len(output_folders)

for folder_name in natsort.natsorted(os.listdir(VIDEO_FILE_PATH)):
    folder_path = os.path.join(VIDEO_FILE_PATH, folder_name)
    if not os.path.isdir(folder_path):
        continue # Skip non-folder entries
        # KR: 폴더가 아닌 항목은 건너뜁니다.
    print(f'\n【　Start Analyzing File {folder_name}　】\n ')
    print('--------------------------------List of Generated Files --------------------------------\n')
    folder_start_time = time.time()
    for extension in VIDEO_FILE_EXTENSION:
        fps_video_path = os.path.join(VIDEO_FILE_PATH,f'{folder_name}.{extension}')
        if os.path.exists(fps_video_path):      # Detect FPS from the original video if the corresponding video file exists
                                                # KR: 현재 분석 폴더와 대응되는 원본 video 파일이 있으면 해당 video에서 FPS를 읽어옵니다.
            cap = cv2.VideoCapture(fps_video_path)
            fps = round(cap.get(cv2.CAP_PROP_FPS), 1)
            cap.release()
            break ## Stop searching other extensions once the matching video is found
            # KR: 대응되는 video 파일을 찾으면 다른 확장자 검색을 중단합니다.
        else:
            fps = VIDEO_FPS
    # Set the reference image path
    # KR: 기준 이미지로 사용할 첫 번째 frame 경로를 설정합니다.
    reference_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}0000.{FRAME_EXTENSION}')
    reference_image = cv2.imread(reference_image_path)
    
    # Verify that the reference image was loaded successfully
    # KR: 기준 이미지가 정상적으로 로드되었는지 확인합니다.
    if reference_image is not None:
        reference_gray = cv2.cvtColor(reference_image, cv2.COLOR_BGR2GRAY)
    else:
        continue  # Skip this folder if the reference image cannot be loaded
        # KR: 기준 이미지 로드에 실패하면 해당 폴더 분석을 건너뜁니다.
    
     # Find the minimum and maximum frame numbers from extracted frame filenames
    # KR: 추출된 frame 파일명에서 가장 작은 번호와 가장 큰 번호를 찾습니다.
    # KR: 이를 통해 분석 시작 frame과 마지막 frame 범위를 자동으로 설정합니다.
    max_number = -1
    min_number = float('inf')
    for filename in os.listdir(folder_path):
        if filename.startswith(f'{FRAME_PREFIX}') and filename.endswith(f'.{FRAME_EXTENSION}'):
            try:
                # Extract the numeric frame index from filenames such as frame_0000.png
                # KR: frame_0000.png 같은 파일명에서 숫자 부분만 추출합니다.
                number = int(filename[len(f'{FRAME_PREFIX}'):-len(f'.{FRAME_EXTENSION}')]) 
                max_number = max(max_number, number)
                min_number = min(min_number, number)
            except ValueError:
                pass # Ignore filenames that cannot be converted to frame numbers
                # KR: 숫자로 변환할 수 없는 파일명은 무시합니다.


    # Define the frame range used for analysis
    # KR: 분석에 사용할 frame 시작 번호와 마지막 번호를 설정합니다.
    start_number = min_number
    # Add 1 to include the last frame in the analysis
    # KR: 마지막 frame까지 포함하기 위해 +1 합니다.
    last_number = max_number+1 
    # Automatically obtain image resolution from the reference image
    # KR: 기준 이미지에서 해상도를 자동으로 가져옵니다.
    resolution_y, resolution_x,_ = reference_image.shape  

    
    try:
        total_time_seconds = int(max_number / fps)
    except ZeroDivisionError:
        pass

    # Initialize lists for image paths and loaded frames
    # KR: 이미지 경로, grayscale frame, RGB frame을 저장할 리스트를 초기화합니다.
    image_paths = []
    current_gray_image_list = []
    current_rgb_image_list = []
    # # Initialize lists for maximum pixel-intensity differences and their coordinates
    max_difference_comparing_0 = []  #max_differences 
    max_difference_comparing_0_locations = [] # max_diff_locations

    max_top30_differences_crop = []
    max_top30_diff_locations_crop = []

    max_top30_differences_crop_2 = []
    max_top30_diff_locations_crop_2 = []

    max_top30_differences_crop_3 = []
    max_top30_diff_locations_crop_3 = []

    max_top30_differences_crop_2_new = []
    max_top30_diff_locations_crop_2_new = []

    max_top30_differences_crop_3_new = []
    max_top30_diff_locations_crop_3_new = []

    valid_pixel_all_counts_min = []
    valid_pixel_all_counts_max = []
    

    # =========================================================
    # (A) Load frame paths and grayscale images
    #     Frames are loaded once and then processed in GPU/CPU batch mode.
    # KR: frame 경로와 grayscale 이미지를 수집합니다.
    # KR: frame은 한 번만 로드한 뒤 이후 GPU/CPU batch 방식으로 처리합니다.
    # =========================================================

    use_gpu = (GPU_MODE.lower() == "on") and torch.cuda.is_available()
    device = torch.device("cuda" if use_gpu else "cpu")

    print(f"Compute mode: {'GPU' if use_gpu else 'CPU'}")

    # =========================================================
    # LSTM CLASSIFICATION MODEL LOAD (PyTorch)
    # =========================================================
    classification_model = HybridLSTMClassifier(
        seq_input_dim=2,
        scalar_input_dim=len(COLUMNS),
        num_classes=NB_CLASSES
    ).to(device)

    classification_model.load_state_dict(
        torch.load(BEST_CLASSIFICATION_CHECKPOINT_FILE, map_location=device)
    )

    classification_model.eval()


    image_paths = []
    gray_frames = []

    for idx in range(start_number, last_number):
        p = os.path.join(folder_path, f'{FRAME_PREFIX}{idx:04d}.{FRAME_EXTENSION}')
        if not os.path.exists(p):
            continue

        img_rgb = cv2.imread(p)
        if img_rgb is None:
            print(f"Image load fail → {p}")
            continue

        g = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)

        image_paths.append(p)
        gray_frames.append(g)
        current_rgb_image_list.append(img_rgb)


    if len(gray_frames) == 0:
        print("No valid frames found. Skip folder.")
        continue

    # =========================================================
    # (B) Compute frame-wise absolute differences from the reference image
    #     using GPU/CPU batch processing
    # KR: reference image와 각 frame 사이의 absolute difference를 계산합니다.
    # KR: 이 계산은 GPU 또는 CPU batch 처리로 한 번에 수행합니다.
    # =========================================================
    
    gray_tensor = torch.from_numpy(np.stack(gray_frames)).to(torch.int16)

    if use_gpu:
        gray_tensor = gray_tensor.pin_memory().to(device, non_blocking=True)
    else:
        gray_tensor = gray_tensor.contiguous()

    ref_tensor = torch.from_numpy(reference_gray).to(torch.int16)

    if use_gpu:
        ref_tensor = ref_tensor.pin_memory().to(device, non_blocking=True)
    else:
        ref_tensor = ref_tensor.contiguous()


    diff_tensor = torch.abs(gray_tensor - ref_tensor)                                 # [T,H,W]
    flat        = diff_tensor.flatten(1)                                              # [T,H*W]

    max_vals, argmax_vals = flat.max(dim=1)                                           # [T], [T]

    H, W = reference_gray.shape
    ys = (argmax_vals // W).cpu().numpy()
    xs = (argmax_vals %  W).cpu().numpy()

    max_difference_comparing_0 = max_vals.cpu().numpy().tolist()
    max_difference_comparing_0_locations = list(zip(xs, ys))

    current_gray_image_list = gray_frames

    image0_path = reference_image_path
    
    image0 = current_gray_image_list[0]

    
    # Search for the maximum intensity change within the initial fps × NUM_FRAMES window.
    # This helps focus the analysis on early contraction-related motion rather than large noise or bubble movement.
    # KR: 초기 fps × NUM_FRAMES 구간 안에서 최대 intensity change를 찾습니다.
    # KR: 큰 noise 또는 bubble 움직임보다 초기 수축 관련 움직임에 분석을 집중하기 위한 과정입니다.
    top30_max_differences_indices = np.arange(len(max_difference_comparing_0))[:int(fps*NUM_FRAMES)]
    top30_max_differences_values = [max_difference_comparing_0[i] for i in top30_max_differences_indices]
    top30_max_differences_values_max = max(top30_max_differences_values)
    
    max_value_index_within_top30 = np.argmax(top30_max_differences_values)
    
    max_difference_index_within_top30 = top30_max_differences_indices[max_value_index_within_top30]
    
    max_difference_value_within_top30 = top30_max_differences_values[max_value_index_within_top30]

    max_diff_location_within_top30 = max_difference_comparing_0_locations[max_difference_index_within_top30]

    x_center_max_diff_location_within_top30 = max_diff_location_within_top30[0]
    y_center_max_diff_location_within_top30 = max_diff_location_within_top30[1]

    
    x11 = x_center_max_diff_location_within_top30 - ANALYSIS_SQUARE_SIZE_REFERENCE
    y11 = y_center_max_diff_location_within_top30 - ANALYSIS_SQUARE_SIZE_REFERENCE
    x22 = x_center_max_diff_location_within_top30 + ANALYSIS_SQUARE_SIZE_REFERENCE
    y22 = y_center_max_diff_location_within_top30 + ANALYSIS_SQUARE_SIZE_REFERENCE
    x11 = max(x11, 0)
    y11 = max(y11, 0)
    x22 = min(x22, resolution_x - 1)
    y22 = min(y22, resolution_y - 1)


    roi_tensor = gray_tensor[:, y11:y22, x11:x22].contiguous()


    image_frame_number = []
    next_image_index = 0   

    visited = set()

    for step in range(min(len(gray_frames), 500)):
        if next_image_index in visited:
            break
        visited.add(next_image_index)

        image_frame_number.append(next_image_index)

        ref_roi = roi_tensor[next_image_index].unsqueeze(0)  # [1,h,w]
        diff_batch = torch.abs(roi_tensor - ref_roi)         # [T,h,w]
        flat = diff_batch.flatten(1)                         # [T,h*w]

        max_vals, _ = flat.max(dim=1)                        # [T]
        best_i = int(max_vals.argmax().item())

        next_image_index = best_i

        if len(image_frame_number) >= 8 and image_frame_number[-1] == image_frame_number[-3]:
            break

    image_frame_n1 = image_frame_number[-1]
    image_frame_n2 = image_frame_number[-2] if len(image_frame_number) >= 2 else image_frame_number[-1]

    n1_max_image_path = image_paths[int(image_frame_n1)]
    n2_max_image_path = image_paths[int(image_frame_n2)]


    roi_full = gray_tensor[:, y11:y22, x11:x22].contiguous()          # [T,h,w]

    ref_n1 = roi_full[int(image_frame_n1)].unsqueeze(0)               # [1,h,w]
    diff_n1 = torch.abs(roi_full - ref_n1)                            # [T,h,w]
    flat_n1 = diff_n1.flatten(1)                                      # [T,h*w]
    max_vals_n1, argmax_n1 = flat_n1.max(dim=1)                       # [T], [T]

    h_roi = roi_full.shape[1]
    w_roi = roi_full.shape[2]
    ys_n1 = (argmax_n1 // w_roi)
    xs_n1 = (argmax_n1 %  w_roi)

    max_top30_differences_crop_2 = max_vals_n1.detach().cpu().numpy().astype(np.float32).tolist()
    max_top30_diff_locations_crop_2 = list(zip(xs_n1.detach().cpu().numpy().tolist(),
                                            ys_n1.detach().cpu().numpy().tolist()))

    max_top30_differences_crop_except0_2 = [v for v in max_top30_differences_crop_2 if v != 0]
    mean_max_top30_differences_crop2 = float(np.mean(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0.0
    mean_round_max_top30_differences_crop2 = round(mean_max_top30_differences_crop2, 2)
    max_max_top30_differences_crop2 = float(np.max(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0.0

    if len(max_top30_differences_crop_except0_2) > 0:
        min_max_top30_differences_crop_except0_2 = float(np.min(max_top30_differences_crop_except0_2))
    else:
        min_max_top30_differences_crop_except0_2 = 0.0

    min_max_top30_differences_crop_except0_2_index = int(np.argmax(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0

    ref_n2 = roi_full[int(image_frame_n2)].unsqueeze(0)               # [1,h,w]
    diff_n2 = torch.abs(roi_full - ref_n2)                            # [T,h,w]
    flat_n2 = diff_n2.flatten(1)                                      # [T,h*w]
    max_vals_n2, argmax_n2 = flat_n2.max(dim=1)                       # [T], [T]

    ys_n2 = (argmax_n2 // w_roi)
    xs_n2 = (argmax_n2 %  w_roi)

    max_top30_differences_crop_3 = max_vals_n2.detach().cpu().numpy().astype(np.float32).tolist()
    max_top30_diff_locations_crop_3 = list(zip(xs_n2.detach().cpu().numpy().tolist(),
                                            ys_n2.detach().cpu().numpy().tolist()))

    max_top30_differences_crop_except0_3 = [v for v in max_top30_differences_crop_3 if v != 0]
    mean_max_top30_differences_crop3 = float(np.mean(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0.0
    mean_round_max_top30_differences_crop3 = round(mean_max_top30_differences_crop3, 2)
    max_max_top30_differences_crop3 = float(np.max(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0.0

    if len(max_top30_differences_crop_except0_3) > 0:
        min_max_top30_differences_crop_except0_3 = float(np.min(max_top30_differences_crop_except0_3))
    else:
        min_max_top30_differences_crop_except0_3 = 0.0

    min_max_top30_differences_crop_except0_3_index = int(np.argmax(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0

    try:
        top30_max_image_path_2 = image_paths[min_max_top30_differences_crop_except0_2_index]
    except:
        top30_max_image_path_2 = ""

    try:
        top30_max_image_path_3 = image_paths[min_max_top30_differences_crop_except0_3_index]
    except:
        top30_max_image_path_3 = ""

    diff_max_all = diff_n2                                           

    diff_min_all = diff_n1                                            

    threshold_max = min_max_top30_differences_crop_except0_3
    threshold_min = min_max_top30_differences_crop_except0_2

    min_threshold_max = threshold_max + 20
    max_threshold_max = max_max_top30_differences_crop3

    min_threshold_min = threshold_min + 20
    max_threshold_min = max_max_top30_differences_crop2

    if max_threshold_max <= 0:
        valid_pixel_counts_max = [0] * len(gray_frames)
    else:
        mask_max = (diff_max_all >= min_threshold_max) & (diff_max_all <= max_threshold_max)
        valid_pixel_counts_max = mask_max.flatten(1).sum(dim=1).detach().cpu().numpy().astype(np.int64).tolist()

    if max_threshold_min <= 0:
        valid_pixel_counts_min = [0] * len(gray_frames)
    else:
        mask_min = (diff_min_all >= min_threshold_min) & (diff_min_all <= max_threshold_min)
        valid_pixel_counts_min = mask_min.flatten(1).sum(dim=1).detach().cpu().numpy().astype(np.int64).tolist()

    valid_pixel_count_mean_min = np.mean([v for v in valid_pixel_counts_min if v not in [0, '']]) if len(valid_pixel_counts_min) else 0
    valid_pixel_count_mean_round_min = round(valid_pixel_count_mean_min) if not np.isnan(valid_pixel_count_mean_min) else 0

    valid_pixel_count_mean_max = np.mean([v for v in valid_pixel_counts_max if v not in [0, '']]) if len(valid_pixel_counts_max) else 0
    valid_pixel_count_mean_round_max = round(valid_pixel_count_mean_max) if not np.isnan(valid_pixel_count_mean_max) else 0

    count_zeros_min = valid_pixel_counts_min.count(0)
    count_zeros_max = valid_pixel_counts_max.count(0)

    sum_count_all_pixel_in_range_min = int(np.sum(np.array(valid_pixel_counts_min) < valid_pixel_count_mean_round_min)) if valid_pixel_count_mean_round_min > 0 else 0
    sum_count_all_pixel_in_range_max = int(np.sum(np.array(valid_pixel_counts_max) < valid_pixel_count_mean_round_max)) if valid_pixel_count_mean_round_max > 0 else 0

    valid_pixel_all_counts_min.append(sum_count_all_pixel_in_range_min)
    valid_pixel_all_counts_max.append(sum_count_all_pixel_in_range_max)

    valid_pixel_all_counts_min_max = max(valid_pixel_all_counts_min) if len(valid_pixel_all_counts_min) else 0
    valid_pixel_all_counts_max_max = max(valid_pixel_all_counts_max) if len(valid_pixel_all_counts_max) else 0

    All_pixel_intensities = []
    
    selected_max_differences = []
    selected_max_differences_location = []
    
    max_image_path_pixel = top30_max_image_path_3   
    min_image_path_pixel = top30_max_image_path_2  

    for i in range(len(gray_frames)):

        if use_gpu:
            image1 = gray_tensor[i].cpu().numpy().astype(np.uint8)
            max_image_gray_pixel = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
            min_image_gray_pixel = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            first_image_gray_pixel = gray_tensor[int(max_difference_index_within_top30)].cpu().numpy().astype(np.uint8)

        else:
            image1 = gray_frames[i].astype(np.uint8)
            max_image_gray_pixel = gray_frames[int(image_frame_n2)].astype(np.uint8)
            min_image_gray_pixel = gray_frames[int(image_frame_n1)].astype(np.uint8)
            first_image_gray_pixel = gray_frames[int(max_difference_index_within_top30)].astype(np.uint8)

        cropped_max_image_pixel = max_image_gray_pixel[y11:y22, x11:x22]
        cropped_min_image_pixel = min_image_gray_pixel[y11:y22, x11:x22]
        cropped_image0_pixel = image0[y11:y22, x11:x22]
        cropped_image1_pixel = image1[y11:y22, x11:x22]

        h = min(cropped_max_image_pixel.shape[0], cropped_image1_pixel.shape[0])
        w = min(cropped_max_image_pixel.shape[1], cropped_image1_pixel.shape[1])

        cropped_max_image_pixel = cropped_max_image_pixel[:h, :w]
        cropped_min_image_pixel = cropped_min_image_pixel[:h, :w]
        cropped_image0_pixel = cropped_image0_pixel[:h, :w]
        cropped_image1_pixel = cropped_image1_pixel[:h, :w]

        cropped_max_image_pixel = cropped_max_image_pixel.astype(np.uint8)
        cropped_min_image_pixel = cropped_min_image_pixel.astype(np.uint8)
        cropped_image0_pixel = cropped_image0_pixel.astype(np.uint8)
        cropped_image1_pixel = cropped_image1_pixel.astype(np.uint8)

        if cropped_image0_pixel is None or cropped_image1_pixel is None:
            print(f"Error cropping image {i}. Skipping...")
            continue

        if valid_pixel_all_counts_min_max >= valid_pixel_all_counts_max_max:
            not_crop_diff = cv2.subtract(image1, min_image_gray_pixel)
            square_image_path = min_image_path_pixel
            square_image_index = image_frame_n1

        else:
            diff = cv2.subtract(cropped_max_image_pixel, cropped_image1_pixel)
            not_crop_diff = cv2.subtract(image1, max_image_gray_pixel)
            square_image_path = max_image_path_pixel
            square_image_index = image_frame_n2

        _, selected_max_diff, _, selected_max_diff_loc = cv2.minMaxLoc(not_crop_diff)

        selected_max_differences.append(selected_max_diff)
        selected_max_differences_location.append(selected_max_diff_loc)

        
        selected_max_differences_indices = np.arange(len(selected_max_differences))[:int(fps*NUM_FRAMES)] 
        selected_max_differences_values = [selected_max_differences[i] for i in selected_max_differences_indices] 

        selected_max_differences_values_max_in_Frames = np.argmax(selected_max_differences_values) 
        selected_max_differences_index_in_Frames = selected_max_differences_indices[selected_max_differences_values_max_in_Frames] 
        selected_max_differences_value_in_Frames = selected_max_differences_values[selected_max_differences_index_in_Frames] 
        selected_max_differences_location_in_Frames = selected_max_differences_location[selected_max_differences_index_in_Frames]

    x_selected = selected_max_differences_location_in_Frames[0]  
    y_selected = selected_max_differences_location_in_Frames[1] 
    
    x1_small = x_selected - ANALYSIS_SQUARE_SIZE_REFERENCE  
    y1_small = y_selected - ANALYSIS_SQUARE_SIZE_REFERENCE
    x2_small = x_selected + ANALYSIS_SQUARE_SIZE_REFERENCE
    y2_small = y_selected + ANALYSIS_SQUARE_SIZE_REFERENCE
    
    x1_small = max(x1_small, 0)
    y1_small = max(y1_small, 0)
    x2_small = min(x2_small, resolution_x -1)
    y2_small = min(y2_small, resolution_y -1)
    
    # Compare all frames again using the cropped ROI-derived maximum frame as a reference.
    # This step is used to identify candidate maximal contraction or relaxation frames.
    # KR: cropped ROI에서 얻은 maximum frame을 기준으로 전체 frame을 다시 비교합니다.
    # KR: 이 과정은 최대 수축 또는 최대 이완 후보 frame을 찾기 위한 단계입니다.

    image_frame_number = []

    next_image_index = start_number
    for n in range(start_number, last_number):
        max_diff_cropped_n_frame_list = []
        max_diff_cropped_loc_n_frame_list = []
    
        for i in range(len(gray_frames)):

            n_frame_path = os.path.join(folder_path, f'{FRAME_PREFIX}{next_image_index + start_number :04d}.{FRAME_EXTENSION}') 

            # Current reference frame path used in this iterative comparison
            # KR: 반복 비교에 사용할 현재 reference frame 경로입니다.
            if use_gpu:
                top30_max_image_gray_new = gray_tensor[next_image_index].cpu().numpy().astype(np.uint8)
            else:
                top30_max_image_gray_new = gray_frames[next_image_index].astype(np.uint8)
            cropped_n_frame_gray = top30_max_image_gray_new[y1_small:y2_small, x1_small:x2_small] 
            current_image_crop = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small] 

            max_diff_loc_n_frame_cropped = cv2.absdiff(cropped_n_frame_gray, current_image_crop)
            _, max_diff_crop_n_frame,_,max_diff_crop_loc_n_frame = cv2.minMaxLoc(max_diff_loc_n_frame_cropped)

            max_diff_cropped_n_frame_list.append(max_diff_crop_n_frame)
            max_diff_cropped_loc_n_frame_list.append(max_diff_crop_loc_n_frame)

            max_diff_cropped_n_indices = np.argsort(max_diff_cropped_n_frame_list)[-1:]            
            max_diff_cropped_n_indices_index = max_diff_cropped_n_indices[0]
            max_diff_cropped_n_indices_path = os.path.join(folder_path, f'{FRAME_PREFIX}{max_diff_cropped_n_indices_index + start_number :04d}.{FRAME_EXTENSION}')
        image_frame_number.append(next_image_index)
        
        # Calculate summary statistics from cropped ROI differences.
        # Zero values are excluded when estimating the nonzero minimum.
        # KR: cropped ROI difference 값에서 요약 통계를 계산합니다.
        # KR: 0이 아닌 최소값을 계산할 때는 zero value를 제외합니다.
        max_diff_cropped_n_except0 = [value for value in max_diff_cropped_n_frame_list if value != 0]
        mean_max_diff_cropped_n_except0 = np.mean(max_diff_cropped_n_frame_list)
        mean_round_max_diff_cropped_n_frame_listew = round(mean_max_diff_cropped_n_except0, 2)
        max_max_diff_cropped_n_frame_list = max(max_diff_cropped_n_frame_list)
        min_max_diff_cropped_n_except0 = min(max_diff_cropped_n_except0)
        min_max_diff_cropped_n_except0_index = max_diff_cropped_n_frame_list.index(max_max_diff_cropped_n_frame_list)
        # Update the reference frame index for the next iteration
        # KR: 다음 반복 비교에 사용할 reference frame index를 업데이트합니다.
        next_image_index = min_max_diff_cropped_n_except0_index
        
        # Stop the iterative search when the selected frame index starts to repeat.
        # KR: 선택된 frame index가 반복되면 탐색이 수렴한 것으로 보고 반복을 중단합니다.
        if len(image_frame_number)>=8 and image_frame_number[n] == image_frame_number[n - 2]:
                image_frame_n1 = image_frame_number[n]
                image_frame_n2 = image_frame_number[n-1]
                n1_max_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}{image_frame_n1 + start_number :04d}.{FRAME_EXTENSION}') 
                n2_max_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}{image_frame_n2 + start_number :04d}.{FRAME_EXTENSION}') 
                break

    # Compare all frames again using the two ROI-derived candidate frames.
    # This step refines the candidate maximal contraction and relaxation frames.
    # KR: ROI에서 얻은 두 후보 frame을 기준으로 전체 frame을 다시 비교합니다.
    # KR: 최대 수축/최대 이완 후보 frame을 더 안정적으로 찾기 위한 단계입니다.
    
    for i in range(len(gray_frames)):
        
        top30_max_image_path_2_new = n1_max_image_path
        # Use the n1 candidate frame as the reference for ROI-based comparison.
        # KR: n1 후보 frame을 기준으로 ROI 기반 frame 비교를 수행합니다.
        
        if use_gpu:
            top30_max_image_gray_2_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
        else:
            top30_max_image_gray_2_new = gray_frames[image_frame_n1].astype(np.uint8)

        
        cropped_top30_max_image_gray_2_new = top30_max_image_gray_2_new[y1_small:y2_small, x1_small:x2_small] 
        current_image_crop_2_new = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small]
        diff_top30_max_2_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_2_new, current_image_crop_2_new)
        _, max_diff_crop_top30_max_2_new, _,max_diff_loc_crop_top30_max_2_new = cv2.minMaxLoc(diff_top30_max_2_cropped_new)
        max_top30_differences_crop_2_new.append(max_diff_crop_top30_max_2_new)
        max_top30_diff_locations_crop_2_new.append(max_diff_loc_crop_top30_max_2_new)
        
        max_top30_differences_crop_indices_2_new = np.argsort(max_top30_differences_crop_2_new)[-1:]
        
        # Search within the initial fps × NUM_FRAMES window to reduce the influence of large noise or bubble movement.
        # KR: 큰 noise 또는 bubble 움직임의 영향을 줄이기 위해 초기 fps × NUM_FRAMES 구간 안에서 탐색합니다.
        max_top30_differences_crop_indices_2_new_num_frames = np.arange(len(max_top30_differences_crop_2_new))[:int(fps*NUM_FRAMES)] 
        
        max_top30_differences_crop_values_2_new = [max_top30_differences_crop_2_new[i] for i in max_top30_differences_crop_indices_2_new_num_frames]

        
        max_top30_differences_crop_values_2_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_2_new) 
        
        index_in_Frames_2 = max_top30_differences_crop_indices_2_new_num_frames[max_top30_differences_crop_values_2_new_max_in_Frames]
        value_in_Frames_2 =  max_top30_differences_crop_values_2_new[index_in_Frames_2]
        location_in_Frames_2 = max_top30_diff_locations_crop_2_new[index_in_Frames_2]

        
        max_top30_differences_crop_indices_2_index_new = max_top30_differences_crop_indices_2_new[0]
        max_top30_differences_crop_indices_2_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_2_index_new + start_number :04d}.{FRAME_EXTENSION}')


    # Calculate summary statistics for ROI differences relative to the n1 candidate frame.
    # KR: n1 후보 frame 기준 ROI difference의 요약 통계를 계산합니다.
    max_top30_differences_crop_except0_2_new = [value for value in max_top30_differences_crop_2_new if value != 0]
    mean_max_top30_differences_crop2_new = np.mean(max_top30_differences_crop_2_new)
    mean_round_max_top30_differences_crop2_new = round(mean_max_top30_differences_crop2_new, 2)
    max_max_top30_differences_crop2_new = max(max_top30_differences_crop_2_new)
    min_max_top30_differences_crop_except0_2_new = min(max_top30_differences_crop_except0_2_new)
    min_max_top30_differences_crop_except0_2_index_new = max_top30_differences_crop_2_new.index(max_max_top30_differences_crop2_new)

    
    for i in range(len(gray_frames)):

        top30_max_image_path_3_new = n2_max_image_path

        # Use the n2 candidate frame as the reference for ROI-based comparison.
        # KR: n2 후보 frame을 기준으로 ROI 기반 frame 비교를 수행합니다.
        if use_gpu:
            top30_max_image_gray_3_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
        else:
            top30_max_image_gray_3_new = gray_frames[image_frame_n2].astype(np.uint8)

        
        cropped_top30_max_image_gray_3_new = top30_max_image_gray_3_new[y1_small:y2_small, x1_small:x2_small] 
        
        current_image_crop_3 = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small] 

        diff_top30_max_3_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_3_new, current_image_crop_3)
        _, max_diff_crop_top30_max_3_new, _,max_diff_loc_crop_top30_max_3_new = cv2.minMaxLoc(diff_top30_max_3_cropped_new)
        
        max_top30_differences_crop_3_new.append(max_diff_crop_top30_max_3_new)
        max_top30_diff_locations_crop_3_new.append(max_diff_loc_crop_top30_max_3_new)
        
        max_top30_differences_crop_indices_3_new = np.argsort(max_top30_differences_crop_3_new)[-1:]
        
        # Search within the initial fps × NUM_FRAMES window to reduce the influence of large noise or bubble movement.
        # KR: 큰 noise 또는 bubble 움직임의 영향을 줄이기 위해 초기 fps × NUM_FRAMES 구간 안에서 탐색합니다.
        max_top30_differences_crop_indices_3_new_num_frames = np.arange(len(max_top30_differences_crop_3_new))[:int(fps*NUM_FRAMES)] 
        
        max_top30_differences_crop_values_3_new = [max_top30_differences_crop_3_new[i] for i in max_top30_differences_crop_indices_3_new_num_frames] 

        
        max_top30_differences_crop_values_3_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_3_new) 
        index_in_Frames_3 = max_top30_differences_crop_indices_3_new_num_frames[max_top30_differences_crop_values_3_new_max_in_Frames]
        value_in_Frames_3 =  max_top30_differences_crop_values_3_new[index_in_Frames_3]
        location_in_Frames_3 = max_top30_diff_locations_crop_3_new[index_in_Frames_3]

        
        max_top30_differences_crop_indices_3_index_new = max_top30_differences_crop_indices_3_new[0]
        max_top30_differences_crop_indices_3_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_3_index_new + start_number :04d}.{FRAME_EXTENSION}')
        

    max_top30_differences_crop_except0_3_new = [value for value in max_top30_differences_crop_3_new if value != 0]
    mean_max_top30_differences_crop3_new = np.mean(max_top30_differences_crop_3_new)
    mean_round_max_top30_differences_crop3_new = round(mean_max_top30_differences_crop3_new, 2)
    max_max_top30_differences_crop3_new = max(max_top30_differences_crop_3_new)
    min_max_top30_differences_crop_except0_3_new = min(max_top30_differences_crop_except0_3_new)
    min_max_top30_differences_crop_except0_3_index_new = max_top30_differences_crop_3_new.index(max_max_top30_differences_crop3_new)
    
    top30_max_image_path_4_new = os.path.join(folder_path, f'{FRAME_PREFIX}{min_max_top30_differences_crop_except0_3_index_new + start_number :04d}.{FRAME_EXTENSION}') 
    



    x_differences = [] 
    y_differences = []
    valid_pixel_counts_min_new = []
    valid_pixel_counts_max_new = []
    valid_pixel_all_counts_max_new = []
    valid_pixel_all_counts_min_new = []

    # Calculate summary statistics for ROI differences relative to the n2 candidate frame.
    # KR: n2 후보 frame 기준 ROI difference의 요약 통계를 계산합니다.
    for i in range(len(gray_frames)):
        image1 = current_gray_image_list[i]
        # Crop local ROIs around the selected motion center.
        # KR: 선택된 motion center 주변의 local ROI를 crop합니다.
        max_image_path_pixel_new = top30_max_image_path_3_new  # Candidate maximal contraction or relaxation frame
        min_image_path_pixel_new = top30_max_image_path_2_new  # Candidate maximal relaxation or contraction frame
        
        if use_gpu:
            max_image_gray_pixel_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
            min_image_gray_pixel_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
        else:
            max_image_gray_pixel_new = gray_frames[image_frame_n2].astype(np.uint8)
            min_image_gray_pixel_new = gray_frames[image_frame_n1].astype(np.uint8)

                
        cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small] 
        cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
        cropped_image0_pixel_new = image0[y1_small:y2_small, x1_small:x2_small] 
        cropped_image1_pixel_new = image1[y1_small:y2_small, x1_small:x2_small] 
        
        if cropped_image0_pixel is None or cropped_image1_pixel is None:
            print(f'Error cropping image{i}. Skipping...')
            continue  # Skip this frame if ROI cropping fails
        
        diff_max_new = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
        diff_min_new = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 

        threshold_max_new = min_max_top30_differences_crop_except0_3_new  
        threshold_min_new = min_max_top30_differences_crop_except0_2_new 
        white_pixel_coords_max_new = np.argwhere(diff_max_new >= threshold_max_new)
        white_pixel_coords_min_new = np.argwhere(diff_min_new >= threshold_min_new)  
        
        # Count pixels whose intensity differences fall within the threshold range.
        # KR: intensity difference가 설정된 threshold 범위 안에 들어오는 pixel을 계산합니다.
        
        # Add a fixed offset to reduce small background or medium-related noise.
        # KR: 작은 background 또는 배지 관련 noise를 줄이기 위해 고정 offset을 더합니다.
        min_threshold_max_new = min_max_top30_differences_crop_except0_3_new + 20  # Fixed offset for small background or medium-related noise
        max_threshold_max_new = max_max_top30_differences_crop3_new  # Maximum ROI intensity difference relative to the candidate frame
        min_threshold_min_new = min_max_top30_differences_crop_except0_2_new  +20 # Fixed offset for small background or medium-related noise
        max_threshold_min_new = max_max_top30_differences_crop2_new # Maximum ROI intensity difference relative to the candidate frame
        
        valid_pixel_coords_max_new = np.argwhere((diff_max_new >= min_threshold_max_new) & (diff_max_new <= max_threshold_max_new))
        valid_pixel_coords_min_new = np.argwhere((diff_min_new >= min_threshold_min_new) & (diff_min_new <= max_threshold_min_new))
        
        valid_pixel_coords_max_all_new = np.argwhere((diff_max_new >= 0) & (diff_max_new <= max_threshold_max))
        valid_pixel_coords_min_all_new = np.argwhere((diff_min_new >= 0) & (diff_min_new <= max_threshold_min))

        valid_pixel_count_max_new = len(valid_pixel_coords_max_new)
        valid_pixel_count_min_new = len(valid_pixel_coords_min_new)
        valid_pixel_counts_max_new.append(valid_pixel_count_max_new)
        valid_pixel_counts_min_new.append(valid_pixel_count_min_new)       

    valid_pixel_count_mean_min_new = np.mean([value for value in valid_pixel_counts_min_new if value not in [0,'']]) #0 과 데이터가 

    valid_pixel_count_mean_round_min_new = round(valid_pixel_count_mean_min_new) if not np.isnan(valid_pixel_count_mean_min_new) else 0 
    valid_pixel_count_mean_max_new = np.mean([value for value in valid_pixel_counts_max_new if value not in [0,'']]) #0 과 데이터가 

    valid_pixel_count_mean_round_max_new = round(valid_pixel_count_mean_max_new) if not np.isnan(valid_pixel_count_mean_max_new) else 0 
    
    count_zeros_min_new = valid_pixel_counts_min_new.count(0)
    count_zeros_max_new = valid_pixel_counts_max_new.count(0)

    count_all_pixel_in_range_min_new = [valid_pixel_counts_min_new.count(i) for i in range(valid_pixel_count_mean_round_min_new)]
    count_all_pixel_in_range_max_new = [valid_pixel_counts_max_new.count(i) for i in range(valid_pixel_count_mean_round_max_new)]
    
    sum_count_all_pixel_in_range_min_new = sum(count_all_pixel_in_range_min_new, 0)
    sum_count_all_pixel_in_range_max_new = sum(count_all_pixel_in_range_max_new, 0)

    valid_pixel_all_counts_min_new.append(sum_count_all_pixel_in_range_min_new)
    valid_pixel_all_counts_max_new.append(sum_count_all_pixel_in_range_max_new)
    
    valid_pixel_all_counts_min_max_new = max(valid_pixel_all_counts_min_new)
    valid_pixel_all_counts_max_max_new = max(valid_pixel_all_counts_max_new)
    
    
    All_pixel_intensities = []  
    valid_pixel_counts = []
    valid_pixel_counts2 = []
    selected_max_differences = []
    selected_max_differences_location = []
    
    x_selected_new, y_selected_new = None, None
    click_done = False  
    



    for i in range(len(gray_frames)):

        max_image_path_pixel_new = top30_max_image_path_3_new
        min_image_path_pixel_new = top30_max_image_path_2_new
        
        if use_gpu:
            max_image_gray_pixel_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
            min_image_gray_pixel_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
        else:
            max_image_gray_pixel_new = gray_frames[image_frame_n2].astype(np.uint8)
            min_image_gray_pixel_new = gray_frames[image_frame_n1].astype(np.uint8)

        
        cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
        cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
        cropped_image0_pixel_new = image0[y1_small:y2_small, x1_small:x2_small]  
        cropped_image1_pixel_new = image1[y1_small:y2_small, x1_small:x2_small]  
        
        if cropped_image0_pixel is None or cropped_image1_pixel is None:
            print(f"Error cropping image {i}. Skipping...")
            continue 
        
        # Select the candidate frame based on the number of valid pixels below the mean threshold.
        # This helps determine which candidate better represents maximal contraction or relaxation.
        # KR: mean threshold보다 작은 valid pixel count를 기준으로 후보 frame을 선택합니다.
        # KR: 최대 수축 또는 최대 이완을 더 잘 나타내는 후보를 고르기 위한 과정입니다.
        
        if valid_pixel_all_counts_min_max_new >= valid_pixel_all_counts_max_max_new:
            loc2 = max_top30_diff_locations_crop_2_new  
            loc3 = location_in_Frames_2  
            
        else:
            loc2 = max_top30_diff_locations_crop_3_new 
            loc3 = location_in_Frames_3 
            
        # =========================================================
        # Manual coordinate selection callback
        # =========================================================
        
        # Mouse-click callback for selecting an analysis coordinate
        # KR: 분석 좌표를 수동으로 선택하기 위한 마우스 클릭 callback 함수입니다.
        def select_point(event, x, y, flags, param):
            global x_selected_new, y_selected_new, click_done  

            if event == cv2.EVENT_LBUTTONDOWN and not click_done: # Left-click once to select a coordinate
                x_selected_new = x
                y_selected_new = y
                click_done = True  # 
                print(f"Selected coordinate: ({x_selected_new}, {y_selected_new})")

                cv2.circle( min_image_gray_pixel_new, (x_selected_new, y_selected_new), 2, (0, 0, 255), -1)
                cv2.imshow("Select Point",  min_image_gray_pixel_new)

                cv2.waitKey(500) 
                cv2.destroyAllWindows()

        # Use predefined coordinates when manual coordinate mode is enabled
        # KR: manual coordinate mode가 켜져 있으면 사용자가 미리 지정한 좌표를 사용합니다.
        if MANUAL_SELECTION.lower() == 'on' and not click_done:
            print("Please click the desired analysis point.")
            

            cv2.namedWindow("Select Point", cv2.WINDOW_NORMAL)
            cv2.resizeWindow("Select Point", resolution_x, resolution_y)  

            cv2.imshow("Select Point", min_image_gray_pixel_new)
            cv2.setMouseCallback("Select Point", select_point)


            while not click_done:
                cv2.waitKey(1)  # 

            print(f"Using manually selected coordinate: ({x_selected_new}, {y_selected_new})")
        
        if MANUAL_COORDINATE_SELECTION.lower() =='on':
            x_selected_new = X_COORDINATES
            y_selected_new = Y_COORDINATES
            

        # Use automatically detected coordinates when manual selection is off,
        # or keep the previously selected coordinate after the first manual click.
        # KR: manual selection이 꺼져 있으면 자동 탐색 좌표를 사용합니다.
        # KR: 첫 번째 수동 클릭 이후에는 선택된 좌표를 유지합니다.
        
        elif MANUAL_SELECTION.lower() == 'off' or (x_selected_new is None and y_selected_new is None): 
            x_selected_new = x1_small + loc3[0]
            y_selected_new = y1_small + loc3[1]
    
        
        x1_small_new = x_selected_new -  resolution_x
        y1_small_new = y_selected_new -  resolution_y
        x2_small_new = x_selected_new +  resolution_x
        y2_small_new = y_selected_new + resolution_y
        
        x1_small_new = max(x1_small_new, 0)
        y1_small_new = max(y1_small_new, 0)
        x2_small_new = min(x2_small_new, resolution_x -1)
        y2_small_new = min(y2_small_new, resolution_y -1)
        
        x1_small_new_reference = max(0, x_selected_new - ANALYSIS_SQUARE_SIZE_REFERENCE)
        y1_small_new_reference = max(0, y_selected_new -  ANALYSIS_SQUARE_SIZE_REFERENCE)
        x2_small_new_reference = x_selected_new + ANALYSIS_SQUARE_SIZE_REFERENCE
        y2_small_new_reference = y_selected_new +  ANALYSIS_SQUARE_SIZE_REFERENCE
        
        x1_small_new_reference = max(x1_small_new_reference, 0)
        y1_small_new_reference = max(y1_small_new_reference, 0)
        x2_small_new_reference = min(x2_small_new_reference, resolution_x -1)
        y2_small_new_reference = min(y2_small_new_reference, resolution_y -1)
        
        x1_small_new_optical = max(0, x_selected_new - resolution_x)
        y1_small_new_optical = max(0, y_selected_new - resolution_y)
        x2_small_new_optical = x_selected_new + resolution_x
        y2_small_new_optical = y_selected_new + resolution_y
        
        x1_small_new_optical = max(x1_small_new_optical, 0)
        y1_small_new_optical = max(y1_small_new_optical, 0)
        x2_small_new_optical = min(x2_small_new_optical, resolution_x -1)
        y2_small_new_optical = min(y2_small_new_optical, resolution_y -1)

    max_top30_differences_crop_2_new = []
    max_top30_diff_locations_crop_2_new = []

    max_top30_differences_crop_3_new = []
    max_top30_diff_locations_crop_3_new = []

    if len(current_rgb_image_list) == 0:
        print("❌ current_rgb_image_list empty → skip")
        continue

    image_frame_n1 = max(0, min(image_frame_n1, len(current_rgb_image_list)-1))
    image_frame_n2 = max(0, min(image_frame_n2, len(current_rgb_image_list)-1))

    top30_max_image_2_new = current_rgb_image_list[image_frame_n1]
    top30_max_image_3_new = current_rgb_image_list[image_frame_n2]


    for i in range(len(gray_frames)):
        
        top30_max_image_path_2_new = n1_max_image_path
        top30_max_image_gray_2_new = cv2.cvtColor(top30_max_image_2_new, cv2.COLOR_BGR2GRAY) 
        
        cropped_top30_max_image_gray_2_new = top30_max_image_gray_2_new[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
        
        current_image_crop_2_new = current_gray_image_list[i][y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 

        diff_top30_max_2_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_2_new, current_image_crop_2_new)
        _, max_diff_crop_top30_max_2_new, _,max_diff_loc_crop_top30_max_2_new = cv2.minMaxLoc(diff_top30_max_2_cropped_new)
        
        max_top30_differences_crop_2_new.append(max_diff_crop_top30_max_2_new)
        max_top30_diff_locations_crop_2_new.append(max_diff_loc_crop_top30_max_2_new)
        
        max_top30_differences_crop_indices_2_new = np.argsort(max_top30_differences_crop_2_new)[-1:]
        
        max_top30_differences_crop_indices_2_new_num_frames = np.arange(len(max_top30_differences_crop_2_new))[:int(fps*NUM_FRAMES)] 
        
        max_top30_differences_crop_values_2_new = [max_top30_differences_crop_2_new[i] for i in max_top30_differences_crop_indices_2_new_num_frames] 
        
        max_top30_differences_crop_values_2_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_2_new)
        
        index_in_Frames_2 = max_top30_differences_crop_indices_2_new_num_frames[max_top30_differences_crop_values_2_new_max_in_Frames]
        
        value_in_Frames_2 =  max_top30_differences_crop_values_2_new[index_in_Frames_2]
        location_in_Frames_2 = max_top30_diff_locations_crop_2_new[index_in_Frames_2]
        
        max_top30_differences_crop_indices_2_index_new = max_top30_differences_crop_indices_2_new[0]
        max_top30_differences_crop_indices_2_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_2_index_new + start_number :04d}.{FRAME_EXTENSION}')



    max_top30_differences_crop_except0_2_new = [value for value in max_top30_differences_crop_2_new if value != 0]
    mean_max_top30_differences_crop2_new = np.mean(max_top30_differences_crop_2_new)
    mean_round_max_top30_differences_crop2_new = round(mean_max_top30_differences_crop2_new, 2)
    max_max_top30_differences_crop2_new = max(max_top30_differences_crop_2_new)
    min_max_top30_differences_crop_except0_2_new = min(max_top30_differences_crop_except0_2_new)
    min_max_top30_differences_crop_except0_2_index_new = max_top30_differences_crop_2_new.index(max_max_top30_differences_crop2_new)

    

    for i in range(len(gray_frames)):

        top30_max_image_path_3_new = n2_max_image_path
        top30_max_image_gray_3_new = cv2.cvtColor(top30_max_image_3_new, cv2.COLOR_BGR2GRAY) 
        
        cropped_top30_max_image_gray_3_new = top30_max_image_gray_3_new[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
        
        current_image_crop_3 = current_gray_image_list[i][y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 

        diff_top30_max_3_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_3_new, current_image_crop_3)
        _, max_diff_crop_top30_max_3_new, _,max_diff_loc_crop_top30_max_3_new = cv2.minMaxLoc(diff_top30_max_3_cropped_new)

        max_top30_differences_crop_3_new.append(max_diff_crop_top30_max_3_new)
        max_top30_diff_locations_crop_3_new.append(max_diff_loc_crop_top30_max_3_new)
        
        max_top30_differences_crop_indices_3_new = np.argsort(max_top30_differences_crop_3_new)[-1:]
        max_top30_differences_crop_indices_3_new_num_frames = np.arange(len(max_top30_differences_crop_3_new))[:int(fps*NUM_FRAMES)] 
        
        max_top30_differences_crop_values_3_new = [max_top30_differences_crop_3_new[i] for i in max_top30_differences_crop_indices_3_new_num_frames] 
    
        max_top30_differences_crop_values_3_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_3_new) 
        index_in_Frames_3 = max_top30_differences_crop_indices_3_new_num_frames[max_top30_differences_crop_values_3_new_max_in_Frames]
        
        value_in_Frames_3 =  max_top30_differences_crop_values_3_new[index_in_Frames_3]
        location_in_Frames_3 = max_top30_diff_locations_crop_3_new[index_in_Frames_3]
        
        max_top30_differences_crop_indices_3_index_new = max_top30_differences_crop_indices_3_new[0]
        max_top30_differences_crop_indices_3_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_3_index_new + start_number :04d}.{FRAME_EXTENSION}')
        

    max_top30_differences_crop_except0_3_new = [value for value in max_top30_differences_crop_3_new if value != 0]
    mean_max_top30_differences_crop3_new = np.mean(max_top30_differences_crop_3_new)
    mean_round_max_top30_differences_crop3_new = round(mean_max_top30_differences_crop3_new, 2)
    max_max_top30_differences_crop3_new = max(max_top30_differences_crop_3_new)
    min_max_top30_differences_crop_except0_3_new = min(max_top30_differences_crop_except0_3_new)
    min_max_top30_differences_crop_except0_3_index_new = max_top30_differences_crop_3_new.index(max_max_top30_differences_crop3_new)
    
    top30_max_image_path_4_new = os.path.join(folder_path, f'{FRAME_PREFIX}{min_max_top30_differences_crop_except0_3_index_new + start_number :04d}.{FRAME_EXTENSION}') 
    
    x_differences = []  
    y_differences = []
    valid_pixel_counts_min_new = []
    valid_pixel_counts_max_new = []
    valid_pixel_all_counts_max_new = []
    valid_pixel_all_counts_min_new = []

    for i in range(len(gray_frames)):

        if use_gpu:
            image1 = gray_tensor[i].cpu().numpy().astype(np.uint8)
        else:
            image1 = gray_frames[i].astype(np.uint8)


        max_image_path_pixel_new = top30_max_image_path_3_new 
        min_image_path_pixel_new = top30_max_image_path_2_new 
        
        max_image_pixel_new = current_rgb_image_list[image_frame_n2]
        min_image_pixel_new = current_rgb_image_list[image_frame_n1]
        max_image_gray_pixel_new = cv2.cvtColor(max_image_pixel_new, cv2.COLOR_BGR2GRAY)
        min_image_gray_pixel_new = cv2.cvtColor(min_image_pixel_new, cv2.COLOR_BGR2GRAY)
        
        cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new] #small_size
        cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
        cropped_image0_pixel_new = image0[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
        cropped_image1_pixel_new = image1[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
        
        if cropped_image0_pixel is None or cropped_image1_pixel is None:
            print(f'Error cropping image{i}. Skipping...')
            continue 
        
        diff_max_new = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
        diff_min_new = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 
        
        # Count pixels with intensity differences above the nonzero minimum threshold.
        # KR: 0을 제외한 최소 threshold 이상으로 변화한 pixel의 좌표를 계산합니다.
        threshold_max_new = min_max_top30_differences_crop_except0_3_new 
        threshold_min_new = min_max_top30_differences_crop_except0_2_new  
        white_pixel_coords_max_new = np.argwhere(diff_max_new >= threshold_max_new)
        white_pixel_coords_min_new = np.argwhere(diff_min_new >= threshold_min_new)
        
        min_threshold_max_new = min_max_top30_differences_crop_except0_3_new +20 
        max_threshold_max_new = max_max_top30_differences_crop3_new 

        min_threshold_min_new = min_max_top30_differences_crop_except0_2_new +20 
        max_threshold_min_new = max_max_top30_differences_crop2_new 
        
        valid_pixel_coords_max_new = np.argwhere((diff_max_new >= min_threshold_max_new) & (diff_max_new <= max_threshold_max_new))
        valid_pixel_coords_min_new = np.argwhere((diff_min_new >= min_threshold_min_new) & (diff_min_new <= max_threshold_min_new))
        
        valid_pixel_coords_max_all_new = np.argwhere((diff_max_new >= 0) & (diff_max_new <= max_threshold_max))
        valid_pixel_coords_min_all_new = np.argwhere((diff_min_new >= 0) & (diff_min_new <= max_threshold_min))

        valid_pixel_count_max_new = len(valid_pixel_coords_max_new)
        valid_pixel_count_min_new = len(valid_pixel_coords_min_new)
        valid_pixel_counts_max_new.append(valid_pixel_count_max_new)
        valid_pixel_counts_min_new.append(valid_pixel_count_min_new)       


    valid_pixel_count_mean_min_new = np.mean([value for value in valid_pixel_counts_min_new if value not in [0,'']]) 

    valid_pixel_count_mean_round_min_new = round(valid_pixel_count_mean_min_new) if not np.isnan(valid_pixel_count_mean_min_new) else 0 
    
    valid_pixel_count_mean_max_new = np.mean([value for value in valid_pixel_counts_max_new if value not in [0,'']]) 

    valid_pixel_count_mean_round_max_new = round(valid_pixel_count_mean_max_new) if not np.isnan(valid_pixel_count_mean_max_new) else 0 

    count_zeros_min_new = valid_pixel_counts_min_new.count(0)
    count_zeros_max_new = valid_pixel_counts_max_new.count(0)


    count_all_pixel_in_range_min_new = [valid_pixel_counts_min_new.count(i) for i in range(valid_pixel_count_mean_round_min_new)]
    count_all_pixel_in_range_max_new = [valid_pixel_counts_max_new.count(i) for i in range(valid_pixel_count_mean_round_max_new)]
    
    sum_count_all_pixel_in_range_min_new = sum(count_all_pixel_in_range_min_new, 0)
    sum_count_all_pixel_in_range_max_new = sum(count_all_pixel_in_range_max_new, 0)

    valid_pixel_all_counts_min_new.append(sum_count_all_pixel_in_range_min_new)
    valid_pixel_all_counts_max_new.append(sum_count_all_pixel_in_range_max_new)
    
    valid_pixel_all_counts_min_max_new = max(valid_pixel_all_counts_min_new)
    valid_pixel_all_counts_max_max_new = max(valid_pixel_all_counts_max_new)

    All_pixel_intensities = []  
    valid_pixel_counts = []
    not_cropped_valid_pixel_counts = []
    selected_max_differences = []
    selected_max_differences_location = []
    Analysis_frames = []
    pixel_intensities_movement_list = []
    pixel_intensities_movement_coordinates_list = []



    for i in range(len(gray_frames)):

        max_image_path_pixel_new = top30_max_image_path_3_new
        min_image_path_pixel_new = top30_max_image_path_2_new
        
        max_image_pixel_new = current_rgb_image_list[image_frame_n2]
        min_image_pixel_new = current_rgb_image_list[image_frame_n1]
        max_image_gray_pixel_new = cv2.cvtColor(max_image_pixel_new, cv2.COLOR_BGR2GRAY)
        min_image_gray_pixel_new = cv2.cvtColor(min_image_pixel_new, cv2.COLOR_BGR2GRAY)
        
        cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
        cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
        cropped_image0_pixel_new = image0[y1_small_new:y2_small_new, x1_small_new:x2_small_new]  
        cropped_image1_pixel_new = image1[y1_small_new:y2_small_new, x1_small_new:x2_small_new]  
        
        if cropped_image0_pixel is None or cropped_image1_pixel is None:
            print(f"Error cropping image {i}. Skipping...")
            continue 
        
        # Select the reference candidate frame based on valid-pixel count statistics.
        # The selected candidate determines which images, thresholds, coordinates,
        # and intensity-difference values are used in downstream displacement analysis.
        # KR: valid-pixel count 통계를 기준으로 기준 후보 frame을 선택합니다.
        # KR: 선택된 후보 frame에 따라 이후 displacement 분석에 사용할 이미지, threshold, 좌표, intensity-difference 값이 결정됩니다.
        
        # Apply default or reversed candidate-selection logic depending on the expected contraction direction.
        # KR: 예상되는 수축 방향에 따라 기본 후보 선택 기준 또는 반전된 후보 선택 기준을 적용합니다.
        if REVERSE_CONTRACTION.lower() == 'off':
            if valid_pixel_all_counts_min_max_new >= valid_pixel_all_counts_max_max_new:
                # Use the candidate relaxation frame as the reference image
                # KR: 후보 이완 frame을 기준 이미지로 사용합니다.
                diff = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new)  
                # Full-frame difference used for visualization
                # KR: visualization을 위해 full-frame difference를 계산합니다.
                not_crop_diff = cv2.subtract(max_image_gray_pixel_new, image1) 
                # Difference between candidate contraction and relaxation frames
                # KR: 후보 수축 frame과 후보 이완 frame 사이의 차이입니다.
                crop_diff = cv2.subtract(cropped_max_image_pixel_new, cropped_min_image_pixel_new) 
                #=================================================================
                # Store paired candidate contraction/relaxation frame information
                # KR: 후보 수축/이완 frame 쌍의 정보를 저장합니다.
                #=================================================================
                # Reference image: candidate relaxation frame
                # KR: 기준 이미지: 후보 이완 frame
                square_image = min_image_pixel_new
                # Path to the reference relaxation frame
                # KR: 기준 이완 frame의 경로
                square_image_path = min_image_path_pixel_new
                # Paired candidate contraction frame
                # KR: 기준 frame과 짝이 되는 후보 수축 frame
                square_image_path_contraction = max_image_path_pixel_new 
                # Loaded RGB image of the reference relaxation frame
                # KR: 기준 이완 frame의 RGB 이미지
                square_image_imread = current_rgb_image_list[image_frame_n1]
                # Loaded RGB image of the paired contraction frame
                # KR: 짝이 되는 후보 수축 frame의 RGB 이미지
                square_image_imread_contraction = current_rgb_image_list[image_frame_n2]
                # Index of the selected reference frame
                # KR: 선택된 기준 frame의 index
                square_image_index = image_frame_n1
                # Image showing the maximum difference from the reference frame
                # KR: 기준 frame과 가장 큰 차이를 보이는 이미지
                square_differences_crop_indices_index_path_image = max_image_pixel_new 
                # Minimum intensity-difference threshold
                # KR: 최소 intensity-difference threshold
                threshold = min_max_top30_differences_crop_except0_2_new 
                # Maximum intensity-difference threshold
                # KR: 최대 intensity-difference threshold
                max_threshold = max_max_top30_differences_crop2_new
                # Maximum intensity difference
                # KR: intensity difference 값
                diff_val = max_diff_crop_top30_max_2_new   
                # List of maximum intensity differences
                # KR: 최대 intensity difference 리스트
                diff2_val = max_top30_differences_crop_2_new
                # Coordinates of the maximum intensity difference
                # KR: 최대 intensity difference 좌표
                loc = max_diff_loc_crop_top30_max_2_new
                # Coordinate of the maximum intensity difference within the initial analysis window
                # KR: 초기 분석 구간 안에서 최대 intensity difference가 나타난 좌표
                loc2 = max_top30_diff_locations_crop_2_new
                # Coordinate of the maximum intensity difference within the initial analysis window
                # KR: 초기 분석 구간 안에서 최대 intensity difference가 나타난 좌표
                loc3 = location_in_Frames_2 
                # Maximum intensity difference from the reference frame
                # KR: 기준 frame과 가장 큰 차이를 보이는 intensity difference 값
                max_values = max_max_top30_differences_crop2_new 
                ## Index of the frame showing the maximum difference from the reference frame
                # KR: 기준 frame과 가장 큰 차이를 보이는 frame의 index
                max_indices = min_max_top30_differences_crop_except0_2_index_new 
                # List of maximum intensity differences
                # KR: 최대 intensity difference 리스트
                adapt = max_top30_differences_crop_2_new 
                # Maximum intensity-difference value
                # KR: 최대 intensity-difference 값
                adapt_max = max_max_top30_differences_crop2_new 
                # Mean intensity-difference value
                # KR: 평균 intensity-difference 값
                adapt_mean = mean_round_max_top30_differences_crop2_new 
                
            else:
                diff = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
                not_crop_diff = cv2.subtract(min_image_gray_pixel_new, image1) 
                crop_diff = cv2.subtract(cropped_min_image_pixel_new, cropped_max_image_pixel_new) 
                square_image = max_image_pixel_new 
                square_image_path = max_image_path_pixel_new 
                square_image_path_contraction = max_image_path_pixel_new 
                square_image_imread = current_rgb_image_list[image_frame_n2] 
                square_image_imread_contraction = current_rgb_image_list[image_frame_n1] 
                square_image_index = image_frame_n2 
                square_differences_crop_indices_index_path_image = min_image_pixel_new 
                threshold = min_max_top30_differences_crop_except0_3_new  
                max_threshold = max_max_top30_differences_crop3_new 
                diff_val = max_diff_crop_top30_max_3_new 
                diff2_val = max_top30_differences_crop_3_new 
                loc = max_diff_loc_crop_top30_max_3_new 
                loc2 = max_top30_diff_locations_crop_3_new 
                loc3 = location_in_Frames_3 
                max_values = max_max_top30_differences_crop3_new 
                max_indices = min_max_top30_differences_crop_except0_3_index_new 
                adapt = max_top30_differences_crop_3_new 
                adapt_max = max_max_top30_differences_crop3_new 
                adapt_mean = mean_round_max_top30_differences_crop3_new 
        else: 
            if valid_pixel_all_counts_min_max_new <= valid_pixel_all_counts_max_max_new:
                diff = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 
                not_crop_diff = cv2.subtract(max_image_gray_pixel_new, image1) 
                crop_diff = cv2.subtract(cropped_max_image_pixel_new, cropped_min_image_pixel_new) 
                square_image = min_image_pixel_new 
                square_image_path = min_image_path_pixel_new 
                square_image_path_contraction = max_image_path_pixel_new 
                square_image_imread = current_rgb_image_list[image_frame_n1] 
                square_image_imread_contraction = current_rgb_image_list[image_frame_n2] 
                square_image_index = image_frame_n1
                square_differences_crop_indices_index_path_image = max_image_pixel_new 
                threshold = min_max_top30_differences_crop_except0_2_new 
                max_threshold = max_max_top30_differences_crop2_new 
                diff_val = max_diff_crop_top30_max_2_new  
                diff2_val = max_top30_differences_crop_2_new 
                loc = max_diff_loc_crop_top30_max_2_new 
                loc2 = max_top30_diff_locations_crop_2_new 
                loc3 = location_in_Frames_2 
                max_values = max_max_top30_differences_crop2_new 
                max_indices = min_max_top30_differences_crop_except0_2_index_new 
                adapt = max_top30_differences_crop_2_new 
                adapt_max = max_max_top30_differences_crop2_new 
                adapt_mean = mean_round_max_top30_differences_crop2_new 
                
            else:
                diff = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
                not_crop_diff = cv2.subtract(min_image_gray_pixel_new, image1) 
                crop_diff = cv2.subtract(cropped_min_image_pixel_new, cropped_max_image_pixel_new) 
                square_image = max_image_pixel_new 
                square_image_path = max_image_path_pixel_new 
                square_image_path_contraction = max_image_path_pixel_new 
                square_image_imread = current_rgb_image_list[image_frame_n2] 
                square_image_imread_contraction = current_rgb_image_list[image_frame_n1] 
                square_image_index = image_frame_n2 
                square_differences_crop_indices_index_path_image = min_image_pixel_new 
                threshold = min_max_top30_differences_crop_except0_3_new  
                max_threshold = max_max_top30_differences_crop3_new 
                diff_val = max_diff_crop_top30_max_3_new 
                diff2_val = max_top30_differences_crop_3_new 
                loc = max_diff_loc_crop_top30_max_3_new 
                loc2 = max_top30_diff_locations_crop_3_new  
                loc3 = location_in_Frames_3 
                max_values = max_max_top30_differences_crop3_new 
                max_indices = min_max_top30_differences_crop_except0_3_index_new 
                adapt = max_top30_differences_crop_3_new 
                adapt_max = max_max_top30_differences_crop3_new 
                adapt_mean = mean_round_max_top30_differences_crop3_new 
            

        max_pixel_intensities = int(not_crop_diff.max())
        All_pixel_intensities.append(max_pixel_intensities)
        max_All_pixel_intensities = int(np.max(All_pixel_intensities))
        All_pixel_intensities.append(max_pixel_intensities)
        max_All_pixel_intensities = np.max(All_pixel_intensities)

        selected_image_path = square_image_path  
        
        selected_image = square_image_imread 
        selected_image_contraction = square_image_imread_contraction 
        selected_image_gray = cv2.cvtColor(selected_image, cv2.COLOR_BGR2GRAY) 
        selected_image_contraction_gray = cv2.cvtColor(selected_image_contraction, cv2.COLOR_BGR2GRAY) 
        
        all_image_gray = current_gray_image_list[i] 
        adapt = np.array(adapt) 
        adapt_min = adapt[adapt >= threshold].tolist()
        try:
            sorted_adapt_min = min(adapt_min)
            sorted_adapt_max = max(adapt_min)
            index_of_sorted_adapt_min = int(np.where(adapt == sorted_adapt_min)[0][0])
            index_of_sorted_adapt_max = int(np.where(adapt == sorted_adapt_max)[0][0])
            coordinate_of_sorted_adapt_min = loc2[index_of_sorted_adapt_min]
            coordinate_of_sorted_adapt_max = loc2[index_of_sorted_adapt_max]
        except ValueError:
            break
        
        cropped_selected_image_gray = selected_image_gray[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
        cropped_all_image_gray = all_image_gray[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
        
        cropped_selected_image_gray_reference = selected_image_gray[y1_small_new_optical:y2_small_new_optical, x1_small_new_optical:x2_small_new_optical] 
        cropped_all_image_gray_reference = all_image_gray[y1_small_new_optical:y2_small_new_optical, x1_small_new_optical:x2_small_new_optical] 
        
        
        diff_selected = cv2.absdiff( cropped_all_image_gray, cropped_selected_image_gray) 
        diff_selected_reference = cv2.subtract( cropped_all_image_gray_reference, cropped_selected_image_gray_reference)

        pixel_intensities_movement = diff_selected.ravel()

        diff_selected_column = np.argwhere(diff_selected >= threshold)

        pixel_intensities_movement_list.append(pixel_intensities_movement)
        pixel_intensities_movement_coordinates_list.append(diff_selected_column)

        max_pixel_intensity = int(diff_selected.max()) if diff_selected.size else 0

        
        reference_selected_image_gray = selected_image_contraction_gray[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference]
        reference_all_image_gray = all_image_gray[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
        
        not_crop_selected_image_gray = selected_image_contraction_gray 
        not_crop_all_image_gray = all_image_gray

        reference_diff_selected = cv2.absdiff(reference_selected_image_gray, reference_all_image_gray) 
        not_crop_diff_selected = cv2.absdiff(not_crop_selected_image_gray, not_crop_all_image_gray) 
        
        diff_loc = np.argwhere((diff_selected_reference >= threshold) & (diff_selected_reference <= max_threshold))
        not_cropped_diff_loc = np.argwhere((reference_diff_selected >= threshold) & (reference_diff_selected <= max_threshold)) 
        
        threshold_pixel_image = np.zeros_like(diff_selected_reference, dtype = np.uint8)
        threshold_pixel_image_color = cv2.cvtColor(threshold_pixel_image, cv2.COLOR_GRAY2BGR)
        for loc in diff_loc:
            threshold_pixel_image_color[loc[0], loc[1]] = diff_selected_reference[loc[0], loc[1]]
        
        result_folder_diff = os.path.join(folder_path, 'result', 'Analysis Square') 
        if not os.path.exists(result_folder_diff):
            os.makedirs(result_folder_diff, exist_ok = True) 
        x_coord, y_coord = x_selected_new - x1_small_new_optical ,y_selected_new - y1_small_new_optical
    
        height, width = diff_selected.shape[:2]
        not_cropped_height, not_cropped_width = reference_diff_selected.shape[:2]

        new_width = 1080

        new_height = int(height * (new_width / width))
        diff_resized = cv2.resize(threshold_pixel_image_color, (new_width, new_height), interpolation=cv2.INTER_LINEAR) 
        result_image_path = os.path.join(result_folder_diff, f'Analysis_Frame_{i:04d}.{FRAME_EXTENSION}')
        Analysis_frames.append(diff_resized)
        white_pixel_coords = np.argwhere(diff_selected >= adapt_mean)
        
        not_cropped_white_pixel_coords = np.argwhere(reference_diff_selected >= adapt_mean)
        
        white_pixel_count = len(white_pixel_coords)
        
        not_cropped_white_pixel_count = len(not_cropped_white_pixel_coords)
        
        min_threshold = threshold
        max_threshold = max_threshold
        
        valid_pixel_coords = np.argwhere((diff_selected >= min_threshold) & (diff_selected <= max_threshold))
        not_cropped_valid_pixel_coords = np.argwhere((reference_diff_selected >= min_threshold) & (reference_diff_selected <= max_threshold))
        
        valid_pixel_count = len(valid_pixel_coords)
        valid_pixel_counts.append(valid_pixel_count)
        total_pixel_count = resolution_x * resolution_y 
        
        not_cropped_valid_pixel_count = len(not_cropped_valid_pixel_coords)
        not_cropped_valid_pixel_counts.append(not_cropped_valid_pixel_count)   



    valid_pixel_counts_np = np.array(valid_pixel_counts, dtype='float32') 

    valid_values = valid_pixel_counts_np[(valid_pixel_counts_np != 0) & ~np.isnan(valid_pixel_counts_np)]

    valid_pixel_count_mean = np.mean(valid_values) if valid_values.size > 0 else 0
    valid_pixel_count_mean_round = round(valid_pixel_count_mean)

    valid_pixel_count_maximum = np.max(valid_values) if valid_values.size > 0 else 0
    valid_pixel_count_maximum_round = round(valid_pixel_count_maximum)

    min_valid_pixel_counts = np.min(valid_values) if valid_values.size > 0 else 0
    if not valid_pixel_counts:
        print("!!!!THIS SAMPLE DOESN'T HAVE ANY MOVEMENT!!!!.")
        print(f"Move to {NOISE_SAMPLE_FOLDER_NAME} Folder")
        classification_dir = os.path.join(VIDEO_FILE_PATH, 'classification')
        destination_folder = os.path.join(classification_dir, NOISE_SAMPLE_FOLDER_NAME)
        folder_path = os.path.join(VIDEO_FILE_PATH, folder_name)
        video_path = os.path.join(VIDEO_FILE_PATH, f'{folder_name}.{extension}')  
        os.makedirs(classification_dir, exist_ok=True)
        os.makedirs(destination_folder, exist_ok=True)
        try:
            shutil.move(folder_path, destination_folder)
            shutil.move(video_path, destination_folder)
        except Exception as e:
            pass
        continue
    
    difference_square_image = cv2.subtract(square_differences_crop_indices_index_path_image, square_image)
    difference_square_image_gray = cv2.cvtColor(difference_square_image, cv2.COLOR_BGR2GRAY)
    

    # =========================================================
    # SEGMENTATION
    # Load the segmentation model once and run inference using GPU or CPU.
    # KR: segmentation model을 한 번 로드하고, GPU 또는 CPU에서 추론을 수행합니다.
    # =========================================================

    # 1) 디바이스 결정 (전역 GPU_MODE 사용)
    if GPU_MODE.lower() == 'on' and torch.cuda.is_available():
        seg_device = torch.device('cuda:0')
    else:
        seg_device = torch.device('cpu')

    print(f"[Segmentation] Using device: {seg_device}")
    hrnet = CustomHRNet(num_classes=2)
    deeplabv3 = deeplabv3_resnet101(weights=None, num_classes=2)

    convert_batchnorm_to_groupnorm(hrnet)
    convert_batchnorm_to_groupnorm(deeplabv3)

    seg_model = DualNet(hrnet, deeplabv3, num_classes=2)

    checkpoint = torch.load(BEST_SEGMENTATION_CHECKPOINT_FILE, map_location=seg_device)
    seg_model.load_state_dict(checkpoint['model_state_dict'])
    seg_model.to(seg_device)
    seg_model.eval()

    # Run segmentation on the selected contraction/relaxation candidate frame
    # KR: 선택된 수축/이완 후보 frame에 대해 segmentation을 수행합니다.

    segmentation_area = preprocess_and_segment(square_image_path_contraction)

    
    plt.figure(figsize = (14,6))
    
    # Display the selected relaxation and contraction candidate frames
    # KR: 선택된 이완 및 수축 후보 frame을 표시합니다.
    plt.subplot(131)
    plt.imshow(cv2.cvtColor(square_image, cv2.COLOR_BGR2RGB))
    plt.title(f'Maximum Relaxation Frame [{square_image_index}]', fontweight='bold')
    plt.axis('off')
    
    plt.subplot(132)
    plt.imshow(cv2.cvtColor(square_differences_crop_indices_index_path_image, cv2.COLOR_BGR2RGB))
    plt.title(f'Maximum Contraction Frame [{max_indices}]', fontweight='bold')
    plt.axis('off')
    
    # Display the difference image between the selected relaxation and contraction frames
    # KR: 선택된 이완 frame과 수축 frame 사이의 difference image를 표시합니다.
    plt.subplot(133)
    plt.imshow(difference_square_image_gray, cmap = 'gray')
    plt.title(f'Difference Frame [{square_image_index}] vs [{max_indices}]', fontweight='bold')
    plt.axis('off')
    
    # Save the result image to the result folder
    # KR: 결과 이미지를 result 폴더에 저장합니다.
    result_folder = os.path.join(folder_path, 'result')
    os.makedirs(result_folder, exist_ok = True)
    result_image_path = os.path.join(result_folder, f'Pixel_Intensity_Difference_Relaxation_vs_Contraction.{FRAME_EXTENSION}')
    plt.savefig(result_image_path, bbox_inches='tight', pad_inches=0.1)
    plt.close()
    
    print("Difference image between relaxation and contraction frames saved.")
    
    
    # Save analysis video to the result folder
    # KR: 분석 결과 video를 result 폴더 안에 저장합니다.
    video_name_base = os.path.join(result_folder_diff, 'Analysis_video')
    video_name = os.path.join(video_name_base, f'{folder_name}_Analysis_video.avi')
    os.makedirs(video_name_base, exist_ok=True)
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(video_name, fourcc, fps, (new_width, new_height))
    for i in range(len(gray_frames)):
        out.write(Analysis_frames[i])

    out.release()
    print("Analysis video export completed.")
    
    gray_color = '#bbbbbd'
    red_color = '#FF6699'
    blue_color = '#009fd6'
    yellow_color = '#d69f00'


    # Plot the frame-wise maximum pixel-intensity difference over time.
    # KR: 시간에 따른 frame별 최대 pixel-intensity difference를 plot합니다.
    fig, ax = plt.subplots(figsize=(8, 8)) 
    ax.plot(np.linspace(0, total_time_seconds, len(adapt)), 
            adapt, 
            marker='o', 
            color=gray_color, 
            linewidth=4, 
            markersize=8, 
            markeredgewidth=1.6)

    ax.set_xlabel("Time (s)", fontsize=40, fontweight='bold', fontname='Arial')
    ax.set_ylabel("Maximum Pixel Intensity Difference", fontsize=25, fontweight='bold', fontname='Arial')

    for spine in ax.spines.values():
        spine.set_linewidth(4) 

    ax.set_xlim(0, total_time_seconds)
    ax.set_xticks(range(0, int(total_time_seconds) + 1, 5))

    ax.set_ylim(bottom=0)  
    ax.set_xlim(left=0) 

    ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=10)
    ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=10)

    ax.spines['left'].set_position(('data', 0))

    ax.spines['right'].set_color('black')  
    ax.spines['top'].set_color('black')   
    plt.xticks(fontweight='bold')

    plt.yticks(fontweight='bold')
    
    Savefig('Maximum_Pixel_Intensity_Difference', 300, 0.05)


    
    # ============================================
    # Optical flow-based displacement analysis
    # KR: optical flow 기반 displacement 분석
    # ============================================

    # Prepare candidate contraction and relaxation frames for ROI-based optical flow.
    # KR: ROI 기반 optical flow 분석을 위해 후보 수축/이완 frame을 준비합니다.
    roi_contraction = cv2.cvtColor(
        square_differences_crop_indices_index_path_image, cv2.COLOR_BGR2GRAY
    )
    roi_relaxation = cv2.cvtColor(
        square_image, cv2.COLOR_BGR2GRAY
    )

    roi_contraction = roi_contraction[
        y1_small_new_optical:y2_small_new_optical,
        x1_small_new_optical:x2_small_new_optical
    ]
    roi_relaxation = roi_relaxation[
        y1_small_new_optical:y2_small_new_optical,
        x1_small_new_optical:x2_small_new_optical
    ]

    # Skip optical flow when the ROI is empty.
    # KR: ROI가 비어 있으면 optical flow 분석을 건너뜁니다.
    if roi_contraction.size == 0:
        print("roi_contraction is empty. Optical flow skipped.")
        target_displacements = [0.0]
        maximum_displacement_list = target_displacements
        max_target_displacements = 0.0
        max_displacement_um = 0.0
        min_target_displacements = 0.0

        # Use the selected coordinate as the fallback position.
        # KR: optical flow를 수행할 수 없을 때 선택 좌표를 fallback 위치로 사용합니다.
        a_global = x_selected_new
        b_global = y_selected_new
        c_global = x_selected_new
        d_global = y_selected_new

        global_max_coords_global = (b_global, a_global, d_global, c_global)
        global_max_coords_global1 = (b_global, a_global)
        global_max_coords_global2 = (d_global, c_global)
        global_max_coords_global1_ab = (a_global, b_global)
        global_max_coords_global2_cd = (c_global, d_global)

        tracked_points = [(x_selected_new - x1_small_new_optical,
                           y_selected_new - y1_small_new_optical)]
        best_x_roi = tracked_points[0][0]
        best_y_roi = tracked_points[0][1]
        best_x_roi_min = best_x_roi
        best_y_roi_min = best_y_roi

    else:
        # Set the output path for the optical flow overlay video.
        # KR: optical flow overlay video의 저장 경로를 설정합니다.
        output_video_path = os.path.join(result_folder, f'{folder_name}_Optical_Flow_Overlay.{extension}')

        fps = fps  
        frame_size = (roi_contraction.shape[1], roi_contraction.shape[0])  # Video frame size: width, height

        fourcc = cv2.VideoWriter_fourcc(*'XVID')
        out = cv2.VideoWriter(output_video_path, fourcc, fps, frame_size)

        # ------------------------------------------------
        # 2) Define the baseline tracking point within the ROI
        #    The selected motion center is clamped to remain inside the ROI.
        # KR: ROI 내부에서 baseline tracking point를 정의합니다.
        # KR: 선택된 motion center가 ROI 밖으로 벗어나지 않도록 좌표를 제한합니다.
        # ------------------------------------------------
        H, W = roi_contraction.shape[:2]

        x0 = int(x_selected_new - x1_small_new_optical)
        y0 = int(y_selected_new - y1_small_new_optical)

        half_win = ANALYSIS_SQUARE_SIZE_REFERENCE // 2    # Half-window size for coordinate clamping
        # Clamp the tracking point so that the local window remains inside the ROI.
        # KR: local tracking window가 ROI 내부에 유지되도록 tracking point 좌표를 제한합니다.
        x0 = max(half_win, min(W - 1 - half_win, x0))
        y0 = max(half_win, min(H - 1 - half_win, y0))

        # Use a single baseline tracking point for Lucas-Kanade optical flow.
        # KR: Lucas-Kanade optical flow에서 하나의 baseline tracking point만 사용합니다.
        p0 = np.array([[[x0, y0]]], dtype=np.float32)

        # Store optical-flow tracking results.
        # KR: optical flow tracking 결과를 저장합니다.
        target_displacements = []   # Displacement for each frame in µm
        tracked_points = []         # Tracked ROI coordinates for each frame

        # Smoothing parameters for visualization only.
        # KR: visualization에만 사용하는 smoothing 설정입니다.
        smoothed_x, smoothed_y = None, None
        alpha = 0.5  # EMA smoothing factor

        # ------------------------------------------------
        # 3) Optical flow tracking from the baseline ROI to each frame ROI
        #    A single point p0 is tracked to define optical flow-derived displacement.
        # KR: baseline ROI에서 각 frame ROI로 optical flow tracking을 수행합니다.
        # KR: 하나의 p0 point를 추적하여 optical flow-derived displacement를 정의합니다.
        # ------------------------------------------------
        for frame_idx, frame_full in enumerate(current_gray_image_list):

            cur_gray = frame_full[
                y1_small_new_optical:y2_small_new_optical,
                x1_small_new_optical:x2_small_new_optical
            ]

            # Keep the previous displacement value when the current ROI is empty.
            # KR: 현재 ROI가 비어 있으면 이전 displacement 값을 유지합니다.
            if cur_gray.size == 0:
                last_disp = target_displacements[-1] if target_displacements else 0.0
                target_displacements.append(float(last_disp))
                tracked_points.append((float(x0), float(y0)))
                continue

            p1, st, err = cv2.calcOpticalFlowPyrLK(
                roi_contraction,
                cur_gray,
                p0,
                None,
                winSize=(15, 15),
                maxLevel=2,
                criteria=(cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, 10, 0.03)
            )

            frame_rgb = cv2.cvtColor(cur_gray, cv2.COLOR_GRAY2BGR)

            # Handle optical-flow tracking failure.
            # KR: optical flow tracking에 실패한 경우를 처리합니다.
            if p1 is None or st is None or st[0, 0] == 0:
                last_disp = target_displacements[-1] if target_displacements else 0.0
                target_displacements.append(float(last_disp))
                tracked_points.append((float(x0), float(y0)))

                # Write the frame even when optical-flow tracking fails.
                # KR: optical flow tracking에 실패해도 output video frame은 저장합니다.
                out.write(frame_rgb)
                continue

            # New tracked coordinate in the baseline ROI coordinate system.
            # KR: baseline ROI 좌표계에서 새롭게 추적된 좌표입니다.
            new_x = float(p1[0, 0, 0])
            new_y = float(p1[0, 0, 1])

            # Calculate displacement from the baseline point p0.
            # This is not cumulative; the same baseline point is used for all frames.
            # KR: baseline point p0로부터의 displacement를 계산합니다.
            # KR: 누적 변위가 아니라 모든 frame에서 동일한 기준점 p0를 사용합니다.
            dx = new_x - x0
            dy = new_y - y0
            disp_um = np.sqrt(dx * dx + dy * dy) * PIXEL_UM

            target_displacements.append(float(disp_um))
            tracked_points.append((new_x, new_y))

            # Smooth the displayed tracking point for visualization.
            # KR: visualization에서 표시되는 tracking point를 부드럽게 보이도록 smoothing합니다.
            if smoothed_x is None:
                smoothed_x, smoothed_y = new_x, new_y
            else:
                smoothed_x = alpha * new_x + (1.0 - alpha) * smoothed_x
                smoothed_y = alpha * new_y + (1.0 - alpha) * smoothed_y

            # Draw a green arrow from the baseline point to the current tracked position,
            # and mark the smoothed tracked point in red.
            # KR: baseline point에서 현재 추적 위치까지 초록색 화살표를 그리고,
            # KR: smoothing된 tracking point를 빨간색으로 표시합니다.
            cv2.arrowedLine(
                frame_rgb,
                (int(x0), int(y0)),
                (int(new_x), int(new_y)),
                (0, 255, 0),
                1
            )
            cv2.circle(
                frame_rgb,
                (int(smoothed_x), int(smoothed_y)),
                2,
                (0, 0, 255),
                -1
            )

            out.write(frame_rgb)
        # Finalize the optical flow overlay video.
        # KR: optical flow overlay video 저장을 종료합니다.
        out.release()

        # ------------------------------------------------
        # 4) Displacement time-series post-processing
        #    - target_displacements are stored in micrometers.
        #    - The analysis window is selected based on start_number, NUM_FRAMES, and fps.
        # KR: displacement time-series를 후처리합니다.
        # KR: target_displacements는 micrometer 단위입니다.
        # KR: 분석 구간은 start_number, NUM_FRAMES, fps를 기준으로 선택합니다.
        # ------------------------------------------------
        if len(target_displacements) == 0:
            target_displacements = [0.0]

        # Shift the minimum displacement to zero.
        # KR: 최소 displacement가 0이 되도록 baseline을 보정합니다.
        min_all = min(target_displacements)
        target_displacements = [float(v - min_all) for v in target_displacements]
        
        # Invert the displacement curve for visualization consistency.
        # KR: visualization 방향을 맞추기 위해 displacement curve를 반전합니다.
        max_all = max(target_displacements)
        target_displacements = [float(max_all - v) for v in target_displacements]

        # Select the initial analysis window.
        # KR: 초기 분석 구간을 선택합니다.
        end_idx = int(fps * NUM_FRAMES)
        end_idx = min(end_idx, len(target_displacements))

        if start_number + 1 < end_idx:
            window_vals = target_displacements[start_number + 1:end_idx]
            window_indices = list(range(start_number + 1, end_idx))
        else:
            window_vals = target_displacements
            window_indices = list(range(len(target_displacements)))

        if len(window_vals) == 0:
            window_vals = target_displacements
            window_indices = list(range(len(target_displacements)))

        # Define the final displacement time series and maximum displacement value.
        # KR: 최종 displacement time series와 maximum displacement 값을 정의합니다.
        maximum_displacement_list = target_displacements
        max_target_displacements = round(max(window_vals), 2)
        min_target_displacements = float(min(window_vals))
        max_displacement_um = max_target_displacements

        max_target_displacements_before_all = round(max(target_displacements), 2)

        # Create the time axis for displacement plotting.
        # KR: displacement plot에 사용할 시간축을 생성합니다.
        if fps > 0:
            total_time_seconds = len(maximum_displacement_list) / fps
        else:
            total_time_seconds = float(len(maximum_displacement_list))
        time_axis = np.linspace(0, total_time_seconds, len(maximum_displacement_list))

        # Frame index where the maximum displacement occurs within the analysis window.
        # KR: 분석 구간 안에서 maximum displacement가 발생한 frame index입니다.
        idx_in_window = int(np.argmax(window_vals))
        frame_idx_max = int(window_indices[idx_in_window])

        # Extract ROI coordinates of the tracked point at maximum and baseline displacement.
        # KR: maximum displacement와 baseline displacement 시점의 tracked point ROI 좌표를 가져옵니다.
        if frame_idx_max < len(tracked_points):
            best_x_roi, best_y_roi = tracked_points[frame_idx_max]
        else:
            best_x_roi, best_y_roi = tracked_points[-1]

        if start_number < len(tracked_points):
            best_x_roi_min, best_y_roi_min = tracked_points[start_number]
        else:
            best_x_roi_min, best_y_roi_min = tracked_points[0]

        # Convert ROI coordinates back to full-frame coordinates.
        # KR: ROI 좌표를 full-frame 좌표로 변환합니다.
        a_global = int(best_x_roi + x1_small_new_optical)
        b_global = int(best_y_roi + y1_small_new_optical)
        c_global = int(best_x_roi_min + x1_small_new_optical)
        d_global = int(best_y_roi_min + y1_small_new_optical)

        global_max_coords_global = (b_global, a_global, d_global, c_global)
        global_max_coords_global1 = (b_global, a_global)
        global_max_coords_global2 = (d_global, c_global)
        global_max_coords_global1_ab = (a_global, b_global)
        global_max_coords_global2_cd = (c_global, d_global)
        
   
    # Convert the displacement time series to a comma-separated string for Excel output.
    # KR: Excel output에 저장하기 위해 displacement time series를 comma-separated string으로 변환합니다.
    maximum_displacement_str = ", ".join(f"{x:.2f}" for x in maximum_displacement_list)
            
    y_max_coordinates = int(y_selected_new - y1_small_new_optical)
    x_max_coordinates = int(x_selected_new - x1_small_new_optical)

    x1_max_coordinates = x_max_coordinates  - ANALYSIS_SQUARE_SIZE_REFERENCE
    y1_max_coordinates = y_max_coordinates  - ANALYSIS_SQUARE_SIZE_REFERENCE
    x2_max_coordinates = x_max_coordinates  + ANALYSIS_SQUARE_SIZE_REFERENCE
    y2_max_coordinates = y_max_coordinates  + ANALYSIS_SQUARE_SIZE_REFERENCE
    
    x1_max_coordinates = max( x1_max_coordinates, 0)
    y1_max_coordinates = max( y1_max_coordinates, 0)
    x2_max_coordinates = min( x2_max_coordinates, resolution_x -1)
    y2_max_coordinates = min( y2_max_coordinates, resolution_y -1)

    max_value_displacement = max(maximum_displacement_list) 
    
    if max_target_displacements == 0 or max_target_displacements is None:
        print("No valid displacement was detected.")
        print("This sample could not be tracked by optical flow.")
        print(f"Moving to the {OPTICAL_NOISE_SAMPLE_FOLDER_NAME} folder.")
        # Prepare the optical-noise output folder.
        # KR: optical-noise sample을 저장할 output folder를 준비합니다.
        classification_dir = os.path.join(VIDEO_FILE_PATH, 'classification')
        destination_folder = os.path.join(classification_dir, OPTICAL_NOISE_SAMPLE_FOLDER_NAME)
        # Define the source folder and video paths to be moved.
        # KR: 이동할 원본 frame folder와 video file 경로를 정의합니다.
        folder_path = os.path.join(VIDEO_FILE_PATH, folder_name)
        video_path = os.path.join(VIDEO_FILE_PATH, f'{folder_name}.{extension}')  
        os.makedirs(classification_dir, exist_ok=True)
        os.makedirs(destination_folder, exist_ok=True)
        try:
            # Move the current analysis folder and corresponding video to the optical-noise folder.
            # KR: 현재 분석 folder와 대응되는 video file을 optical-noise folder로 이동합니다.
            shutil.move(folder_path, destination_folder)
            shutil.move(video_path, destination_folder)
            continue
        except Exception as e:
            print(f"[WARN] Failed to move optical-noise sample: {e}")
            continue

    # =====================================================
    # Analysis ROI and tracked-point visualization
    # =====================================================
    # This figure provides a quality-control view of the selected analysis ROI
    # and tracked points on the maximum relaxation and maximum contraction frames.
    # The green point indicates the tracked point at maximum displacement,
    # the orange point indicates the baseline/minimum-displacement point,
    # and the red point indicates the initially selected analysis point.
    # The purple rectangle shows the local ROI used for optical-flow tracking.
    #
    # KR:
    # 이 figure는 maximum relaxation frame과 maximum contraction frame에서
    # 선택된 analysis ROI와 tracked point를 확인하기 위한 QC 시각화입니다.
    # 초록색 점은 maximum displacement 시점의 tracked point,
    # 주황색 점은 baseline/minimum-displacement 시점의 point,
    # 빨간색 점은 처음 선택된 analysis point를 의미합니다.
    # 보라색 사각형은 optical-flow tracking에 사용된 local ROI를 나타냅니다.
    # =====================================================

    best_x_roi_min, best_y_roi_min = tracked_points[start_number]

    # Convert ROI-local coordinates to full-image coordinates.
    # KR: ROI 내부 좌표를 원본 이미지 기준 좌표로 변환합니다.
    cx_max = safe_int(best_x_roi + x1_small_new_optical)
    cy_max = safe_int(best_y_roi + y1_small_new_optical)

    cx_min = safe_int(best_x_roi_min + x1_small_new_optical)
    cy_min = safe_int(best_y_roi_min + y1_small_new_optical)

    cx_selected = safe_int(x_selected_new)
    cy_selected = safe_int(y_selected_new)

    roi_top_left = (
        safe_int(x1_small_new_optical),
        safe_int(y1_small_new_optical)
    )
    roi_bottom_right = (
        safe_int(x2_small_new_optical),
        safe_int(y2_small_new_optical)
    )

    # Create visualization copies to avoid modifying the original analysis images.
    # KR: 원본 분석 이미지가 바뀌지 않도록 visualization용 복사본을 만듭니다.
    contraction_vis = square_differences_crop_indices_index_path_image.copy()
    relaxation_vis = square_image.copy()

    for vis_img in [contraction_vis, relaxation_vis]:

        # Maximum-displacement tracked point
        # KR: maximum displacement 시점의 tracked point
        cv2.circle(
            vis_img,
            (cx_max, cy_max),
            radius=3,
            color=(0, 255, 0),
            thickness=-1
        )

        # Baseline or minimum-displacement tracked point
        # KR: baseline 또는 minimum-displacement 시점의 tracked point
        cv2.circle(
            vis_img,
            (cx_min, cy_min),
            radius=3,
            color=(0, 159, 214),
            thickness=-1
        )

        # Initially selected analysis point
        # KR: 처음 선택된 analysis point
        cv2.circle(
            vis_img,
            (cx_selected, cy_selected),
            radius=3,
            color=(0, 0, 255),
            thickness=-1
        )

        cv2.rectangle(
            vis_img,
            roi_top_left,
            roi_bottom_right,
            (153, 102, 255),
            2
        )

    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    axes[0].imshow(cv2.cvtColor(relaxation_vis, cv2.COLOR_BGR2RGB))
    axes[0].set_title('Maximum Relaxation Frame', fontsize=20, fontweight='bold')
    axes[0].axis('off')

    axes[1].imshow(cv2.cvtColor(contraction_vis, cv2.COLOR_BGR2RGB))
    axes[1].set_title('Maximum Contraction Frame', fontsize=20, fontweight='bold')
    axes[1].axis('off')

    plt.tight_layout()
    Savefig('Analysis_ROI_and_Tracked_Points', 300, 0.1)


    
    # Compute summary statistics after excluding zero-count frames.
    # KR: pixel count가 0인 frame을 제외하고 moved-region 통계값을 계산합니다.
    valid_pixel_count_mean = np.mean([value for value in valid_pixel_counts if value not in [0, '']])
    valid_pixel_count_mean_round = round(valid_pixel_count_mean) if not np.isnan(valid_pixel_count_mean) else 0
    valid_pixel_count_maximum = max([value for value in valid_pixel_counts if value not in [0, '']], default=0)
    valid_pixel_count_maximum_round = round(valid_pixel_count_maximum) if not np.isnan(valid_pixel_count_maximum) else 0

    # Convert moved-region pixel counts to physical area.
    # KR: moved region의 pixel count를 실제 면적으로 변환합니다.
    valid_pixel_count_mean_area = valid_pixel_count_mean_round * AREA_MOVED
    valid_pixel_count_maximum_area = valid_pixel_count_maximum_round * AREA_MOVED
    valid_pixel_count_mean_area_round = round(valid_pixel_count_mean_area,2)
    valid_pixel_count_maximum_area_round = round(valid_pixel_count_maximum_area, 2)
    
    # Frame-wise moved-region area and length-equivalent values.
    # KR: 각 frame별 moved-region area와 length-equivalent 값을 계산합니다.

    displacement_list = [count * PIXEL_UM for count in valid_pixel_counts]

    # =========================================================
    # Moved-region area time-series visualization
    # =========================================================
    # This plot shows the frame-wise moved-region area estimated from
    # thresholded motion pixels. The moved-region area is calculated by
    # multiplying the number of thresholded pixels by the calibrated pixel area.
    # This output is used as an auxiliary quality-control visualization and is
    # separate from optical-flow displacement.
    #
    # KR:
    # 이 plot은 threshold 조건을 통과한 moved pixel 수를 기반으로
    # frame별 moved-region area 변화를 보여줍니다.
    # moved-region area는 thresholded pixel count에 pixel area 보정값을
    # 곱하여 계산됩니다.
    # 이 결과는 보조적인 QC 시각화이며, optical-flow displacement 값과는
    # 별개의 지표입니다.
    # =========================================================
    if EXPORT_EXTRA_FIGURES.lower() == 'on':
        fig, ax = plt.subplots(figsize=(8, 8)) 

        
        # Prepare moved-region area values.
        # KR: valid pixel count를 실제 면적으로 변환합니다.
        area_moved = [count * AREA_MOVED for count in valid_pixel_counts]

        # Draw bar and line plots for moved-region area.
        # KR: moved-region area를 bar plot과 line plot으로 함께 표시합니다.
        ax.bar(np.linspace(0, total_time_seconds, len(area_moved)), 
            area_moved, 
            align='center', 
            alpha=1, 
            width=0.125, 
            color=red_color)

        ax.plot(np.linspace(0, total_time_seconds, len(area_moved)), 
                area_moved, 
                marker='o', 
                color=blue_color, 
                linewidth=4, 
                markersize=8, 
                markeredgewidth=1.6)

        # Set axis labels.
        ax.set_xlabel("Time (s)", fontsize=40, fontweight='bold', fontname='Arial')
        ax.set_ylabel(f"Moved Region Area ({AREA_UNIT})", fontsize=35, fontweight='bold', fontname='Arial')
        

        # Increase plot border thickness.
        for spine in ax.spines.values():
            spine.set_linewidth(4)  # Increase plot border thickness.
            
        # Set the x-axis range to the full video duration.
        ax.set_xlim(0, total_time_seconds)
        
        ax.set_xticks(range(0, int(total_time_seconds) + 1, 5))

        # Configure axis ticks.
        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=10)
        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=10)

        # Keep the axis positions at the default origin.
        ax.spines['left'].set_position(('data', 0))
        ax.spines['bottom'].set_position(('data', 0))

        # Keep the top and right plot borders visible.
        ax.spines['right'].set_color('black')  
        ax.spines['top'].set_color('black')    

        # Use bold x-axis tick labels.
        plt.xticks(fontweight='bold')

        # Use bold y-axis tick labels.
        plt.yticks(fontweight='bold')

        Savefig(f'Moved Region Area', 300, 0.05)
        

        # =========================================================
        # Moved pixel count time-series visualization
        # =========================================================
        # This plot shows the number of pixels that passed the
        # intensity-difference threshold in each frame.
        # The maximum and mean moved-pixel counts are shown as dashed lines.
        # This output is used as a quality-control visualization for
        # pixel-intensity-based motion detection, and is separate from
        # optical-flow displacement.
        #
        # KR:
        # 이 plot은 각 frame에서 intensity-difference threshold를 통과한
        # moved pixel 수의 시간 변화를 보여줍니다.
        # 최대값과 평균값은 dashed line으로 표시됩니다.
        # 이 결과는 pixel-intensity 기반 motion detection의 QC용 시각화이며,
        # optical-flow displacement 값과는 별개의 보조 지표입니다.
        # =========================================================
        
        fig, ax = plt.subplots(figsize=(8, 8))  

        ax.plot(np.linspace(0, total_time_seconds, len(valid_pixel_counts)), 
                valid_pixel_counts, 
                marker='o', 
                color=gray_color, 
                linewidth=4, 
                markersize=8, 
                markeredgewidth=1.6)

        ax.set_xlabel("Time (s)", fontsize=40, fontweight='bold', fontname='Arial')
        ax.set_ylabel("Number of Moved Pixels", fontsize=35, fontweight='bold', fontname='Arial')

        for spine in ax.spines.values():
            spine.set_linewidth(4)  

        ax.set_xlim(0, total_time_seconds)
        ax.set_xticks(range(0, int(total_time_seconds) + 1, 5))
        
        ax.set_ylim(bottom=0) 
        ax.set_xlim(left=0)  

        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=10)
        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=10)

        ax.axhline(y=valid_pixel_count_maximum_round, color=red_color, linewidth=3.5, linestyle='--')
        ax.axhline(y=valid_pixel_count_mean_round, color=blue_color, linewidth=3.5, linestyle='--')

        ax.text(ax.get_xlim()[0] - 0.25, 
                valid_pixel_count_maximum_round, 
                f'{valid_pixel_count_maximum_round}', 
                ha='right', va='center', 
                color=red_color, 
                fontweight='bold', 
                fontsize=25)

        ax.text(ax.get_xlim()[0] - 0.25, 
                valid_pixel_count_mean_round, 
                f'{valid_pixel_count_mean_round}', 
                ha='right', va='center', 
                color=blue_color, 
                fontweight='bold', 
                fontsize=25)

        ax.spines['left'].set_position(('data', 0))
        ax.spines['bottom'].set_position(('data', 0))

        ax.spines['right'].set_color('black')  
        ax.spines['top'].set_color('black')    

        plt.xticks(fontweight='bold')

        plt.yticks(fontweight='bold')

        Savefig('Moved_Pixel_Count_Time_Series', 300, 0.05)

        # =========================================================
        # Optical-flow displacement time-series visualization
        # =========================================================
        # This section exports optical-flow displacement time-series plots
        # in several visualization formats. The displacement values are
        # calculated from the tracked point movement and converted from pixels
        # to micrometers using the calibrated PIXEL_UM value.
        #
        # maximum_displacement_list contains the frame-wise optical-flow
        # displacement values used for visualization and feature extraction.
        # target_displacements represents the selected displacement sequence
        # used for final analysis and A/IA/N classification.
        #
        # KR:
        # 이 section은 optical flow로 계산된 displacement time-series를
        # 여러 figure 형식으로 저장합니다. displacement 값은 tracked point의
        # 이동량을 기반으로 계산되며, PIXEL_UM 보정값을 이용해 pixel 단위에서
        # micrometer 단위로 변환됩니다.
        #
        # maximum_displacement_list는 frame별 optical-flow displacement 값을
        # 저장한 리스트이며, 시각화 및 feature extraction에 사용됩니다.
        # target_displacements는 최종 분석 및 A/IA/N classification에 사용되는
        # 선택된 displacement sequence입니다.
        # =========================================================
        
        # Export optional manuscript-style displacement plot.
        # KR: 논문 figure 제작용 displacement plot을 추가로 저장합니다.
        # =====================================================
        
        fig, ax = plt.subplots(figsize=(8, 8))

        ax.plot(np.linspace(0, total_time_seconds, len(maximum_displacement_list)), 
                maximum_displacement_list, 
                color=blue_color, 
                linewidth=4)

        ax.set_xlabel('Time (s)',  fontsize=40, fontweight='bold', fontname='Arial')
        ax.set_ylabel(f'Displacement ({LENGTH_UNIT})',  fontsize=40,fontweight='bold', fontname='Arial')

        ax.set_ylim(0, Y_LIM)

        ax.set_yticks(range(0, Y_LIM + 1, 1))
        
        for spine in ax.spines.values():
            spine.set_linewidth(4)
            
        ax.set_xlim(0, total_time_seconds)
        
        ax.set_xticks(range(0, int(total_time_seconds) + 1, X_TICKS))

        ax.axhline(y=max_target_displacements, color=red_color, linewidth=3.5, linestyle='--')


        ax.grid(False)
        plt.text(total_time_seconds - (total_time_seconds*0.0125), max_target_displacements + 0.625, f'Max Displacement ({LENGTH_UNIT}) : {max_target_displacements}',
                verticalalignment='top', horizontalalignment='right', color=red_color,fontweight='bold', fontsize=25)

        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=10)
        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=20)

        ax.spines['left'].set_position(('data', 0))  
        ax.spines['bottom'].set_position(('data', 0))  

        ax.spines['right'].set_color('black') 
        ax.spines['top'].set_color('black')    

        plt.xticks( fontweight='bold')

        plt.yticks( fontweight='bold')
        
        Savefig('Optical_Flow_Displacement', 300, 0.2)
        
        # =====================================================
        
        fig, ax = plt.subplots(figsize=(8, 8))

        ax.plot(np.linspace(0, total_time_seconds, len(maximum_displacement_list)), 
                maximum_displacement_list, 
                color=blue_color, 
                linewidth=4)
        
        ax.set_xlabel('Time (s)',  fontsize=50, fontname='Arial',fontweight='bold')
        ax.set_ylabel(f'Displacement ({LENGTH_UNIT})',  fontsize=50, fontname='Arial',fontweight='bold')

        ax.set_ylim(0, Y_LIM)
        
        ax.set_yticks(range(0, Y_LIM + 1, 1))
        
        for spine in ax.spines.values():
            spine.set_linewidth(4)  
            
        ax.set_xlim(0, total_time_seconds)
        
        ax.set_xticks(range(0, int(total_time_seconds) + 1, X_TICKS))

        # 격자 제거
        ax.grid(False)


        ax.tick_params(axis='y', direction='in', length=10, width=4, pad=10)
        ax.tick_params(axis='x', direction='in', length=10, width=4, pad=10)

        ax.spines['left'].set_position(('data', 0))  
        ax.spines['bottom'].set_position(('data', 0))  

        ax.spines['right'].set_color('black')  
        ax.spines['top'].set_color('black')    

        plt.xticks( fontsize=50,fontweight='bold')

        plt.yticks( fontsize=50,fontweight='bold')
        
        Savefig('Optical_Flow_Displacement_Time_Series_Narrow', 300, 0.2)
        
        # =====================================================
        fig, ax = plt.subplots(figsize=(8, 8))

        ax.plot(np.linspace(0, total_time_seconds, len(maximum_displacement_list)), 
                maximum_displacement_list, 
                color=blue_color, 
                linewidth=4)

        ax.set_ylim(0, Y_LIM)
        
        ax.set_yticks(range(0, Y_LIM + 1, 1)) 

        for spine in ax.spines.values():
            spine.set_linewidth(4)  

        ax.set_xlim(0, total_time_seconds)
        
        ax.set_xticks(range(0, int(total_time_seconds) + 1, 5))

        ax.grid(False)

        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=10)
        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=10)

        ax.spines['left'].set_position(('data', 0))  
        ax.spines['bottom'].set_position(('data', 0))  

        ax.spines['right'].set_color('black')  
        ax.spines['top'].set_color('black')    

        plt.xticks( fontweight='bold')

        plt.yticks( fontweight='bold')
        
        Savefig('Optical_Flow_Displacement_Time_Series_Paper', 300, 0.05)

        # =====================================================
        
        fig, ax = plt.subplots(figsize=(20, 8))

        ax.plot(np.linspace(0, total_time_seconds, len(target_displacements)), 
                target_displacements, 
                color=blue_color, 
                linewidth=4)

        ax.set_xlabel('Time (s)',  fontsize=50, fontname='Arial', fontweight='bold')
        ax.set_ylabel(f'Displacement ({LENGTH_UNIT})',  fontsize=50, fontname='Arial', fontweight='bold')

        ax.set_ylim(0, int(np.ceil(max_target_displacements)+1))
        
        ax.set_yticks(range(0, int(np.ceil(max_target_displacements)+1) + 1, 1))

        for spine in ax.spines.values():
            spine.set_linewidth(4)  
            
        ax.set_xlim(0, total_time_seconds)
        
        ax.set_xticks(range(0, int(total_time_seconds) + 1, X_TICKS))
        
        average_displacement_one_pixel = round(np.mean([value for value in maximum_displacement_list if value not in [0,'']]),2)
        ax.axhline(y=max_target_displacements, color=red_color, linewidth=3.5, linestyle='--')

        ax.grid(False)

        ax.text(total_time_seconds - (total_time_seconds * 0.0125), 
                max_target_displacements + (max_target_displacements*0.1),
                f'Max Displacement ({LENGTH_UNIT}) : {max_target_displacements}', 
                ha='right', va='top', color=red_color, fontweight='bold', fontsize=35)
        
        ax.tick_params(axis='y', direction='in', length=10, width=4, pad=10)
        ax.tick_params(axis='x', direction='in', length=10, width=4, pad=10)

        ax.spines['left'].set_position(('data', 0))  
        ax.spines['bottom'].set_position(('data', 0))  

        ax.spines['right'].set_color('black')  
        ax.spines['top'].set_color('black')   


        plt.xticks( fontsize=50,fontweight='bold')

        plt.yticks( fontsize=50,fontweight='bold')
        
        Savefig('Optical_Flow_Target_Displacement_Time_Series', 300, 0.2)
        
    else:
        fig, ax = plt.subplots(figsize=(8, 8))
        ax.plot(np.linspace(0, total_time_seconds, len(maximum_displacement_list)), 
                maximum_displacement_list, 
                color=blue_color, 
                linewidth=4)

        ax.set_xlabel('Time (s)',  fontsize=40, fontweight='bold', fontname='Arial')
        ax.set_ylabel(f'Displacement ({LENGTH_UNIT})',  fontsize=40,fontweight='bold', fontname='Arial')

        ax.set_ylim(0, Y_LIM)
        
        ax.set_yticks(range(0, Y_LIM + 1, 1))
        
        for spine in ax.spines.values():
            spine.set_linewidth(4) 
            
        ax.set_xlim(0, total_time_seconds)
        
        ax.set_xticks(range(0, int(total_time_seconds) + 1, X_TICKS))

        ax.axhline(y=max_target_displacements, color=red_color, linewidth=3.5, linestyle='--')

        ax.grid(False)
        plt.text(total_time_seconds - (total_time_seconds*0.0125), max_target_displacements + 0.625, f'Max Displacement ({LENGTH_UNIT}) : {max_target_displacements}',
                verticalalignment='top', horizontalalignment='right', color=red_color,fontweight='bold', fontsize=25)

        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=10)
        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=20)

        ax.spines['left'].set_position(('data', 0))  
        ax.spines['bottom'].set_position(('data', 0))  


        ax.spines['right'].set_color('black') 
        ax.spines['top'].set_color('black')    


        plt.xticks( fontweight='bold')
        plt.yticks( fontweight='bold')
        
        Savefig('Optical_Flow_Displacement', 300, 0.2)
    
    
    
    # ============================================================
    # MyoCAS composite video generation
    # ============================================================
    # This section generates a composite video containing the original
    # phase-contrast video, optical-flow overlay, displacement time-series plot,
    # and analysis ROI video. Video export is performed only when VIDEO_MODE is "on".
    #
    # KR:
    # 이 section은 원본 phase-contrast video, optical-flow overlay,
    # displacement time-series plot, analysis ROI video를 결합한
    # MyoCAS composite video를 생성합니다.
    # VIDEO_MODE가 "on"일 때만 video export를 수행합니다.
    # ============================================================

    if VIDEO_MODE.lower() == 'on':

        print("[INFO] Generating MyoCAS composite video...")

        # ------------------------------------------------------------
        # # Input and output paths
        # ------------------------------------------------------------
        displacement_image_path = os.path.join(
            result_folder,
            f'Optical_Flow_Displacement.{FRAME_EXTENSION}'
        )

        base_video_path     = fps_video_path          # original phase contrast video
        optical_video_path  = output_video_path       # optical flow overlay video
        analysis_video_path = os.path.join(result_folder, video_name)

        # ------------------------------------------------------------
        # Layout and size settings
        # ------------------------------------------------------------
        FINAL_SIZE = (1920, 1080)

        MAIN_VIDEO_SIZE = (1280, 1080)

        RIGHT_PANEL_X = 1280

        DISPLACEMENT_SIZE = (640, 690)
        ANALYSIS_VIDEO_SIZE = (390, 390)

        # ------------------------------------------------------------
        # Load source videos
        # ------------------------------------------------------------
        base_video_clip = VideoFileClip(base_video_path)
        optical_video_clip = VideoFileClip(optical_video_path)

        optical_video_position = (x1_small_new_optical, y1_small_new_optical)

        analysis_video_clip = (
            VideoFileClip(analysis_video_path)
            .resize(ANALYSIS_VIDEO_SIZE)
        )
        analysis_x = RIGHT_PANEL_X + (DISPLACEMENT_SIZE[0] - ANALYSIS_VIDEO_SIZE[0]) // 2
        analysis_y = DISPLACEMENT_SIZE[1] + (FINAL_SIZE[1] - DISPLACEMENT_SIZE[1] - ANALYSIS_VIDEO_SIZE[1]) // 2

        analysis_video_clip = analysis_video_clip.set_position(
            (analysis_x, analysis_y)
        )
        # ------------------------------------------------------------
        # Composite main video
        # ------------------------------------------------------------
        composite_main_clip = (
            CompositeVideoClip([
                base_video_clip,
                optical_video_clip.set_position(optical_video_position)
            ])
            .resize(MAIN_VIDEO_SIZE)
        )

        video_duration = composite_main_clip.duration

        # ------------------------------------------------------------
        # Displacement plot clip
        # ------------------------------------------------------------
        displacement_clip = (
            ImageClip(displacement_image_path)
            .set_duration(video_duration)
            .resize(DISPLACEMENT_SIZE)
            .set_position((RIGHT_PANEL_X, 0))
        )

        # ------------------------------------------------------------
        # Time-synchronized vertical bar
        # ------------------------------------------------------------
        DISP_X_START = 1386      # start x of displacement plot
        DISP_X_RANGE = 496       # width corresponding to full duration

        def create_red_bar(duration):
            def position_func(t):
                bar_x = DISP_X_START + t * (DISP_X_RANGE / duration)
                return (bar_x, 32)

            return (
                ColorClip(size=(4, 542), color=(255, 102, 153))
                .set_position(position_func)
                .set_duration(duration)
            )

        red_bar = create_red_bar(video_duration)

        # ------------------------------------------------------------
        # Final video composition
        # ------------------------------------------------------------
        final_clip = CompositeVideoClip(
            [
                composite_main_clip.set_position((0, 0)),
                displacement_clip,
                analysis_video_clip,   
                red_bar
            ],
            size=FINAL_SIZE
        )


        # ------------------------------------------------------------
        # Export
        # ------------------------------------------------------------
        final_output_path = os.path.join(
            result_folder,
            f"{folder_name}_MyoCAS_Composite_Video.{extension}"
        )

        final_clip.write_videofile(
            final_output_path,
            codec="libx264",
            fps=fps,
            logger=None
        )

        print("[INFO] MyoCAS composite video export completed.")
    
    if FRAME_REMOVE.lower() == 'on':
        paths_to_clean = [
            (folder_path, "result"),  
        ]
        
        for base_path, exclude_name in paths_to_clean:
            if os.path.exists(base_path): 
                for item_name in os.listdir(base_path):
                    item_path = os.path.join(base_path, item_name)
                    if item_name != exclude_name:  
                        if os.path.isfile(item_path):
                            os.remove(item_path)
                        elif os.path.isdir(item_path):
                            shutil.rmtree(item_path)

    
    band_energy_optical_displacement, max_value_optical_displacement_band_energy, max_optical_amp_at_max_freq, optical_freq_at_max_amp_number  = plot_fft(target_displacements, X_LIM_AMPLITUDE, Y_LIM_AMPLITUDE,'Optical_Flow_Displacement_FFT')


    classification_dir = os.path.join(VIDEO_FILE_PATH, 'classification')
    if not os.path.exists(classification_dir):
        os.makedirs(classification_dir, exist_ok=True)
        
    folder_path = os.path.join(VIDEO_FILE_PATH, folder_name)
    video_path = os.path.join(VIDEO_FILE_PATH, f'{folder_name}.{extension}')
    
    # ======================================================
    # Scalar feature extraction for Excel export and A/IA/N classification
    # ======================================================
    # This section computes displacement-derived scalar features from the
    # optical-flow displacement time series. These features are saved to Excel
    # and used together with the displacement sequence for A/IA/N classification.
    #
    # KR:
    # 이 section은 optical-flow displacement time series에서 scalar feature를 계산합니다.
    # 계산된 feature는 Excel output에 저장되며, displacement sequence와 함께
    # A/IA/N classification 입력으로 사용됩니다.
    # ======================================================
    seq_raw_np = np.array(maximum_displacement_list, dtype=np.float32)

    if len(seq_raw_np) == 0:
        max_abs_disp = 0.0
        abs_net_change_ratio = 0.0
        sign_changes = 0
        diff_std = 0.0
    else:
        max_abs_disp = float(np.max(np.abs(seq_raw_np)))

        peak_to_peak = float(np.max(seq_raw_np) - np.min(seq_raw_np))

        abs_net_change_ratio = float(
            np.abs(seq_raw_np[-1] - seq_raw_np[0]) / (peak_to_peak + 1e-8)
        )

        diff_seq = np.diff(seq_raw_np)
        if len(diff_seq) >= 2:
            sign_changes = int(np.sum(np.diff(np.sign(diff_seq)) != 0))
        else:
            sign_changes = 0

        diff_std = float(np.std(diff_seq)) if len(diff_seq) > 0 else 0.0

    
    # Assemble one row of analysis results for Excel export.
    # KR: Excel output에 저장할 분석 결과 한 줄을 구성합니다.
    Excel_result_data = [
        folder_name,
        adapt_max,
        segmentation_area,
        max_target_displacements,
        band_energy_optical_displacement,
        max_value_optical_displacement_band_energy,
        x_selected_new,
        y_selected_new,
        optical_freq_at_max_amp_number,
        maximum_displacement_str,
        max_abs_disp,
        abs_net_change_ratio,
        sign_changes,
        diff_std
    ]
    # ======================================================
    # A/IA/N classification
    # ======================================================
    # When CLASSIFICATION_MODE is "on", the trained classifier predicts whether
    # the analyzed region is inactive, active, or noise using the displacement
    # sequence and scalar features.
    #
    # KR:
    # CLASSIFICATION_MODE가 "on"이면 학습된 classifier를 사용하여 분석 영역을
    # inactive, active, noise 중 하나로 분류합니다.
    # 입력으로 displacement sequence와 scalar feature를 함께 사용합니다.
    # ======================================================
    if CLASSIFICATION_MODE.lower() == 'on':

        if len(seq_raw_np) == 0:
            print("[WARN] maximum_displacement_list is empty. Treating this sample as NOISE.")
            predicted_class = 2
            predicted_class2 = 2
            
        elif max_optical_amp_at_max_freq == 0:
            print("[WARN] No dominant optical-flow frequency was detected. Treating this sample as NOISE.")
            predicted_class = 2
            predicted_class2 = 2

        else:
            scalar_values = [
                np.log1p(float(band_energy_optical_displacement)),
                float(max_value_optical_displacement_band_energy),
                max_abs_disp,
                abs_net_change_ratio,
                float(sign_changes),
                diff_std
            ]

            predicted_class, probs_np = predict_single_sample_torch(
                classification_model,
                maximum_displacement_list,
                scalar_values,
                device
            )
            print(f"Maximum displacement: {max_target_displacements} {LENGTH_UNIT}")
            print(f"Predicted class: {CLASS_NAMES[predicted_class]}")
            print(f"Prediction probabilities: {probs_np}")
            # Optional second prediction pass.
            # In the current implementation, this uses the same input sequence and scalar
            # features as the first prediction. Keep this option "off" unless a custom
            # second-check routine is added.
            #
            # KR:
            # 선택적 second prediction 단계입니다.
            # 현재 구현에서는 첫 번째 예측과 동일한 displacement sequence 및 scalar feature를 사용합니다.
            # 별도의 custom double-check logic을 추가하지 않았다면 "off"로 두는 것이 좋습니다.

            if DOUBLE_CHECK_DISPLACEMENT.lower() == 'on':
                predicted_class2, probs_np2 = predict_single_sample_torch(
                    classification_model,
                    maximum_displacement_list,
                    scalar_values,
                    device
                )
            else:
                predicted_class2 = predicted_class

        # ======================================================
        # Classification result handling and output sorting
        # ======================================================
        # Based on the predicted A/IA/N class, this section appends the result row
        # to the corresponding Excel file and moves the analyzed folder/video to
        # the appropriate classification output folder.
        #
        # KR:
        # 예측된 A/IA/N class에 따라 결과를 해당 Excel 파일에 저장하고,
        # 분석된 folder/video를 classification 결과별 output folder로 이동합니다.
        # ======================================================

        if predicted_class == 1 and predicted_class2 == 1:

            Hz_folder = f'{max_optical_amp_at_max_freq} Hz'
            destination_folder = os.path.join(
                classification_dir,
                STRONG_SAMPLE_FOLDER_NAME,
                Hz_folder
            )

            print("[INFO] Classification result: ACTIVE")
            good_ws.append(Excel_result_data)
            good_wb.save(strong_sample_excel_path)

            handle_folder_video(
                folder_path,
                video_path,
                destination_folder,
                remove_noise_flag=False
            )

        elif predicted_class in [0, 1] and predicted_class2 in [0, 1]:

            destination_folder = os.path.join(
                classification_dir,
                WEAK_SAMPLE_FOLDER_NAME
            )

            print("[INFO] Classification result: INACTIVE")
            weak_ws.append(Excel_result_data)
            weak_wb.save(weak_sample_excel_path)

            handle_folder_video(
                folder_path,
                video_path,
                destination_folder,
                remove_noise_flag=INACTIVE_NOISE_FRAME_REMOVE.lower() == 'on'
            )

        else:

            destination_folder = os.path.join(
                classification_dir,
                NOISE_SAMPLE_FOLDER_NAME
            )

            print("[INFO] Classification result: NOISE")
            noise_ws.append(Excel_result_data)
            noise_wb.save(noise_sample_excel_path)

            handle_folder_video(
                folder_path,
                video_path,
                destination_folder,
                remove_noise_flag=INACTIVE_NOISE_FRAME_REMOVE.lower() == 'on'
            )

    else:
        # If classifier mode is disabled, save the result to the check folder.
        # KR: classifier mode가 꺼져 있으면 결과를 check folder에 저장합니다.
        destination_folder = os.path.join(
            classification_dir,
            CHECK_SAMPLE_FOLDER_NAME
        )

        check_ws.append(Excel_result_data)
        check_wb.save(check_sample_excel_path)

        handle_folder_video(
            folder_path,
            video_path,
            destination_folder,
            remove_noise_flag=INACTIVE_NOISE_FRAME_REMOVE.lower() == 'on'
        )

    # Record processing time for the current folder.
    # KR: 현재 folder의 처리 시간을 기록합니다.
    folder_end_time = time.time()
    

    folder_execution_time = round((folder_end_time - folder_start_time),2)
    estimated_execution_times.append(folder_execution_time)  

    # Estimate the remaining batch-processing time.
    # KR: 전체 batch analysis의 예상 남은 시간을 계산합니다.
    average_execution_time = round(sum(estimated_execution_times) / len(estimated_execution_times),2)
    estimated_left_time = round((((average_execution_time * number_of_folders) - sum(estimated_execution_times))/60),1)
    
    print(f"\n[INFO] Folder completed: {folder_name}")
    print(f"[INFO] Processing time: {folder_execution_time} seconds")
    print(f"[INFO] Estimated remaining time: {estimated_left_time} minutes\n")
    
    # ======================================================
    # Full memory cleanup after each video/folder
    # ======================================================
    # This block releases large arrays, figures, and GPU memory after each
    # analyzed folder to reduce memory accumulation during batch processing.
    #
    # KR:
    # 이 block은 각 folder/video 분석이 끝난 뒤 큰 배열, figure, GPU memory를
    # 정리하여 batch processing 중 memory 누적을 줄이기 위한 단계입니다.
    # ======================================================

    print("\nFULL MEMORY CLEANUP START")

    try:
       
        plt.close('all')
        plt.clf()
        plt.cla()
    except:
        pass

    try:
        for scope in (locals(), globals()):
            for name, val in list(scope.items()):
                if isinstance(val, np.ndarray):
                    try:
                        del scope[name]
                    except:
                        pass
    except:
        pass


    try:
        torch.cuda.empty_cache()
    except:
        pass

    try:
        cv2.destroyAllWindows()
    except:
        pass

    try:
        cap.release()
    except:
        pass

    gc.collect()

    print("MEMORY CLEANUP DONE\n")

plt.close()
good_wb.close()
weak_wb.close()
noise_wb.close()