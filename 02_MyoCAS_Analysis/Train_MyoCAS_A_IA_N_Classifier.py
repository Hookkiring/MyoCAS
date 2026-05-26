### MyoCAS (Myotubes Contraction Analysis System) ###

### 25-03-23 수정 ###
import cv2
import os
import numpy as np
import matplotlib.pyplot as plt
import shutil
from mpl_toolkits.mplot3d import Axes3D
import time
from openpyxl import Workbook, load_workbook
import natsort
import math
import gc
import pandas as pd
from sklearn.model_selection import train_test_split
import glob
from sklearn.metrics import confusion_matrix
import itertools
from io import BytesIO
import datetime
from scipy.fft import fft, fftfreq
from matplotlib import colors
import plotly.graph_objects as go
import matplotlib.ticker as ticker
from matplotlib import rcParams
import pywt
import torch
import torch.nn as nn
import torch.nn.functional as F
import albumentations as A
from albumentations.pytorch import ToTensorV2
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm
import multiprocessing
from moviepy.editor import VideoFileClip, CompositeVideoClip, ImageClip, ColorClip
import ast    
from torch.utils.data import Dataset, DataLoader

# =========================================================
# USER SETTINGS
# =========================================================

# =========================================================
# Execution mode
# =========================================================
# "TRAIN":
#     Train the MyoCAS A/IA/N classifier using the TRAIN and VAL datasets.
#     Model checkpoints and training logs are saved during training.
#
# "TEST":
#     Evaluate saved checkpoints using the TEST dataset.
#     This mode is useful for comparing multiple trained checkpoints and
#     selecting the best-performing model.
#
# "PREDICTION":
#     Load BEST_CHECKPOINT_FILE and evaluate prediction performance on the
#     TEST dataset without training a new model.
#
# "CLASSIFICATION":
#     Load BEST_CHECKPOINT_FILE and classify samples into ACTIVE, INACTIVE,
#     or NOISE output folders.
#
# Note:
#     Keep MODE = "TRAIN" when training a new A/IA/N classifier.
#     Use the other modes only when the corresponding evaluation,
#     prediction, or classification logic is implemented below.
#
# KR:
# 실행 모드를 설정합니다.
#
# "TRAIN":
#     TRAIN/VAL dataset을 이용해 MyoCAS A/IA/N classifier를 학습합니다.
#     학습 중 checkpoint와 log가 저장됩니다.
#
# "TEST":
#     TEST dataset을 이용해 저장된 여러 checkpoint를 평가합니다.
#     여러 epoch/checkpoint의 성능을 비교하여 가장 좋은 model을 고를 때 사용합니다.
#
# "PREDICTION":
#     BEST_CHECKPOINT_FILE을 불러와 새로 학습하지 않고 TEST dataset에 대한
#     예측 성능을 평가합니다.
#
# "CLASSIFICATION":
#     BEST_CHECKPOINT_FILE을 불러와 sample을 ACTIVE, INACTIVE, NOISE
#     output folder로 분류합니다.
#
# 주의:
#     새 A/IA/N classifier를 학습할 때는 MODE = "TRAIN"으로 둡니다.
#     다른 mode는 아래에 해당 evaluation, prediction, classification logic이
#     구현되어 있을 때만 사용하세요.
MODE = 'Test'

# =========================================================
# DATASET AND OUTPUT PATH SETTINGS
# =========================================================

# Root directory containing TRAIN, VAL, and TEST folders.
# Expected structure:
#   dataset/
#       TRAIN/
#           ACTIVE/
#           INACTIVE/
#           NOISE/
#       VAL/
#           ACTIVE/
#           INACTIVE/
#           NOISE/
#       TEST/
#           ACTIVE/
#           INACTIVE/
#           NOISE/
REFERENCE_FILE_PATH = r"C:\Users\Hyeon Jun\Desktop\test2\test\github test\Train_results"

TRAIN_FILE_PATH = os.path.join(REFERENCE_FILE_PATH, 'TRAIN')
VAL_FILE_PATH = os.path.join(REFERENCE_FILE_PATH, 'VAL')
TEST_FILE_PATH = os.path.join(REFERENCE_FILE_PATH, 'TEST')

# Output directory for model checkpoints and logs.
CHECKPOINT_DIR = os.path.join(REFERENCE_FILE_PATH, 'result', 'checkpoints')
current_time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
log_dir = os.path.join(REFERENCE_FILE_PATH, 'result', 'logs', current_time)

# Checkpoint file used when continuing training.
# This is used only when KEEP_TRAINING is "on".
BEST_CHECKPOINT_FILE = r"weights/MyoCAS_AIN_LSTM_classifier_checkpoint.pth"


# =========================================================
# TRAINING SETTINGS
# =========================================================

EPOCHS = 100
BATCH_SIZE = 8
LEARNING_RATE = 0.001

# Continue training from BEST_CHECKPOINT_FILE.
# "on": load BEST_CHECKPOINT_FILE and continue training.
# "off": train from scratch.
KEEP_TRAINING = 'on'

# GPU acceleration mode.
# "on": use CUDA-enabled GPU acceleration when available.
# "off": force CPU-based processing.
# If CUDA is not available, the script automatically falls back to CPU.
GPU_MODE = 'on'


# =========================================================
# MYOCAS ANALYSIS PARAMETERS USED FOR FEATURE EXTRACTION
# =========================================================

# Half-size of the square ROI used for pixel-intensity-based motion search.
# This should match the value used in Run_MyoCAS_Analysis.py when generating
# training samples.
ANALYSIS_SQUARE_SIZE_REFERENCE = 9

# Pixel-to-micrometer calibration factor.
# This should match the microscope setting used for the analysis data.
PIXEL_UM = 0.78125

# Area represented by one pixel, calculated from PIXEL_UM.
AREA_MOVED = round((PIXEL_UM * PIXEL_UM), 4)

# Initial analysis window in seconds.
# The actual number of frames is calculated as fps × NUM_FRAMES.
NUM_FRAMES = 4

# Reference contraction/stimulation frequency in Hz.
# The dominant frequency is estimated from the displacement signal during FFT analysis;
# this value is used as a fallback/reference when needed.
Hz = 1

# Fallback video frame rate.
# The actual FPS is automatically read from each input video when available.
VIDEO_FPS = 30


# =========================================================
# FEATURE AND CLASS SETTINGS
# =========================================================

# Number of output classes for A/IA/N classification.
NB_CLASSES = 3

# Scalar features used by the fully connected scalar branch.
# The displacement time-series itself is processed separately by the BiLSTM branch.
COLUMNS = [
    'MAX BAND ENERGY',
    'MAX Apeak',
    'MAX ABS DISP',
    'ABS NET CHANGE RATIO',
    'NUM SIGN CHANGES DIFF',
    'DIFF STD'
]

NB_INPUT = len(COLUMNS)

# Class labels used by the classifier.
CLASS_NAMES = ['INACTIVE', 'ACTIVE', 'NOISE']


# =========================================================
# DATASET CLASS FOLDERS
# =========================================================

BAD_FILE_PATH_TRAIN = os.path.join(TRAIN_FILE_PATH, 'INACTIVE')
TARGET_FILE_PATH_TRAIN = os.path.join(TRAIN_FILE_PATH, 'ACTIVE')
NOISE_FILE_PATH_TRAIN = os.path.join(TRAIN_FILE_PATH, 'NOISE')

BAD_FILE_PATH_VAL = os.path.join(VAL_FILE_PATH, 'INACTIVE')
TARGET_FILE_PATH_VAL = os.path.join(VAL_FILE_PATH, 'ACTIVE')
NOISE_FILE_PATH_VAL = os.path.join(VAL_FILE_PATH, 'NOISE')

BAD_FILE_PATH_TEST = os.path.join(TEST_FILE_PATH, 'INACTIVE')
TARGET_FILE_PATH_TEST = os.path.join(TEST_FILE_PATH, 'ACTIVE')
NOISE_FILE_PATH_TEST = os.path.join(TEST_FILE_PATH, 'NOISE')


# =========================================================
# FILE, OUTPUT, AND VISUALIZATION SETTINGS
# =========================================================

FRAME_PREFIX = 'frame_'
VIDEO_FILE_EXTENSION = ['mp4', 'mov', 'wmv', 'mkv', 'mpeg', 'flv', 'webm', 'avi']
FRAME_EXTENSION = 'png'
VIDEO_CODEX = 'mp4v'

PARAMETER_FILE = 'PARAMETER.xlsx'

BAD_SAMPLE_EXCEL_NAME = 'INACTIVE_SAMPLE.xlsx'
TARGET_SAMPLE_EXCEL_NAME = 'ACTIVE_SAMPLE.xlsx'
NOISE_SAMPLE_EXCEL_NAME = 'NOISE_SAMPLE.xlsx'

bad_sample_excel_path_train = os.path.join(BAD_FILE_PATH_TRAIN, BAD_SAMPLE_EXCEL_NAME)
target_sample_excel_path_train = os.path.join(TARGET_FILE_PATH_TRAIN, TARGET_SAMPLE_EXCEL_NAME)
noise_sample_excel_path_train = os.path.join(NOISE_FILE_PATH_TRAIN, NOISE_SAMPLE_EXCEL_NAME)

bad_sample_excel_path_val = os.path.join(BAD_FILE_PATH_VAL, BAD_SAMPLE_EXCEL_NAME)
target_sample_excel_path_val = os.path.join(TARGET_FILE_PATH_VAL, TARGET_SAMPLE_EXCEL_NAME)
noise_sample_excel_path_val = os.path.join(NOISE_FILE_PATH_VAL, NOISE_SAMPLE_EXCEL_NAME)

bad_sample_excel_path_test = os.path.join(BAD_FILE_PATH_TEST, BAD_SAMPLE_EXCEL_NAME)
target_sample_excel_path_test = os.path.join(TARGET_FILE_PATH_TEST, TARGET_SAMPLE_EXCEL_NAME)
noise_sample_excel_path_test = os.path.join(NOISE_FILE_PATH_TEST, NOISE_SAMPLE_EXCEL_NAME)


# =========================================================
# OPTIONAL ANALYSIS / OUTPUT MODES
# =========================================================

# Manual region selection mode.
# Usually "off" for training-data processing.
MANUAL_SELECTION = 'off'

# Frame removal mode.
# "on": remove or reorganize frame folders after analysis.
# "off": keep extracted frames.
FRAME_REMOVE = 'on'

# Reverse contraction direction mode.
# "on": reverse displacement direction when needed.
# "off": keep the displacement direction as calculated.
REVERSE_CONTRACTION = 'off'

# Analysis video export mode.
# Usually "off" for classifier training to reduce file size and processing time.
VIDEO_MODE = 'off'

# Frame extraction skip mode.
# "on": reuse existing extracted frames if validation passes.
# "off": always extract frames again.
FRAME_EXTRACTION_SKIP = 'on'


# =========================================================
# PLOT SETTINGS
# =========================================================

Y_LIM = 8
X_LIM_AMPLITUDE = 3
Y_LIM_AMPLITUDE = 500
X_TICKS = 5  

LENGTH_UNIT = 'μm'
AREA_UNIT = 'μm²'

gray_color = '#bbbbbd'
red_color = '#FF6699'
blue_color = '#009fd6'
yellow_color = '#d69f00'

# ==========================================================
# Folder / Video cleanup handler
# ==========================================================
def handle_folder_video(folder_path, video_path, destination_folder, remove_noise_flag):
    try:
        os.makedirs(destination_folder, exist_ok=True)

        if remove_noise_flag:
            # 폴더 삭제 + 영상 이동
            if os.path.exists(folder_path):
                shutil.rmtree(folder_path, ignore_errors=True)

            if os.path.exists(video_path):
                shutil.move(video_path, destination_folder)

        else:
            # 둘 다 이동
            if os.path.exists(folder_path):
                shutil.move(folder_path, destination_folder)

            if os.path.exists(video_path):
                shutil.move(video_path, destination_folder)

    except Exception as e:
        print(f"[WARN] Folder/Video 처리 실패: {e}")



# ==========================================================
# Classification Pipeline (Refactored Full Block)
# ==========================================================

def preprocess_single_sample(x_data):
    x_data[0][0] *= 100
    return x_data.astype(np.float32)


        
# Save figure to the result folder
# Save figure to the result folder of the current sample.
def Savefig(folder_path, name, dpi, pad_inches):
    result_folder = os.path.join(folder_path, 'result')
    os.makedirs(result_folder, exist_ok=True)

    result_image_path = os.path.join(
        result_folder,
        f'{name}.{FRAME_EXTENSION}'
    )

    plt.savefig(
        result_image_path,
        dpi=dpi,
        bbox_inches='tight',
        pad_inches=pad_inches
    )
    plt.close()




def extract_frames(base_path, video_file):

    video_path = os.path.join(base_path, video_file)
    folder_name = os.path.splitext(video_file)[0]
    output_dir = os.path.join(base_path, folder_name)

    # ---------------------------------------------------------
    # SKIP MODE CHECK + Frame Validation 추가
    # ---------------------------------------------------------
    if FRAME_EXTRACTION_SKIP.lower() == 'on' and os.path.exists(output_dir):

        cap = cv2.VideoCapture(video_path, cv2.CAP_FFMPEG)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 2)

        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        cap.release()
        del cap

        saved_frames = sum(
            1 for f in os.scandir(output_dir)
            if f.name.lower().endswith("." + FRAME_EXTENSION)
        )

        if total_frames > 0 and saved_frames == total_frames:
            print(f"🔎 Frame validation check... ({video_file})")

            try:
                sample_frames = sorted(os.listdir(output_dir))[:5]
                for sf in sample_frames:
                    img = cv2.imread(os.path.join(output_dir, sf))
                    if img is None:
                        raise Exception("Frame corrupted")
                    del img

                print(f"✔ Frame validation OK: {video_file}")
                return saved_frames

            except:
                print(f"⚠ Frame validation failed → re-extract: {video_file}")

        print(f"⚠ Frame mismatch → re-extract: {video_file}")

    # ---------------------------------------------------------
    # FRAME EXTRACTION
    # ---------------------------------------------------------
    os.makedirs(output_dir, exist_ok=True)

    cap = cv2.VideoCapture(video_path)
    frame_number = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        image_filename = os.path.join(
            output_dir,
            f"{FRAME_PREFIX}{frame_number:04d}.{FRAME_EXTENSION}"
        )

        cv2.imwrite(
            image_filename,
            frame,
            [cv2.IMWRITE_PNG_COMPRESSION, 1]
        )

        del frame
        if frame_number % 100 == 0:
            gc.collect()

        frame_number += 1

    cap.release()
    del cap

    print(f"Video {video_file}: extracted {frame_number} frames")
    return frame_number



# ---------- BatchNorm → GroupNorm ----------
def convert_batchnorm_to_groupnorm(module):
    for name, child in module.named_children():
        if isinstance(child, torch.nn.BatchNorm2d):
            setattr(module, name, torch.nn.GroupNorm(32, child.num_features))
        else:
            convert_batchnorm_to_groupnorm(child)

# =========================================================
# MULTI EXTRACTION RUN
# =========================================================
print("\n Parallel Frame Extraction START\n")



def plot_fft( folder_path, fps,
    data_list, X_LIM_AMPLITUDE = X_LIM_AMPLITUDE, Y_LIM_AMPLITUDE = Y_LIM_AMPLITUDE,
    save_prefix='save_prefix'):
    sampling_rate = fps

    # FFT
    fft_values = fft(data_list)
    frequencies = fftfreq(len(data_list), d=1/sampling_rate)

    positive_frequencies = frequencies[:len(frequencies)//2]
    positive_fft_values = np.abs(fft_values[:len(fft_values)//2])
    
    # Full frequency band above 0.5 Hz
    total_band = (positive_frequencies > 0.5)

    # Frequencies and FFT amplitudes within the selected band
    freqs_band = positive_frequencies[total_band]
    amps_band = positive_fft_values[total_band]

    # Find the frequency with the maximum amplitude
    max_idx = np.argmax(amps_band)  # Index of the maximum amplitude
    freq_at_max_amp = freqs_band[max_idx]  # Frequency at maximum amplitude
    freq_at_max_amp_number = round(freqs_band[max_idx],2)  
    amp_at_max_freq = amps_band[max_idx]    # Amplitude at the selected frequency
    if freq_at_max_amp < Hz-0.1:
        freq_at_max_amp = np.floor(Hz-0.1)
    else:
        freq_at_max_amp = round(freq_at_max_amp)
    half_width = 0.5  # ±0.5 Hz frequency window
    lower_bound = round(max(freq_at_max_amp - half_width, 0.5),2)
    upper_bound = round(freq_at_max_amp + half_width,2)

    target_band = (positive_frequencies >= lower_bound) & (positive_frequencies <= upper_bound)
    energy_peak_window = np.sum(positive_fft_values[target_band] ** 2)
    band_energy = round(energy_peak_window)  

    # Round and print FFT summary values
    freq_at_max_amp_round = round(freq_at_max_amp,2)
    amp_at_max_freq_round = round(amp_at_max_freq)
  
    print(f'Max amplitude {amp_at_max_freq_round} occurs at {freq_at_max_amp_number} Hz | Energy {band_energy} | Range {lower_bound} ~ {upper_bound}')
    

    # ---------- First plot ----------
    fig, ax = plt.subplots(figsize=(10, 10))
    ax.fill_between(positive_frequencies, 0, positive_fft_values, where=target_band, color=yellow_color, alpha=0.5)
    ax.plot(positive_frequencies, positive_fft_values, color=red_color, linewidth=4, zorder=10)
    ax.set_xlabel("Frequency (Hz)", fontweight='bold', fontsize=50)
    ax.set_ylabel("Amplitude", fontweight='bold', fontsize=50)
    ax.grid(False)
    ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=20)
    ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=20)
    ax.set_xlim(0, max(positive_frequencies)+1)
    ax.set_xticks(np.arange(0, max(positive_frequencies)+1, 1))
    ax.set_ylim(0, max(positive_fft_values) + 50)
    ax.set_yticks(np.arange(0, max(positive_fft_values) + 50, 100))
    for spine in ax.spines.values():
        spine.set_linewidth(4)
        spine.set_color('black')
    ax.spines['left'].set_position(('outward', 0))
    ax.spines['bottom'].set_position(('outward', 0))
    plt.xticks(fontweight='bold')
    plt.yticks(fontweight='bold')
    ax.text(0.95, 0.95, f'E [{lower_bound , upper_bound}]: {band_energy}', ha='right', va='top', transform=ax.transAxes, fontsize=20, color=yellow_color , weight='bold')
    ax.text(0.9, 0.9, f'Apeak: {amp_at_max_freq_round}', ha='right', va='top', transform=ax.transAxes, fontsize=20, color=blue_color , weight='bold')
    ax.axhline(y=amp_at_max_freq_round, color=blue_color, linestyle='--', linewidth=2)

    plt.tight_layout()
    Savefig(folder_path, f'{save_prefix}', 300, 0.05)
    
    return band_energy, amp_at_max_freq_round, freq_at_max_amp_round, freq_at_max_amp_number



rcParams['font.family'] = 'Arial'
plt.rcParams['figure.dpi'] = 300

    
def TRAIN_MODE(BAD_FILE_PATH, TARGET_FILE_PATH, NOISE_FILE_PATH, BAD_EXCEL, TARGET_EXCEL, NOISE_EXCEL):
    print('============================ INACTIVE FILE START ============================')
    video_files = []
    
    
    if not os.path.exists(BAD_FILE_PATH):
        os.makedirs(BAD_FILE_PATH, exist_ok=True)

    # 먼저 비디오 파일 목록 수집
    for filename in natsort.natsorted(os.listdir(BAD_FILE_PATH)):
        for extension in VIDEO_FILE_EXTENSION:
            if filename.endswith(f'.{extension}'):
                video_files.append(filename)
                break

    cpu_count = multiprocessing.cpu_count()
    max_workers = round(max(1, int(cpu_count * 2/3)))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(extract_frames, BAD_FILE_PATH, v) for v in video_files]

        for f in tqdm(futures, total=len(futures), desc="Frame Extraction"):
            f.result()


    
    # 예상 처리 시간을 저장할 리스트
    estimated_execution_times = []


    #기준 이미지 초기화
    reference_image = None

    # 기존 엑셀 파일 열기
    try:
        data_wb = load_workbook(filename=BAD_EXCEL)
        data_ws = data_wb.active
    except FileNotFoundError:
        # 기존 파일이 없는 경우 새로운 워크북 생성
        data_wb = Workbook()
        data_ws = data_wb.active
        
    headers = [
    'Folder name',
    'Maximum Intensity of Pixels',
    f'Maximum of Displacement ({LENGTH_UNIT})',
    'MAX BAND ENERGY',
    'MAX Apeak',
    'Coordinate X',
    'Coordinate Y',
    'Hz',
    'Displacement List',
    'MAX ABS DISP',
    'ABS NET CHANGE RATIO',
    'NUM SIGN CHANGES DIFF',
    'DIFF STD'
]

    data_ws.append(headers)

    def safe_int(v, default=0):
        try:
            if v is None:
                return default
            v = float(v)
            if np.isnan(v):
                return default
            return int(round(v))
        except:
            return default
        
    output_folders = [f for f in natsort.natsorted(os.listdir(BAD_FILE_PATH)) if os.path.isdir(os.path.join(BAD_FILE_PATH, f))]

    number_of_folders = len(output_folders)

    for folder_name in natsort.natsorted(os.listdir(BAD_FILE_PATH)):
        folder_path = os.path.join(BAD_FILE_PATH, folder_name)
        if not os.path.isdir(folder_path):
            continue # 폴더가 아니라면 스킵
        # Create the result folder for the current sample.
        result_folder = os.path.join(folder_path, 'result')
        os.makedirs(result_folder, exist_ok=True)
        print(f'\n【　Starting to Analyze the File {folder_name}　】\n ')
        print('--------------------------------List of Generated Files --------------------------------\n')
        folder_start_time = time.time()
        for extension in VIDEO_FILE_EXTENSION:
            fps_video_path = os.path.join(BAD_FILE_PATH,f'{folder_name}.{extension}')
            if os.path.exists(fps_video_path):  
                cap = cv2.VideoCapture(fps_video_path)
                fps = round(cap.get(cv2.CAP_PROP_FPS), 1)
                cap.release()
                break 
            else:
                fps = VIDEO_FPS   

        reference_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}0000.{FRAME_EXTENSION}')
        reference_image = cv2.imread(reference_image_path)
        

        if reference_image is not None:
            reference_gray = cv2.cvtColor(reference_image, cv2.COLOR_BGR2GRAY)
        else:
            continue  
        

        max_number = -1
        min_number = float('inf')
        for filename in os.listdir(folder_path):
            if filename.startswith(f'{FRAME_PREFIX}') and filename.endswith(f'.{FRAME_EXTENSION}'):
                try:
                   
                    number = int(filename[len(f'{FRAME_PREFIX}'):-len(f'.{FRAME_EXTENSION}')]) 
                    max_number = max(max_number, number)
                    min_number = min(min_number, number)
                except ValueError:
                    pass 

        

        start_number = min_number
        last_number = max_number+1 
        resolution_y, resolution_x,_ = reference_image.shape 
        
        try:
            total_time_seconds = int(max_number / fps)
        except ZeroDivisionError:
            pass
            
        # Initialize lists for image paths and loaded frames
        # KR: 이미지 경로, grayscale frame, RGB frame을 저장할 리스트를 초기화합니다.
        image_paths = []
        current_gray_image_list = []
        current_rgb_image_list = []
        # # Initialize lists for maximum pixel-intensity differences and their coordinates
        max_difference_comparing_0 = []  #max_differences 
        max_difference_comparing_0_locations = [] # max_diff_locations

        max_top30_differences_crop = []
        max_top30_diff_locations_crop = []

        max_top30_differences_crop_2 = []
        max_top30_diff_locations_crop_2 = []

        max_top30_differences_crop_3 = []
        max_top30_diff_locations_crop_3 = []

        max_top30_differences_crop_2_new = []
        max_top30_diff_locations_crop_2_new = []

        max_top30_differences_crop_3_new = []
        max_top30_diff_locations_crop_3_new = []

        valid_pixel_all_counts_min = []
        valid_pixel_all_counts_max = []
        

        # =========================================================
        # (A) Load frame paths and grayscale images
        #     Frames are loaded once and then processed in GPU/CPU batch mode.
        # KR: frame 경로와 grayscale 이미지를 수집합니다.
        # KR: frame은 한 번만 로드한 뒤 이후 GPU/CPU batch 방식으로 처리합니다.
        # =========================================================

        use_gpu = (GPU_MODE.lower() == "on") and torch.cuda.is_available()
        device = torch.device("cuda" if use_gpu else "cpu")

        print(f"Compute mode: {'GPU' if use_gpu else 'CPU'}")




        image_paths = []
        gray_frames = []

        for idx in range(start_number, last_number):
            p = os.path.join(folder_path, f'{FRAME_PREFIX}{idx:04d}.{FRAME_EXTENSION}')
            if not os.path.exists(p):
                continue

            img_rgb = cv2.imread(p)
            if img_rgb is None:
                print(f"Image load fail → {p}")
                continue

            g = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)

            image_paths.append(p)
            gray_frames.append(g)
            current_rgb_image_list.append(img_rgb)


        if len(gray_frames) == 0:
            print("No valid frames found. Skip folder.")
            continue

        # =========================================================
        # (B) Compute frame-wise absolute differences from the reference image
        #     using GPU/CPU batch processing
        # KR: reference image와 각 frame 사이의 absolute difference를 계산합니다.
        # KR: 이 계산은 GPU 또는 CPU batch 처리로 한 번에 수행합니다.
        # =========================================================
        
        gray_tensor = torch.from_numpy(np.stack(gray_frames)).to(torch.int16)

        if use_gpu:
            gray_tensor = gray_tensor.pin_memory().to(device, non_blocking=True)
        else:
            gray_tensor = gray_tensor.contiguous()

        ref_tensor = torch.from_numpy(reference_gray).to(torch.int16)

        if use_gpu:
            ref_tensor = ref_tensor.pin_memory().to(device, non_blocking=True)
        else:
            ref_tensor = ref_tensor.contiguous()


        diff_tensor = torch.abs(gray_tensor - ref_tensor)                                 # [T,H,W]
        flat        = diff_tensor.flatten(1)                                              # [T,H*W]

        max_vals, argmax_vals = flat.max(dim=1)                                           # [T], [T]

        H, W = reference_gray.shape
        ys = (argmax_vals // W).cpu().numpy()
        xs = (argmax_vals %  W).cpu().numpy()

        max_difference_comparing_0 = max_vals.cpu().numpy().tolist()
        max_difference_comparing_0_locations = list(zip(xs, ys))

        current_gray_image_list = gray_frames

        image0_path = reference_image_path
        
        image0 = current_gray_image_list[0]

        
        # Search for the maximum intensity change within the initial fps × NUM_FRAMES window.
        # This helps focus the analysis on early contraction-related motion rather than large noise or bubble movement.
        # KR: 초기 fps × NUM_FRAMES 구간 안에서 최대 intensity change를 찾습니다.
        # KR: 큰 noise 또는 bubble 움직임보다 초기 수축 관련 움직임에 분석을 집중하기 위한 과정입니다.
        top30_max_differences_indices = np.arange(len(max_difference_comparing_0))[:int(fps*NUM_FRAMES)]
        top30_max_differences_values = [max_difference_comparing_0[i] for i in top30_max_differences_indices]
        top30_max_differences_values_max = max(top30_max_differences_values)
        
        max_value_index_within_top30 = np.argmax(top30_max_differences_values)
        
        max_difference_index_within_top30 = top30_max_differences_indices[max_value_index_within_top30]
        
        max_difference_value_within_top30 = top30_max_differences_values[max_value_index_within_top30]

        max_diff_location_within_top30 = max_difference_comparing_0_locations[max_difference_index_within_top30]

        x_center_max_diff_location_within_top30 = max_diff_location_within_top30[0]
        y_center_max_diff_location_within_top30 = max_diff_location_within_top30[1]

        
        x11 = x_center_max_diff_location_within_top30 - ANALYSIS_SQUARE_SIZE_REFERENCE
        y11 = y_center_max_diff_location_within_top30 - ANALYSIS_SQUARE_SIZE_REFERENCE
        x22 = x_center_max_diff_location_within_top30 + ANALYSIS_SQUARE_SIZE_REFERENCE
        y22 = y_center_max_diff_location_within_top30 + ANALYSIS_SQUARE_SIZE_REFERENCE
        x11 = max(x11, 0)
        y11 = max(y11, 0)
        x22 = min(x22, resolution_x - 1)
        y22 = min(y22, resolution_y - 1)


        roi_tensor = gray_tensor[:, y11:y22, x11:x22].contiguous()


        image_frame_number = []
        next_image_index = 0   

        visited = set()

        for step in range(min(len(gray_frames), 500)):
            if next_image_index in visited:
                break
            visited.add(next_image_index)

            image_frame_number.append(next_image_index)

            ref_roi = roi_tensor[next_image_index].unsqueeze(0)  # [1,h,w]
            diff_batch = torch.abs(roi_tensor - ref_roi)         # [T,h,w]
            flat = diff_batch.flatten(1)                         # [T,h*w]

            max_vals, _ = flat.max(dim=1)                        # [T]
            best_i = int(max_vals.argmax().item())

            next_image_index = best_i

            if len(image_frame_number) >= 8 and image_frame_number[-1] == image_frame_number[-3]:
                break

        image_frame_n1 = image_frame_number[-1]
        image_frame_n2 = image_frame_number[-2] if len(image_frame_number) >= 2 else image_frame_number[-1]

        n1_max_image_path = image_paths[int(image_frame_n1)]
        n2_max_image_path = image_paths[int(image_frame_n2)]


        roi_full = gray_tensor[:, y11:y22, x11:x22].contiguous()          # [T,h,w]

        ref_n1 = roi_full[int(image_frame_n1)].unsqueeze(0)               # [1,h,w]
        diff_n1 = torch.abs(roi_full - ref_n1)                            # [T,h,w]
        flat_n1 = diff_n1.flatten(1)                                      # [T,h*w]
        max_vals_n1, argmax_n1 = flat_n1.max(dim=1)                       # [T], [T]

        h_roi = roi_full.shape[1]
        w_roi = roi_full.shape[2]
        ys_n1 = (argmax_n1 // w_roi)
        xs_n1 = (argmax_n1 %  w_roi)

        max_top30_differences_crop_2 = max_vals_n1.detach().cpu().numpy().astype(np.float32).tolist()
        max_top30_diff_locations_crop_2 = list(zip(xs_n1.detach().cpu().numpy().tolist(),
                                                ys_n1.detach().cpu().numpy().tolist()))

        max_top30_differences_crop_except0_2 = [v for v in max_top30_differences_crop_2 if v != 0]
        mean_max_top30_differences_crop2 = float(np.mean(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0.0
        mean_round_max_top30_differences_crop2 = round(mean_max_top30_differences_crop2, 2)
        max_max_top30_differences_crop2 = float(np.max(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0.0

        if len(max_top30_differences_crop_except0_2) > 0:
            min_max_top30_differences_crop_except0_2 = float(np.min(max_top30_differences_crop_except0_2))
        else:
            min_max_top30_differences_crop_except0_2 = 0.0

        min_max_top30_differences_crop_except0_2_index = int(np.argmax(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0

        ref_n2 = roi_full[int(image_frame_n2)].unsqueeze(0)               # [1,h,w]
        diff_n2 = torch.abs(roi_full - ref_n2)                            # [T,h,w]
        flat_n2 = diff_n2.flatten(1)                                      # [T,h*w]
        max_vals_n2, argmax_n2 = flat_n2.max(dim=1)                       # [T], [T]

        ys_n2 = (argmax_n2 // w_roi)
        xs_n2 = (argmax_n2 %  w_roi)

        max_top30_differences_crop_3 = max_vals_n2.detach().cpu().numpy().astype(np.float32).tolist()
        max_top30_diff_locations_crop_3 = list(zip(xs_n2.detach().cpu().numpy().tolist(),
                                                ys_n2.detach().cpu().numpy().tolist()))

        max_top30_differences_crop_except0_3 = [v for v in max_top30_differences_crop_3 if v != 0]
        mean_max_top30_differences_crop3 = float(np.mean(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0.0
        mean_round_max_top30_differences_crop3 = round(mean_max_top30_differences_crop3, 2)
        max_max_top30_differences_crop3 = float(np.max(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0.0

        if len(max_top30_differences_crop_except0_3) > 0:
            min_max_top30_differences_crop_except0_3 = float(np.min(max_top30_differences_crop_except0_3))
        else:
            min_max_top30_differences_crop_except0_3 = 0.0

        min_max_top30_differences_crop_except0_3_index = int(np.argmax(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0

        try:
            top30_max_image_path_2 = image_paths[min_max_top30_differences_crop_except0_2_index]
        except:
            top30_max_image_path_2 = ""

        try:
            top30_max_image_path_3 = image_paths[min_max_top30_differences_crop_except0_3_index]
        except:
            top30_max_image_path_3 = ""

        diff_max_all = diff_n2                                           

        diff_min_all = diff_n1                                            

        threshold_max = min_max_top30_differences_crop_except0_3
        threshold_min = min_max_top30_differences_crop_except0_2

        min_threshold_max = threshold_max + 20
        max_threshold_max = max_max_top30_differences_crop3

        min_threshold_min = threshold_min + 20
        max_threshold_min = max_max_top30_differences_crop2

        if max_threshold_max <= 0:
            valid_pixel_counts_max = [0] * len(gray_frames)
        else:
            mask_max = (diff_max_all >= min_threshold_max) & (diff_max_all <= max_threshold_max)
            valid_pixel_counts_max = mask_max.flatten(1).sum(dim=1).detach().cpu().numpy().astype(np.int64).tolist()

        if max_threshold_min <= 0:
            valid_pixel_counts_min = [0] * len(gray_frames)
        else:
            mask_min = (diff_min_all >= min_threshold_min) & (diff_min_all <= max_threshold_min)
            valid_pixel_counts_min = mask_min.flatten(1).sum(dim=1).detach().cpu().numpy().astype(np.int64).tolist()

        valid_pixel_count_mean_min = np.mean([v for v in valid_pixel_counts_min if v not in [0, '']]) if len(valid_pixel_counts_min) else 0
        valid_pixel_count_mean_round_min = round(valid_pixel_count_mean_min) if not np.isnan(valid_pixel_count_mean_min) else 0

        valid_pixel_count_mean_max = np.mean([v for v in valid_pixel_counts_max if v not in [0, '']]) if len(valid_pixel_counts_max) else 0
        valid_pixel_count_mean_round_max = round(valid_pixel_count_mean_max) if not np.isnan(valid_pixel_count_mean_max) else 0

        count_zeros_min = valid_pixel_counts_min.count(0)
        count_zeros_max = valid_pixel_counts_max.count(0)

        sum_count_all_pixel_in_range_min = int(np.sum(np.array(valid_pixel_counts_min) < valid_pixel_count_mean_round_min)) if valid_pixel_count_mean_round_min > 0 else 0
        sum_count_all_pixel_in_range_max = int(np.sum(np.array(valid_pixel_counts_max) < valid_pixel_count_mean_round_max)) if valid_pixel_count_mean_round_max > 0 else 0

        valid_pixel_all_counts_min.append(sum_count_all_pixel_in_range_min)
        valid_pixel_all_counts_max.append(sum_count_all_pixel_in_range_max)

        valid_pixel_all_counts_min_max = max(valid_pixel_all_counts_min) if len(valid_pixel_all_counts_min) else 0
        valid_pixel_all_counts_max_max = max(valid_pixel_all_counts_max) if len(valid_pixel_all_counts_max) else 0

        All_pixel_intensities = []
        
        selected_max_differences = []
        selected_max_differences_location = []
        
        max_image_path_pixel = top30_max_image_path_3   
        min_image_path_pixel = top30_max_image_path_2  

        for i in range(len(gray_frames)):

            if use_gpu:
                image1 = gray_tensor[i].cpu().numpy().astype(np.uint8)
                max_image_gray_pixel = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
                min_image_gray_pixel = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
                first_image_gray_pixel = gray_tensor[int(max_difference_index_within_top30)].cpu().numpy().astype(np.uint8)

            else:
                image1 = gray_frames[i].astype(np.uint8)
                max_image_gray_pixel = gray_frames[int(image_frame_n2)].astype(np.uint8)
                min_image_gray_pixel = gray_frames[int(image_frame_n1)].astype(np.uint8)
                first_image_gray_pixel = gray_frames[int(max_difference_index_within_top30)].astype(np.uint8)

            cropped_max_image_pixel = max_image_gray_pixel[y11:y22, x11:x22]
            cropped_min_image_pixel = min_image_gray_pixel[y11:y22, x11:x22]
            cropped_image0_pixel = image0[y11:y22, x11:x22]
            cropped_image1_pixel = image1[y11:y22, x11:x22]

            h = min(cropped_max_image_pixel.shape[0], cropped_image1_pixel.shape[0])
            w = min(cropped_max_image_pixel.shape[1], cropped_image1_pixel.shape[1])

            cropped_max_image_pixel = cropped_max_image_pixel[:h, :w]
            cropped_min_image_pixel = cropped_min_image_pixel[:h, :w]
            cropped_image0_pixel = cropped_image0_pixel[:h, :w]
            cropped_image1_pixel = cropped_image1_pixel[:h, :w]

            cropped_max_image_pixel = cropped_max_image_pixel.astype(np.uint8)
            cropped_min_image_pixel = cropped_min_image_pixel.astype(np.uint8)
            cropped_image0_pixel = cropped_image0_pixel.astype(np.uint8)
            cropped_image1_pixel = cropped_image1_pixel.astype(np.uint8)

            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f"Error cropping image {i}. Skipping...")
                continue

            if valid_pixel_all_counts_min_max >= valid_pixel_all_counts_max_max:
                not_crop_diff = cv2.subtract(image1, min_image_gray_pixel)
                square_image_path = min_image_path_pixel
                square_image_index = image_frame_n1

            else:
                diff = cv2.subtract(cropped_max_image_pixel, cropped_image1_pixel)
                not_crop_diff = cv2.subtract(image1, max_image_gray_pixel)
                square_image_path = max_image_path_pixel
                square_image_index = image_frame_n2

            _, selected_max_diff, _, selected_max_diff_loc = cv2.minMaxLoc(not_crop_diff)

            selected_max_differences.append(selected_max_diff)
            selected_max_differences_location.append(selected_max_diff_loc)

            
            selected_max_differences_indices = np.arange(len(selected_max_differences))[:int(fps*NUM_FRAMES)] 
            selected_max_differences_values = [selected_max_differences[i] for i in selected_max_differences_indices] 

            selected_max_differences_values_max_in_Frames = np.argmax(selected_max_differences_values) 
            selected_max_differences_index_in_Frames = selected_max_differences_indices[selected_max_differences_values_max_in_Frames] 
            selected_max_differences_value_in_Frames = selected_max_differences_values[selected_max_differences_index_in_Frames] 
            selected_max_differences_location_in_Frames = selected_max_differences_location[selected_max_differences_index_in_Frames]

        x_selected = selected_max_differences_location_in_Frames[0]  
        y_selected = selected_max_differences_location_in_Frames[1] 
        
        x1_small = x_selected - ANALYSIS_SQUARE_SIZE_REFERENCE  
        y1_small = y_selected - ANALYSIS_SQUARE_SIZE_REFERENCE
        x2_small = x_selected + ANALYSIS_SQUARE_SIZE_REFERENCE
        y2_small = y_selected + ANALYSIS_SQUARE_SIZE_REFERENCE
        
        x1_small = max(x1_small, 0)
        y1_small = max(y1_small, 0)
        x2_small = min(x2_small, resolution_x -1)
        y2_small = min(y2_small, resolution_y -1)
        
        # Compare all frames again using the cropped ROI-derived maximum frame as a reference.
        # This step is used to identify candidate maximal contraction or relaxation frames.
        # KR: cropped ROI에서 얻은 maximum frame을 기준으로 전체 frame을 다시 비교합니다.
        # KR: 이 과정은 최대 수축 또는 최대 이완 후보 frame을 찾기 위한 단계입니다.

        image_frame_number = []

        next_image_index = start_number
        for n in range(start_number, last_number):
            max_diff_cropped_n_frame_list = []
            max_diff_cropped_loc_n_frame_list = []
        
            for i in range(len(gray_frames)):

                n_frame_path = os.path.join(folder_path, f'{FRAME_PREFIX}{next_image_index + start_number :04d}.{FRAME_EXTENSION}') 

                # Current reference frame path used in this iterative comparison
                # KR: 반복 비교에 사용할 현재 reference frame 경로입니다.
                if use_gpu:
                    top30_max_image_gray_new = gray_tensor[next_image_index].cpu().numpy().astype(np.uint8)
                else:
                    top30_max_image_gray_new = gray_frames[next_image_index].astype(np.uint8)
                cropped_n_frame_gray = top30_max_image_gray_new[y1_small:y2_small, x1_small:x2_small] 
                current_image_crop = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small] 

                max_diff_loc_n_frame_cropped = cv2.absdiff(cropped_n_frame_gray, current_image_crop)
                _, max_diff_crop_n_frame,_,max_diff_crop_loc_n_frame = cv2.minMaxLoc(max_diff_loc_n_frame_cropped)

                max_diff_cropped_n_frame_list.append(max_diff_crop_n_frame)
                max_diff_cropped_loc_n_frame_list.append(max_diff_crop_loc_n_frame)

                max_diff_cropped_n_indices = np.argsort(max_diff_cropped_n_frame_list)[-1:]            
                max_diff_cropped_n_indices_index = max_diff_cropped_n_indices[0]
                max_diff_cropped_n_indices_path = os.path.join(folder_path, f'{FRAME_PREFIX}{max_diff_cropped_n_indices_index + start_number :04d}.{FRAME_EXTENSION}')
            image_frame_number.append(next_image_index)
            
            # Calculate summary statistics from cropped ROI differences.
            # Zero values are excluded when estimating the nonzero minimum.
            # KR: cropped ROI difference 값에서 요약 통계를 계산합니다.
            # KR: 0이 아닌 최소값을 계산할 때는 zero value를 제외합니다.
            max_diff_cropped_n_except0 = [value for value in max_diff_cropped_n_frame_list if value != 0]
            mean_max_diff_cropped_n_except0 = np.mean(max_diff_cropped_n_frame_list)
            mean_round_max_diff_cropped_n_frame_listew = round(mean_max_diff_cropped_n_except0, 2)
            max_max_diff_cropped_n_frame_list = max(max_diff_cropped_n_frame_list)
            min_max_diff_cropped_n_except0 = min(max_diff_cropped_n_except0)
            min_max_diff_cropped_n_except0_index = max_diff_cropped_n_frame_list.index(max_max_diff_cropped_n_frame_list)
            # Update the reference frame index for the next iteration
            # KR: 다음 반복 비교에 사용할 reference frame index를 업데이트합니다.
            next_image_index = min_max_diff_cropped_n_except0_index
            
            # Stop the iterative search when the selected frame index starts to repeat.
            # KR: 선택된 frame index가 반복되면 탐색이 수렴한 것으로 보고 반복을 중단합니다.
            if len(image_frame_number)>=8 and image_frame_number[n] == image_frame_number[n - 2]:
                    image_frame_n1 = image_frame_number[n]
                    image_frame_n2 = image_frame_number[n-1]
                    n1_max_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}{image_frame_n1 + start_number :04d}.{FRAME_EXTENSION}') 
                    n2_max_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}{image_frame_n2 + start_number :04d}.{FRAME_EXTENSION}') 
                    break

        # Compare all frames again using the two ROI-derived candidate frames.
        # This step refines the candidate maximal contraction and relaxation frames.
        # KR: ROI에서 얻은 두 후보 frame을 기준으로 전체 frame을 다시 비교합니다.
        # KR: 최대 수축/최대 이완 후보 frame을 더 안정적으로 찾기 위한 단계입니다.
        
        for i in range(len(gray_frames)):
            
            top30_max_image_path_2_new = n1_max_image_path
            # Use the n1 candidate frame as the reference for ROI-based comparison.
            # KR: n1 후보 frame을 기준으로 ROI 기반 frame 비교를 수행합니다.
            
            if use_gpu:
                top30_max_image_gray_2_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            else:
                top30_max_image_gray_2_new = gray_frames[image_frame_n1].astype(np.uint8)

            
            cropped_top30_max_image_gray_2_new = top30_max_image_gray_2_new[y1_small:y2_small, x1_small:x2_small] 
            current_image_crop_2_new = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small]
            diff_top30_max_2_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_2_new, current_image_crop_2_new)
            _, max_diff_crop_top30_max_2_new, _,max_diff_loc_crop_top30_max_2_new = cv2.minMaxLoc(diff_top30_max_2_cropped_new)
            max_top30_differences_crop_2_new.append(max_diff_crop_top30_max_2_new)
            max_top30_diff_locations_crop_2_new.append(max_diff_loc_crop_top30_max_2_new)
            
            max_top30_differences_crop_indices_2_new = np.argsort(max_top30_differences_crop_2_new)[-1:]
            
            # Search within the initial fps × NUM_FRAMES window to reduce the influence of large noise or bubble movement.
            # KR: 큰 noise 또는 bubble 움직임의 영향을 줄이기 위해 초기 fps × NUM_FRAMES 구간 안에서 탐색합니다.
            max_top30_differences_crop_indices_2_new_num_frames = np.arange(len(max_top30_differences_crop_2_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_2_new = [max_top30_differences_crop_2_new[i] for i in max_top30_differences_crop_indices_2_new_num_frames]

            
            max_top30_differences_crop_values_2_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_2_new) 
            
            index_in_Frames_2 = max_top30_differences_crop_indices_2_new_num_frames[max_top30_differences_crop_values_2_new_max_in_Frames]
            value_in_Frames_2 =  max_top30_differences_crop_values_2_new[index_in_Frames_2]
            location_in_Frames_2 = max_top30_diff_locations_crop_2_new[index_in_Frames_2]

            
            max_top30_differences_crop_indices_2_index_new = max_top30_differences_crop_indices_2_new[0]
            max_top30_differences_crop_indices_2_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_2_index_new + start_number :04d}.{FRAME_EXTENSION}')


        # Calculate summary statistics for ROI differences relative to the n1 candidate frame.
        # KR: n1 후보 frame 기준 ROI difference의 요약 통계를 계산합니다.
        max_top30_differences_crop_except0_2_new = [value for value in max_top30_differences_crop_2_new if value != 0]
        mean_max_top30_differences_crop2_new = np.mean(max_top30_differences_crop_2_new)
        mean_round_max_top30_differences_crop2_new = round(mean_max_top30_differences_crop2_new, 2)
        max_max_top30_differences_crop2_new = max(max_top30_differences_crop_2_new)
        min_max_top30_differences_crop_except0_2_new = min(max_top30_differences_crop_except0_2_new)
        min_max_top30_differences_crop_except0_2_index_new = max_top30_differences_crop_2_new.index(max_max_top30_differences_crop2_new)

        
        for i in range(len(gray_frames)):

            top30_max_image_path_3_new = n2_max_image_path

            # Use the n2 candidate frame as the reference for ROI-based comparison.
            # KR: n2 후보 frame을 기준으로 ROI 기반 frame 비교를 수행합니다.
            if use_gpu:
                top30_max_image_gray_3_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
            else:
                top30_max_image_gray_3_new = gray_frames[image_frame_n2].astype(np.uint8)

            
            cropped_top30_max_image_gray_3_new = top30_max_image_gray_3_new[y1_small:y2_small, x1_small:x2_small] 
            
            current_image_crop_3 = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small] 

            diff_top30_max_3_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_3_new, current_image_crop_3)
            _, max_diff_crop_top30_max_3_new, _,max_diff_loc_crop_top30_max_3_new = cv2.minMaxLoc(diff_top30_max_3_cropped_new)
            
            max_top30_differences_crop_3_new.append(max_diff_crop_top30_max_3_new)
            max_top30_diff_locations_crop_3_new.append(max_diff_loc_crop_top30_max_3_new)
            
            max_top30_differences_crop_indices_3_new = np.argsort(max_top30_differences_crop_3_new)[-1:]
            
            # Search within the initial fps × NUM_FRAMES window to reduce the influence of large noise or bubble movement.
            # KR: 큰 noise 또는 bubble 움직임의 영향을 줄이기 위해 초기 fps × NUM_FRAMES 구간 안에서 탐색합니다.
            max_top30_differences_crop_indices_3_new_num_frames = np.arange(len(max_top30_differences_crop_3_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_3_new = [max_top30_differences_crop_3_new[i] for i in max_top30_differences_crop_indices_3_new_num_frames] 

            
            max_top30_differences_crop_values_3_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_3_new) 
            index_in_Frames_3 = max_top30_differences_crop_indices_3_new_num_frames[max_top30_differences_crop_values_3_new_max_in_Frames]
            value_in_Frames_3 =  max_top30_differences_crop_values_3_new[index_in_Frames_3]
            location_in_Frames_3 = max_top30_diff_locations_crop_3_new[index_in_Frames_3]

            
            max_top30_differences_crop_indices_3_index_new = max_top30_differences_crop_indices_3_new[0]
            max_top30_differences_crop_indices_3_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_3_index_new + start_number :04d}.{FRAME_EXTENSION}')
            

        max_top30_differences_crop_except0_3_new = [value for value in max_top30_differences_crop_3_new if value != 0]
        mean_max_top30_differences_crop3_new = np.mean(max_top30_differences_crop_3_new)
        mean_round_max_top30_differences_crop3_new = round(mean_max_top30_differences_crop3_new, 2)
        max_max_top30_differences_crop3_new = max(max_top30_differences_crop_3_new)
        min_max_top30_differences_crop_except0_3_new = min(max_top30_differences_crop_except0_3_new)
        min_max_top30_differences_crop_except0_3_index_new = max_top30_differences_crop_3_new.index(max_max_top30_differences_crop3_new)
        
        top30_max_image_path_4_new = os.path.join(folder_path, f'{FRAME_PREFIX}{min_max_top30_differences_crop_except0_3_index_new + start_number :04d}.{FRAME_EXTENSION}') 
        



        x_differences = [] 
        y_differences = []
        valid_pixel_counts_min_new = []
        valid_pixel_counts_max_new = []
        valid_pixel_all_counts_max_new = []
        valid_pixel_all_counts_min_new = []

        # Calculate summary statistics for ROI differences relative to the n2 candidate frame.
        # KR: n2 후보 frame 기준 ROI difference의 요약 통계를 계산합니다.
        for i in range(len(gray_frames)):
            image1 = current_gray_image_list[i]
            # Crop local ROIs around the selected motion center.
            # KR: 선택된 motion center 주변의 local ROI를 crop합니다.
            max_image_path_pixel_new = top30_max_image_path_3_new  # Candidate maximal contraction or relaxation frame
            min_image_path_pixel_new = top30_max_image_path_2_new  # Candidate maximal relaxation or contraction frame
            
            if use_gpu:
                max_image_gray_pixel_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
                min_image_gray_pixel_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            else:
                max_image_gray_pixel_new = gray_frames[image_frame_n2].astype(np.uint8)
                min_image_gray_pixel_new = gray_frames[image_frame_n1].astype(np.uint8)

                    
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small] 
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
            cropped_image0_pixel_new = image0[y1_small:y2_small, x1_small:x2_small] 
            cropped_image1_pixel_new = image1[y1_small:y2_small, x1_small:x2_small] 
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f'Error cropping image{i}. Skipping...')
                continue  # Skip this frame if ROI cropping fails
            
            diff_max_new = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
            diff_min_new = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 

            threshold_max_new = min_max_top30_differences_crop_except0_3_new  
            threshold_min_new = min_max_top30_differences_crop_except0_2_new 
            white_pixel_coords_max_new = np.argwhere(diff_max_new >= threshold_max_new)
            white_pixel_coords_min_new = np.argwhere(diff_min_new >= threshold_min_new)  
            
            # Count pixels whose intensity differences fall within the threshold range.
            # KR: intensity difference가 설정된 threshold 범위 안에 들어오는 pixel을 계산합니다.
            
            # Add a fixed offset to reduce small background or medium-related noise.
            # KR: 작은 background 또는 배지 관련 noise를 줄이기 위해 고정 offset을 더합니다.
            min_threshold_max_new = min_max_top30_differences_crop_except0_3_new + 20  # Fixed offset for small background or medium-related noise
            max_threshold_max_new = max_max_top30_differences_crop3_new  # Maximum ROI intensity difference relative to the candidate frame
            min_threshold_min_new = min_max_top30_differences_crop_except0_2_new  +20 # Fixed offset for small background or medium-related noise
            max_threshold_min_new = max_max_top30_differences_crop2_new # Maximum ROI intensity difference relative to the candidate frame
            
            valid_pixel_coords_max_new = np.argwhere((diff_max_new >= min_threshold_max_new) & (diff_max_new <= max_threshold_max_new))
            valid_pixel_coords_min_new = np.argwhere((diff_min_new >= min_threshold_min_new) & (diff_min_new <= max_threshold_min_new))
            
            valid_pixel_coords_max_all_new = np.argwhere((diff_max_new >= 0) & (diff_max_new <= max_threshold_max))
            valid_pixel_coords_min_all_new = np.argwhere((diff_min_new >= 0) & (diff_min_new <= max_threshold_min))

            valid_pixel_count_max_new = len(valid_pixel_coords_max_new)
            valid_pixel_count_min_new = len(valid_pixel_coords_min_new)
            valid_pixel_counts_max_new.append(valid_pixel_count_max_new)
            valid_pixel_counts_min_new.append(valid_pixel_count_min_new)       

        valid_pixel_count_mean_min_new = np.mean([value for value in valid_pixel_counts_min_new if value not in [0,'']]) #0 과 데이터가 

        valid_pixel_count_mean_round_min_new = round(valid_pixel_count_mean_min_new) if not np.isnan(valid_pixel_count_mean_min_new) else 0 
        valid_pixel_count_mean_max_new = np.mean([value for value in valid_pixel_counts_max_new if value not in [0,'']]) #0 과 데이터가 

        valid_pixel_count_mean_round_max_new = round(valid_pixel_count_mean_max_new) if not np.isnan(valid_pixel_count_mean_max_new) else 0 
        
        count_zeros_min_new = valid_pixel_counts_min_new.count(0)
        count_zeros_max_new = valid_pixel_counts_max_new.count(0)

        count_all_pixel_in_range_min_new = [valid_pixel_counts_min_new.count(i) for i in range(valid_pixel_count_mean_round_min_new)]
        count_all_pixel_in_range_max_new = [valid_pixel_counts_max_new.count(i) for i in range(valid_pixel_count_mean_round_max_new)]
        
        sum_count_all_pixel_in_range_min_new = sum(count_all_pixel_in_range_min_new, 0)
        sum_count_all_pixel_in_range_max_new = sum(count_all_pixel_in_range_max_new, 0)

        valid_pixel_all_counts_min_new.append(sum_count_all_pixel_in_range_min_new)
        valid_pixel_all_counts_max_new.append(sum_count_all_pixel_in_range_max_new)
        
        valid_pixel_all_counts_min_max_new = max(valid_pixel_all_counts_min_new)
        valid_pixel_all_counts_max_max_new = max(valid_pixel_all_counts_max_new)
        
        
        All_pixel_intensities = []  
        valid_pixel_counts = []
        valid_pixel_counts2 = []
        selected_max_differences = []
        selected_max_differences_location = []
        
        x_selected_new, y_selected_new = None, None
        click_done = False  
        



        for i in range(len(gray_frames)):

            max_image_path_pixel_new = top30_max_image_path_3_new
            min_image_path_pixel_new = top30_max_image_path_2_new
            
            if use_gpu:
                max_image_gray_pixel_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
                min_image_gray_pixel_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            else:
                max_image_gray_pixel_new = gray_frames[image_frame_n2].astype(np.uint8)
                min_image_gray_pixel_new = gray_frames[image_frame_n1].astype(np.uint8)

            
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
            cropped_image0_pixel_new = image0[y1_small:y2_small, x1_small:x2_small]  
            cropped_image1_pixel_new = image1[y1_small:y2_small, x1_small:x2_small]  
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f"Error cropping image {i}. Skipping...")
                continue 
            
            # Select the candidate frame based on the number of valid pixels below the mean threshold.
            # This helps determine which candidate better represents maximal contraction or relaxation.
            # KR: mean threshold보다 작은 valid pixel count를 기준으로 후보 frame을 선택합니다.
            # KR: 최대 수축 또는 최대 이완을 더 잘 나타내는 후보를 고르기 위한 과정입니다.
            
            if valid_pixel_all_counts_min_max_new >= valid_pixel_all_counts_max_max_new:
                loc2 = max_top30_diff_locations_crop_2_new  
                loc3 = location_in_Frames_2  
                
            else:
                loc2 = max_top30_diff_locations_crop_3_new 
                loc3 = location_in_Frames_3 
                
            # =========================================================
            # Manual coordinate selection callback
            # =========================================================
            
            # Mouse-click callback for selecting an analysis coordinate
            # KR: 분석 좌표를 수동으로 선택하기 위한 마우스 클릭 callback 함수입니다.
            def select_point(event, x, y, flags, param):
                global x_selected_new, y_selected_new, click_done  

                if event == cv2.EVENT_LBUTTONDOWN and not click_done: # Left-click once to select a coordinate
                    x_selected_new = x
                    y_selected_new = y
                    click_done = True  # 
                    print(f"Selected coordinate: ({x_selected_new}, {y_selected_new})")

                    cv2.circle( min_image_gray_pixel_new, (x_selected_new, y_selected_new), 2, (0, 0, 255), -1)
                    cv2.imshow("Select Point",  min_image_gray_pixel_new)

                    cv2.waitKey(500) 
                    cv2.destroyAllWindows()

            # Use predefined coordinates when manual coordinate mode is enabled
            # KR: manual coordinate mode가 켜져 있으면 사용자가 미리 지정한 좌표를 사용합니다.
            if MANUAL_SELECTION.lower() == 'on' and not click_done:
                print("Please click the desired analysis point.")
                

                cv2.namedWindow("Select Point", cv2.WINDOW_NORMAL)
                cv2.resizeWindow("Select Point", resolution_x, resolution_y)  

                cv2.imshow("Select Point", min_image_gray_pixel_new)
                cv2.setMouseCallback("Select Point", select_point)


                while not click_done:
                    cv2.waitKey(1)  # 

                print(f"Using manually selected coordinate: ({x_selected_new}, {y_selected_new})")

            # Use automatically detected coordinates when manual selection is off,
            # or keep the previously selected coordinate after the first manual click.
            # KR: manual selection이 꺼져 있으면 자동 탐색 좌표를 사용합니다.
            # KR: 첫 번째 수동 클릭 이후에는 선택된 좌표를 유지합니다.
            
            elif MANUAL_SELECTION.lower() == 'off' or (x_selected_new is None and y_selected_new is None): 
                x_selected_new = x1_small + loc3[0]
                y_selected_new = y1_small + loc3[1]
        
            
            x1_small_new = x_selected_new -  resolution_x
            y1_small_new = y_selected_new -  resolution_y
            x2_small_new = x_selected_new +  resolution_x
            y2_small_new = y_selected_new + resolution_y
            
            x1_small_new = max(x1_small_new, 0)
            y1_small_new = max(y1_small_new, 0)
            x2_small_new = min(x2_small_new, resolution_x -1)
            y2_small_new = min(y2_small_new, resolution_y -1)
            
            x1_small_new_reference = max(0, x_selected_new - ANALYSIS_SQUARE_SIZE_REFERENCE)
            y1_small_new_reference = max(0, y_selected_new -  ANALYSIS_SQUARE_SIZE_REFERENCE)
            x2_small_new_reference = x_selected_new + ANALYSIS_SQUARE_SIZE_REFERENCE
            y2_small_new_reference = y_selected_new +  ANALYSIS_SQUARE_SIZE_REFERENCE
            
            x1_small_new_reference = max(x1_small_new_reference, 0)
            y1_small_new_reference = max(y1_small_new_reference, 0)
            x2_small_new_reference = min(x2_small_new_reference, resolution_x -1)
            y2_small_new_reference = min(y2_small_new_reference, resolution_y -1)
            
            x1_small_new_optical = max(0, x_selected_new - resolution_x)
            y1_small_new_optical = max(0, y_selected_new - resolution_y)
            x2_small_new_optical = x_selected_new + resolution_x
            y2_small_new_optical = y_selected_new + resolution_y
            
            x1_small_new_optical = max(x1_small_new_optical, 0)
            y1_small_new_optical = max(y1_small_new_optical, 0)
            x2_small_new_optical = min(x2_small_new_optical, resolution_x -1)
            y2_small_new_optical = min(y2_small_new_optical, resolution_y -1)

        max_top30_differences_crop_2_new = []
        max_top30_diff_locations_crop_2_new = []

        max_top30_differences_crop_3_new = []
        max_top30_diff_locations_crop_3_new = []

        if len(current_rgb_image_list) == 0:
            print("❌ current_rgb_image_list empty → skip")
            continue

        image_frame_n1 = max(0, min(image_frame_n1, len(current_rgb_image_list)-1))
        image_frame_n2 = max(0, min(image_frame_n2, len(current_rgb_image_list)-1))

        top30_max_image_2_new = current_rgb_image_list[image_frame_n1]
        top30_max_image_3_new = current_rgb_image_list[image_frame_n2]


        for i in range(len(gray_frames)):
            
            top30_max_image_path_2_new = n1_max_image_path
            top30_max_image_gray_2_new = cv2.cvtColor(top30_max_image_2_new, cv2.COLOR_BGR2GRAY) 
            
            cropped_top30_max_image_gray_2_new = top30_max_image_gray_2_new[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
            
            current_image_crop_2_new = current_gray_image_list[i][y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 

            diff_top30_max_2_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_2_new, current_image_crop_2_new)
            _, max_diff_crop_top30_max_2_new, _,max_diff_loc_crop_top30_max_2_new = cv2.minMaxLoc(diff_top30_max_2_cropped_new)
            
            max_top30_differences_crop_2_new.append(max_diff_crop_top30_max_2_new)
            max_top30_diff_locations_crop_2_new.append(max_diff_loc_crop_top30_max_2_new)
            
            max_top30_differences_crop_indices_2_new = np.argsort(max_top30_differences_crop_2_new)[-1:]
            
            max_top30_differences_crop_indices_2_new_num_frames = np.arange(len(max_top30_differences_crop_2_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_2_new = [max_top30_differences_crop_2_new[i] for i in max_top30_differences_crop_indices_2_new_num_frames] 
            
            max_top30_differences_crop_values_2_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_2_new)
            
            index_in_Frames_2 = max_top30_differences_crop_indices_2_new_num_frames[max_top30_differences_crop_values_2_new_max_in_Frames]
            
            value_in_Frames_2 =  max_top30_differences_crop_values_2_new[index_in_Frames_2]
            location_in_Frames_2 = max_top30_diff_locations_crop_2_new[index_in_Frames_2]
            
            max_top30_differences_crop_indices_2_index_new = max_top30_differences_crop_indices_2_new[0]
            max_top30_differences_crop_indices_2_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_2_index_new + start_number :04d}.{FRAME_EXTENSION}')



        max_top30_differences_crop_except0_2_new = [value for value in max_top30_differences_crop_2_new if value != 0]
        mean_max_top30_differences_crop2_new = np.mean(max_top30_differences_crop_2_new)
        mean_round_max_top30_differences_crop2_new = round(mean_max_top30_differences_crop2_new, 2)
        max_max_top30_differences_crop2_new = max(max_top30_differences_crop_2_new)
        min_max_top30_differences_crop_except0_2_new = min(max_top30_differences_crop_except0_2_new)
        min_max_top30_differences_crop_except0_2_index_new = max_top30_differences_crop_2_new.index(max_max_top30_differences_crop2_new)

        

        for i in range(len(gray_frames)):

            top30_max_image_path_3_new = n2_max_image_path
            top30_max_image_gray_3_new = cv2.cvtColor(top30_max_image_3_new, cv2.COLOR_BGR2GRAY) 
            
            cropped_top30_max_image_gray_3_new = top30_max_image_gray_3_new[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
            
            current_image_crop_3 = current_gray_image_list[i][y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 

            diff_top30_max_3_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_3_new, current_image_crop_3)
            _, max_diff_crop_top30_max_3_new, _,max_diff_loc_crop_top30_max_3_new = cv2.minMaxLoc(diff_top30_max_3_cropped_new)

            max_top30_differences_crop_3_new.append(max_diff_crop_top30_max_3_new)
            max_top30_diff_locations_crop_3_new.append(max_diff_loc_crop_top30_max_3_new)
            
            max_top30_differences_crop_indices_3_new = np.argsort(max_top30_differences_crop_3_new)[-1:]
            max_top30_differences_crop_indices_3_new_num_frames = np.arange(len(max_top30_differences_crop_3_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_3_new = [max_top30_differences_crop_3_new[i] for i in max_top30_differences_crop_indices_3_new_num_frames] 
        
            max_top30_differences_crop_values_3_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_3_new) 
            index_in_Frames_3 = max_top30_differences_crop_indices_3_new_num_frames[max_top30_differences_crop_values_3_new_max_in_Frames]
            
            value_in_Frames_3 =  max_top30_differences_crop_values_3_new[index_in_Frames_3]
            location_in_Frames_3 = max_top30_diff_locations_crop_3_new[index_in_Frames_3]
            
            max_top30_differences_crop_indices_3_index_new = max_top30_differences_crop_indices_3_new[0]
            max_top30_differences_crop_indices_3_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_3_index_new + start_number :04d}.{FRAME_EXTENSION}')
            

        max_top30_differences_crop_except0_3_new = [value for value in max_top30_differences_crop_3_new if value != 0]
        mean_max_top30_differences_crop3_new = np.mean(max_top30_differences_crop_3_new)
        mean_round_max_top30_differences_crop3_new = round(mean_max_top30_differences_crop3_new, 2)
        max_max_top30_differences_crop3_new = max(max_top30_differences_crop_3_new)
        min_max_top30_differences_crop_except0_3_new = min(max_top30_differences_crop_except0_3_new)
        min_max_top30_differences_crop_except0_3_index_new = max_top30_differences_crop_3_new.index(max_max_top30_differences_crop3_new)
        
        top30_max_image_path_4_new = os.path.join(folder_path, f'{FRAME_PREFIX}{min_max_top30_differences_crop_except0_3_index_new + start_number :04d}.{FRAME_EXTENSION}') 
        
        x_differences = []  
        y_differences = []
        valid_pixel_counts_min_new = []
        valid_pixel_counts_max_new = []
        valid_pixel_all_counts_max_new = []
        valid_pixel_all_counts_min_new = []

        for i in range(len(gray_frames)):

            if use_gpu:
                image1 = gray_tensor[i].cpu().numpy().astype(np.uint8)
            else:
                image1 = gray_frames[i].astype(np.uint8)


            max_image_path_pixel_new = top30_max_image_path_3_new 
            min_image_path_pixel_new = top30_max_image_path_2_new 
            
            max_image_pixel_new = current_rgb_image_list[image_frame_n2]
            min_image_pixel_new = current_rgb_image_list[image_frame_n1]
            max_image_gray_pixel_new = cv2.cvtColor(max_image_pixel_new, cv2.COLOR_BGR2GRAY)
            min_image_gray_pixel_new = cv2.cvtColor(min_image_pixel_new, cv2.COLOR_BGR2GRAY)
            
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new] #small_size
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_image0_pixel_new = image0[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_image1_pixel_new = image1[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f'Error cropping image{i}. Skipping...')
                continue 
            
            diff_max_new = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
            diff_min_new = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 
            
            # Count pixels with intensity differences above the nonzero minimum threshold.
            # KR: 0을 제외한 최소 threshold 이상으로 변화한 pixel의 좌표를 계산합니다.
            threshold_max_new = min_max_top30_differences_crop_except0_3_new 
            threshold_min_new = min_max_top30_differences_crop_except0_2_new  
            white_pixel_coords_max_new = np.argwhere(diff_max_new >= threshold_max_new)
            white_pixel_coords_min_new = np.argwhere(diff_min_new >= threshold_min_new)
            
            min_threshold_max_new = min_max_top30_differences_crop_except0_3_new +20 
            max_threshold_max_new = max_max_top30_differences_crop3_new 

            min_threshold_min_new = min_max_top30_differences_crop_except0_2_new +20 
            max_threshold_min_new = max_max_top30_differences_crop2_new 
            
            valid_pixel_coords_max_new = np.argwhere((diff_max_new >= min_threshold_max_new) & (diff_max_new <= max_threshold_max_new))
            valid_pixel_coords_min_new = np.argwhere((diff_min_new >= min_threshold_min_new) & (diff_min_new <= max_threshold_min_new))
            
            valid_pixel_coords_max_all_new = np.argwhere((diff_max_new >= 0) & (diff_max_new <= max_threshold_max))
            valid_pixel_coords_min_all_new = np.argwhere((diff_min_new >= 0) & (diff_min_new <= max_threshold_min))

            valid_pixel_count_max_new = len(valid_pixel_coords_max_new)
            valid_pixel_count_min_new = len(valid_pixel_coords_min_new)
            valid_pixel_counts_max_new.append(valid_pixel_count_max_new)
            valid_pixel_counts_min_new.append(valid_pixel_count_min_new)       


        valid_pixel_count_mean_min_new = np.mean([value for value in valid_pixel_counts_min_new if value not in [0,'']]) 

        valid_pixel_count_mean_round_min_new = round(valid_pixel_count_mean_min_new) if not np.isnan(valid_pixel_count_mean_min_new) else 0 
        
        valid_pixel_count_mean_max_new = np.mean([value for value in valid_pixel_counts_max_new if value not in [0,'']]) 

        valid_pixel_count_mean_round_max_new = round(valid_pixel_count_mean_max_new) if not np.isnan(valid_pixel_count_mean_max_new) else 0 

        count_zeros_min_new = valid_pixel_counts_min_new.count(0)
        count_zeros_max_new = valid_pixel_counts_max_new.count(0)


        count_all_pixel_in_range_min_new = [valid_pixel_counts_min_new.count(i) for i in range(valid_pixel_count_mean_round_min_new)]
        count_all_pixel_in_range_max_new = [valid_pixel_counts_max_new.count(i) for i in range(valid_pixel_count_mean_round_max_new)]
        
        sum_count_all_pixel_in_range_min_new = sum(count_all_pixel_in_range_min_new, 0)
        sum_count_all_pixel_in_range_max_new = sum(count_all_pixel_in_range_max_new, 0)

        valid_pixel_all_counts_min_new.append(sum_count_all_pixel_in_range_min_new)
        valid_pixel_all_counts_max_new.append(sum_count_all_pixel_in_range_max_new)
        
        valid_pixel_all_counts_min_max_new = max(valid_pixel_all_counts_min_new)
        valid_pixel_all_counts_max_max_new = max(valid_pixel_all_counts_max_new)

        All_pixel_intensities = []  
        valid_pixel_counts = []
        not_cropped_valid_pixel_counts = []
        selected_max_differences = []
        selected_max_differences_location = []
        Analysis_frames = []
        pixel_intensities_movement_list = []
        pixel_intensities_movement_coordinates_list = []



        for i in range(len(gray_frames)):

            max_image_path_pixel_new = top30_max_image_path_3_new
            min_image_path_pixel_new = top30_max_image_path_2_new
            
            max_image_pixel_new = current_rgb_image_list[image_frame_n2]
            min_image_pixel_new = current_rgb_image_list[image_frame_n1]
            max_image_gray_pixel_new = cv2.cvtColor(max_image_pixel_new, cv2.COLOR_BGR2GRAY)
            min_image_gray_pixel_new = cv2.cvtColor(min_image_pixel_new, cv2.COLOR_BGR2GRAY)
            
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_image0_pixel_new = image0[y1_small_new:y2_small_new, x1_small_new:x2_small_new]  
            cropped_image1_pixel_new = image1[y1_small_new:y2_small_new, x1_small_new:x2_small_new]  
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f"Error cropping image {i}. Skipping...")
                continue 
            
            # Select the reference candidate frame based on valid-pixel count statistics.
            # The selected candidate determines which images, thresholds, coordinates,
            # and intensity-difference values are used in downstream displacement analysis.
            # KR: valid-pixel count 통계를 기준으로 기준 후보 frame을 선택합니다.
            # KR: 선택된 후보 frame에 따라 이후 displacement 분석에 사용할 이미지, threshold, 좌표, intensity-difference 값이 결정됩니다.
            
            # Apply default or reversed candidate-selection logic depending on the expected contraction direction.
            # KR: 예상되는 수축 방향에 따라 기본 후보 선택 기준 또는 반전된 후보 선택 기준을 적용합니다.
            if REVERSE_CONTRACTION.lower() == 'off':
                if valid_pixel_all_counts_min_max_new >= valid_pixel_all_counts_max_max_new:
                    # Use the candidate relaxation frame as the reference image
                    # KR: 후보 이완 frame을 기준 이미지로 사용합니다.
                    diff = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new)  
                    # Full-frame difference used for visualization
                    # KR: visualization을 위해 full-frame difference를 계산합니다.
                    not_crop_diff = cv2.subtract(max_image_gray_pixel_new, image1) 
                    # Difference between candidate contraction and relaxation frames
                    # KR: 후보 수축 frame과 후보 이완 frame 사이의 차이입니다.
                    crop_diff = cv2.subtract(cropped_max_image_pixel_new, cropped_min_image_pixel_new) 
                    #=================================================================
                    # Store paired candidate contraction/relaxation frame information
                    # KR: 후보 수축/이완 frame 쌍의 정보를 저장합니다.
                    #=================================================================
                    # Reference image: candidate relaxation frame
                    # KR: 기준 이미지: 후보 이완 frame
                    square_image = min_image_pixel_new
                    # Path to the reference relaxation frame
                    # KR: 기준 이완 frame의 경로
                    square_image_path = min_image_path_pixel_new
                    # Paired candidate contraction frame
                    # KR: 기준 frame과 짝이 되는 후보 수축 frame
                    square_image_path_contraction = max_image_path_pixel_new 
                    # Loaded RGB image of the reference relaxation frame
                    # KR: 기준 이완 frame의 RGB 이미지
                    square_image_imread = current_rgb_image_list[image_frame_n1]
                    # Loaded RGB image of the paired contraction frame
                    # KR: 짝이 되는 후보 수축 frame의 RGB 이미지
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n2]
                    # Index of the selected reference frame
                    # KR: 선택된 기준 frame의 index
                    square_image_index = image_frame_n1
                    # Image showing the maximum difference from the reference frame
                    # KR: 기준 frame과 가장 큰 차이를 보이는 이미지
                    square_differences_crop_indices_index_path_image = max_image_pixel_new 
                    # Minimum intensity-difference threshold
                    # KR: 최소 intensity-difference threshold
                    threshold = min_max_top30_differences_crop_except0_2_new 
                    # Maximum intensity-difference threshold
                    # KR: 최대 intensity-difference threshold
                    max_threshold = max_max_top30_differences_crop2_new
                    # Maximum intensity difference
                    # KR: intensity difference 값
                    diff_val = max_diff_crop_top30_max_2_new   
                    # List of maximum intensity differences
                    # KR: 최대 intensity difference 리스트
                    diff2_val = max_top30_differences_crop_2_new
                    # Coordinates of the maximum intensity difference
                    # KR: 최대 intensity difference 좌표
                    loc = max_diff_loc_crop_top30_max_2_new
                    # Coordinate of the maximum intensity difference within the initial analysis window
                    # KR: 초기 분석 구간 안에서 최대 intensity difference가 나타난 좌표
                    loc2 = max_top30_diff_locations_crop_2_new
                    # Coordinate of the maximum intensity difference within the initial analysis window
                    # KR: 초기 분석 구간 안에서 최대 intensity difference가 나타난 좌표
                    loc3 = location_in_Frames_2 
                    # Maximum intensity difference from the reference frame
                    # KR: 기준 frame과 가장 큰 차이를 보이는 intensity difference 값
                    max_values = max_max_top30_differences_crop2_new 
                    ## Index of the frame showing the maximum difference from the reference frame
                    # KR: 기준 frame과 가장 큰 차이를 보이는 frame의 index
                    max_indices = min_max_top30_differences_crop_except0_2_index_new 
                    # List of maximum intensity differences
                    # KR: 최대 intensity difference 리스트
                    adapt = max_top30_differences_crop_2_new 
                    # Maximum intensity-difference value
                    # KR: 최대 intensity-difference 값
                    adapt_max = max_max_top30_differences_crop2_new 
                    # Mean intensity-difference value
                    # KR: 평균 intensity-difference 값
                    adapt_mean = mean_round_max_top30_differences_crop2_new 
                    
                else:
                    diff = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
                    not_crop_diff = cv2.subtract(min_image_gray_pixel_new, image1) 
                    crop_diff = cv2.subtract(cropped_min_image_pixel_new, cropped_max_image_pixel_new) 
                    square_image = max_image_pixel_new 
                    square_image_path = max_image_path_pixel_new 
                    square_image_path_contraction = max_image_path_pixel_new 
                    square_image_imread = current_rgb_image_list[image_frame_n2] 
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n1] 
                    square_image_index = image_frame_n2 
                    square_differences_crop_indices_index_path_image = min_image_pixel_new 
                    threshold = min_max_top30_differences_crop_except0_3_new  
                    max_threshold = max_max_top30_differences_crop3_new 
                    diff_val = max_diff_crop_top30_max_3_new 
                    diff2_val = max_top30_differences_crop_3_new 
                    loc = max_diff_loc_crop_top30_max_3_new 
                    loc2 = max_top30_diff_locations_crop_3_new 
                    loc3 = location_in_Frames_3 
                    max_values = max_max_top30_differences_crop3_new 
                    max_indices = min_max_top30_differences_crop_except0_3_index_new 
                    adapt = max_top30_differences_crop_3_new 
                    adapt_max = max_max_top30_differences_crop3_new 
                    adapt_mean = mean_round_max_top30_differences_crop3_new 
            else: 
                if valid_pixel_all_counts_min_max_new <= valid_pixel_all_counts_max_max_new:
                    diff = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 
                    not_crop_diff = cv2.subtract(max_image_gray_pixel_new, image1) 
                    crop_diff = cv2.subtract(cropped_max_image_pixel_new, cropped_min_image_pixel_new) 
                    square_image = min_image_pixel_new 
                    square_image_path = min_image_path_pixel_new 
                    square_image_path_contraction = max_image_path_pixel_new 
                    square_image_imread = current_rgb_image_list[image_frame_n1] 
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n2] 
                    square_image_index = image_frame_n1
                    square_differences_crop_indices_index_path_image = max_image_pixel_new 
                    threshold = min_max_top30_differences_crop_except0_2_new 
                    max_threshold = max_max_top30_differences_crop2_new 
                    diff_val = max_diff_crop_top30_max_2_new  
                    diff2_val = max_top30_differences_crop_2_new 
                    loc = max_diff_loc_crop_top30_max_2_new 
                    loc2 = max_top30_diff_locations_crop_2_new 
                    loc3 = location_in_Frames_2 
                    max_values = max_max_top30_differences_crop2_new 
                    max_indices = min_max_top30_differences_crop_except0_2_index_new 
                    adapt = max_top30_differences_crop_2_new 
                    adapt_max = max_max_top30_differences_crop2_new 
                    adapt_mean = mean_round_max_top30_differences_crop2_new 
                    
                else:
                    diff = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
                    not_crop_diff = cv2.subtract(min_image_gray_pixel_new, image1) 
                    crop_diff = cv2.subtract(cropped_min_image_pixel_new, cropped_max_image_pixel_new) 
                    square_image = max_image_pixel_new 
                    square_image_path = max_image_path_pixel_new 
                    square_image_path_contraction = max_image_path_pixel_new 
                    square_image_imread = current_rgb_image_list[image_frame_n2] 
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n1] 
                    square_image_index = image_frame_n2 
                    square_differences_crop_indices_index_path_image = min_image_pixel_new 
                    threshold = min_max_top30_differences_crop_except0_3_new  
                    max_threshold = max_max_top30_differences_crop3_new 
                    diff_val = max_diff_crop_top30_max_3_new 
                    diff2_val = max_top30_differences_crop_3_new 
                    loc = max_diff_loc_crop_top30_max_3_new 
                    loc2 = max_top30_diff_locations_crop_3_new  
                    loc3 = location_in_Frames_3 
                    max_values = max_max_top30_differences_crop3_new 
                    max_indices = min_max_top30_differences_crop_except0_3_index_new 
                    adapt = max_top30_differences_crop_3_new 
                    adapt_max = max_max_top30_differences_crop3_new 
                    adapt_mean = mean_round_max_top30_differences_crop3_new 
                

            max_pixel_intensities = int(not_crop_diff.max())
            All_pixel_intensities.append(max_pixel_intensities)
            max_All_pixel_intensities = int(np.max(All_pixel_intensities))
            All_pixel_intensities.append(max_pixel_intensities)
            max_All_pixel_intensities = np.max(All_pixel_intensities)

            selected_image_path = square_image_path  
            
            selected_image = square_image_imread 
            selected_image_contraction = square_image_imread_contraction 
            selected_image_gray = cv2.cvtColor(selected_image, cv2.COLOR_BGR2GRAY) 
            selected_image_contraction_gray = cv2.cvtColor(selected_image_contraction, cv2.COLOR_BGR2GRAY) 
            
            all_image_gray = current_gray_image_list[i] 
            adapt = np.array(adapt) 
            adapt_min = adapt[adapt >= threshold].tolist()
            try:
                sorted_adapt_min = min(adapt_min)
                sorted_adapt_max = max(adapt_min)
                index_of_sorted_adapt_min = int(np.where(adapt == sorted_adapt_min)[0][0])
                index_of_sorted_adapt_max = int(np.where(adapt == sorted_adapt_max)[0][0])
                coordinate_of_sorted_adapt_min = loc2[index_of_sorted_adapt_min]
                coordinate_of_sorted_adapt_max = loc2[index_of_sorted_adapt_max]
            except ValueError:
                break
            
            cropped_selected_image_gray = selected_image_gray[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
            cropped_all_image_gray = all_image_gray[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
            
            cropped_selected_image_gray_reference = selected_image_gray[y1_small_new_optical:y2_small_new_optical, x1_small_new_optical:x2_small_new_optical] 
            cropped_all_image_gray_reference = all_image_gray[y1_small_new_optical:y2_small_new_optical, x1_small_new_optical:x2_small_new_optical] 
            
            
            diff_selected = cv2.absdiff( cropped_all_image_gray, cropped_selected_image_gray) 
            diff_selected_reference = cv2.subtract( cropped_all_image_gray_reference, cropped_selected_image_gray_reference)

            pixel_intensities_movement = diff_selected.ravel()

            diff_selected_column = np.argwhere(diff_selected >= threshold)

            pixel_intensities_movement_list.append(pixel_intensities_movement)
            pixel_intensities_movement_coordinates_list.append(diff_selected_column)

            max_pixel_intensity = int(diff_selected.max()) if diff_selected.size else 0

            
            reference_selected_image_gray = selected_image_contraction_gray[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference]
            reference_all_image_gray = all_image_gray[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
            
            not_crop_selected_image_gray = selected_image_contraction_gray 
            not_crop_all_image_gray = all_image_gray

            reference_diff_selected = cv2.absdiff(reference_selected_image_gray, reference_all_image_gray) 
            not_crop_diff_selected = cv2.absdiff(not_crop_selected_image_gray, not_crop_all_image_gray) 
            
            diff_loc = np.argwhere((diff_selected_reference >= threshold) & (diff_selected_reference <= max_threshold))
            not_cropped_diff_loc = np.argwhere((reference_diff_selected >= threshold) & (reference_diff_selected <= max_threshold)) 
            
            threshold_pixel_image = np.zeros_like(diff_selected_reference, dtype = np.uint8)
            threshold_pixel_image_color = cv2.cvtColor(threshold_pixel_image, cv2.COLOR_GRAY2BGR)
            for loc in diff_loc:
                threshold_pixel_image_color[loc[0], loc[1]] = diff_selected_reference[loc[0], loc[1]]
            
            result_folder_diff = os.path.join(folder_path, 'result', 'Analysis Square') 
            if not os.path.exists(result_folder_diff):
                os.makedirs(result_folder_diff, exist_ok = True) 
            x_coord, y_coord = x_selected_new - x1_small_new_optical ,y_selected_new - y1_small_new_optical
        
            height, width = diff_selected.shape[:2]
            not_cropped_height, not_cropped_width = reference_diff_selected.shape[:2]

            new_width = 1080

            new_height = int(height * (new_width / width))
            diff_resized = cv2.resize(threshold_pixel_image_color, (new_width, new_height), interpolation=cv2.INTER_LINEAR) 
            result_image_path = os.path.join(result_folder_diff, f'Analysis_Frame_{i:04d}.{FRAME_EXTENSION}')
            Analysis_frames.append(diff_resized)
            white_pixel_coords = np.argwhere(diff_selected >= adapt_mean)
            
            not_cropped_white_pixel_coords = np.argwhere(reference_diff_selected >= adapt_mean)
            
            white_pixel_count = len(white_pixel_coords)
            
            not_cropped_white_pixel_count = len(not_cropped_white_pixel_coords)
            
            min_threshold = threshold
            max_threshold = max_threshold
            
            valid_pixel_coords = np.argwhere((diff_selected >= min_threshold) & (diff_selected <= max_threshold))
            not_cropped_valid_pixel_coords = np.argwhere((reference_diff_selected >= min_threshold) & (reference_diff_selected <= max_threshold))
            
            valid_pixel_count = len(valid_pixel_coords)
            valid_pixel_counts.append(valid_pixel_count)
            total_pixel_count = resolution_x * resolution_y 
            
            not_cropped_valid_pixel_count = len(not_cropped_valid_pixel_coords)
            not_cropped_valid_pixel_counts.append(not_cropped_valid_pixel_count)   



        valid_pixel_counts_np = np.array(valid_pixel_counts, dtype='float32') 

        valid_values = valid_pixel_counts_np[(valid_pixel_counts_np != 0) & ~np.isnan(valid_pixel_counts_np)]

        valid_pixel_count_mean = np.mean(valid_values) if valid_values.size > 0 else 0
        valid_pixel_count_mean_round = round(valid_pixel_count_mean)

        valid_pixel_count_maximum = np.max(valid_values) if valid_values.size > 0 else 0
        valid_pixel_count_maximum_round = round(valid_pixel_count_maximum)

        min_valid_pixel_counts = np.min(valid_values) if valid_values.size > 0 else 0
        
        if not valid_pixel_counts:
            print("!!!!THIS SAMPLE DOESN'T HAVE ANY MOVEMENT!!!!.")
            classification_dir = os.path.join(BAD_FILE_PATH, 'FINISHED')
            destination_folder = os.path.join(classification_dir, 'BAD')
            folder_path = os.path.join(BAD_FILE_PATH, folder_name)
            video_path = os.path.join(BAD_FILE_PATH, f'{folder_name}.{extension}')  
            os.makedirs(classification_dir, exist_ok=True)
            os.makedirs(destination_folder, exist_ok=True)
            try:
                shutil.move(folder_path, destination_folder)
                shutil.move(video_path, destination_folder)
            except Exception as e:
                pass
            continue
        

        difference_square_image = cv2.subtract(square_differences_crop_indices_index_path_image, square_image)
        difference_square_image_gray = cv2.cvtColor(difference_square_image, cv2.COLOR_BGR2GRAY)
        


        
   
        
        # 동영상 파일명과 확장자 설정
        video_name_base = f'{result_folder_diff}\\Analysis_Video'
        video_name = f'{video_name_base}\\{folder_name}_Analysis_Video.avi'
        if not os.path.exists(video_name_base):
            os.makedirs(video_name_base)
        # VideoWriter 객체 생성
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(video_name, fourcc, fps, (new_width, new_height))
        # 프레임 이미지들을 읽어와 동영상에 추가
        for i in range(len(gray_frames)):
            out.write(Analysis_frames[i])

        # VideoWriter 객체 해제
        out.release()

        gray_color = '#bbbbbd'
        red_color = '#FF6699'
        blue_color = '#009fd6'
        yellow_color = '#d69f00'



        # ============================================
        # Optical flow-based displacement analysis
        # KR: optical flow 기반 displacement 분석
        # ============================================

        # Prepare candidate contraction and relaxation frames for ROI-based optical flow.
        # KR: ROI 기반 optical flow 분석을 위해 후보 수축/이완 frame을 준비합니다.
        roi_contraction = cv2.cvtColor(
            square_differences_crop_indices_index_path_image, cv2.COLOR_BGR2GRAY
        )
        roi_relaxation = cv2.cvtColor(
            square_image, cv2.COLOR_BGR2GRAY
        )

        roi_contraction = roi_contraction[
            y1_small_new_optical:y2_small_new_optical,
            x1_small_new_optical:x2_small_new_optical
        ]
        roi_relaxation = roi_relaxation[
            y1_small_new_optical:y2_small_new_optical,
            x1_small_new_optical:x2_small_new_optical
        ]

        # Skip optical flow when the ROI is empty.
        # KR: ROI가 비어 있으면 optical flow 분석을 건너뜁니다.
        if roi_contraction.size == 0:
            print("roi_contraction is empty. Optical flow skipped.")
            target_displacements = [0.0]
            maximum_displacement_list = target_displacements
            max_target_displacements = 0.0
            max_displacement_um = 0.0
            min_target_displacements = 0.0

            # Use the selected coordinate as the fallback position.
            # KR: optical flow를 수행할 수 없을 때 선택 좌표를 fallback 위치로 사용합니다.
            a_global = x_selected_new
            b_global = y_selected_new
            c_global = x_selected_new
            d_global = y_selected_new

            global_max_coords_global = (b_global, a_global, d_global, c_global)
            global_max_coords_global1 = (b_global, a_global)
            global_max_coords_global2 = (d_global, c_global)
            global_max_coords_global1_ab = (a_global, b_global)
            global_max_coords_global2_cd = (c_global, d_global)

            tracked_points = [(x_selected_new - x1_small_new_optical,
                            y_selected_new - y1_small_new_optical)]
            best_x_roi = tracked_points[0][0]
            best_y_roi = tracked_points[0][1]
            best_x_roi_min = best_x_roi
            best_y_roi_min = best_y_roi

        else:
            # Set the output path for the optical flow overlay video.
            # KR: optical flow overlay video의 저장 경로를 설정합니다.
            output_video_path = os.path.join(result_folder, f'{folder_name}_Optical_Flow_Overlay.{extension}')

            fps = fps  
            frame_size = (roi_contraction.shape[1], roi_contraction.shape[0])  # Video frame size: width, height

            fourcc = cv2.VideoWriter_fourcc(*'XVID')
            out = cv2.VideoWriter(output_video_path, fourcc, fps, frame_size)

            # ------------------------------------------------
            # 2) Define the baseline tracking point within the ROI
            #    The selected motion center is clamped to remain inside the ROI.
            # KR: ROI 내부에서 baseline tracking point를 정의합니다.
            # KR: 선택된 motion center가 ROI 밖으로 벗어나지 않도록 좌표를 제한합니다.
            # ------------------------------------------------
            H, W = roi_contraction.shape[:2]

            x0 = int(x_selected_new - x1_small_new_optical)
            y0 = int(y_selected_new - y1_small_new_optical)

            half_win = ANALYSIS_SQUARE_SIZE_REFERENCE // 2    # Half-window size for coordinate clamping
            # Clamp the tracking point so that the local window remains inside the ROI.
            # KR: local tracking window가 ROI 내부에 유지되도록 tracking point 좌표를 제한합니다.
            x0 = max(half_win, min(W - 1 - half_win, x0))
            y0 = max(half_win, min(H - 1 - half_win, y0))

            # Use a single baseline tracking point for Lucas-Kanade optical flow.
            # KR: Lucas-Kanade optical flow에서 하나의 baseline tracking point만 사용합니다.
            p0 = np.array([[[x0, y0]]], dtype=np.float32)

            # Store optical-flow tracking results.
            # KR: optical flow tracking 결과를 저장합니다.
            target_displacements = []   # Displacement for each frame in µm
            tracked_points = []         # Tracked ROI coordinates for each frame

            # Smoothing parameters for visualization only.
            # KR: visualization에만 사용하는 smoothing 설정입니다.
            smoothed_x, smoothed_y = None, None
            alpha = 0.5  # EMA smoothing factor

            # ------------------------------------------------
            # 3) Optical flow tracking from the baseline ROI to each frame ROI
            #    A single point p0 is tracked to define optical flow-derived displacement.
            # KR: baseline ROI에서 각 frame ROI로 optical flow tracking을 수행합니다.
            # KR: 하나의 p0 point를 추적하여 optical flow-derived displacement를 정의합니다.
            # ------------------------------------------------
            for frame_idx, frame_full in enumerate(current_gray_image_list):

                cur_gray = frame_full[
                    y1_small_new_optical:y2_small_new_optical,
                    x1_small_new_optical:x2_small_new_optical
                ]

                # Keep the previous displacement value when the current ROI is empty.
                # KR: 현재 ROI가 비어 있으면 이전 displacement 값을 유지합니다.
                if cur_gray.size == 0:
                    last_disp = target_displacements[-1] if target_displacements else 0.0
                    target_displacements.append(float(last_disp))
                    tracked_points.append((float(x0), float(y0)))
                    continue

                p1, st, err = cv2.calcOpticalFlowPyrLK(
                    roi_contraction,
                    cur_gray,
                    p0,
                    None,
                    winSize=(15, 15),
                    maxLevel=2,
                    criteria=(cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, 10, 0.03)
                )

                frame_rgb = cv2.cvtColor(cur_gray, cv2.COLOR_GRAY2BGR)

                # Handle optical-flow tracking failure.
                # KR: optical flow tracking에 실패한 경우를 처리합니다.
                if p1 is None or st is None or st[0, 0] == 0:
                    last_disp = target_displacements[-1] if target_displacements else 0.0
                    target_displacements.append(float(last_disp))
                    tracked_points.append((float(x0), float(y0)))

                    # Write the frame even when optical-flow tracking fails.
                    # KR: optical flow tracking에 실패해도 output video frame은 저장합니다.
                    out.write(frame_rgb)
                    continue

                # New tracked coordinate in the baseline ROI coordinate system.
                # KR: baseline ROI 좌표계에서 새롭게 추적된 좌표입니다.
                new_x = float(p1[0, 0, 0])
                new_y = float(p1[0, 0, 1])

                # Calculate displacement from the baseline point p0.
                # This is not cumulative; the same baseline point is used for all frames.
                # KR: baseline point p0로부터의 displacement를 계산합니다.
                # KR: 누적 변위가 아니라 모든 frame에서 동일한 기준점 p0를 사용합니다.
                dx = new_x - x0
                dy = new_y - y0
                disp_um = np.sqrt(dx * dx + dy * dy) * PIXEL_UM

                target_displacements.append(float(disp_um))
                tracked_points.append((new_x, new_y))

                # Smooth the displayed tracking point for visualization.
                # KR: visualization에서 표시되는 tracking point를 부드럽게 보이도록 smoothing합니다.
                if smoothed_x is None:
                    smoothed_x, smoothed_y = new_x, new_y
                else:
                    smoothed_x = alpha * new_x + (1.0 - alpha) * smoothed_x
                    smoothed_y = alpha * new_y + (1.0 - alpha) * smoothed_y

                # Draw a green arrow from the baseline point to the current tracked position,
                # and mark the smoothed tracked point in red.
                # KR: baseline point에서 현재 추적 위치까지 초록색 화살표를 그리고,
                # KR: smoothing된 tracking point를 빨간색으로 표시합니다.
                cv2.arrowedLine(
                    frame_rgb,
                    (int(x0), int(y0)),
                    (int(new_x), int(new_y)),
                    (0, 255, 0),
                    1
                )
                cv2.circle(
                    frame_rgb,
                    (int(smoothed_x), int(smoothed_y)),
                    2,
                    (0, 0, 255),
                    -1
                )

                out.write(frame_rgb)
            # Finalize the optical flow overlay video.
            # KR: optical flow overlay video 저장을 종료합니다.
            out.release()

            # ------------------------------------------------
            # 4) Displacement time-series post-processing
            #    - target_displacements are stored in micrometers.
            #    - The analysis window is selected based on start_number, NUM_FRAMES, and fps.
            # KR: displacement time-series를 후처리합니다.
            # KR: target_displacements는 micrometer 단위입니다.
            # KR: 분석 구간은 start_number, NUM_FRAMES, fps를 기준으로 선택합니다.
            # ------------------------------------------------
            if len(target_displacements) == 0:
                target_displacements = [0.0]

            # Shift the minimum displacement to zero.
            # KR: 최소 displacement가 0이 되도록 baseline을 보정합니다.
            min_all = min(target_displacements)
            target_displacements = [float(v - min_all) for v in target_displacements]
            
            # Invert the displacement curve for visualization consistency.
            # KR: visualization 방향을 맞추기 위해 displacement curve를 반전합니다.
            max_all = max(target_displacements)
            target_displacements = [float(max_all - v) for v in target_displacements]

            # Select the initial analysis window.
            # KR: 초기 분석 구간을 선택합니다.
            end_idx = int(fps * NUM_FRAMES)
            end_idx = min(end_idx, len(target_displacements))

            if start_number + 1 < end_idx:
                window_vals = target_displacements[start_number + 1:end_idx]
                window_indices = list(range(start_number + 1, end_idx))
            else:
                window_vals = target_displacements
                window_indices = list(range(len(target_displacements)))

            if len(window_vals) == 0:
                window_vals = target_displacements
                window_indices = list(range(len(target_displacements)))

            # Define the final displacement time series and maximum displacement value.
            # KR: 최종 displacement time series와 maximum displacement 값을 정의합니다.
            maximum_displacement_list = target_displacements
            max_target_displacements = round(max(window_vals), 2)
            min_target_displacements = float(min(window_vals))
            max_displacement_um = max_target_displacements

            max_target_displacements_before_all = round(max(target_displacements), 2)

            # Create the time axis for displacement plotting.
            # KR: displacement plot에 사용할 시간축을 생성합니다.
            if fps > 0:
                total_time_seconds = len(maximum_displacement_list) / fps
            else:
                total_time_seconds = float(len(maximum_displacement_list))
            time_axis = np.linspace(0, total_time_seconds, len(maximum_displacement_list))

            # Frame index where the maximum displacement occurs within the analysis window.
            # KR: 분석 구간 안에서 maximum displacement가 발생한 frame index입니다.
            idx_in_window = int(np.argmax(window_vals))
            frame_idx_max = int(window_indices[idx_in_window])

            # Extract ROI coordinates of the tracked point at maximum and baseline displacement.
            # KR: maximum displacement와 baseline displacement 시점의 tracked point ROI 좌표를 가져옵니다.
            if frame_idx_max < len(tracked_points):
                best_x_roi, best_y_roi = tracked_points[frame_idx_max]
            else:
                best_x_roi, best_y_roi = tracked_points[-1]

            if start_number < len(tracked_points):
                best_x_roi_min, best_y_roi_min = tracked_points[start_number]
            else:
                best_x_roi_min, best_y_roi_min = tracked_points[0]

            # Convert ROI coordinates back to full-frame coordinates.
            # KR: ROI 좌표를 full-frame 좌표로 변환합니다.
            a_global = int(best_x_roi + x1_small_new_optical)
            b_global = int(best_y_roi + y1_small_new_optical)
            c_global = int(best_x_roi_min + x1_small_new_optical)
            d_global = int(best_y_roi_min + y1_small_new_optical)

            global_max_coords_global = (b_global, a_global, d_global, c_global)
            global_max_coords_global1 = (b_global, a_global)
            global_max_coords_global2 = (d_global, c_global)
            global_max_coords_global1_ab = (a_global, b_global)
            global_max_coords_global2_cd = (c_global, d_global)
            
    
        # Convert the displacement time series to a comma-separated string for Excel output.
        # KR: Excel output에 저장하기 위해 displacement time series를 comma-separated string으로 변환합니다.
        maximum_displacement_str = ", ".join(f"{x:.2f}" for x in maximum_displacement_list)
                
        y_max_coordinates = int(y_selected_new - y1_small_new_optical)
        x_max_coordinates = int(x_selected_new - x1_small_new_optical)

        x1_max_coordinates = x_max_coordinates  - ANALYSIS_SQUARE_SIZE_REFERENCE
        y1_max_coordinates = y_max_coordinates  - ANALYSIS_SQUARE_SIZE_REFERENCE
        x2_max_coordinates = x_max_coordinates  + ANALYSIS_SQUARE_SIZE_REFERENCE
        y2_max_coordinates = y_max_coordinates  + ANALYSIS_SQUARE_SIZE_REFERENCE
        
        x1_max_coordinates = max( x1_max_coordinates, 0)
        y1_max_coordinates = max( y1_max_coordinates, 0)
        x2_max_coordinates = min( x2_max_coordinates, resolution_x -1)
        y2_max_coordinates = min( y2_max_coordinates, resolution_y -1)

        max_value_displacement = max(maximum_displacement_list) 
        
        if max_target_displacements == 0 or max_target_displacements is None:
            print("No valid displacement was detected.")
            print("This sample could not be tracked by optical flow.")
            print(f"Move to BAD Folder")
            classification_dir = os.path.join(BAD_FILE_PATH, 'FINISHED')
            destination_folder = os.path.join(classification_dir, 'BAD')
            folder_path = os.path.join(BAD_FILE_PATH, folder_name)
            video_path = os.path.join(BAD_FILE_PATH, f'{folder_name}.{extension}')  
            os.makedirs(classification_dir, exist_ok=True)
            os.makedirs(destination_folder, exist_ok=True)
            try:
                shutil.move(folder_path, destination_folder)
                shutil.move(video_path, destination_folder)
                continue
            except Exception as e:
                continue
        else:
            pass


        #=====================================================   image   ========================================================= 

            
        # =====================================================
        # Analysis ROI and tracked-point visualization
        # =====================================================
        # This figure provides a quality-control view of the selected analysis ROI
        # and tracked points on the maximum relaxation and maximum contraction frames.
        # The green point indicates the tracked point at maximum displacement,
        # the orange point indicates the baseline/minimum-displacement point,
        # and the red point indicates the initially selected analysis point.
        # The purple rectangle shows the local ROI used for optical-flow tracking.
        #
        # KR:
        # 이 figure는 maximum relaxation frame과 maximum contraction frame에서
        # 선택된 analysis ROI와 tracked point를 확인하기 위한 QC 시각화입니다.
        # 초록색 점은 maximum displacement 시점의 tracked point,
        # 주황색 점은 baseline/minimum-displacement 시점의 point,
        # 빨간색 점은 처음 선택된 analysis point를 의미합니다.
        # 보라색 사각형은 optical-flow tracking에 사용된 local ROI를 나타냅니다.
        # =====================================================

        best_x_roi_min, best_y_roi_min = tracked_points[start_number]

        # Convert ROI-local coordinates to full-image coordinates.
        # KR: ROI 내부 좌표를 원본 이미지 기준 좌표로 변환합니다.
        cx_max = safe_int(best_x_roi + x1_small_new_optical)
        cy_max = safe_int(best_y_roi + y1_small_new_optical)

        cx_min = safe_int(best_x_roi_min + x1_small_new_optical)
        cy_min = safe_int(best_y_roi_min + y1_small_new_optical)

        cx_selected = safe_int(x_selected_new)
        cy_selected = safe_int(y_selected_new)

        roi_top_left = (
            safe_int(x1_small_new_optical),
            safe_int(y1_small_new_optical)
        )
        roi_bottom_right = (
            safe_int(x2_small_new_optical),
            safe_int(y2_small_new_optical)
        )

        # Create visualization copies to avoid modifying the original analysis images.
        # KR: 원본 분석 이미지가 바뀌지 않도록 visualization용 복사본을 만듭니다.
        contraction_vis = square_differences_crop_indices_index_path_image.copy()
        relaxation_vis = square_image.copy()



        
        # Compute summary statistics after excluding zero-count frames.
        # KR: pixel count가 0인 frame을 제외하고 moved-region 통계값을 계산합니다.
        valid_pixel_count_mean = np.mean([value for value in valid_pixel_counts if value not in [0, '']])
        valid_pixel_count_mean_round = round(valid_pixel_count_mean) if not np.isnan(valid_pixel_count_mean) else 0
        valid_pixel_count_maximum = max([value for value in valid_pixel_counts if value not in [0, '']], default=0)
        valid_pixel_count_maximum_round = round(valid_pixel_count_maximum) if not np.isnan(valid_pixel_count_maximum) else 0

        # Convert moved-region pixel counts to physical area.
        # KR: moved region의 pixel count를 실제 면적으로 변환합니다.
        valid_pixel_count_mean_area = valid_pixel_count_mean_round * AREA_MOVED
        valid_pixel_count_maximum_area = valid_pixel_count_maximum_round * AREA_MOVED
        valid_pixel_count_mean_area_round = round(valid_pixel_count_mean_area,2)
        valid_pixel_count_maximum_area_round = round(valid_pixel_count_maximum_area, 2)
        
        # Frame-wise moved-region area and length-equivalent values.
        # KR: 각 frame별 moved-region area와 length-equivalent 값을 계산합니다.

        displacement_list = [count * PIXEL_UM for count in valid_pixel_counts]

        # =========================================================
        # Moved-region area time-series visualization
        # =========================================================
        # This plot shows the frame-wise moved-region area estimated from
        # thresholded motion pixels. The moved-region area is calculated by
        # multiplying the number of thresholded pixels by the calibrated pixel area.
        # This output is used as an auxiliary quality-control visualization and is
        # separate from optical-flow displacement.
        #
        # KR:
        # 이 plot은 threshold 조건을 통과한 moved pixel 수를 기반으로
        # frame별 moved-region area 변화를 보여줍니다.
        # moved-region area는 thresholded pixel count에 pixel area 보정값을
        # 곱하여 계산됩니다.
        # 이 결과는 보조적인 QC 시각화이며, optical-flow displacement 값과는
        # 별개의 지표입니다.
        # =========================================================
     
        

        fig, ax = plt.subplots(figsize=(8, 8))
        ax.plot(np.linspace(0, total_time_seconds, len(maximum_displacement_list)), 
                maximum_displacement_list, 
                color=blue_color, 
                linewidth=4)

        ax.set_xlabel('Time (s)',  fontsize=40, fontweight='bold', fontname='Arial')
        ax.set_ylabel(f'Displacement ({LENGTH_UNIT})',  fontsize=40,fontweight='bold', fontname='Arial')

        ax.set_ylim(0, Y_LIM)
        
        ax.set_yticks(range(0, Y_LIM + 1, 1))
        
        for spine in ax.spines.values():
            spine.set_linewidth(4) 
            
        ax.set_xlim(0, total_time_seconds)
        
        ax.set_xticks(range(0, int(total_time_seconds) + 1, X_TICKS))

        ax.axhline(y=max_target_displacements, color=red_color, linewidth=3.5, linestyle='--')

        ax.grid(False)
        plt.text(total_time_seconds - (total_time_seconds*0.0125), max_target_displacements + 0.625, f'Max Displacement ({LENGTH_UNIT}) : {max_target_displacements}',
                verticalalignment='top', horizontalalignment='right', color=red_color,fontweight='bold', fontsize=25)

        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=10)
        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=20)

        ax.spines['left'].set_position(('data', 0))  
        ax.spines['bottom'].set_position(('data', 0))  


        ax.spines['right'].set_color('black') 
        ax.spines['top'].set_color('black')    


        plt.xticks( fontweight='bold')
        plt.yticks( fontweight='bold')
        
        Savefig(folder_path, 'Optical_Flow_Displacement', 300, 0.2)
        
        seq_raw_np = np.array(maximum_displacement_list, dtype=np.float32)

        if len(seq_raw_np) == 0:
            max_abs_disp = 0.0
            abs_net_change_ratio = 0.0
            sign_changes = 0
            diff_std = 0.0
        else:
            max_abs_disp = float(np.max(np.abs(seq_raw_np)))

            peak_to_peak = float(np.max(seq_raw_np) - np.min(seq_raw_np))

            abs_net_change_ratio = float(
                np.abs(seq_raw_np[-1] - seq_raw_np[0]) / (peak_to_peak + 1e-8)
            )

            diff_seq = np.diff(seq_raw_np)

            if len(diff_seq) >= 2:
                sign_changes = int(np.sum(np.diff(np.sign(diff_seq)) != 0))
            else:
                sign_changes = 0

            diff_std = float(np.std(diff_seq)) if len(diff_seq) > 0 else 0.0
        
        band_energy_optical_displacement, max_value_optical_displacement_band_energy, max_optical_amp_at_max_freq, optical_freq_at_max_amp_number  = plot_fft(folder_path, fps, target_displacements, X_LIM_AMPLITUDE, Y_LIM_AMPLITUDE,'Optical_Flow_Displacement_FFT')
        
    
        Excel_result_data_bad = [
            folder_name,
            adapt_max,
            max_target_displacements,
            band_energy_optical_displacement,
            max_value_optical_displacement_band_energy,
            x_selected_new,
            y_selected_new,
            optical_freq_at_max_amp_number,
            maximum_displacement_str,
            max_abs_disp,
            abs_net_change_ratio,
            sign_changes,
            diff_std
        ]

        # 엑셀에 데이터 추가
        data_ws.append(Excel_result_data_bad)
        data_wb.save(BAD_EXCEL)
        print(f'Save The  『 {folder_name} 』 Information to {BAD_SAMPLE_EXCEL_NAME}')

        finish_dir = os.path.join(BAD_FILE_PATH, 'FINISHED')
        if not os.path.exists(finish_dir):
            os.makedirs(finish_dir, exist_ok=True)
        folder_path = os.path.join(BAD_FILE_PATH, folder_name)
        video_path = os.path.join(BAD_FILE_PATH, f'{folder_name}.{extension}')
        try:
            shutil.move(folder_path, finish_dir)
            shutil.move(video_path, finish_dir)
        except Exception as e:
            pass 

        # 현재 폴더의 처리 종료 시간 기록
        folder_end_time = time.time()
        
        # 현재 폴더의 실행 시간 계산 및 출력
        folder_execution_time = round((folder_end_time - folder_start_time),2)
        estimated_execution_times.append(folder_execution_time)  # 폴더의 실행 시간을 리스트에 추가

        # 각 폴더의 예상 처리 시간 계산 및 출력
        average_execution_time = round(sum(estimated_execution_times) / len(estimated_execution_times),2)
        # estimated_total_time = round(((average_execution_time * number_of_folders) / 60),1)
        estimated_left_time = round((((average_execution_time * number_of_folders) - sum(estimated_execution_times))/60),1)
        
        del Excel_result_data_bad
        
        print(f"\n[INFO] Folder completed: {folder_name}")
        print(f"[INFO] Processing time: {folder_execution_time} seconds")
        print(f"[INFO] Estimated remaining time: {estimated_left_time} minutes\n")
        plt.close()
        data_wb.close()


    
    try:
        plt.close()
        # Excel 파일 사용이 끝났으므로 파일을 닫습니다.
        data_wb.close()
    except :
        # 이미 삭제된 변수에 대한 NameError는 무시하고 다음 코드를 실행합니다.
        pass

    print('============================ INACTIVE FILE FINISHED ============================')
    print('============================ ACTIVE FILE START ============================')
    #만약 결과를 저장할 기본 폴더가 없으면 생성
    if not os.path.exists(TARGET_FILE_PATH):
        os.makedirs(TARGET_FILE_PATH, exist_ok=True)
        
    video_files = []
    
    

    # 먼저 비디오 파일 목록 수집
    for filename in natsort.natsorted(os.listdir(TARGET_FILE_PATH)):
        for extension in VIDEO_FILE_EXTENSION:
            if filename.endswith(f'.{extension}'):
                video_files.append(filename)
                break

    cpu_count = multiprocessing.cpu_count()
    max_workers = round(max(1, int(cpu_count * 2/3)))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(extract_frames, TARGET_FILE_PATH, v) for v in video_files]

        for f in tqdm(futures, total=len(futures), desc="Frame Extraction"):
            f.result()

        
    # 예상 처리 시간을 저장할 리스트
    estimated_execution_times = []


    #기준 이미지 초기화
    reference_image = None

    # 기존 엑셀 파일 열기
    try:
        good_wb = load_workbook(filename=TARGET_EXCEL)
        good_ws = good_wb.active
    except FileNotFoundError:
        # 기존 파일이 없는 경우 새로운 워크북 생성
        good_wb = Workbook()
        good_ws = good_wb.active
        
    # 엑셀 파일에 헤더 추가
    good_ws.append(headers)

    # parameter_ws['A5'] = 'MAX_PIXEL_MEAN_INDEX'

    # parameter_ws.append(['NOISE_INDEX, MAX_NOISE_INDEX, PIXEL_MEAN_INDEX, MAX_PIXEL_MEAN_INDEX, DIFF_MAX_MIN_INDEX'])

    # 저장 위치의 폴더만 가져오기
    output_folders = [f for f in natsort.natsorted(os.listdir(TARGET_FILE_PATH)) if os.path.isdir(os.path.join(TARGET_FILE_PATH, f))]

    # 확인된 저장 위치의 폴더 갯수
    number_of_folders = len(output_folders)

    for folder_name in natsort.natsorted(os.listdir(TARGET_FILE_PATH)):
        folder_path = os.path.join(TARGET_FILE_PATH, folder_name)
        if not os.path.isdir(folder_path):
            continue # 폴더가 아니라면 스킵
        # Create the result folder for the current sample.
        result_folder = os.path.join(folder_path, 'result')
        os.makedirs(result_folder, exist_ok=True)
        print(f'\n【　Starting to Analyze the File {folder_name}　】\n ')
        print('--------------------------------List of Generated Files --------------------------------\n')
        
        folder_start_time = time.time()
        for extension in VIDEO_FILE_EXTENSION:
            fps_video_path = os.path.join(TARGET_FILE_PATH,f'{folder_name}.{extension}')
            if os.path.exists(fps_video_path):  # 파일이 존재하는 경우에만 FPS 계산
                cap = cv2.VideoCapture(fps_video_path)
                fps = round(cap.get(cv2.CAP_PROP_FPS), 1)
                cap.release()
                break # 파일이 존재하는 경우에는 이후 확장자에 대한 반복을 중지하고 다음 폴더로 넘어감
            else:
                fps = VIDEO_FPS   
        # 기준 이미지를 설정
        reference_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}0000.{FRAME_EXTENSION}')
        reference_image = cv2.imread(reference_image_path)
        
        # 기준 이미지를 성공적으로 로드 했는지 확인
        if reference_image is not None:
            reference_gray = cv2.cvtColor(reference_image, cv2.COLOR_BGR2GRAY)
        else:
            continue  # 기준 이미지 로드에 실패하면 다음 폴더로 건너뛸 수 있도록 continue 사용
        
        # 폴더 내 파일명에 있는 숫자 중 최댓값 찾기 (자동으로 시작, 끝 프레임 갯수 찾기 위해)
        max_number = -1
        min_number = float('inf')
        for filename in os.listdir(folder_path):
            if filename.startswith(f'{FRAME_PREFIX}') and filename.endswith(f'.{FRAME_EXTENSION}'):
                try:
                    # 파일명에서 'frame_'과 이미지 확장자를 제외한 부분을 추출하고 숫자로 변환
                    number = int(filename[len(f'{FRAME_PREFIX}'):-len(f'.{FRAME_EXTENSION}')]) #frame_0000.png 에서  0000만 추출 
                    max_number = max(max_number, number)
                    min_number = min(min_number, number)
                except ValueError:
                    pass # 숫자로 변환할 수 없는 경우 무시

        
        # 최댓값 출력
        start_number = min_number
        last_number = max_number+1 #마지막 프레임까지 계산 하기 위해 +1
        resolution_y, resolution_x,_ = reference_image.shape #해상도 자동 계산.
        
        try:
            total_time_seconds = int(max_number / fps)
        except ZeroDivisionError:
            pass
        
    
        # Initialize lists for image paths and loaded frames
        # KR: 이미지 경로, grayscale frame, RGB frame을 저장할 리스트를 초기화합니다.
        image_paths = []
        current_gray_image_list = []
        current_rgb_image_list = []
        # # Initialize lists for maximum pixel-intensity differences and their coordinates
        max_difference_comparing_0 = []  #max_differences 
        max_difference_comparing_0_locations = [] # max_diff_locations

        max_top30_differences_crop = []
        max_top30_diff_locations_crop = []

        max_top30_differences_crop_2 = []
        max_top30_diff_locations_crop_2 = []

        max_top30_differences_crop_3 = []
        max_top30_diff_locations_crop_3 = []

        max_top30_differences_crop_2_new = []
        max_top30_diff_locations_crop_2_new = []

        max_top30_differences_crop_3_new = []
        max_top30_diff_locations_crop_3_new = []

        valid_pixel_all_counts_min = []
        valid_pixel_all_counts_max = []
        

        # =========================================================
        # (A) Load frame paths and grayscale images
        #     Frames are loaded once and then processed in GPU/CPU batch mode.
        # KR: frame 경로와 grayscale 이미지를 수집합니다.
        # KR: frame은 한 번만 로드한 뒤 이후 GPU/CPU batch 방식으로 처리합니다.
        # =========================================================

        use_gpu = (GPU_MODE.lower() == "on") and torch.cuda.is_available()
        device = torch.device("cuda" if use_gpu else "cpu")

        print(f"Compute mode: {'GPU' if use_gpu else 'CPU'}")




        image_paths = []
        gray_frames = []

        for idx in range(start_number, last_number):
            p = os.path.join(folder_path, f'{FRAME_PREFIX}{idx:04d}.{FRAME_EXTENSION}')
            if not os.path.exists(p):
                continue

            img_rgb = cv2.imread(p)
            if img_rgb is None:
                print(f"Image load fail → {p}")
                continue

            g = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)

            image_paths.append(p)
            gray_frames.append(g)
            current_rgb_image_list.append(img_rgb)


        if len(gray_frames) == 0:
            print("No valid frames found. Skip folder.")
            continue

        # =========================================================
        # (B) Compute frame-wise absolute differences from the reference image
        #     using GPU/CPU batch processing
        # KR: reference image와 각 frame 사이의 absolute difference를 계산합니다.
        # KR: 이 계산은 GPU 또는 CPU batch 처리로 한 번에 수행합니다.
        # =========================================================
        
        gray_tensor = torch.from_numpy(np.stack(gray_frames)).to(torch.int16)

        if use_gpu:
            gray_tensor = gray_tensor.pin_memory().to(device, non_blocking=True)
        else:
            gray_tensor = gray_tensor.contiguous()

        ref_tensor = torch.from_numpy(reference_gray).to(torch.int16)

        if use_gpu:
            ref_tensor = ref_tensor.pin_memory().to(device, non_blocking=True)
        else:
            ref_tensor = ref_tensor.contiguous()


        diff_tensor = torch.abs(gray_tensor - ref_tensor)                                 # [T,H,W]
        flat        = diff_tensor.flatten(1)                                              # [T,H*W]

        max_vals, argmax_vals = flat.max(dim=1)                                           # [T], [T]

        H, W = reference_gray.shape
        ys = (argmax_vals // W).cpu().numpy()
        xs = (argmax_vals %  W).cpu().numpy()

        max_difference_comparing_0 = max_vals.cpu().numpy().tolist()
        max_difference_comparing_0_locations = list(zip(xs, ys))

        current_gray_image_list = gray_frames

        image0_path = reference_image_path
        
        image0 = current_gray_image_list[0]

        
        # Search for the maximum intensity change within the initial fps × NUM_FRAMES window.
        # This helps focus the analysis on early contraction-related motion rather than large noise or bubble movement.
        # KR: 초기 fps × NUM_FRAMES 구간 안에서 최대 intensity change를 찾습니다.
        # KR: 큰 noise 또는 bubble 움직임보다 초기 수축 관련 움직임에 분석을 집중하기 위한 과정입니다.
        top30_max_differences_indices = np.arange(len(max_difference_comparing_0))[:int(fps*NUM_FRAMES)]
        top30_max_differences_values = [max_difference_comparing_0[i] for i in top30_max_differences_indices]
        top30_max_differences_values_max = max(top30_max_differences_values)
        
        max_value_index_within_top30 = np.argmax(top30_max_differences_values)
        
        max_difference_index_within_top30 = top30_max_differences_indices[max_value_index_within_top30]
        
        max_difference_value_within_top30 = top30_max_differences_values[max_value_index_within_top30]

        max_diff_location_within_top30 = max_difference_comparing_0_locations[max_difference_index_within_top30]

        x_center_max_diff_location_within_top30 = max_diff_location_within_top30[0]
        y_center_max_diff_location_within_top30 = max_diff_location_within_top30[1]

        
        x11 = x_center_max_diff_location_within_top30 - ANALYSIS_SQUARE_SIZE_REFERENCE
        y11 = y_center_max_diff_location_within_top30 - ANALYSIS_SQUARE_SIZE_REFERENCE
        x22 = x_center_max_diff_location_within_top30 + ANALYSIS_SQUARE_SIZE_REFERENCE
        y22 = y_center_max_diff_location_within_top30 + ANALYSIS_SQUARE_SIZE_REFERENCE
        x11 = max(x11, 0)
        y11 = max(y11, 0)
        x22 = min(x22, resolution_x - 1)
        y22 = min(y22, resolution_y - 1)


        roi_tensor = gray_tensor[:, y11:y22, x11:x22].contiguous()


        image_frame_number = []
        next_image_index = 0   

        visited = set()

        for step in range(min(len(gray_frames), 500)):
            if next_image_index in visited:
                break
            visited.add(next_image_index)

            image_frame_number.append(next_image_index)

            ref_roi = roi_tensor[next_image_index].unsqueeze(0)  # [1,h,w]
            diff_batch = torch.abs(roi_tensor - ref_roi)         # [T,h,w]
            flat = diff_batch.flatten(1)                         # [T,h*w]

            max_vals, _ = flat.max(dim=1)                        # [T]
            best_i = int(max_vals.argmax().item())

            next_image_index = best_i

            if len(image_frame_number) >= 8 and image_frame_number[-1] == image_frame_number[-3]:
                break

        image_frame_n1 = image_frame_number[-1]
        image_frame_n2 = image_frame_number[-2] if len(image_frame_number) >= 2 else image_frame_number[-1]

        n1_max_image_path = image_paths[int(image_frame_n1)]
        n2_max_image_path = image_paths[int(image_frame_n2)]


        roi_full = gray_tensor[:, y11:y22, x11:x22].contiguous()          # [T,h,w]

        ref_n1 = roi_full[int(image_frame_n1)].unsqueeze(0)               # [1,h,w]
        diff_n1 = torch.abs(roi_full - ref_n1)                            # [T,h,w]
        flat_n1 = diff_n1.flatten(1)                                      # [T,h*w]
        max_vals_n1, argmax_n1 = flat_n1.max(dim=1)                       # [T], [T]

        h_roi = roi_full.shape[1]
        w_roi = roi_full.shape[2]
        ys_n1 = (argmax_n1 // w_roi)
        xs_n1 = (argmax_n1 %  w_roi)

        max_top30_differences_crop_2 = max_vals_n1.detach().cpu().numpy().astype(np.float32).tolist()
        max_top30_diff_locations_crop_2 = list(zip(xs_n1.detach().cpu().numpy().tolist(),
                                                ys_n1.detach().cpu().numpy().tolist()))

        max_top30_differences_crop_except0_2 = [v for v in max_top30_differences_crop_2 if v != 0]
        mean_max_top30_differences_crop2 = float(np.mean(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0.0
        mean_round_max_top30_differences_crop2 = round(mean_max_top30_differences_crop2, 2)
        max_max_top30_differences_crop2 = float(np.max(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0.0

        if len(max_top30_differences_crop_except0_2) > 0:
            min_max_top30_differences_crop_except0_2 = float(np.min(max_top30_differences_crop_except0_2))
        else:
            min_max_top30_differences_crop_except0_2 = 0.0

        min_max_top30_differences_crop_except0_2_index = int(np.argmax(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0

        ref_n2 = roi_full[int(image_frame_n2)].unsqueeze(0)               # [1,h,w]
        diff_n2 = torch.abs(roi_full - ref_n2)                            # [T,h,w]
        flat_n2 = diff_n2.flatten(1)                                      # [T,h*w]
        max_vals_n2, argmax_n2 = flat_n2.max(dim=1)                       # [T], [T]

        ys_n2 = (argmax_n2 // w_roi)
        xs_n2 = (argmax_n2 %  w_roi)

        max_top30_differences_crop_3 = max_vals_n2.detach().cpu().numpy().astype(np.float32).tolist()
        max_top30_diff_locations_crop_3 = list(zip(xs_n2.detach().cpu().numpy().tolist(),
                                                ys_n2.detach().cpu().numpy().tolist()))

        max_top30_differences_crop_except0_3 = [v for v in max_top30_differences_crop_3 if v != 0]
        mean_max_top30_differences_crop3 = float(np.mean(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0.0
        mean_round_max_top30_differences_crop3 = round(mean_max_top30_differences_crop3, 2)
        max_max_top30_differences_crop3 = float(np.max(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0.0

        if len(max_top30_differences_crop_except0_3) > 0:
            min_max_top30_differences_crop_except0_3 = float(np.min(max_top30_differences_crop_except0_3))
        else:
            min_max_top30_differences_crop_except0_3 = 0.0

        min_max_top30_differences_crop_except0_3_index = int(np.argmax(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0

        try:
            top30_max_image_path_2 = image_paths[min_max_top30_differences_crop_except0_2_index]
        except:
            top30_max_image_path_2 = ""

        try:
            top30_max_image_path_3 = image_paths[min_max_top30_differences_crop_except0_3_index]
        except:
            top30_max_image_path_3 = ""

        diff_max_all = diff_n2                                           

        diff_min_all = diff_n1                                            

        threshold_max = min_max_top30_differences_crop_except0_3
        threshold_min = min_max_top30_differences_crop_except0_2

        min_threshold_max = threshold_max + 20
        max_threshold_max = max_max_top30_differences_crop3

        min_threshold_min = threshold_min + 20
        max_threshold_min = max_max_top30_differences_crop2

        if max_threshold_max <= 0:
            valid_pixel_counts_max = [0] * len(gray_frames)
        else:
            mask_max = (diff_max_all >= min_threshold_max) & (diff_max_all <= max_threshold_max)
            valid_pixel_counts_max = mask_max.flatten(1).sum(dim=1).detach().cpu().numpy().astype(np.int64).tolist()

        if max_threshold_min <= 0:
            valid_pixel_counts_min = [0] * len(gray_frames)
        else:
            mask_min = (diff_min_all >= min_threshold_min) & (diff_min_all <= max_threshold_min)
            valid_pixel_counts_min = mask_min.flatten(1).sum(dim=1).detach().cpu().numpy().astype(np.int64).tolist()

        valid_pixel_count_mean_min = np.mean([v for v in valid_pixel_counts_min if v not in [0, '']]) if len(valid_pixel_counts_min) else 0
        valid_pixel_count_mean_round_min = round(valid_pixel_count_mean_min) if not np.isnan(valid_pixel_count_mean_min) else 0

        valid_pixel_count_mean_max = np.mean([v for v in valid_pixel_counts_max if v not in [0, '']]) if len(valid_pixel_counts_max) else 0
        valid_pixel_count_mean_round_max = round(valid_pixel_count_mean_max) if not np.isnan(valid_pixel_count_mean_max) else 0

        count_zeros_min = valid_pixel_counts_min.count(0)
        count_zeros_max = valid_pixel_counts_max.count(0)

        sum_count_all_pixel_in_range_min = int(np.sum(np.array(valid_pixel_counts_min) < valid_pixel_count_mean_round_min)) if valid_pixel_count_mean_round_min > 0 else 0
        sum_count_all_pixel_in_range_max = int(np.sum(np.array(valid_pixel_counts_max) < valid_pixel_count_mean_round_max)) if valid_pixel_count_mean_round_max > 0 else 0

        valid_pixel_all_counts_min.append(sum_count_all_pixel_in_range_min)
        valid_pixel_all_counts_max.append(sum_count_all_pixel_in_range_max)

        valid_pixel_all_counts_min_max = max(valid_pixel_all_counts_min) if len(valid_pixel_all_counts_min) else 0
        valid_pixel_all_counts_max_max = max(valid_pixel_all_counts_max) if len(valid_pixel_all_counts_max) else 0

        All_pixel_intensities = []
        
        selected_max_differences = []
        selected_max_differences_location = []
        
        max_image_path_pixel = top30_max_image_path_3   
        min_image_path_pixel = top30_max_image_path_2  

        for i in range(len(gray_frames)):

            if use_gpu:
                image1 = gray_tensor[i].cpu().numpy().astype(np.uint8)
                max_image_gray_pixel = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
                min_image_gray_pixel = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
                first_image_gray_pixel = gray_tensor[int(max_difference_index_within_top30)].cpu().numpy().astype(np.uint8)

            else:
                image1 = gray_frames[i].astype(np.uint8)
                max_image_gray_pixel = gray_frames[int(image_frame_n2)].astype(np.uint8)
                min_image_gray_pixel = gray_frames[int(image_frame_n1)].astype(np.uint8)
                first_image_gray_pixel = gray_frames[int(max_difference_index_within_top30)].astype(np.uint8)

            cropped_max_image_pixel = max_image_gray_pixel[y11:y22, x11:x22]
            cropped_min_image_pixel = min_image_gray_pixel[y11:y22, x11:x22]
            cropped_image0_pixel = image0[y11:y22, x11:x22]
            cropped_image1_pixel = image1[y11:y22, x11:x22]

            h = min(cropped_max_image_pixel.shape[0], cropped_image1_pixel.shape[0])
            w = min(cropped_max_image_pixel.shape[1], cropped_image1_pixel.shape[1])

            cropped_max_image_pixel = cropped_max_image_pixel[:h, :w]
            cropped_min_image_pixel = cropped_min_image_pixel[:h, :w]
            cropped_image0_pixel = cropped_image0_pixel[:h, :w]
            cropped_image1_pixel = cropped_image1_pixel[:h, :w]

            cropped_max_image_pixel = cropped_max_image_pixel.astype(np.uint8)
            cropped_min_image_pixel = cropped_min_image_pixel.astype(np.uint8)
            cropped_image0_pixel = cropped_image0_pixel.astype(np.uint8)
            cropped_image1_pixel = cropped_image1_pixel.astype(np.uint8)

            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f"Error cropping image {i}. Skipping...")
                continue

            if valid_pixel_all_counts_min_max >= valid_pixel_all_counts_max_max:
                not_crop_diff = cv2.subtract(image1, min_image_gray_pixel)
                square_image_path = min_image_path_pixel
                square_image_index = image_frame_n1

            else:
                diff = cv2.subtract(cropped_max_image_pixel, cropped_image1_pixel)
                not_crop_diff = cv2.subtract(image1, max_image_gray_pixel)
                square_image_path = max_image_path_pixel
                square_image_index = image_frame_n2

            _, selected_max_diff, _, selected_max_diff_loc = cv2.minMaxLoc(not_crop_diff)

            selected_max_differences.append(selected_max_diff)
            selected_max_differences_location.append(selected_max_diff_loc)

            
            selected_max_differences_indices = np.arange(len(selected_max_differences))[:int(fps*NUM_FRAMES)] 
            selected_max_differences_values = [selected_max_differences[i] for i in selected_max_differences_indices] 

            selected_max_differences_values_max_in_Frames = np.argmax(selected_max_differences_values) 
            selected_max_differences_index_in_Frames = selected_max_differences_indices[selected_max_differences_values_max_in_Frames] 
            selected_max_differences_value_in_Frames = selected_max_differences_values[selected_max_differences_index_in_Frames] 
            selected_max_differences_location_in_Frames = selected_max_differences_location[selected_max_differences_index_in_Frames]

        x_selected = selected_max_differences_location_in_Frames[0]  
        y_selected = selected_max_differences_location_in_Frames[1] 
        
        x1_small = x_selected - ANALYSIS_SQUARE_SIZE_REFERENCE  
        y1_small = y_selected - ANALYSIS_SQUARE_SIZE_REFERENCE
        x2_small = x_selected + ANALYSIS_SQUARE_SIZE_REFERENCE
        y2_small = y_selected + ANALYSIS_SQUARE_SIZE_REFERENCE
        
        x1_small = max(x1_small, 0)
        y1_small = max(y1_small, 0)
        x2_small = min(x2_small, resolution_x -1)
        y2_small = min(y2_small, resolution_y -1)
        
        # Compare all frames again using the cropped ROI-derived maximum frame as a reference.
        # This step is used to identify candidate maximal contraction or relaxation frames.
        # KR: cropped ROI에서 얻은 maximum frame을 기준으로 전체 frame을 다시 비교합니다.
        # KR: 이 과정은 최대 수축 또는 최대 이완 후보 frame을 찾기 위한 단계입니다.

        image_frame_number = []

        next_image_index = start_number
        for n in range(start_number, last_number):
            max_diff_cropped_n_frame_list = []
            max_diff_cropped_loc_n_frame_list = []
        
            for i in range(len(gray_frames)):

                n_frame_path = os.path.join(folder_path, f'{FRAME_PREFIX}{next_image_index + start_number :04d}.{FRAME_EXTENSION}') 

                # Current reference frame path used in this iterative comparison
                # KR: 반복 비교에 사용할 현재 reference frame 경로입니다.
                if use_gpu:
                    top30_max_image_gray_new = gray_tensor[next_image_index].cpu().numpy().astype(np.uint8)
                else:
                    top30_max_image_gray_new = gray_frames[next_image_index].astype(np.uint8)
                cropped_n_frame_gray = top30_max_image_gray_new[y1_small:y2_small, x1_small:x2_small] 
                current_image_crop = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small] 

                max_diff_loc_n_frame_cropped = cv2.absdiff(cropped_n_frame_gray, current_image_crop)
                _, max_diff_crop_n_frame,_,max_diff_crop_loc_n_frame = cv2.minMaxLoc(max_diff_loc_n_frame_cropped)

                max_diff_cropped_n_frame_list.append(max_diff_crop_n_frame)
                max_diff_cropped_loc_n_frame_list.append(max_diff_crop_loc_n_frame)

                max_diff_cropped_n_indices = np.argsort(max_diff_cropped_n_frame_list)[-1:]            
                max_diff_cropped_n_indices_index = max_diff_cropped_n_indices[0]
                max_diff_cropped_n_indices_path = os.path.join(folder_path, f'{FRAME_PREFIX}{max_diff_cropped_n_indices_index + start_number :04d}.{FRAME_EXTENSION}')
            image_frame_number.append(next_image_index)
            
            # Calculate summary statistics from cropped ROI differences.
            # Zero values are excluded when estimating the nonzero minimum.
            # KR: cropped ROI difference 값에서 요약 통계를 계산합니다.
            # KR: 0이 아닌 최소값을 계산할 때는 zero value를 제외합니다.
            max_diff_cropped_n_except0 = [value for value in max_diff_cropped_n_frame_list if value != 0]
            mean_max_diff_cropped_n_except0 = np.mean(max_diff_cropped_n_frame_list)
            mean_round_max_diff_cropped_n_frame_listew = round(mean_max_diff_cropped_n_except0, 2)
            max_max_diff_cropped_n_frame_list = max(max_diff_cropped_n_frame_list)
            min_max_diff_cropped_n_except0 = min(max_diff_cropped_n_except0)
            min_max_diff_cropped_n_except0_index = max_diff_cropped_n_frame_list.index(max_max_diff_cropped_n_frame_list)
            # Update the reference frame index for the next iteration
            # KR: 다음 반복 비교에 사용할 reference frame index를 업데이트합니다.
            next_image_index = min_max_diff_cropped_n_except0_index
            
            # Stop the iterative search when the selected frame index starts to repeat.
            # KR: 선택된 frame index가 반복되면 탐색이 수렴한 것으로 보고 반복을 중단합니다.
            if len(image_frame_number)>=8 and image_frame_number[n] == image_frame_number[n - 2]:
                    image_frame_n1 = image_frame_number[n]
                    image_frame_n2 = image_frame_number[n-1]
                    n1_max_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}{image_frame_n1 + start_number :04d}.{FRAME_EXTENSION}') 
                    n2_max_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}{image_frame_n2 + start_number :04d}.{FRAME_EXTENSION}') 
                    break

        # Compare all frames again using the two ROI-derived candidate frames.
        # This step refines the candidate maximal contraction and relaxation frames.
        # KR: ROI에서 얻은 두 후보 frame을 기준으로 전체 frame을 다시 비교합니다.
        # KR: 최대 수축/최대 이완 후보 frame을 더 안정적으로 찾기 위한 단계입니다.
        
        for i in range(len(gray_frames)):
            
            top30_max_image_path_2_new = n1_max_image_path
            # Use the n1 candidate frame as the reference for ROI-based comparison.
            # KR: n1 후보 frame을 기준으로 ROI 기반 frame 비교를 수행합니다.
            
            if use_gpu:
                top30_max_image_gray_2_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            else:
                top30_max_image_gray_2_new = gray_frames[image_frame_n1].astype(np.uint8)

            
            cropped_top30_max_image_gray_2_new = top30_max_image_gray_2_new[y1_small:y2_small, x1_small:x2_small] 
            current_image_crop_2_new = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small]
            diff_top30_max_2_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_2_new, current_image_crop_2_new)
            _, max_diff_crop_top30_max_2_new, _,max_diff_loc_crop_top30_max_2_new = cv2.minMaxLoc(diff_top30_max_2_cropped_new)
            max_top30_differences_crop_2_new.append(max_diff_crop_top30_max_2_new)
            max_top30_diff_locations_crop_2_new.append(max_diff_loc_crop_top30_max_2_new)
            
            max_top30_differences_crop_indices_2_new = np.argsort(max_top30_differences_crop_2_new)[-1:]
            
            # Search within the initial fps × NUM_FRAMES window to reduce the influence of large noise or bubble movement.
            # KR: 큰 noise 또는 bubble 움직임의 영향을 줄이기 위해 초기 fps × NUM_FRAMES 구간 안에서 탐색합니다.
            max_top30_differences_crop_indices_2_new_num_frames = np.arange(len(max_top30_differences_crop_2_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_2_new = [max_top30_differences_crop_2_new[i] for i in max_top30_differences_crop_indices_2_new_num_frames]

            
            max_top30_differences_crop_values_2_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_2_new) 
            
            index_in_Frames_2 = max_top30_differences_crop_indices_2_new_num_frames[max_top30_differences_crop_values_2_new_max_in_Frames]
            value_in_Frames_2 =  max_top30_differences_crop_values_2_new[index_in_Frames_2]
            location_in_Frames_2 = max_top30_diff_locations_crop_2_new[index_in_Frames_2]

            
            max_top30_differences_crop_indices_2_index_new = max_top30_differences_crop_indices_2_new[0]
            max_top30_differences_crop_indices_2_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_2_index_new + start_number :04d}.{FRAME_EXTENSION}')


        # Calculate summary statistics for ROI differences relative to the n1 candidate frame.
        # KR: n1 후보 frame 기준 ROI difference의 요약 통계를 계산합니다.
        max_top30_differences_crop_except0_2_new = [value for value in max_top30_differences_crop_2_new if value != 0]
        mean_max_top30_differences_crop2_new = np.mean(max_top30_differences_crop_2_new)
        mean_round_max_top30_differences_crop2_new = round(mean_max_top30_differences_crop2_new, 2)
        max_max_top30_differences_crop2_new = max(max_top30_differences_crop_2_new)
        min_max_top30_differences_crop_except0_2_new = min(max_top30_differences_crop_except0_2_new)
        min_max_top30_differences_crop_except0_2_index_new = max_top30_differences_crop_2_new.index(max_max_top30_differences_crop2_new)

        
        for i in range(len(gray_frames)):

            top30_max_image_path_3_new = n2_max_image_path

            # Use the n2 candidate frame as the reference for ROI-based comparison.
            # KR: n2 후보 frame을 기준으로 ROI 기반 frame 비교를 수행합니다.
            if use_gpu:
                top30_max_image_gray_3_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
            else:
                top30_max_image_gray_3_new = gray_frames[image_frame_n2].astype(np.uint8)

            
            cropped_top30_max_image_gray_3_new = top30_max_image_gray_3_new[y1_small:y2_small, x1_small:x2_small] 
            
            current_image_crop_3 = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small] 

            diff_top30_max_3_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_3_new, current_image_crop_3)
            _, max_diff_crop_top30_max_3_new, _,max_diff_loc_crop_top30_max_3_new = cv2.minMaxLoc(diff_top30_max_3_cropped_new)
            
            max_top30_differences_crop_3_new.append(max_diff_crop_top30_max_3_new)
            max_top30_diff_locations_crop_3_new.append(max_diff_loc_crop_top30_max_3_new)
            
            max_top30_differences_crop_indices_3_new = np.argsort(max_top30_differences_crop_3_new)[-1:]
            
            # Search within the initial fps × NUM_FRAMES window to reduce the influence of large noise or bubble movement.
            # KR: 큰 noise 또는 bubble 움직임의 영향을 줄이기 위해 초기 fps × NUM_FRAMES 구간 안에서 탐색합니다.
            max_top30_differences_crop_indices_3_new_num_frames = np.arange(len(max_top30_differences_crop_3_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_3_new = [max_top30_differences_crop_3_new[i] for i in max_top30_differences_crop_indices_3_new_num_frames] 

            
            max_top30_differences_crop_values_3_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_3_new) 
            index_in_Frames_3 = max_top30_differences_crop_indices_3_new_num_frames[max_top30_differences_crop_values_3_new_max_in_Frames]
            value_in_Frames_3 =  max_top30_differences_crop_values_3_new[index_in_Frames_3]
            location_in_Frames_3 = max_top30_diff_locations_crop_3_new[index_in_Frames_3]

            
            max_top30_differences_crop_indices_3_index_new = max_top30_differences_crop_indices_3_new[0]
            max_top30_differences_crop_indices_3_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_3_index_new + start_number :04d}.{FRAME_EXTENSION}')
            

        max_top30_differences_crop_except0_3_new = [value for value in max_top30_differences_crop_3_new if value != 0]
        mean_max_top30_differences_crop3_new = np.mean(max_top30_differences_crop_3_new)
        mean_round_max_top30_differences_crop3_new = round(mean_max_top30_differences_crop3_new, 2)
        max_max_top30_differences_crop3_new = max(max_top30_differences_crop_3_new)
        min_max_top30_differences_crop_except0_3_new = min(max_top30_differences_crop_except0_3_new)
        min_max_top30_differences_crop_except0_3_index_new = max_top30_differences_crop_3_new.index(max_max_top30_differences_crop3_new)
        
        top30_max_image_path_4_new = os.path.join(folder_path, f'{FRAME_PREFIX}{min_max_top30_differences_crop_except0_3_index_new + start_number :04d}.{FRAME_EXTENSION}') 
        



        x_differences = [] 
        y_differences = []
        valid_pixel_counts_min_new = []
        valid_pixel_counts_max_new = []
        valid_pixel_all_counts_max_new = []
        valid_pixel_all_counts_min_new = []

        # Calculate summary statistics for ROI differences relative to the n2 candidate frame.
        # KR: n2 후보 frame 기준 ROI difference의 요약 통계를 계산합니다.
        for i in range(len(gray_frames)):
            image1 = current_gray_image_list[i]
            # Crop local ROIs around the selected motion center.
            # KR: 선택된 motion center 주변의 local ROI를 crop합니다.
            max_image_path_pixel_new = top30_max_image_path_3_new  # Candidate maximal contraction or relaxation frame
            min_image_path_pixel_new = top30_max_image_path_2_new  # Candidate maximal relaxation or contraction frame
            
            if use_gpu:
                max_image_gray_pixel_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
                min_image_gray_pixel_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            else:
                max_image_gray_pixel_new = gray_frames[image_frame_n2].astype(np.uint8)
                min_image_gray_pixel_new = gray_frames[image_frame_n1].astype(np.uint8)

                    
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small] 
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
            cropped_image0_pixel_new = image0[y1_small:y2_small, x1_small:x2_small] 
            cropped_image1_pixel_new = image1[y1_small:y2_small, x1_small:x2_small] 
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f'Error cropping image{i}. Skipping...')
                continue  # Skip this frame if ROI cropping fails
            
            diff_max_new = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
            diff_min_new = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 

            threshold_max_new = min_max_top30_differences_crop_except0_3_new  
            threshold_min_new = min_max_top30_differences_crop_except0_2_new 
            white_pixel_coords_max_new = np.argwhere(diff_max_new >= threshold_max_new)
            white_pixel_coords_min_new = np.argwhere(diff_min_new >= threshold_min_new)  
            
            # Count pixels whose intensity differences fall within the threshold range.
            # KR: intensity difference가 설정된 threshold 범위 안에 들어오는 pixel을 계산합니다.
            
            # Add a fixed offset to reduce small background or medium-related noise.
            # KR: 작은 background 또는 배지 관련 noise를 줄이기 위해 고정 offset을 더합니다.
            min_threshold_max_new = min_max_top30_differences_crop_except0_3_new + 20  # Fixed offset for small background or medium-related noise
            max_threshold_max_new = max_max_top30_differences_crop3_new  # Maximum ROI intensity difference relative to the candidate frame
            min_threshold_min_new = min_max_top30_differences_crop_except0_2_new  +20 # Fixed offset for small background or medium-related noise
            max_threshold_min_new = max_max_top30_differences_crop2_new # Maximum ROI intensity difference relative to the candidate frame
            
            valid_pixel_coords_max_new = np.argwhere((diff_max_new >= min_threshold_max_new) & (diff_max_new <= max_threshold_max_new))
            valid_pixel_coords_min_new = np.argwhere((diff_min_new >= min_threshold_min_new) & (diff_min_new <= max_threshold_min_new))
            
            valid_pixel_coords_max_all_new = np.argwhere((diff_max_new >= 0) & (diff_max_new <= max_threshold_max))
            valid_pixel_coords_min_all_new = np.argwhere((diff_min_new >= 0) & (diff_min_new <= max_threshold_min))

            valid_pixel_count_max_new = len(valid_pixel_coords_max_new)
            valid_pixel_count_min_new = len(valid_pixel_coords_min_new)
            valid_pixel_counts_max_new.append(valid_pixel_count_max_new)
            valid_pixel_counts_min_new.append(valid_pixel_count_min_new)       

        valid_pixel_count_mean_min_new = np.mean([value for value in valid_pixel_counts_min_new if value not in [0,'']]) #0 과 데이터가 

        valid_pixel_count_mean_round_min_new = round(valid_pixel_count_mean_min_new) if not np.isnan(valid_pixel_count_mean_min_new) else 0 
        valid_pixel_count_mean_max_new = np.mean([value for value in valid_pixel_counts_max_new if value not in [0,'']]) #0 과 데이터가 

        valid_pixel_count_mean_round_max_new = round(valid_pixel_count_mean_max_new) if not np.isnan(valid_pixel_count_mean_max_new) else 0 
        
        count_zeros_min_new = valid_pixel_counts_min_new.count(0)
        count_zeros_max_new = valid_pixel_counts_max_new.count(0)

        count_all_pixel_in_range_min_new = [valid_pixel_counts_min_new.count(i) for i in range(valid_pixel_count_mean_round_min_new)]
        count_all_pixel_in_range_max_new = [valid_pixel_counts_max_new.count(i) for i in range(valid_pixel_count_mean_round_max_new)]
        
        sum_count_all_pixel_in_range_min_new = sum(count_all_pixel_in_range_min_new, 0)
        sum_count_all_pixel_in_range_max_new = sum(count_all_pixel_in_range_max_new, 0)

        valid_pixel_all_counts_min_new.append(sum_count_all_pixel_in_range_min_new)
        valid_pixel_all_counts_max_new.append(sum_count_all_pixel_in_range_max_new)
        
        valid_pixel_all_counts_min_max_new = max(valid_pixel_all_counts_min_new)
        valid_pixel_all_counts_max_max_new = max(valid_pixel_all_counts_max_new)
        
        
        All_pixel_intensities = []  
        valid_pixel_counts = []
        valid_pixel_counts2 = []
        selected_max_differences = []
        selected_max_differences_location = []
        
        x_selected_new, y_selected_new = None, None
        click_done = False  
        



        for i in range(len(gray_frames)):

            max_image_path_pixel_new = top30_max_image_path_3_new
            min_image_path_pixel_new = top30_max_image_path_2_new
            
            if use_gpu:
                max_image_gray_pixel_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
                min_image_gray_pixel_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            else:
                max_image_gray_pixel_new = gray_frames[image_frame_n2].astype(np.uint8)
                min_image_gray_pixel_new = gray_frames[image_frame_n1].astype(np.uint8)

            
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
            cropped_image0_pixel_new = image0[y1_small:y2_small, x1_small:x2_small]  
            cropped_image1_pixel_new = image1[y1_small:y2_small, x1_small:x2_small]  
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f"Error cropping image {i}. Skipping...")
                continue 
            
            # Select the candidate frame based on the number of valid pixels below the mean threshold.
            # This helps determine which candidate better represents maximal contraction or relaxation.
            # KR: mean threshold보다 작은 valid pixel count를 기준으로 후보 frame을 선택합니다.
            # KR: 최대 수축 또는 최대 이완을 더 잘 나타내는 후보를 고르기 위한 과정입니다.
            
            if valid_pixel_all_counts_min_max_new >= valid_pixel_all_counts_max_max_new:
                loc2 = max_top30_diff_locations_crop_2_new  
                loc3 = location_in_Frames_2  
                
            else:
                loc2 = max_top30_diff_locations_crop_3_new 
                loc3 = location_in_Frames_3 
                
            # =========================================================
            # Manual coordinate selection callback
            # =========================================================
            
            # Mouse-click callback for selecting an analysis coordinate
            # KR: 분석 좌표를 수동으로 선택하기 위한 마우스 클릭 callback 함수입니다.
            def select_point(event, x, y, flags, param):
                global x_selected_new, y_selected_new, click_done  

                if event == cv2.EVENT_LBUTTONDOWN and not click_done: # Left-click once to select a coordinate
                    x_selected_new = x
                    y_selected_new = y
                    click_done = True  # 
                    print(f"Selected coordinate: ({x_selected_new}, {y_selected_new})")

                    cv2.circle( min_image_gray_pixel_new, (x_selected_new, y_selected_new), 2, (0, 0, 255), -1)
                    cv2.imshow("Select Point",  min_image_gray_pixel_new)

                    cv2.waitKey(500) 
                    cv2.destroyAllWindows()

            # Use predefined coordinates when manual coordinate mode is enabled
            # KR: manual coordinate mode가 켜져 있으면 사용자가 미리 지정한 좌표를 사용합니다.
            if MANUAL_SELECTION.lower() == 'on' and not click_done:
                print("Please click the desired analysis point.")
                

                cv2.namedWindow("Select Point", cv2.WINDOW_NORMAL)
                cv2.resizeWindow("Select Point", resolution_x, resolution_y)  

                cv2.imshow("Select Point", min_image_gray_pixel_new)
                cv2.setMouseCallback("Select Point", select_point)


                while not click_done:
                    cv2.waitKey(1)  # 

                print(f"Using manually selected coordinate: ({x_selected_new}, {y_selected_new})")
            

            # Use automatically detected coordinates when manual selection is off,
            # or keep the previously selected coordinate after the first manual click.
            # KR: manual selection이 꺼져 있으면 자동 탐색 좌표를 사용합니다.
            # KR: 첫 번째 수동 클릭 이후에는 선택된 좌표를 유지합니다.
            
            elif MANUAL_SELECTION.lower() == 'off' or (x_selected_new is None and y_selected_new is None): 
                x_selected_new = x1_small + loc3[0]
                y_selected_new = y1_small + loc3[1]
        
            
            x1_small_new = x_selected_new -  resolution_x
            y1_small_new = y_selected_new -  resolution_y
            x2_small_new = x_selected_new +  resolution_x
            y2_small_new = y_selected_new + resolution_y
            
            x1_small_new = max(x1_small_new, 0)
            y1_small_new = max(y1_small_new, 0)
            x2_small_new = min(x2_small_new, resolution_x -1)
            y2_small_new = min(y2_small_new, resolution_y -1)
            
            x1_small_new_reference = max(0, x_selected_new - ANALYSIS_SQUARE_SIZE_REFERENCE)
            y1_small_new_reference = max(0, y_selected_new -  ANALYSIS_SQUARE_SIZE_REFERENCE)
            x2_small_new_reference = x_selected_new + ANALYSIS_SQUARE_SIZE_REFERENCE
            y2_small_new_reference = y_selected_new +  ANALYSIS_SQUARE_SIZE_REFERENCE
            
            x1_small_new_reference = max(x1_small_new_reference, 0)
            y1_small_new_reference = max(y1_small_new_reference, 0)
            x2_small_new_reference = min(x2_small_new_reference, resolution_x -1)
            y2_small_new_reference = min(y2_small_new_reference, resolution_y -1)
            
            x1_small_new_optical = max(0, x_selected_new - resolution_x)
            y1_small_new_optical = max(0, y_selected_new - resolution_y)
            x2_small_new_optical = x_selected_new + resolution_x
            y2_small_new_optical = y_selected_new + resolution_y
            
            x1_small_new_optical = max(x1_small_new_optical, 0)
            y1_small_new_optical = max(y1_small_new_optical, 0)
            x2_small_new_optical = min(x2_small_new_optical, resolution_x -1)
            y2_small_new_optical = min(y2_small_new_optical, resolution_y -1)

        max_top30_differences_crop_2_new = []
        max_top30_diff_locations_crop_2_new = []

        max_top30_differences_crop_3_new = []
        max_top30_diff_locations_crop_3_new = []

        if len(current_rgb_image_list) == 0:
            print("❌ current_rgb_image_list empty → skip")
            continue

        image_frame_n1 = max(0, min(image_frame_n1, len(current_rgb_image_list)-1))
        image_frame_n2 = max(0, min(image_frame_n2, len(current_rgb_image_list)-1))

        top30_max_image_2_new = current_rgb_image_list[image_frame_n1]
        top30_max_image_3_new = current_rgb_image_list[image_frame_n2]


        for i in range(len(gray_frames)):
            
            top30_max_image_path_2_new = n1_max_image_path
            top30_max_image_gray_2_new = cv2.cvtColor(top30_max_image_2_new, cv2.COLOR_BGR2GRAY) 
            
            cropped_top30_max_image_gray_2_new = top30_max_image_gray_2_new[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
            
            current_image_crop_2_new = current_gray_image_list[i][y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 

            diff_top30_max_2_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_2_new, current_image_crop_2_new)
            _, max_diff_crop_top30_max_2_new, _,max_diff_loc_crop_top30_max_2_new = cv2.minMaxLoc(diff_top30_max_2_cropped_new)
            
            max_top30_differences_crop_2_new.append(max_diff_crop_top30_max_2_new)
            max_top30_diff_locations_crop_2_new.append(max_diff_loc_crop_top30_max_2_new)
            
            max_top30_differences_crop_indices_2_new = np.argsort(max_top30_differences_crop_2_new)[-1:]
            
            max_top30_differences_crop_indices_2_new_num_frames = np.arange(len(max_top30_differences_crop_2_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_2_new = [max_top30_differences_crop_2_new[i] for i in max_top30_differences_crop_indices_2_new_num_frames] 
            
            max_top30_differences_crop_values_2_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_2_new)
            
            index_in_Frames_2 = max_top30_differences_crop_indices_2_new_num_frames[max_top30_differences_crop_values_2_new_max_in_Frames]
            
            value_in_Frames_2 =  max_top30_differences_crop_values_2_new[index_in_Frames_2]
            location_in_Frames_2 = max_top30_diff_locations_crop_2_new[index_in_Frames_2]
            
            max_top30_differences_crop_indices_2_index_new = max_top30_differences_crop_indices_2_new[0]
            max_top30_differences_crop_indices_2_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_2_index_new + start_number :04d}.{FRAME_EXTENSION}')



        max_top30_differences_crop_except0_2_new = [value for value in max_top30_differences_crop_2_new if value != 0]
        mean_max_top30_differences_crop2_new = np.mean(max_top30_differences_crop_2_new)
        mean_round_max_top30_differences_crop2_new = round(mean_max_top30_differences_crop2_new, 2)
        max_max_top30_differences_crop2_new = max(max_top30_differences_crop_2_new)
        min_max_top30_differences_crop_except0_2_new = min(max_top30_differences_crop_except0_2_new)
        min_max_top30_differences_crop_except0_2_index_new = max_top30_differences_crop_2_new.index(max_max_top30_differences_crop2_new)

        

        for i in range(len(gray_frames)):

            top30_max_image_path_3_new = n2_max_image_path
            top30_max_image_gray_3_new = cv2.cvtColor(top30_max_image_3_new, cv2.COLOR_BGR2GRAY) 
            
            cropped_top30_max_image_gray_3_new = top30_max_image_gray_3_new[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
            
            current_image_crop_3 = current_gray_image_list[i][y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 

            diff_top30_max_3_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_3_new, current_image_crop_3)
            _, max_diff_crop_top30_max_3_new, _,max_diff_loc_crop_top30_max_3_new = cv2.minMaxLoc(diff_top30_max_3_cropped_new)

            max_top30_differences_crop_3_new.append(max_diff_crop_top30_max_3_new)
            max_top30_diff_locations_crop_3_new.append(max_diff_loc_crop_top30_max_3_new)
            
            max_top30_differences_crop_indices_3_new = np.argsort(max_top30_differences_crop_3_new)[-1:]
            max_top30_differences_crop_indices_3_new_num_frames = np.arange(len(max_top30_differences_crop_3_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_3_new = [max_top30_differences_crop_3_new[i] for i in max_top30_differences_crop_indices_3_new_num_frames] 
        
            max_top30_differences_crop_values_3_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_3_new) 
            index_in_Frames_3 = max_top30_differences_crop_indices_3_new_num_frames[max_top30_differences_crop_values_3_new_max_in_Frames]
            
            value_in_Frames_3 =  max_top30_differences_crop_values_3_new[index_in_Frames_3]
            location_in_Frames_3 = max_top30_diff_locations_crop_3_new[index_in_Frames_3]
            
            max_top30_differences_crop_indices_3_index_new = max_top30_differences_crop_indices_3_new[0]
            max_top30_differences_crop_indices_3_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_3_index_new + start_number :04d}.{FRAME_EXTENSION}')
            

        max_top30_differences_crop_except0_3_new = [value for value in max_top30_differences_crop_3_new if value != 0]
        mean_max_top30_differences_crop3_new = np.mean(max_top30_differences_crop_3_new)
        mean_round_max_top30_differences_crop3_new = round(mean_max_top30_differences_crop3_new, 2)
        max_max_top30_differences_crop3_new = max(max_top30_differences_crop_3_new)
        min_max_top30_differences_crop_except0_3_new = min(max_top30_differences_crop_except0_3_new)
        min_max_top30_differences_crop_except0_3_index_new = max_top30_differences_crop_3_new.index(max_max_top30_differences_crop3_new)
        
        top30_max_image_path_4_new = os.path.join(folder_path, f'{FRAME_PREFIX}{min_max_top30_differences_crop_except0_3_index_new + start_number :04d}.{FRAME_EXTENSION}') 
        
        x_differences = []  
        y_differences = []
        valid_pixel_counts_min_new = []
        valid_pixel_counts_max_new = []
        valid_pixel_all_counts_max_new = []
        valid_pixel_all_counts_min_new = []

        for i in range(len(gray_frames)):

            if use_gpu:
                image1 = gray_tensor[i].cpu().numpy().astype(np.uint8)
            else:
                image1 = gray_frames[i].astype(np.uint8)


            max_image_path_pixel_new = top30_max_image_path_3_new 
            min_image_path_pixel_new = top30_max_image_path_2_new 
            
            max_image_pixel_new = current_rgb_image_list[image_frame_n2]
            min_image_pixel_new = current_rgb_image_list[image_frame_n1]
            max_image_gray_pixel_new = cv2.cvtColor(max_image_pixel_new, cv2.COLOR_BGR2GRAY)
            min_image_gray_pixel_new = cv2.cvtColor(min_image_pixel_new, cv2.COLOR_BGR2GRAY)
            
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new] #small_size
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_image0_pixel_new = image0[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_image1_pixel_new = image1[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f'Error cropping image{i}. Skipping...')
                continue 
            
            diff_max_new = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
            diff_min_new = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 
            
            # Count pixels with intensity differences above the nonzero minimum threshold.
            # KR: 0을 제외한 최소 threshold 이상으로 변화한 pixel의 좌표를 계산합니다.
            threshold_max_new = min_max_top30_differences_crop_except0_3_new 
            threshold_min_new = min_max_top30_differences_crop_except0_2_new  
            white_pixel_coords_max_new = np.argwhere(diff_max_new >= threshold_max_new)
            white_pixel_coords_min_new = np.argwhere(diff_min_new >= threshold_min_new)
            
            min_threshold_max_new = min_max_top30_differences_crop_except0_3_new +20 
            max_threshold_max_new = max_max_top30_differences_crop3_new 

            min_threshold_min_new = min_max_top30_differences_crop_except0_2_new +20 
            max_threshold_min_new = max_max_top30_differences_crop2_new 
            
            valid_pixel_coords_max_new = np.argwhere((diff_max_new >= min_threshold_max_new) & (diff_max_new <= max_threshold_max_new))
            valid_pixel_coords_min_new = np.argwhere((diff_min_new >= min_threshold_min_new) & (diff_min_new <= max_threshold_min_new))
            
            valid_pixel_coords_max_all_new = np.argwhere((diff_max_new >= 0) & (diff_max_new <= max_threshold_max))
            valid_pixel_coords_min_all_new = np.argwhere((diff_min_new >= 0) & (diff_min_new <= max_threshold_min))

            valid_pixel_count_max_new = len(valid_pixel_coords_max_new)
            valid_pixel_count_min_new = len(valid_pixel_coords_min_new)
            valid_pixel_counts_max_new.append(valid_pixel_count_max_new)
            valid_pixel_counts_min_new.append(valid_pixel_count_min_new)       


        valid_pixel_count_mean_min_new = np.mean([value for value in valid_pixel_counts_min_new if value not in [0,'']]) 

        valid_pixel_count_mean_round_min_new = round(valid_pixel_count_mean_min_new) if not np.isnan(valid_pixel_count_mean_min_new) else 0 
        
        valid_pixel_count_mean_max_new = np.mean([value for value in valid_pixel_counts_max_new if value not in [0,'']]) 

        valid_pixel_count_mean_round_max_new = round(valid_pixel_count_mean_max_new) if not np.isnan(valid_pixel_count_mean_max_new) else 0 

        count_zeros_min_new = valid_pixel_counts_min_new.count(0)
        count_zeros_max_new = valid_pixel_counts_max_new.count(0)


        count_all_pixel_in_range_min_new = [valid_pixel_counts_min_new.count(i) for i in range(valid_pixel_count_mean_round_min_new)]
        count_all_pixel_in_range_max_new = [valid_pixel_counts_max_new.count(i) for i in range(valid_pixel_count_mean_round_max_new)]
        
        sum_count_all_pixel_in_range_min_new = sum(count_all_pixel_in_range_min_new, 0)
        sum_count_all_pixel_in_range_max_new = sum(count_all_pixel_in_range_max_new, 0)

        valid_pixel_all_counts_min_new.append(sum_count_all_pixel_in_range_min_new)
        valid_pixel_all_counts_max_new.append(sum_count_all_pixel_in_range_max_new)
        
        valid_pixel_all_counts_min_max_new = max(valid_pixel_all_counts_min_new)
        valid_pixel_all_counts_max_max_new = max(valid_pixel_all_counts_max_new)

        All_pixel_intensities = []  
        valid_pixel_counts = []
        not_cropped_valid_pixel_counts = []
        selected_max_differences = []
        selected_max_differences_location = []
        Analysis_frames = []
        pixel_intensities_movement_list = []
        pixel_intensities_movement_coordinates_list = []



        for i in range(len(gray_frames)):

            max_image_path_pixel_new = top30_max_image_path_3_new
            min_image_path_pixel_new = top30_max_image_path_2_new
            
            max_image_pixel_new = current_rgb_image_list[image_frame_n2]
            min_image_pixel_new = current_rgb_image_list[image_frame_n1]
            max_image_gray_pixel_new = cv2.cvtColor(max_image_pixel_new, cv2.COLOR_BGR2GRAY)
            min_image_gray_pixel_new = cv2.cvtColor(min_image_pixel_new, cv2.COLOR_BGR2GRAY)
            
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_image0_pixel_new = image0[y1_small_new:y2_small_new, x1_small_new:x2_small_new]  
            cropped_image1_pixel_new = image1[y1_small_new:y2_small_new, x1_small_new:x2_small_new]  
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f"Error cropping image {i}. Skipping...")
                continue 
            
            # Select the reference candidate frame based on valid-pixel count statistics.
            # The selected candidate determines which images, thresholds, coordinates,
            # and intensity-difference values are used in downstream displacement analysis.
            # KR: valid-pixel count 통계를 기준으로 기준 후보 frame을 선택합니다.
            # KR: 선택된 후보 frame에 따라 이후 displacement 분석에 사용할 이미지, threshold, 좌표, intensity-difference 값이 결정됩니다.
            
            # Apply default or reversed candidate-selection logic depending on the expected contraction direction.
            # KR: 예상되는 수축 방향에 따라 기본 후보 선택 기준 또는 반전된 후보 선택 기준을 적용합니다.
            if REVERSE_CONTRACTION.lower() == 'off':
                if valid_pixel_all_counts_min_max_new >= valid_pixel_all_counts_max_max_new:
                    # Use the candidate relaxation frame as the reference image
                    # KR: 후보 이완 frame을 기준 이미지로 사용합니다.
                    diff = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new)  
                    # Full-frame difference used for visualization
                    # KR: visualization을 위해 full-frame difference를 계산합니다.
                    not_crop_diff = cv2.subtract(max_image_gray_pixel_new, image1) 
                    # Difference between candidate contraction and relaxation frames
                    # KR: 후보 수축 frame과 후보 이완 frame 사이의 차이입니다.
                    crop_diff = cv2.subtract(cropped_max_image_pixel_new, cropped_min_image_pixel_new) 
                    #=================================================================
                    # Store paired candidate contraction/relaxation frame information
                    # KR: 후보 수축/이완 frame 쌍의 정보를 저장합니다.
                    #=================================================================
                    # Reference image: candidate relaxation frame
                    # KR: 기준 이미지: 후보 이완 frame
                    square_image = min_image_pixel_new
                    # Path to the reference relaxation frame
                    # KR: 기준 이완 frame의 경로
                    square_image_path = min_image_path_pixel_new
                    # Paired candidate contraction frame
                    # KR: 기준 frame과 짝이 되는 후보 수축 frame
                    square_image_path_contraction = max_image_path_pixel_new 
                    # Loaded RGB image of the reference relaxation frame
                    # KR: 기준 이완 frame의 RGB 이미지
                    square_image_imread = current_rgb_image_list[image_frame_n1]
                    # Loaded RGB image of the paired contraction frame
                    # KR: 짝이 되는 후보 수축 frame의 RGB 이미지
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n2]
                    # Index of the selected reference frame
                    # KR: 선택된 기준 frame의 index
                    square_image_index = image_frame_n1
                    # Image showing the maximum difference from the reference frame
                    # KR: 기준 frame과 가장 큰 차이를 보이는 이미지
                    square_differences_crop_indices_index_path_image = max_image_pixel_new 
                    # Minimum intensity-difference threshold
                    # KR: 최소 intensity-difference threshold
                    threshold = min_max_top30_differences_crop_except0_2_new 
                    # Maximum intensity-difference threshold
                    # KR: 최대 intensity-difference threshold
                    max_threshold = max_max_top30_differences_crop2_new
                    # Maximum intensity difference
                    # KR: intensity difference 값
                    diff_val = max_diff_crop_top30_max_2_new   
                    # List of maximum intensity differences
                    # KR: 최대 intensity difference 리스트
                    diff2_val = max_top30_differences_crop_2_new
                    # Coordinates of the maximum intensity difference
                    # KR: 최대 intensity difference 좌표
                    loc = max_diff_loc_crop_top30_max_2_new
                    # Coordinate of the maximum intensity difference within the initial analysis window
                    # KR: 초기 분석 구간 안에서 최대 intensity difference가 나타난 좌표
                    loc2 = max_top30_diff_locations_crop_2_new
                    # Coordinate of the maximum intensity difference within the initial analysis window
                    # KR: 초기 분석 구간 안에서 최대 intensity difference가 나타난 좌표
                    loc3 = location_in_Frames_2 
                    # Maximum intensity difference from the reference frame
                    # KR: 기준 frame과 가장 큰 차이를 보이는 intensity difference 값
                    max_values = max_max_top30_differences_crop2_new 
                    ## Index of the frame showing the maximum difference from the reference frame
                    # KR: 기준 frame과 가장 큰 차이를 보이는 frame의 index
                    max_indices = min_max_top30_differences_crop_except0_2_index_new 
                    # List of maximum intensity differences
                    # KR: 최대 intensity difference 리스트
                    adapt = max_top30_differences_crop_2_new 
                    # Maximum intensity-difference value
                    # KR: 최대 intensity-difference 값
                    adapt_max = max_max_top30_differences_crop2_new 
                    # Mean intensity-difference value
                    # KR: 평균 intensity-difference 값
                    adapt_mean = mean_round_max_top30_differences_crop2_new 
                    
                else:
                    diff = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
                    not_crop_diff = cv2.subtract(min_image_gray_pixel_new, image1) 
                    crop_diff = cv2.subtract(cropped_min_image_pixel_new, cropped_max_image_pixel_new) 
                    square_image = max_image_pixel_new 
                    square_image_path = max_image_path_pixel_new 
                    square_image_path_contraction = max_image_path_pixel_new 
                    square_image_imread = current_rgb_image_list[image_frame_n2] 
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n1] 
                    square_image_index = image_frame_n2 
                    square_differences_crop_indices_index_path_image = min_image_pixel_new 
                    threshold = min_max_top30_differences_crop_except0_3_new  
                    max_threshold = max_max_top30_differences_crop3_new 
                    diff_val = max_diff_crop_top30_max_3_new 
                    diff2_val = max_top30_differences_crop_3_new 
                    loc = max_diff_loc_crop_top30_max_3_new 
                    loc2 = max_top30_diff_locations_crop_3_new 
                    loc3 = location_in_Frames_3 
                    max_values = max_max_top30_differences_crop3_new 
                    max_indices = min_max_top30_differences_crop_except0_3_index_new 
                    adapt = max_top30_differences_crop_3_new 
                    adapt_max = max_max_top30_differences_crop3_new 
                    adapt_mean = mean_round_max_top30_differences_crop3_new 
            else: 
                if valid_pixel_all_counts_min_max_new <= valid_pixel_all_counts_max_max_new:
                    diff = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 
                    not_crop_diff = cv2.subtract(max_image_gray_pixel_new, image1) 
                    crop_diff = cv2.subtract(cropped_max_image_pixel_new, cropped_min_image_pixel_new) 
                    square_image = min_image_pixel_new 
                    square_image_path = min_image_path_pixel_new 
                    square_image_path_contraction = max_image_path_pixel_new 
                    square_image_imread = current_rgb_image_list[image_frame_n1] 
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n2] 
                    square_image_index = image_frame_n1
                    square_differences_crop_indices_index_path_image = max_image_pixel_new 
                    threshold = min_max_top30_differences_crop_except0_2_new 
                    max_threshold = max_max_top30_differences_crop2_new 
                    diff_val = max_diff_crop_top30_max_2_new  
                    diff2_val = max_top30_differences_crop_2_new 
                    loc = max_diff_loc_crop_top30_max_2_new 
                    loc2 = max_top30_diff_locations_crop_2_new 
                    loc3 = location_in_Frames_2 
                    max_values = max_max_top30_differences_crop2_new 
                    max_indices = min_max_top30_differences_crop_except0_2_index_new 
                    adapt = max_top30_differences_crop_2_new 
                    adapt_max = max_max_top30_differences_crop2_new 
                    adapt_mean = mean_round_max_top30_differences_crop2_new 
                    
                else:
                    diff = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
                    not_crop_diff = cv2.subtract(min_image_gray_pixel_new, image1) 
                    crop_diff = cv2.subtract(cropped_min_image_pixel_new, cropped_max_image_pixel_new) 
                    square_image = max_image_pixel_new 
                    square_image_path = max_image_path_pixel_new 
                    square_image_path_contraction = max_image_path_pixel_new 
                    square_image_imread = current_rgb_image_list[image_frame_n2] 
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n1] 
                    square_image_index = image_frame_n2 
                    square_differences_crop_indices_index_path_image = min_image_pixel_new 
                    threshold = min_max_top30_differences_crop_except0_3_new  
                    max_threshold = max_max_top30_differences_crop3_new 
                    diff_val = max_diff_crop_top30_max_3_new 
                    diff2_val = max_top30_differences_crop_3_new 
                    loc = max_diff_loc_crop_top30_max_3_new 
                    loc2 = max_top30_diff_locations_crop_3_new  
                    loc3 = location_in_Frames_3 
                    max_values = max_max_top30_differences_crop3_new 
                    max_indices = min_max_top30_differences_crop_except0_3_index_new 
                    adapt = max_top30_differences_crop_3_new 
                    adapt_max = max_max_top30_differences_crop3_new 
                    adapt_mean = mean_round_max_top30_differences_crop3_new 
                

            max_pixel_intensities = int(not_crop_diff.max())
            All_pixel_intensities.append(max_pixel_intensities)
            max_All_pixel_intensities = int(np.max(All_pixel_intensities))
            All_pixel_intensities.append(max_pixel_intensities)
            max_All_pixel_intensities = np.max(All_pixel_intensities)

            selected_image_path = square_image_path  
            
            selected_image = square_image_imread 
            selected_image_contraction = square_image_imread_contraction 
            selected_image_gray = cv2.cvtColor(selected_image, cv2.COLOR_BGR2GRAY) 
            selected_image_contraction_gray = cv2.cvtColor(selected_image_contraction, cv2.COLOR_BGR2GRAY) 
            
            all_image_gray = current_gray_image_list[i] 
            adapt = np.array(adapt) 
            adapt_min = adapt[adapt >= threshold].tolist()
            try:
                sorted_adapt_min = min(adapt_min)
                sorted_adapt_max = max(adapt_min)
                index_of_sorted_adapt_min = int(np.where(adapt == sorted_adapt_min)[0][0])
                index_of_sorted_adapt_max = int(np.where(adapt == sorted_adapt_max)[0][0])
                coordinate_of_sorted_adapt_min = loc2[index_of_sorted_adapt_min]
                coordinate_of_sorted_adapt_max = loc2[index_of_sorted_adapt_max]
            except ValueError:
                break
            
            cropped_selected_image_gray = selected_image_gray[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
            cropped_all_image_gray = all_image_gray[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
            
            cropped_selected_image_gray_reference = selected_image_gray[y1_small_new_optical:y2_small_new_optical, x1_small_new_optical:x2_small_new_optical] 
            cropped_all_image_gray_reference = all_image_gray[y1_small_new_optical:y2_small_new_optical, x1_small_new_optical:x2_small_new_optical] 
            
            
            diff_selected = cv2.absdiff( cropped_all_image_gray, cropped_selected_image_gray) 
            diff_selected_reference = cv2.subtract( cropped_all_image_gray_reference, cropped_selected_image_gray_reference)

            pixel_intensities_movement = diff_selected.ravel()

            diff_selected_column = np.argwhere(diff_selected >= threshold)

            pixel_intensities_movement_list.append(pixel_intensities_movement)
            pixel_intensities_movement_coordinates_list.append(diff_selected_column)

            max_pixel_intensity = int(diff_selected.max()) if diff_selected.size else 0

            
            reference_selected_image_gray = selected_image_contraction_gray[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference]
            reference_all_image_gray = all_image_gray[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
            
            not_crop_selected_image_gray = selected_image_contraction_gray 
            not_crop_all_image_gray = all_image_gray

            reference_diff_selected = cv2.absdiff(reference_selected_image_gray, reference_all_image_gray) 
            not_crop_diff_selected = cv2.absdiff(not_crop_selected_image_gray, not_crop_all_image_gray) 
            
            diff_loc = np.argwhere((diff_selected_reference >= threshold) & (diff_selected_reference <= max_threshold))
            not_cropped_diff_loc = np.argwhere((reference_diff_selected >= threshold) & (reference_diff_selected <= max_threshold)) 
            
            threshold_pixel_image = np.zeros_like(diff_selected_reference, dtype = np.uint8)
            threshold_pixel_image_color = cv2.cvtColor(threshold_pixel_image, cv2.COLOR_GRAY2BGR)
            for loc in diff_loc:
                threshold_pixel_image_color[loc[0], loc[1]] = diff_selected_reference[loc[0], loc[1]]
            
            result_folder_diff = os.path.join(folder_path, 'result', 'Analysis Square') 
            if not os.path.exists(result_folder_diff):
                os.makedirs(result_folder_diff, exist_ok = True) 
            x_coord, y_coord = x_selected_new - x1_small_new_optical ,y_selected_new - y1_small_new_optical
        
            height, width = diff_selected.shape[:2]
            not_cropped_height, not_cropped_width = reference_diff_selected.shape[:2]

            new_width = 1080

            new_height = int(height * (new_width / width))
            diff_resized = cv2.resize(threshold_pixel_image_color, (new_width, new_height), interpolation=cv2.INTER_LINEAR) 
            result_image_path = os.path.join(result_folder_diff, f'Analysis_Frame_{i:04d}.{FRAME_EXTENSION}')
            Analysis_frames.append(diff_resized)
            white_pixel_coords = np.argwhere(diff_selected >= adapt_mean)
            
            not_cropped_white_pixel_coords = np.argwhere(reference_diff_selected >= adapt_mean)
            
            white_pixel_count = len(white_pixel_coords)
            
            not_cropped_white_pixel_count = len(not_cropped_white_pixel_coords)
            
            min_threshold = threshold
            max_threshold = max_threshold
            
            valid_pixel_coords = np.argwhere((diff_selected >= min_threshold) & (diff_selected <= max_threshold))
            not_cropped_valid_pixel_coords = np.argwhere((reference_diff_selected >= min_threshold) & (reference_diff_selected <= max_threshold))
            
            valid_pixel_count = len(valid_pixel_coords)
            valid_pixel_counts.append(valid_pixel_count)
            total_pixel_count = resolution_x * resolution_y 
            
            not_cropped_valid_pixel_count = len(not_cropped_valid_pixel_coords)
            not_cropped_valid_pixel_counts.append(not_cropped_valid_pixel_count)   



        valid_pixel_counts_np = np.array(valid_pixel_counts, dtype='float32') 

        valid_values = valid_pixel_counts_np[(valid_pixel_counts_np != 0) & ~np.isnan(valid_pixel_counts_np)]

        valid_pixel_count_mean = np.mean(valid_values) if valid_values.size > 0 else 0
        valid_pixel_count_mean_round = round(valid_pixel_count_mean)

        valid_pixel_count_maximum = np.max(valid_values) if valid_values.size > 0 else 0
        valid_pixel_count_maximum_round = round(valid_pixel_count_maximum)

        min_valid_pixel_counts = np.min(valid_values) if valid_values.size > 0 else 0
        
        if not valid_pixel_counts:
            print("!!!!THIS SAMPLE DOESN'T HAVE ANY MOVEMENT!!!!.")
            classification_dir = os.path.join(BAD_FILE_PATH, 'FINISHED')
            destination_folder = os.path.join(classification_dir, 'BAD')
            folder_path = os.path.join(BAD_FILE_PATH, folder_name)
            video_path = os.path.join(BAD_FILE_PATH, f'{folder_name}.{extension}')  
            os.makedirs(classification_dir, exist_ok=True)
            os.makedirs(destination_folder, exist_ok=True)
            try:
                shutil.move(folder_path, destination_folder)
                shutil.move(video_path, destination_folder)
            except Exception as e:
                pass
            continue
        

        difference_square_image = cv2.subtract(square_differences_crop_indices_index_path_image, square_image)
        difference_square_image_gray = cv2.cvtColor(difference_square_image, cv2.COLOR_BGR2GRAY)
        

        
        
        # 동영상 파일명과 확장자 설정
        video_name_base = f'{result_folder_diff}\\Analysis_Video'
        video_name = f'{video_name_base}\\{folder_name}_Analysis_Video.avi'
        if not os.path.exists(video_name_base):
            os.makedirs(video_name_base)
        # VideoWriter 객체 생성
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(video_name, fourcc, fps, (new_width, new_height))
        # 프레임 이미지들을 읽어와 동영상에 추가
        for i in range(len(gray_frames)):
            out.write(Analysis_frames[i])

        # VideoWriter 객체 해제
        out.release()
        print("Analysis Video Completed.")
        
        gray_color = '#bbbbbd'
        red_color = '#FF6699'
        blue_color = '#009fd6'
        yellow_color = '#d69f00'
        # 가장 많이 움직인 곳 사각형 표시 #214,159,0 (파란색)


        # ============================================
        # Optical flow-based displacement analysis
        # KR: optical flow 기반 displacement 분석
        # ============================================

        # Prepare candidate contraction and relaxation frames for ROI-based optical flow.
        # KR: ROI 기반 optical flow 분석을 위해 후보 수축/이완 frame을 준비합니다.
        roi_contraction = cv2.cvtColor(
            square_differences_crop_indices_index_path_image, cv2.COLOR_BGR2GRAY
        )
        roi_relaxation = cv2.cvtColor(
            square_image, cv2.COLOR_BGR2GRAY
        )

        roi_contraction = roi_contraction[
            y1_small_new_optical:y2_small_new_optical,
            x1_small_new_optical:x2_small_new_optical
        ]
        roi_relaxation = roi_relaxation[
            y1_small_new_optical:y2_small_new_optical,
            x1_small_new_optical:x2_small_new_optical
        ]

        # Skip optical flow when the ROI is empty.
        # KR: ROI가 비어 있으면 optical flow 분석을 건너뜁니다.
        if roi_contraction.size == 0:
            print("roi_contraction is empty. Optical flow skipped.")
            target_displacements = [0.0]
            maximum_displacement_list = target_displacements
            max_target_displacements = 0.0
            max_displacement_um = 0.0
            min_target_displacements = 0.0

            # Use the selected coordinate as the fallback position.
            # KR: optical flow를 수행할 수 없을 때 선택 좌표를 fallback 위치로 사용합니다.
            a_global = x_selected_new
            b_global = y_selected_new
            c_global = x_selected_new
            d_global = y_selected_new

            global_max_coords_global = (b_global, a_global, d_global, c_global)
            global_max_coords_global1 = (b_global, a_global)
            global_max_coords_global2 = (d_global, c_global)
            global_max_coords_global1_ab = (a_global, b_global)
            global_max_coords_global2_cd = (c_global, d_global)

            tracked_points = [(x_selected_new - x1_small_new_optical,
                            y_selected_new - y1_small_new_optical)]
            best_x_roi = tracked_points[0][0]
            best_y_roi = tracked_points[0][1]
            best_x_roi_min = best_x_roi
            best_y_roi_min = best_y_roi

        else:
            # Set the output path for the optical flow overlay video.
            # KR: optical flow overlay video의 저장 경로를 설정합니다.
            output_video_path = os.path.join(result_folder, f'{folder_name}_Optical_Flow_Overlay.{extension}')

            fps = fps  
            frame_size = (roi_contraction.shape[1], roi_contraction.shape[0])  # Video frame size: width, height

            fourcc = cv2.VideoWriter_fourcc(*'XVID')
            out = cv2.VideoWriter(output_video_path, fourcc, fps, frame_size)

            # ------------------------------------------------
            # 2) Define the baseline tracking point within the ROI
            #    The selected motion center is clamped to remain inside the ROI.
            # KR: ROI 내부에서 baseline tracking point를 정의합니다.
            # KR: 선택된 motion center가 ROI 밖으로 벗어나지 않도록 좌표를 제한합니다.
            # ------------------------------------------------
            H, W = roi_contraction.shape[:2]

            x0 = int(x_selected_new - x1_small_new_optical)
            y0 = int(y_selected_new - y1_small_new_optical)

            half_win = ANALYSIS_SQUARE_SIZE_REFERENCE // 2    # Half-window size for coordinate clamping
            # Clamp the tracking point so that the local window remains inside the ROI.
            # KR: local tracking window가 ROI 내부에 유지되도록 tracking point 좌표를 제한합니다.
            x0 = max(half_win, min(W - 1 - half_win, x0))
            y0 = max(half_win, min(H - 1 - half_win, y0))

            # Use a single baseline tracking point for Lucas-Kanade optical flow.
            # KR: Lucas-Kanade optical flow에서 하나의 baseline tracking point만 사용합니다.
            p0 = np.array([[[x0, y0]]], dtype=np.float32)

            # Store optical-flow tracking results.
            # KR: optical flow tracking 결과를 저장합니다.
            target_displacements = []   # Displacement for each frame in µm
            tracked_points = []         # Tracked ROI coordinates for each frame

            # Smoothing parameters for visualization only.
            # KR: visualization에만 사용하는 smoothing 설정입니다.
            smoothed_x, smoothed_y = None, None
            alpha = 0.5  # EMA smoothing factor

            # ------------------------------------------------
            # 3) Optical flow tracking from the baseline ROI to each frame ROI
            #    A single point p0 is tracked to define optical flow-derived displacement.
            # KR: baseline ROI에서 각 frame ROI로 optical flow tracking을 수행합니다.
            # KR: 하나의 p0 point를 추적하여 optical flow-derived displacement를 정의합니다.
            # ------------------------------------------------
            for frame_idx, frame_full in enumerate(current_gray_image_list):

                cur_gray = frame_full[
                    y1_small_new_optical:y2_small_new_optical,
                    x1_small_new_optical:x2_small_new_optical
                ]

                # Keep the previous displacement value when the current ROI is empty.
                # KR: 현재 ROI가 비어 있으면 이전 displacement 값을 유지합니다.
                if cur_gray.size == 0:
                    last_disp = target_displacements[-1] if target_displacements else 0.0
                    target_displacements.append(float(last_disp))
                    tracked_points.append((float(x0), float(y0)))
                    continue

                p1, st, err = cv2.calcOpticalFlowPyrLK(
                    roi_contraction,
                    cur_gray,
                    p0,
                    None,
                    winSize=(15, 15),
                    maxLevel=2,
                    criteria=(cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, 10, 0.03)
                )

                frame_rgb = cv2.cvtColor(cur_gray, cv2.COLOR_GRAY2BGR)

                # Handle optical-flow tracking failure.
                # KR: optical flow tracking에 실패한 경우를 처리합니다.
                if p1 is None or st is None or st[0, 0] == 0:
                    last_disp = target_displacements[-1] if target_displacements else 0.0
                    target_displacements.append(float(last_disp))
                    tracked_points.append((float(x0), float(y0)))

                    # Write the frame even when optical-flow tracking fails.
                    # KR: optical flow tracking에 실패해도 output video frame은 저장합니다.
                    out.write(frame_rgb)
                    continue

                # New tracked coordinate in the baseline ROI coordinate system.
                # KR: baseline ROI 좌표계에서 새롭게 추적된 좌표입니다.
                new_x = float(p1[0, 0, 0])
                new_y = float(p1[0, 0, 1])

                # Calculate displacement from the baseline point p0.
                # This is not cumulative; the same baseline point is used for all frames.
                # KR: baseline point p0로부터의 displacement를 계산합니다.
                # KR: 누적 변위가 아니라 모든 frame에서 동일한 기준점 p0를 사용합니다.
                dx = new_x - x0
                dy = new_y - y0
                disp_um = np.sqrt(dx * dx + dy * dy) * PIXEL_UM

                target_displacements.append(float(disp_um))
                tracked_points.append((new_x, new_y))

                # Smooth the displayed tracking point for visualization.
                # KR: visualization에서 표시되는 tracking point를 부드럽게 보이도록 smoothing합니다.
                if smoothed_x is None:
                    smoothed_x, smoothed_y = new_x, new_y
                else:
                    smoothed_x = alpha * new_x + (1.0 - alpha) * smoothed_x
                    smoothed_y = alpha * new_y + (1.0 - alpha) * smoothed_y

                # Draw a green arrow from the baseline point to the current tracked position,
                # and mark the smoothed tracked point in red.
                # KR: baseline point에서 현재 추적 위치까지 초록색 화살표를 그리고,
                # KR: smoothing된 tracking point를 빨간색으로 표시합니다.
                cv2.arrowedLine(
                    frame_rgb,
                    (int(x0), int(y0)),
                    (int(new_x), int(new_y)),
                    (0, 255, 0),
                    1
                )
                cv2.circle(
                    frame_rgb,
                    (int(smoothed_x), int(smoothed_y)),
                    2,
                    (0, 0, 255),
                    -1
                )

                out.write(frame_rgb)
            # Finalize the optical flow overlay video.
            # KR: optical flow overlay video 저장을 종료합니다.
            out.release()

            # ------------------------------------------------
            # 4) Displacement time-series post-processing
            #    - target_displacements are stored in micrometers.
            #    - The analysis window is selected based on start_number, NUM_FRAMES, and fps.
            # KR: displacement time-series를 후처리합니다.
            # KR: target_displacements는 micrometer 단위입니다.
            # KR: 분석 구간은 start_number, NUM_FRAMES, fps를 기준으로 선택합니다.
            # ------------------------------------------------
            if len(target_displacements) == 0:
                target_displacements = [0.0]

            # Shift the minimum displacement to zero.
            # KR: 최소 displacement가 0이 되도록 baseline을 보정합니다.
            min_all = min(target_displacements)
            target_displacements = [float(v - min_all) for v in target_displacements]
            
            # Invert the displacement curve for visualization consistency.
            # KR: visualization 방향을 맞추기 위해 displacement curve를 반전합니다.
            max_all = max(target_displacements)
            target_displacements = [float(max_all - v) for v in target_displacements]

            # Select the initial analysis window.
            # KR: 초기 분석 구간을 선택합니다.
            end_idx = int(fps * NUM_FRAMES)
            end_idx = min(end_idx, len(target_displacements))

            if start_number + 1 < end_idx:
                window_vals = target_displacements[start_number + 1:end_idx]
                window_indices = list(range(start_number + 1, end_idx))
            else:
                window_vals = target_displacements
                window_indices = list(range(len(target_displacements)))

            if len(window_vals) == 0:
                window_vals = target_displacements
                window_indices = list(range(len(target_displacements)))

            # Define the final displacement time series and maximum displacement value.
            # KR: 최종 displacement time series와 maximum displacement 값을 정의합니다.
            maximum_displacement_list = target_displacements
            max_target_displacements = round(max(window_vals), 2)
            min_target_displacements = float(min(window_vals))
            max_displacement_um = max_target_displacements

            max_target_displacements_before_all = round(max(target_displacements), 2)

            # Create the time axis for displacement plotting.
            # KR: displacement plot에 사용할 시간축을 생성합니다.
            if fps > 0:
                total_time_seconds = len(maximum_displacement_list) / fps
            else:
                total_time_seconds = float(len(maximum_displacement_list))
            time_axis = np.linspace(0, total_time_seconds, len(maximum_displacement_list))

            # Frame index where the maximum displacement occurs within the analysis window.
            # KR: 분석 구간 안에서 maximum displacement가 발생한 frame index입니다.
            idx_in_window = int(np.argmax(window_vals))
            frame_idx_max = int(window_indices[idx_in_window])

            # Extract ROI coordinates of the tracked point at maximum and baseline displacement.
            # KR: maximum displacement와 baseline displacement 시점의 tracked point ROI 좌표를 가져옵니다.
            if frame_idx_max < len(tracked_points):
                best_x_roi, best_y_roi = tracked_points[frame_idx_max]
            else:
                best_x_roi, best_y_roi = tracked_points[-1]

            if start_number < len(tracked_points):
                best_x_roi_min, best_y_roi_min = tracked_points[start_number]
            else:
                best_x_roi_min, best_y_roi_min = tracked_points[0]

            # Convert ROI coordinates back to full-frame coordinates.
            # KR: ROI 좌표를 full-frame 좌표로 변환합니다.
            a_global = int(best_x_roi + x1_small_new_optical)
            b_global = int(best_y_roi + y1_small_new_optical)
            c_global = int(best_x_roi_min + x1_small_new_optical)
            d_global = int(best_y_roi_min + y1_small_new_optical)

            global_max_coords_global = (b_global, a_global, d_global, c_global)
            global_max_coords_global1 = (b_global, a_global)
            global_max_coords_global2 = (d_global, c_global)
            global_max_coords_global1_ab = (a_global, b_global)
            global_max_coords_global2_cd = (c_global, d_global)
            
    
        # Convert the displacement time series to a comma-separated string for Excel output.
        # KR: Excel output에 저장하기 위해 displacement time series를 comma-separated string으로 변환합니다.
        maximum_displacement_str = ", ".join(f"{x:.2f}" for x in maximum_displacement_list)
                
        y_max_coordinates = int(y_selected_new - y1_small_new_optical)
        x_max_coordinates = int(x_selected_new - x1_small_new_optical)

        x1_max_coordinates = x_max_coordinates  - ANALYSIS_SQUARE_SIZE_REFERENCE
        y1_max_coordinates = y_max_coordinates  - ANALYSIS_SQUARE_SIZE_REFERENCE
        x2_max_coordinates = x_max_coordinates  + ANALYSIS_SQUARE_SIZE_REFERENCE
        y2_max_coordinates = y_max_coordinates  + ANALYSIS_SQUARE_SIZE_REFERENCE
        
        x1_max_coordinates = max( x1_max_coordinates, 0)
        y1_max_coordinates = max( y1_max_coordinates, 0)
        x2_max_coordinates = min( x2_max_coordinates, resolution_x -1)
        y2_max_coordinates = min( y2_max_coordinates, resolution_y -1)

        max_value_displacement = max(maximum_displacement_list) 
        
        if max_target_displacements == 0 or max_target_displacements is None:
            print("No valid displacement was detected.")
            print("This sample could not be tracked by optical flow.")
            print(f"Move to BAD Folder")
            classification_dir = os.path.join(BAD_FILE_PATH, 'FINISHED')
            destination_folder = os.path.join(classification_dir, 'BAD')
            folder_path = os.path.join(BAD_FILE_PATH, folder_name)
            video_path = os.path.join(BAD_FILE_PATH, f'{folder_name}.{extension}')  
            os.makedirs(classification_dir, exist_ok=True)
            os.makedirs(destination_folder, exist_ok=True)
            try:
                shutil.move(folder_path, destination_folder)
                shutil.move(video_path, destination_folder)
                continue
            except Exception as e:
                continue
        else:
            pass




        
        # Compute summary statistics after excluding zero-count frames.
        # KR: pixel count가 0인 frame을 제외하고 moved-region 통계값을 계산합니다.
        valid_pixel_count_mean = np.mean([value for value in valid_pixel_counts if value not in [0, '']])
        valid_pixel_count_mean_round = round(valid_pixel_count_mean) if not np.isnan(valid_pixel_count_mean) else 0
        valid_pixel_count_maximum = max([value for value in valid_pixel_counts if value not in [0, '']], default=0)
        valid_pixel_count_maximum_round = round(valid_pixel_count_maximum) if not np.isnan(valid_pixel_count_maximum) else 0

        # Convert moved-region pixel counts to physical area.
        # KR: moved region의 pixel count를 실제 면적으로 변환합니다.
        valid_pixel_count_mean_area = valid_pixel_count_mean_round * AREA_MOVED
        valid_pixel_count_maximum_area = valid_pixel_count_maximum_round * AREA_MOVED
        valid_pixel_count_mean_area_round = round(valid_pixel_count_mean_area,2)
        valid_pixel_count_maximum_area_round = round(valid_pixel_count_maximum_area, 2)
        
        # Frame-wise moved-region area and length-equivalent values.
        # KR: 각 frame별 moved-region area와 length-equivalent 값을 계산합니다.

        displacement_list = [count * PIXEL_UM for count in valid_pixel_counts]

        # =========================================================
        # Moved-region area time-series visualization
        # =========================================================
        # This plot shows the frame-wise moved-region area estimated from
        # thresholded motion pixels. The moved-region area is calculated by
        # multiplying the number of thresholded pixels by the calibrated pixel area.
        # This output is used as an auxiliary quality-control visualization and is
        # separate from optical-flow displacement.
        #
        # KR:
        # 이 plot은 threshold 조건을 통과한 moved pixel 수를 기반으로
        # frame별 moved-region area 변화를 보여줍니다.
        # moved-region area는 thresholded pixel count에 pixel area 보정값을
        # 곱하여 계산됩니다.
        # 이 결과는 보조적인 QC 시각화이며, optical-flow displacement 값과는
        # 별개의 지표입니다.
        # =========================================================
       
        

        fig, ax = plt.subplots(figsize=(8, 8))
        ax.plot(np.linspace(0, total_time_seconds, len(maximum_displacement_list)), 
                maximum_displacement_list, 
                color=blue_color, 
                linewidth=4)

        ax.set_xlabel('Time (s)',  fontsize=40, fontweight='bold', fontname='Arial')
        ax.set_ylabel(f'Displacement ({LENGTH_UNIT})',  fontsize=40,fontweight='bold', fontname='Arial')

        ax.set_ylim(0, Y_LIM)
        
        ax.set_yticks(range(0, Y_LIM + 1, 1))
        
        for spine in ax.spines.values():
            spine.set_linewidth(4) 
            
        ax.set_xlim(0, total_time_seconds)
        
        ax.set_xticks(range(0, int(total_time_seconds) + 1, X_TICKS))

        ax.axhline(y=max_target_displacements, color=red_color, linewidth=3.5, linestyle='--')

        ax.grid(False)
        plt.text(total_time_seconds - (total_time_seconds*0.0125), max_target_displacements + 0.625, f'Max Displacement ({LENGTH_UNIT}) : {max_target_displacements}',
                verticalalignment='top', horizontalalignment='right', color=red_color,fontweight='bold', fontsize=25)

        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=10)
        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=20)

        ax.spines['left'].set_position(('data', 0))  
        ax.spines['bottom'].set_position(('data', 0))  


        ax.spines['right'].set_color('black') 
        ax.spines['top'].set_color('black')    


        plt.xticks( fontweight='bold')
        plt.yticks( fontweight='bold')
        
        Savefig(folder_path, 'Optical_Flow_Displacement', 300, 0.2)
        
        
        if FRAME_REMOVE.lower() == 'on':
            paths_to_clean = [
                (folder_path, "result"),  
            ]
            
            for base_path, exclude_name in paths_to_clean:
                if os.path.exists(base_path): 
                    for item_name in os.listdir(base_path):
                        item_path = os.path.join(base_path, item_name)
                        if item_name != exclude_name:  
                            if os.path.isfile(item_path):
                                os.remove(item_path)
                            elif os.path.isdir(item_path):
                                shutil.rmtree(item_path)

        seq_raw_np = np.array(maximum_displacement_list, dtype=np.float32)

        if len(seq_raw_np) == 0:
            max_abs_disp = 0.0
            abs_net_change_ratio = 0.0
            sign_changes = 0
            diff_std = 0.0
        else:
            max_abs_disp = float(np.max(np.abs(seq_raw_np)))

            peak_to_peak = float(np.max(seq_raw_np) - np.min(seq_raw_np))

            abs_net_change_ratio = float(
                np.abs(seq_raw_np[-1] - seq_raw_np[0]) / (peak_to_peak + 1e-8)
            )

            diff_seq = np.diff(seq_raw_np)

            if len(diff_seq) >= 2:
                sign_changes = int(np.sum(np.diff(np.sign(diff_seq)) != 0))
            else:
                sign_changes = 0

            diff_std = float(np.std(diff_seq)) if len(diff_seq) > 0 else 0.0
    
        band_energy_optical_displacement, max_value_optical_displacement_band_energy, max_optical_amp_at_max_freq, optical_freq_at_max_amp_number  = plot_fft(folder_path, fps, target_displacements, X_LIM_AMPLITUDE, Y_LIM_AMPLITUDE,'Optical_Flow_Displacement_FFT')
        
    
    
        Excel_result_data_good = [
        folder_name,
        adapt_max,
        max_target_displacements,
        band_energy_optical_displacement,
        max_value_optical_displacement_band_energy,
        x_selected_new,
        y_selected_new,
        optical_freq_at_max_amp_number,
        maximum_displacement_str,
        max_abs_disp,
        abs_net_change_ratio,
        sign_changes,
        diff_std
    ]

        # 엑셀에 데이터 추가
        good_ws.append(Excel_result_data_good)
        good_wb.save(TARGET_EXCEL)
        print(f'Save The  『 {folder_name} 』 Information to {TARGET_SAMPLE_EXCEL_NAME}')
        
        finish_dir = os.path.join(TARGET_FILE_PATH, 'FINISHED')
        if not os.path.exists(finish_dir):
            os.makedirs(finish_dir, exist_ok=True)
        folder_path = os.path.join(TARGET_FILE_PATH, folder_name)
        video_path = os.path.join(TARGET_FILE_PATH, f'{folder_name}.{extension}')
        try:
            shutil.move(folder_path, finish_dir)
            shutil.move(video_path, finish_dir)
        except Exception as e:
            pass 

        # 현재 폴더의 처리 종료 시간 기록
        folder_end_time = time.time()
        
        # 현재 폴더의 실행 시간 계산 및 출력
        folder_execution_time = round((folder_end_time - folder_start_time),2)
        estimated_execution_times.append(folder_execution_time)  # 폴더의 실행 시간을 리스트에 추가

        # 각 폴더의 예상 처리 시간 계산 및 출력
        average_execution_time = round(sum(estimated_execution_times) / len(estimated_execution_times),2)
        # estimated_total_time = round(((average_execution_time * number_of_folders) / 60),1)
        estimated_left_time = round((((average_execution_time * number_of_folders) - sum(estimated_execution_times))/60),1)
        del Excel_result_data_good

        print(f"\n[INFO] Folder completed: {folder_name}")
        print(f"[INFO] Processing time: {folder_execution_time} seconds")
        print(f"[INFO] Estimated remaining time: {estimated_left_time} minutes\n")
        
        good_wb.close()
        plt.close()


    try:
        plt.close()
        # Excel 파일 사용이 끝났으므로 파일을 닫습니다.
        good_wb.close()
        gc.collect()
    except:
        # 이미 삭제된 변수에 대한 NameError는 무시하고 다음 코드를 실행합니다.
        pass


    print('============================ ACTIVE FILE FINISHED ============================')
    print('============================ NOISE FILE START ============================')
    #만약 결과를 저장할 기본 폴더가 없으면 생성
    if not os.path.exists(NOISE_FILE_PATH):
        os.makedirs(NOISE_FILE_PATH)
        
    video_files = []
    
    
    # 먼저 비디오 파일 목록 수집
    for filename in natsort.natsorted(os.listdir(NOISE_FILE_PATH)):
        for extension in VIDEO_FILE_EXTENSION:
            if filename.endswith(f'.{extension}'):
                video_files.append(filename)
                break

    cpu_count = multiprocessing.cpu_count()
    max_workers = round(max(1, int(cpu_count * 2/3)))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(extract_frames, NOISE_FILE_PATH, v) for v in video_files]

        for f in tqdm(futures, total=len(futures), desc="Frame Extraction"):
            f.result()

    # 예상 처리 시간을 저장할 리스트
    estimated_execution_times = []


    #기준 이미지 초기화
    reference_image = None

    # 기존 엑셀 파일 열기
    try:
        noise_wb = load_workbook(filename=NOISE_EXCEL)
        noise_ws = noise_wb.active
    except FileNotFoundError:
        # 기존 파일이 없는 경우 새로운 워크북 생성
        noise_wb = Workbook()
        noise_ws = noise_wb.active

        
    # 엑셀 파일에 헤더 추가
    noise_ws.append(headers)

    # parameter_ws.append(['NOISE_INDEX, MAX_NOISE_INDEX, PIXEL_MEAN_INDEX, MAX_PIXEL_MEAN_INDEX, DIFF_MAX_MIN_INDEX'])

    # 저장 위치의 폴더만 가져오기
    output_folders = [f for f in natsort.natsorted(os.listdir(NOISE_FILE_PATH)) if os.path.isdir(os.path.join(NOISE_FILE_PATH, f))]

    # 확인된 저장 위치의 폴더 갯수
    number_of_folders = len(output_folders)

    for folder_name in natsort.natsorted(os.listdir(NOISE_FILE_PATH)):
        folder_path = os.path.join(NOISE_FILE_PATH, folder_name)
        if not os.path.isdir(folder_path):
            continue # 폴더가 아니라면 스킵
        
        # Create the result folder for the current sample.
        result_folder = os.path.join(folder_path, 'result')
        os.makedirs(result_folder, exist_ok=True)
        
        print(f'\n【　Starting to Analyze the File {folder_name}　】\n ')
        print('--------------------------------List of Generated Files --------------------------------\n')
        folder_start_time = time.time()
        for extension in VIDEO_FILE_EXTENSION:
            fps_video_path = os.path.join(NOISE_FILE_PATH,f'{folder_name}.{extension}')
            if os.path.exists(fps_video_path):  # 파일이 존재하는 경우에만 FPS 계산
                cap = cv2.VideoCapture(fps_video_path)
                fps = round(cap.get(cv2.CAP_PROP_FPS), 1)
                cap.release()
                break # 파일이 존재하는 경우에는 이후 확장자에 대한 반복을 중지하고 다음 폴더로 넘어감
            else:
                fps = VIDEO_FPS   
        # 기준 이미지를 설정
        reference_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}0000.{FRAME_EXTENSION}')
        reference_image = cv2.imread(reference_image_path)
        
        # 기준 이미지를 성공적으로 로드 했는지 확인
        if reference_image is not None:
            reference_gray = cv2.cvtColor(reference_image, cv2.COLOR_BGR2GRAY)
        else:
            continue  # 기준 이미지 로드에 실패하면 다음 폴더로 건너뛸 수 있도록 continue 사용
        
        # 폴더 내 파일명에 있는 숫자 중 최댓값 찾기 (자동으로 시작, 끝 프레임 갯수 찾기 위해)
        max_number = -1
        min_number = float('inf')
        for filename in os.listdir(folder_path):
            if filename.startswith(f'{FRAME_PREFIX}') and filename.endswith(f'.{FRAME_EXTENSION}'):
                try:
                    # 파일명에서 'frame_'과 이미지 확장자를 제외한 부분을 추출하고 숫자로 변환
                    number = int(filename[len(f'{FRAME_PREFIX}'):-len(f'.{FRAME_EXTENSION}')]) #frame_0000.png 에서  0000만 추출 
                    max_number = max(max_number, number)
                    min_number = min(min_number, number)
                except ValueError:
                    pass # 숫자로 변환할 수 없는 경우 무시

        
        # 최댓값 출력
        start_number = min_number
        last_number = max_number+1 #마지막 프레임까지 계산 하기 위해 +1
        resolution_y, resolution_x,_ = reference_image.shape #해상도 자동 계산.
        
        try:
            total_time_seconds = int(max_number / fps)
        except ZeroDivisionError:
            pass
        
        

        # 이미지 파일 경로 리스트 초기화
        image_paths = []
            
        # Initialize lists for image paths and loaded frames
        # KR: 이미지 경로, grayscale frame, RGB frame을 저장할 리스트를 초기화합니다.
        image_paths = []
        current_gray_image_list = []
        current_rgb_image_list = []
        # # Initialize lists for maximum pixel-intensity differences and their coordinates
        max_difference_comparing_0 = []  #max_differences 
        max_difference_comparing_0_locations = [] # max_diff_locations

        max_top30_differences_crop = []
        max_top30_diff_locations_crop = []

        max_top30_differences_crop_2 = []
        max_top30_diff_locations_crop_2 = []

        max_top30_differences_crop_3 = []
        max_top30_diff_locations_crop_3 = []

        max_top30_differences_crop_2_new = []
        max_top30_diff_locations_crop_2_new = []

        max_top30_differences_crop_3_new = []
        max_top30_diff_locations_crop_3_new = []

        valid_pixel_all_counts_min = []
        valid_pixel_all_counts_max = []
        

        # =========================================================
        # (A) Load frame paths and grayscale images
        #     Frames are loaded once and then processed in GPU/CPU batch mode.
        # KR: frame 경로와 grayscale 이미지를 수집합니다.
        # KR: frame은 한 번만 로드한 뒤 이후 GPU/CPU batch 방식으로 처리합니다.
        # =========================================================

        use_gpu = (GPU_MODE.lower() == "on") and torch.cuda.is_available()
        device = torch.device("cuda" if use_gpu else "cpu")

        print(f"Compute mode: {'GPU' if use_gpu else 'CPU'}")


        image_paths = []
        gray_frames = []

        for idx in range(start_number, last_number):
            p = os.path.join(folder_path, f'{FRAME_PREFIX}{idx:04d}.{FRAME_EXTENSION}')
            if not os.path.exists(p):
                continue

            img_rgb = cv2.imread(p)
            if img_rgb is None:
                print(f"Image load fail → {p}")
                continue

            g = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)

            image_paths.append(p)
            gray_frames.append(g)
            current_rgb_image_list.append(img_rgb)


        if len(gray_frames) == 0:
            print("No valid frames found. Skip folder.")
            continue

        # =========================================================
        # (B) Compute frame-wise absolute differences from the reference image
        #     using GPU/CPU batch processing
        # KR: reference image와 각 frame 사이의 absolute difference를 계산합니다.
        # KR: 이 계산은 GPU 또는 CPU batch 처리로 한 번에 수행합니다.
        # =========================================================
        
        gray_tensor = torch.from_numpy(np.stack(gray_frames)).to(torch.int16)

        if use_gpu:
            gray_tensor = gray_tensor.pin_memory().to(device, non_blocking=True)
        else:
            gray_tensor = gray_tensor.contiguous()

        ref_tensor = torch.from_numpy(reference_gray).to(torch.int16)

        if use_gpu:
            ref_tensor = ref_tensor.pin_memory().to(device, non_blocking=True)
        else:
            ref_tensor = ref_tensor.contiguous()


        diff_tensor = torch.abs(gray_tensor - ref_tensor)                                 # [T,H,W]
        flat        = diff_tensor.flatten(1)                                              # [T,H*W]

        max_vals, argmax_vals = flat.max(dim=1)                                           # [T], [T]

        H, W = reference_gray.shape
        ys = (argmax_vals // W).cpu().numpy()
        xs = (argmax_vals %  W).cpu().numpy()

        max_difference_comparing_0 = max_vals.cpu().numpy().tolist()
        max_difference_comparing_0_locations = list(zip(xs, ys))

        current_gray_image_list = gray_frames

        image0_path = reference_image_path
        
        image0 = current_gray_image_list[0]

        
        # Search for the maximum intensity change within the initial fps × NUM_FRAMES window.
        # This helps focus the analysis on early contraction-related motion rather than large noise or bubble movement.
        # KR: 초기 fps × NUM_FRAMES 구간 안에서 최대 intensity change를 찾습니다.
        # KR: 큰 noise 또는 bubble 움직임보다 초기 수축 관련 움직임에 분석을 집중하기 위한 과정입니다.
        top30_max_differences_indices = np.arange(len(max_difference_comparing_0))[:int(fps*NUM_FRAMES)]
        top30_max_differences_values = [max_difference_comparing_0[i] for i in top30_max_differences_indices]
        top30_max_differences_values_max = max(top30_max_differences_values)
        
        max_value_index_within_top30 = np.argmax(top30_max_differences_values)
        
        max_difference_index_within_top30 = top30_max_differences_indices[max_value_index_within_top30]
        
        max_difference_value_within_top30 = top30_max_differences_values[max_value_index_within_top30]

        max_diff_location_within_top30 = max_difference_comparing_0_locations[max_difference_index_within_top30]

        x_center_max_diff_location_within_top30 = max_diff_location_within_top30[0]
        y_center_max_diff_location_within_top30 = max_diff_location_within_top30[1]

        
        x11 = x_center_max_diff_location_within_top30 - ANALYSIS_SQUARE_SIZE_REFERENCE
        y11 = y_center_max_diff_location_within_top30 - ANALYSIS_SQUARE_SIZE_REFERENCE
        x22 = x_center_max_diff_location_within_top30 + ANALYSIS_SQUARE_SIZE_REFERENCE
        y22 = y_center_max_diff_location_within_top30 + ANALYSIS_SQUARE_SIZE_REFERENCE
        x11 = max(x11, 0)
        y11 = max(y11, 0)
        x22 = min(x22, resolution_x - 1)
        y22 = min(y22, resolution_y - 1)


        roi_tensor = gray_tensor[:, y11:y22, x11:x22].contiguous()


        image_frame_number = []
        next_image_index = 0   

        visited = set()

        for step in range(min(len(gray_frames), 500)):
            if next_image_index in visited:
                break
            visited.add(next_image_index)

            image_frame_number.append(next_image_index)

            ref_roi = roi_tensor[next_image_index].unsqueeze(0)  # [1,h,w]
            diff_batch = torch.abs(roi_tensor - ref_roi)         # [T,h,w]
            flat = diff_batch.flatten(1)                         # [T,h*w]

            max_vals, _ = flat.max(dim=1)                        # [T]
            best_i = int(max_vals.argmax().item())

            next_image_index = best_i

            if len(image_frame_number) >= 8 and image_frame_number[-1] == image_frame_number[-3]:
                break

        image_frame_n1 = image_frame_number[-1]
        image_frame_n2 = image_frame_number[-2] if len(image_frame_number) >= 2 else image_frame_number[-1]

        n1_max_image_path = image_paths[int(image_frame_n1)]
        n2_max_image_path = image_paths[int(image_frame_n2)]


        roi_full = gray_tensor[:, y11:y22, x11:x22].contiguous()          # [T,h,w]

        ref_n1 = roi_full[int(image_frame_n1)].unsqueeze(0)               # [1,h,w]
        diff_n1 = torch.abs(roi_full - ref_n1)                            # [T,h,w]
        flat_n1 = diff_n1.flatten(1)                                      # [T,h*w]
        max_vals_n1, argmax_n1 = flat_n1.max(dim=1)                       # [T], [T]

        h_roi = roi_full.shape[1]
        w_roi = roi_full.shape[2]
        ys_n1 = (argmax_n1 // w_roi)
        xs_n1 = (argmax_n1 %  w_roi)

        max_top30_differences_crop_2 = max_vals_n1.detach().cpu().numpy().astype(np.float32).tolist()
        max_top30_diff_locations_crop_2 = list(zip(xs_n1.detach().cpu().numpy().tolist(),
                                                ys_n1.detach().cpu().numpy().tolist()))

        max_top30_differences_crop_except0_2 = [v for v in max_top30_differences_crop_2 if v != 0]
        mean_max_top30_differences_crop2 = float(np.mean(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0.0
        mean_round_max_top30_differences_crop2 = round(mean_max_top30_differences_crop2, 2)
        max_max_top30_differences_crop2 = float(np.max(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0.0

        if len(max_top30_differences_crop_except0_2) > 0:
            min_max_top30_differences_crop_except0_2 = float(np.min(max_top30_differences_crop_except0_2))
        else:
            min_max_top30_differences_crop_except0_2 = 0.0

        min_max_top30_differences_crop_except0_2_index = int(np.argmax(max_top30_differences_crop_2)) if len(max_top30_differences_crop_2) else 0

        ref_n2 = roi_full[int(image_frame_n2)].unsqueeze(0)               # [1,h,w]
        diff_n2 = torch.abs(roi_full - ref_n2)                            # [T,h,w]
        flat_n2 = diff_n2.flatten(1)                                      # [T,h*w]
        max_vals_n2, argmax_n2 = flat_n2.max(dim=1)                       # [T], [T]

        ys_n2 = (argmax_n2 // w_roi)
        xs_n2 = (argmax_n2 %  w_roi)

        max_top30_differences_crop_3 = max_vals_n2.detach().cpu().numpy().astype(np.float32).tolist()
        max_top30_diff_locations_crop_3 = list(zip(xs_n2.detach().cpu().numpy().tolist(),
                                                ys_n2.detach().cpu().numpy().tolist()))

        max_top30_differences_crop_except0_3 = [v for v in max_top30_differences_crop_3 if v != 0]
        mean_max_top30_differences_crop3 = float(np.mean(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0.0
        mean_round_max_top30_differences_crop3 = round(mean_max_top30_differences_crop3, 2)
        max_max_top30_differences_crop3 = float(np.max(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0.0

        if len(max_top30_differences_crop_except0_3) > 0:
            min_max_top30_differences_crop_except0_3 = float(np.min(max_top30_differences_crop_except0_3))
        else:
            min_max_top30_differences_crop_except0_3 = 0.0

        min_max_top30_differences_crop_except0_3_index = int(np.argmax(max_top30_differences_crop_3)) if len(max_top30_differences_crop_3) else 0

        try:
            top30_max_image_path_2 = image_paths[min_max_top30_differences_crop_except0_2_index]
        except:
            top30_max_image_path_2 = ""

        try:
            top30_max_image_path_3 = image_paths[min_max_top30_differences_crop_except0_3_index]
        except:
            top30_max_image_path_3 = ""

        diff_max_all = diff_n2                                           

        diff_min_all = diff_n1                                            

        threshold_max = min_max_top30_differences_crop_except0_3
        threshold_min = min_max_top30_differences_crop_except0_2

        min_threshold_max = threshold_max + 20
        max_threshold_max = max_max_top30_differences_crop3

        min_threshold_min = threshold_min + 20
        max_threshold_min = max_max_top30_differences_crop2

        if max_threshold_max <= 0:
            valid_pixel_counts_max = [0] * len(gray_frames)
        else:
            mask_max = (diff_max_all >= min_threshold_max) & (diff_max_all <= max_threshold_max)
            valid_pixel_counts_max = mask_max.flatten(1).sum(dim=1).detach().cpu().numpy().astype(np.int64).tolist()

        if max_threshold_min <= 0:
            valid_pixel_counts_min = [0] * len(gray_frames)
        else:
            mask_min = (diff_min_all >= min_threshold_min) & (diff_min_all <= max_threshold_min)
            valid_pixel_counts_min = mask_min.flatten(1).sum(dim=1).detach().cpu().numpy().astype(np.int64).tolist()

        valid_pixel_count_mean_min = np.mean([v for v in valid_pixel_counts_min if v not in [0, '']]) if len(valid_pixel_counts_min) else 0
        valid_pixel_count_mean_round_min = round(valid_pixel_count_mean_min) if not np.isnan(valid_pixel_count_mean_min) else 0

        valid_pixel_count_mean_max = np.mean([v for v in valid_pixel_counts_max if v not in [0, '']]) if len(valid_pixel_counts_max) else 0
        valid_pixel_count_mean_round_max = round(valid_pixel_count_mean_max) if not np.isnan(valid_pixel_count_mean_max) else 0

        count_zeros_min = valid_pixel_counts_min.count(0)
        count_zeros_max = valid_pixel_counts_max.count(0)

        sum_count_all_pixel_in_range_min = int(np.sum(np.array(valid_pixel_counts_min) < valid_pixel_count_mean_round_min)) if valid_pixel_count_mean_round_min > 0 else 0
        sum_count_all_pixel_in_range_max = int(np.sum(np.array(valid_pixel_counts_max) < valid_pixel_count_mean_round_max)) if valid_pixel_count_mean_round_max > 0 else 0

        valid_pixel_all_counts_min.append(sum_count_all_pixel_in_range_min)
        valid_pixel_all_counts_max.append(sum_count_all_pixel_in_range_max)

        valid_pixel_all_counts_min_max = max(valid_pixel_all_counts_min) if len(valid_pixel_all_counts_min) else 0
        valid_pixel_all_counts_max_max = max(valid_pixel_all_counts_max) if len(valid_pixel_all_counts_max) else 0

        All_pixel_intensities = []
        
        selected_max_differences = []
        selected_max_differences_location = []
        
        max_image_path_pixel = top30_max_image_path_3   
        min_image_path_pixel = top30_max_image_path_2  

        for i in range(len(gray_frames)):

            if use_gpu:
                image1 = gray_tensor[i].cpu().numpy().astype(np.uint8)
                max_image_gray_pixel = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
                min_image_gray_pixel = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
                first_image_gray_pixel = gray_tensor[int(max_difference_index_within_top30)].cpu().numpy().astype(np.uint8)

            else:
                image1 = gray_frames[i].astype(np.uint8)
                max_image_gray_pixel = gray_frames[int(image_frame_n2)].astype(np.uint8)
                min_image_gray_pixel = gray_frames[int(image_frame_n1)].astype(np.uint8)
                first_image_gray_pixel = gray_frames[int(max_difference_index_within_top30)].astype(np.uint8)

            cropped_max_image_pixel = max_image_gray_pixel[y11:y22, x11:x22]
            cropped_min_image_pixel = min_image_gray_pixel[y11:y22, x11:x22]
            cropped_image0_pixel = image0[y11:y22, x11:x22]
            cropped_image1_pixel = image1[y11:y22, x11:x22]

            h = min(cropped_max_image_pixel.shape[0], cropped_image1_pixel.shape[0])
            w = min(cropped_max_image_pixel.shape[1], cropped_image1_pixel.shape[1])

            cropped_max_image_pixel = cropped_max_image_pixel[:h, :w]
            cropped_min_image_pixel = cropped_min_image_pixel[:h, :w]
            cropped_image0_pixel = cropped_image0_pixel[:h, :w]
            cropped_image1_pixel = cropped_image1_pixel[:h, :w]

            cropped_max_image_pixel = cropped_max_image_pixel.astype(np.uint8)
            cropped_min_image_pixel = cropped_min_image_pixel.astype(np.uint8)
            cropped_image0_pixel = cropped_image0_pixel.astype(np.uint8)
            cropped_image1_pixel = cropped_image1_pixel.astype(np.uint8)

            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f"Error cropping image {i}. Skipping...")
                continue

            if valid_pixel_all_counts_min_max >= valid_pixel_all_counts_max_max:
                not_crop_diff = cv2.subtract(image1, min_image_gray_pixel)
                square_image_path = min_image_path_pixel
                square_image_index = image_frame_n1

            else:
                diff = cv2.subtract(cropped_max_image_pixel, cropped_image1_pixel)
                not_crop_diff = cv2.subtract(image1, max_image_gray_pixel)
                square_image_path = max_image_path_pixel
                square_image_index = image_frame_n2

            _, selected_max_diff, _, selected_max_diff_loc = cv2.minMaxLoc(not_crop_diff)

            selected_max_differences.append(selected_max_diff)
            selected_max_differences_location.append(selected_max_diff_loc)

            
            selected_max_differences_indices = np.arange(len(selected_max_differences))[:int(fps*NUM_FRAMES)] 
            selected_max_differences_values = [selected_max_differences[i] for i in selected_max_differences_indices] 

            selected_max_differences_values_max_in_Frames = np.argmax(selected_max_differences_values) 
            selected_max_differences_index_in_Frames = selected_max_differences_indices[selected_max_differences_values_max_in_Frames] 
            selected_max_differences_value_in_Frames = selected_max_differences_values[selected_max_differences_index_in_Frames] 
            selected_max_differences_location_in_Frames = selected_max_differences_location[selected_max_differences_index_in_Frames]

        x_selected = selected_max_differences_location_in_Frames[0]  
        y_selected = selected_max_differences_location_in_Frames[1] 
        
        x1_small = x_selected - ANALYSIS_SQUARE_SIZE_REFERENCE  
        y1_small = y_selected - ANALYSIS_SQUARE_SIZE_REFERENCE
        x2_small = x_selected + ANALYSIS_SQUARE_SIZE_REFERENCE
        y2_small = y_selected + ANALYSIS_SQUARE_SIZE_REFERENCE
        
        x1_small = max(x1_small, 0)
        y1_small = max(y1_small, 0)
        x2_small = min(x2_small, resolution_x -1)
        y2_small = min(y2_small, resolution_y -1)
        
        # Compare all frames again using the cropped ROI-derived maximum frame as a reference.
        # This step is used to identify candidate maximal contraction or relaxation frames.
        # KR: cropped ROI에서 얻은 maximum frame을 기준으로 전체 frame을 다시 비교합니다.
        # KR: 이 과정은 최대 수축 또는 최대 이완 후보 frame을 찾기 위한 단계입니다.

        image_frame_number = []

        next_image_index = start_number
        for n in range(start_number, last_number):
            max_diff_cropped_n_frame_list = []
            max_diff_cropped_loc_n_frame_list = []
        
            for i in range(len(gray_frames)):

                n_frame_path = os.path.join(folder_path, f'{FRAME_PREFIX}{next_image_index + start_number :04d}.{FRAME_EXTENSION}') 

                # Current reference frame path used in this iterative comparison
                # KR: 반복 비교에 사용할 현재 reference frame 경로입니다.
                if use_gpu:
                    top30_max_image_gray_new = gray_tensor[next_image_index].cpu().numpy().astype(np.uint8)
                else:
                    top30_max_image_gray_new = gray_frames[next_image_index].astype(np.uint8)
                cropped_n_frame_gray = top30_max_image_gray_new[y1_small:y2_small, x1_small:x2_small] 
                current_image_crop = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small] 

                max_diff_loc_n_frame_cropped = cv2.absdiff(cropped_n_frame_gray, current_image_crop)
                _, max_diff_crop_n_frame,_,max_diff_crop_loc_n_frame = cv2.minMaxLoc(max_diff_loc_n_frame_cropped)

                max_diff_cropped_n_frame_list.append(max_diff_crop_n_frame)
                max_diff_cropped_loc_n_frame_list.append(max_diff_crop_loc_n_frame)

                max_diff_cropped_n_indices = np.argsort(max_diff_cropped_n_frame_list)[-1:]            
                max_diff_cropped_n_indices_index = max_diff_cropped_n_indices[0]
                max_diff_cropped_n_indices_path = os.path.join(folder_path, f'{FRAME_PREFIX}{max_diff_cropped_n_indices_index + start_number :04d}.{FRAME_EXTENSION}')
            image_frame_number.append(next_image_index)
            
            # Calculate summary statistics from cropped ROI differences.
            # Zero values are excluded when estimating the nonzero minimum.
            # KR: cropped ROI difference 값에서 요약 통계를 계산합니다.
            # KR: 0이 아닌 최소값을 계산할 때는 zero value를 제외합니다.
            max_diff_cropped_n_except0 = [value for value in max_diff_cropped_n_frame_list if value != 0]
            mean_max_diff_cropped_n_except0 = np.mean(max_diff_cropped_n_frame_list)
            mean_round_max_diff_cropped_n_frame_listew = round(mean_max_diff_cropped_n_except0, 2)
            max_max_diff_cropped_n_frame_list = max(max_diff_cropped_n_frame_list)
            min_max_diff_cropped_n_except0 = min(max_diff_cropped_n_except0)
            min_max_diff_cropped_n_except0_index = max_diff_cropped_n_frame_list.index(max_max_diff_cropped_n_frame_list)
            # Update the reference frame index for the next iteration
            # KR: 다음 반복 비교에 사용할 reference frame index를 업데이트합니다.
            next_image_index = min_max_diff_cropped_n_except0_index
            
            # Stop the iterative search when the selected frame index starts to repeat.
            # KR: 선택된 frame index가 반복되면 탐색이 수렴한 것으로 보고 반복을 중단합니다.
            if len(image_frame_number)>=8 and image_frame_number[n] == image_frame_number[n - 2]:
                    image_frame_n1 = image_frame_number[n]
                    image_frame_n2 = image_frame_number[n-1]
                    n1_max_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}{image_frame_n1 + start_number :04d}.{FRAME_EXTENSION}') 
                    n2_max_image_path = os.path.join(folder_path, f'{FRAME_PREFIX}{image_frame_n2 + start_number :04d}.{FRAME_EXTENSION}') 
                    break

        # Compare all frames again using the two ROI-derived candidate frames.
        # This step refines the candidate maximal contraction and relaxation frames.
        # KR: ROI에서 얻은 두 후보 frame을 기준으로 전체 frame을 다시 비교합니다.
        # KR: 최대 수축/최대 이완 후보 frame을 더 안정적으로 찾기 위한 단계입니다.
        
        for i in range(len(gray_frames)):
            
            top30_max_image_path_2_new = n1_max_image_path
            # Use the n1 candidate frame as the reference for ROI-based comparison.
            # KR: n1 후보 frame을 기준으로 ROI 기반 frame 비교를 수행합니다.
            
            if use_gpu:
                top30_max_image_gray_2_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            else:
                top30_max_image_gray_2_new = gray_frames[image_frame_n1].astype(np.uint8)

            
            cropped_top30_max_image_gray_2_new = top30_max_image_gray_2_new[y1_small:y2_small, x1_small:x2_small] 
            current_image_crop_2_new = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small]
            diff_top30_max_2_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_2_new, current_image_crop_2_new)
            _, max_diff_crop_top30_max_2_new, _,max_diff_loc_crop_top30_max_2_new = cv2.minMaxLoc(diff_top30_max_2_cropped_new)
            max_top30_differences_crop_2_new.append(max_diff_crop_top30_max_2_new)
            max_top30_diff_locations_crop_2_new.append(max_diff_loc_crop_top30_max_2_new)
            
            max_top30_differences_crop_indices_2_new = np.argsort(max_top30_differences_crop_2_new)[-1:]
            
            # Search within the initial fps × NUM_FRAMES window to reduce the influence of large noise or bubble movement.
            # KR: 큰 noise 또는 bubble 움직임의 영향을 줄이기 위해 초기 fps × NUM_FRAMES 구간 안에서 탐색합니다.
            max_top30_differences_crop_indices_2_new_num_frames = np.arange(len(max_top30_differences_crop_2_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_2_new = [max_top30_differences_crop_2_new[i] for i in max_top30_differences_crop_indices_2_new_num_frames]

            
            max_top30_differences_crop_values_2_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_2_new) 
            
            index_in_Frames_2 = max_top30_differences_crop_indices_2_new_num_frames[max_top30_differences_crop_values_2_new_max_in_Frames]
            value_in_Frames_2 =  max_top30_differences_crop_values_2_new[index_in_Frames_2]
            location_in_Frames_2 = max_top30_diff_locations_crop_2_new[index_in_Frames_2]

            
            max_top30_differences_crop_indices_2_index_new = max_top30_differences_crop_indices_2_new[0]
            max_top30_differences_crop_indices_2_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_2_index_new + start_number :04d}.{FRAME_EXTENSION}')


        # Calculate summary statistics for ROI differences relative to the n1 candidate frame.
        # KR: n1 후보 frame 기준 ROI difference의 요약 통계를 계산합니다.
        max_top30_differences_crop_except0_2_new = [value for value in max_top30_differences_crop_2_new if value != 0]
        mean_max_top30_differences_crop2_new = np.mean(max_top30_differences_crop_2_new)
        mean_round_max_top30_differences_crop2_new = round(mean_max_top30_differences_crop2_new, 2)
        max_max_top30_differences_crop2_new = max(max_top30_differences_crop_2_new)
        min_max_top30_differences_crop_except0_2_new = min(max_top30_differences_crop_except0_2_new)
        min_max_top30_differences_crop_except0_2_index_new = max_top30_differences_crop_2_new.index(max_max_top30_differences_crop2_new)

        
        for i in range(len(gray_frames)):

            top30_max_image_path_3_new = n2_max_image_path

            # Use the n2 candidate frame as the reference for ROI-based comparison.
            # KR: n2 후보 frame을 기준으로 ROI 기반 frame 비교를 수행합니다.
            if use_gpu:
                top30_max_image_gray_3_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
            else:
                top30_max_image_gray_3_new = gray_frames[image_frame_n2].astype(np.uint8)

            
            cropped_top30_max_image_gray_3_new = top30_max_image_gray_3_new[y1_small:y2_small, x1_small:x2_small] 
            
            current_image_crop_3 = current_gray_image_list[i][y1_small:y2_small, x1_small:x2_small] 

            diff_top30_max_3_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_3_new, current_image_crop_3)
            _, max_diff_crop_top30_max_3_new, _,max_diff_loc_crop_top30_max_3_new = cv2.minMaxLoc(diff_top30_max_3_cropped_new)
            
            max_top30_differences_crop_3_new.append(max_diff_crop_top30_max_3_new)
            max_top30_diff_locations_crop_3_new.append(max_diff_loc_crop_top30_max_3_new)
            
            max_top30_differences_crop_indices_3_new = np.argsort(max_top30_differences_crop_3_new)[-1:]
            
            # Search within the initial fps × NUM_FRAMES window to reduce the influence of large noise or bubble movement.
            # KR: 큰 noise 또는 bubble 움직임의 영향을 줄이기 위해 초기 fps × NUM_FRAMES 구간 안에서 탐색합니다.
            max_top30_differences_crop_indices_3_new_num_frames = np.arange(len(max_top30_differences_crop_3_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_3_new = [max_top30_differences_crop_3_new[i] for i in max_top30_differences_crop_indices_3_new_num_frames] 

            
            max_top30_differences_crop_values_3_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_3_new) 
            index_in_Frames_3 = max_top30_differences_crop_indices_3_new_num_frames[max_top30_differences_crop_values_3_new_max_in_Frames]
            value_in_Frames_3 =  max_top30_differences_crop_values_3_new[index_in_Frames_3]
            location_in_Frames_3 = max_top30_diff_locations_crop_3_new[index_in_Frames_3]

            
            max_top30_differences_crop_indices_3_index_new = max_top30_differences_crop_indices_3_new[0]
            max_top30_differences_crop_indices_3_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_3_index_new + start_number :04d}.{FRAME_EXTENSION}')
            

        max_top30_differences_crop_except0_3_new = [value for value in max_top30_differences_crop_3_new if value != 0]
        mean_max_top30_differences_crop3_new = np.mean(max_top30_differences_crop_3_new)
        mean_round_max_top30_differences_crop3_new = round(mean_max_top30_differences_crop3_new, 2)
        max_max_top30_differences_crop3_new = max(max_top30_differences_crop_3_new)
        min_max_top30_differences_crop_except0_3_new = min(max_top30_differences_crop_except0_3_new)
        min_max_top30_differences_crop_except0_3_index_new = max_top30_differences_crop_3_new.index(max_max_top30_differences_crop3_new)
        
        top30_max_image_path_4_new = os.path.join(folder_path, f'{FRAME_PREFIX}{min_max_top30_differences_crop_except0_3_index_new + start_number :04d}.{FRAME_EXTENSION}') 
        



        x_differences = [] 
        y_differences = []
        valid_pixel_counts_min_new = []
        valid_pixel_counts_max_new = []
        valid_pixel_all_counts_max_new = []
        valid_pixel_all_counts_min_new = []

        # Calculate summary statistics for ROI differences relative to the n2 candidate frame.
        # KR: n2 후보 frame 기준 ROI difference의 요약 통계를 계산합니다.
        for i in range(len(gray_frames)):
            image1 = current_gray_image_list[i]
            # Crop local ROIs around the selected motion center.
            # KR: 선택된 motion center 주변의 local ROI를 crop합니다.
            max_image_path_pixel_new = top30_max_image_path_3_new  # Candidate maximal contraction or relaxation frame
            min_image_path_pixel_new = top30_max_image_path_2_new  # Candidate maximal relaxation or contraction frame
            
            if use_gpu:
                max_image_gray_pixel_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
                min_image_gray_pixel_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            else:
                max_image_gray_pixel_new = gray_frames[image_frame_n2].astype(np.uint8)
                min_image_gray_pixel_new = gray_frames[image_frame_n1].astype(np.uint8)

                    
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small] 
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
            cropped_image0_pixel_new = image0[y1_small:y2_small, x1_small:x2_small] 
            cropped_image1_pixel_new = image1[y1_small:y2_small, x1_small:x2_small] 
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f'Error cropping image{i}. Skipping...')
                continue  # Skip this frame if ROI cropping fails
            
            diff_max_new = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
            diff_min_new = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 

            threshold_max_new = min_max_top30_differences_crop_except0_3_new  
            threshold_min_new = min_max_top30_differences_crop_except0_2_new 
            white_pixel_coords_max_new = np.argwhere(diff_max_new >= threshold_max_new)
            white_pixel_coords_min_new = np.argwhere(diff_min_new >= threshold_min_new)  
            
            # Count pixels whose intensity differences fall within the threshold range.
            # KR: intensity difference가 설정된 threshold 범위 안에 들어오는 pixel을 계산합니다.
            
            # Add a fixed offset to reduce small background or medium-related noise.
            # KR: 작은 background 또는 배지 관련 noise를 줄이기 위해 고정 offset을 더합니다.
            min_threshold_max_new = min_max_top30_differences_crop_except0_3_new + 20  # Fixed offset for small background or medium-related noise
            max_threshold_max_new = max_max_top30_differences_crop3_new  # Maximum ROI intensity difference relative to the candidate frame
            min_threshold_min_new = min_max_top30_differences_crop_except0_2_new  +20 # Fixed offset for small background or medium-related noise
            max_threshold_min_new = max_max_top30_differences_crop2_new # Maximum ROI intensity difference relative to the candidate frame
            
            valid_pixel_coords_max_new = np.argwhere((diff_max_new >= min_threshold_max_new) & (diff_max_new <= max_threshold_max_new))
            valid_pixel_coords_min_new = np.argwhere((diff_min_new >= min_threshold_min_new) & (diff_min_new <= max_threshold_min_new))
            
            valid_pixel_coords_max_all_new = np.argwhere((diff_max_new >= 0) & (diff_max_new <= max_threshold_max))
            valid_pixel_coords_min_all_new = np.argwhere((diff_min_new >= 0) & (diff_min_new <= max_threshold_min))

            valid_pixel_count_max_new = len(valid_pixel_coords_max_new)
            valid_pixel_count_min_new = len(valid_pixel_coords_min_new)
            valid_pixel_counts_max_new.append(valid_pixel_count_max_new)
            valid_pixel_counts_min_new.append(valid_pixel_count_min_new)       

        valid_pixel_count_mean_min_new = np.mean([value for value in valid_pixel_counts_min_new if value not in [0,'']]) #0 과 데이터가 

        valid_pixel_count_mean_round_min_new = round(valid_pixel_count_mean_min_new) if not np.isnan(valid_pixel_count_mean_min_new) else 0 
        valid_pixel_count_mean_max_new = np.mean([value for value in valid_pixel_counts_max_new if value not in [0,'']]) #0 과 데이터가 

        valid_pixel_count_mean_round_max_new = round(valid_pixel_count_mean_max_new) if not np.isnan(valid_pixel_count_mean_max_new) else 0 
        
        count_zeros_min_new = valid_pixel_counts_min_new.count(0)
        count_zeros_max_new = valid_pixel_counts_max_new.count(0)

        count_all_pixel_in_range_min_new = [valid_pixel_counts_min_new.count(i) for i in range(valid_pixel_count_mean_round_min_new)]
        count_all_pixel_in_range_max_new = [valid_pixel_counts_max_new.count(i) for i in range(valid_pixel_count_mean_round_max_new)]
        
        sum_count_all_pixel_in_range_min_new = sum(count_all_pixel_in_range_min_new, 0)
        sum_count_all_pixel_in_range_max_new = sum(count_all_pixel_in_range_max_new, 0)

        valid_pixel_all_counts_min_new.append(sum_count_all_pixel_in_range_min_new)
        valid_pixel_all_counts_max_new.append(sum_count_all_pixel_in_range_max_new)
        
        valid_pixel_all_counts_min_max_new = max(valid_pixel_all_counts_min_new)
        valid_pixel_all_counts_max_max_new = max(valid_pixel_all_counts_max_new)
        
        
        All_pixel_intensities = []  
        valid_pixel_counts = []
        valid_pixel_counts2 = []
        selected_max_differences = []
        selected_max_differences_location = []
        
        x_selected_new, y_selected_new = None, None
        click_done = False  
        



        for i in range(len(gray_frames)):

            max_image_path_pixel_new = top30_max_image_path_3_new
            min_image_path_pixel_new = top30_max_image_path_2_new
            
            if use_gpu:
                max_image_gray_pixel_new = gray_tensor[image_frame_n2].cpu().numpy().astype(np.uint8)
                min_image_gray_pixel_new = gray_tensor[image_frame_n1].cpu().numpy().astype(np.uint8)
            else:
                max_image_gray_pixel_new = gray_frames[image_frame_n2].astype(np.uint8)
                min_image_gray_pixel_new = gray_frames[image_frame_n1].astype(np.uint8)

            
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small:y2_small, x1_small:x2_small]
            cropped_image0_pixel_new = image0[y1_small:y2_small, x1_small:x2_small]  
            cropped_image1_pixel_new = image1[y1_small:y2_small, x1_small:x2_small]  
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f"Error cropping image {i}. Skipping...")
                continue 
            
            # Select the candidate frame based on the number of valid pixels below the mean threshold.
            # This helps determine which candidate better represents maximal contraction or relaxation.
            # KR: mean threshold보다 작은 valid pixel count를 기준으로 후보 frame을 선택합니다.
            # KR: 최대 수축 또는 최대 이완을 더 잘 나타내는 후보를 고르기 위한 과정입니다.
            
            if valid_pixel_all_counts_min_max_new >= valid_pixel_all_counts_max_max_new:
                loc2 = max_top30_diff_locations_crop_2_new  
                loc3 = location_in_Frames_2  
                
            else:
                loc2 = max_top30_diff_locations_crop_3_new 
                loc3 = location_in_Frames_3 
                
            # =========================================================
            # Manual coordinate selection callback
            # =========================================================
            
            # Mouse-click callback for selecting an analysis coordinate
            # KR: 분석 좌표를 수동으로 선택하기 위한 마우스 클릭 callback 함수입니다.
            def select_point(event, x, y, flags, param):
                global x_selected_new, y_selected_new, click_done  

                if event == cv2.EVENT_LBUTTONDOWN and not click_done: # Left-click once to select a coordinate
                    x_selected_new = x
                    y_selected_new = y
                    click_done = True  # 
                    print(f"Selected coordinate: ({x_selected_new}, {y_selected_new})")

                    cv2.circle( min_image_gray_pixel_new, (x_selected_new, y_selected_new), 2, (0, 0, 255), -1)
                    cv2.imshow("Select Point",  min_image_gray_pixel_new)

                    cv2.waitKey(500) 
                    cv2.destroyAllWindows()

            # Use predefined coordinates when manual coordinate mode is enabled
            # KR: manual coordinate mode가 켜져 있으면 사용자가 미리 지정한 좌표를 사용합니다.
            if MANUAL_SELECTION.lower() == 'on' and not click_done:
                print("Please click the desired analysis point.")
                

                cv2.namedWindow("Select Point", cv2.WINDOW_NORMAL)
                cv2.resizeWindow("Select Point", resolution_x, resolution_y)  

                cv2.imshow("Select Point", min_image_gray_pixel_new)
                cv2.setMouseCallback("Select Point", select_point)


                while not click_done:
                    cv2.waitKey(1)  # 

                print(f"Using manually selected coordinate: ({x_selected_new}, {y_selected_new})")


            # Use automatically detected coordinates when manual selection is off,
            # or keep the previously selected coordinate after the first manual click.
            # KR: manual selection이 꺼져 있으면 자동 탐색 좌표를 사용합니다.
            # KR: 첫 번째 수동 클릭 이후에는 선택된 좌표를 유지합니다.
            
            elif MANUAL_SELECTION.lower() == 'off' or (x_selected_new is None and y_selected_new is None): 
                x_selected_new = x1_small + loc3[0]
                y_selected_new = y1_small + loc3[1]
        
            
            x1_small_new = x_selected_new -  resolution_x
            y1_small_new = y_selected_new -  resolution_y
            x2_small_new = x_selected_new +  resolution_x
            y2_small_new = y_selected_new + resolution_y
            
            x1_small_new = max(x1_small_new, 0)
            y1_small_new = max(y1_small_new, 0)
            x2_small_new = min(x2_small_new, resolution_x -1)
            y2_small_new = min(y2_small_new, resolution_y -1)
            
            x1_small_new_reference = max(0, x_selected_new - ANALYSIS_SQUARE_SIZE_REFERENCE)
            y1_small_new_reference = max(0, y_selected_new -  ANALYSIS_SQUARE_SIZE_REFERENCE)
            x2_small_new_reference = x_selected_new + ANALYSIS_SQUARE_SIZE_REFERENCE
            y2_small_new_reference = y_selected_new +  ANALYSIS_SQUARE_SIZE_REFERENCE
            
            x1_small_new_reference = max(x1_small_new_reference, 0)
            y1_small_new_reference = max(y1_small_new_reference, 0)
            x2_small_new_reference = min(x2_small_new_reference, resolution_x -1)
            y2_small_new_reference = min(y2_small_new_reference, resolution_y -1)
            
            x1_small_new_optical = max(0, x_selected_new - resolution_x)
            y1_small_new_optical = max(0, y_selected_new - resolution_y)
            x2_small_new_optical = x_selected_new + resolution_x
            y2_small_new_optical = y_selected_new + resolution_y
            
            x1_small_new_optical = max(x1_small_new_optical, 0)
            y1_small_new_optical = max(y1_small_new_optical, 0)
            x2_small_new_optical = min(x2_small_new_optical, resolution_x -1)
            y2_small_new_optical = min(y2_small_new_optical, resolution_y -1)

        max_top30_differences_crop_2_new = []
        max_top30_diff_locations_crop_2_new = []

        max_top30_differences_crop_3_new = []
        max_top30_diff_locations_crop_3_new = []

        if len(current_rgb_image_list) == 0:
            print("❌ current_rgb_image_list empty → skip")
            continue

        image_frame_n1 = max(0, min(image_frame_n1, len(current_rgb_image_list)-1))
        image_frame_n2 = max(0, min(image_frame_n2, len(current_rgb_image_list)-1))

        top30_max_image_2_new = current_rgb_image_list[image_frame_n1]
        top30_max_image_3_new = current_rgb_image_list[image_frame_n2]


        for i in range(len(gray_frames)):
            
            top30_max_image_path_2_new = n1_max_image_path
            top30_max_image_gray_2_new = cv2.cvtColor(top30_max_image_2_new, cv2.COLOR_BGR2GRAY) 
            
            cropped_top30_max_image_gray_2_new = top30_max_image_gray_2_new[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
            
            current_image_crop_2_new = current_gray_image_list[i][y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 

            diff_top30_max_2_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_2_new, current_image_crop_2_new)
            _, max_diff_crop_top30_max_2_new, _,max_diff_loc_crop_top30_max_2_new = cv2.minMaxLoc(diff_top30_max_2_cropped_new)
            
            max_top30_differences_crop_2_new.append(max_diff_crop_top30_max_2_new)
            max_top30_diff_locations_crop_2_new.append(max_diff_loc_crop_top30_max_2_new)
            
            max_top30_differences_crop_indices_2_new = np.argsort(max_top30_differences_crop_2_new)[-1:]
            
            max_top30_differences_crop_indices_2_new_num_frames = np.arange(len(max_top30_differences_crop_2_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_2_new = [max_top30_differences_crop_2_new[i] for i in max_top30_differences_crop_indices_2_new_num_frames] 
            
            max_top30_differences_crop_values_2_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_2_new)
            
            index_in_Frames_2 = max_top30_differences_crop_indices_2_new_num_frames[max_top30_differences_crop_values_2_new_max_in_Frames]
            
            value_in_Frames_2 =  max_top30_differences_crop_values_2_new[index_in_Frames_2]
            location_in_Frames_2 = max_top30_diff_locations_crop_2_new[index_in_Frames_2]
            
            max_top30_differences_crop_indices_2_index_new = max_top30_differences_crop_indices_2_new[0]
            max_top30_differences_crop_indices_2_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_2_index_new + start_number :04d}.{FRAME_EXTENSION}')



        max_top30_differences_crop_except0_2_new = [value for value in max_top30_differences_crop_2_new if value != 0]
        mean_max_top30_differences_crop2_new = np.mean(max_top30_differences_crop_2_new)
        mean_round_max_top30_differences_crop2_new = round(mean_max_top30_differences_crop2_new, 2)
        max_max_top30_differences_crop2_new = max(max_top30_differences_crop_2_new)
        min_max_top30_differences_crop_except0_2_new = min(max_top30_differences_crop_except0_2_new)
        min_max_top30_differences_crop_except0_2_index_new = max_top30_differences_crop_2_new.index(max_max_top30_differences_crop2_new)

        

        for i in range(len(gray_frames)):

            top30_max_image_path_3_new = n2_max_image_path
            top30_max_image_gray_3_new = cv2.cvtColor(top30_max_image_3_new, cv2.COLOR_BGR2GRAY) 
            
            cropped_top30_max_image_gray_3_new = top30_max_image_gray_3_new[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
            
            current_image_crop_3 = current_gray_image_list[i][y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 

            diff_top30_max_3_cropped_new = cv2.absdiff(cropped_top30_max_image_gray_3_new, current_image_crop_3)
            _, max_diff_crop_top30_max_3_new, _,max_diff_loc_crop_top30_max_3_new = cv2.minMaxLoc(diff_top30_max_3_cropped_new)

            max_top30_differences_crop_3_new.append(max_diff_crop_top30_max_3_new)
            max_top30_diff_locations_crop_3_new.append(max_diff_loc_crop_top30_max_3_new)
            
            max_top30_differences_crop_indices_3_new = np.argsort(max_top30_differences_crop_3_new)[-1:]
            max_top30_differences_crop_indices_3_new_num_frames = np.arange(len(max_top30_differences_crop_3_new))[:int(fps*NUM_FRAMES)] 
            
            max_top30_differences_crop_values_3_new = [max_top30_differences_crop_3_new[i] for i in max_top30_differences_crop_indices_3_new_num_frames] 
        
            max_top30_differences_crop_values_3_new_max_in_Frames = np.argmax(max_top30_differences_crop_values_3_new) 
            index_in_Frames_3 = max_top30_differences_crop_indices_3_new_num_frames[max_top30_differences_crop_values_3_new_max_in_Frames]
            
            value_in_Frames_3 =  max_top30_differences_crop_values_3_new[index_in_Frames_3]
            location_in_Frames_3 = max_top30_diff_locations_crop_3_new[index_in_Frames_3]
            
            max_top30_differences_crop_indices_3_index_new = max_top30_differences_crop_indices_3_new[0]
            max_top30_differences_crop_indices_3_index_path_new = os.path.join(folder_path, f'{FRAME_PREFIX}{max_top30_differences_crop_indices_3_index_new + start_number :04d}.{FRAME_EXTENSION}')
            

        max_top30_differences_crop_except0_3_new = [value for value in max_top30_differences_crop_3_new if value != 0]
        mean_max_top30_differences_crop3_new = np.mean(max_top30_differences_crop_3_new)
        mean_round_max_top30_differences_crop3_new = round(mean_max_top30_differences_crop3_new, 2)
        max_max_top30_differences_crop3_new = max(max_top30_differences_crop_3_new)
        min_max_top30_differences_crop_except0_3_new = min(max_top30_differences_crop_except0_3_new)
        min_max_top30_differences_crop_except0_3_index_new = max_top30_differences_crop_3_new.index(max_max_top30_differences_crop3_new)
        
        top30_max_image_path_4_new = os.path.join(folder_path, f'{FRAME_PREFIX}{min_max_top30_differences_crop_except0_3_index_new + start_number :04d}.{FRAME_EXTENSION}') 
        
        x_differences = []  
        y_differences = []
        valid_pixel_counts_min_new = []
        valid_pixel_counts_max_new = []
        valid_pixel_all_counts_max_new = []
        valid_pixel_all_counts_min_new = []

        for i in range(len(gray_frames)):

            if use_gpu:
                image1 = gray_tensor[i].cpu().numpy().astype(np.uint8)
            else:
                image1 = gray_frames[i].astype(np.uint8)


            max_image_path_pixel_new = top30_max_image_path_3_new 
            min_image_path_pixel_new = top30_max_image_path_2_new 
            
            max_image_pixel_new = current_rgb_image_list[image_frame_n2]
            min_image_pixel_new = current_rgb_image_list[image_frame_n1]
            max_image_gray_pixel_new = cv2.cvtColor(max_image_pixel_new, cv2.COLOR_BGR2GRAY)
            min_image_gray_pixel_new = cv2.cvtColor(min_image_pixel_new, cv2.COLOR_BGR2GRAY)
            
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new] #small_size
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_image0_pixel_new = image0[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_image1_pixel_new = image1[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f'Error cropping image{i}. Skipping...')
                continue 
            
            diff_max_new = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
            diff_min_new = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 
            
            # Count pixels with intensity differences above the nonzero minimum threshold.
            # KR: 0을 제외한 최소 threshold 이상으로 변화한 pixel의 좌표를 계산합니다.
            threshold_max_new = min_max_top30_differences_crop_except0_3_new 
            threshold_min_new = min_max_top30_differences_crop_except0_2_new  
            white_pixel_coords_max_new = np.argwhere(diff_max_new >= threshold_max_new)
            white_pixel_coords_min_new = np.argwhere(diff_min_new >= threshold_min_new)
            
            min_threshold_max_new = min_max_top30_differences_crop_except0_3_new +20 
            max_threshold_max_new = max_max_top30_differences_crop3_new 

            min_threshold_min_new = min_max_top30_differences_crop_except0_2_new +20 
            max_threshold_min_new = max_max_top30_differences_crop2_new 
            
            valid_pixel_coords_max_new = np.argwhere((diff_max_new >= min_threshold_max_new) & (diff_max_new <= max_threshold_max_new))
            valid_pixel_coords_min_new = np.argwhere((diff_min_new >= min_threshold_min_new) & (diff_min_new <= max_threshold_min_new))
            
            valid_pixel_coords_max_all_new = np.argwhere((diff_max_new >= 0) & (diff_max_new <= max_threshold_max))
            valid_pixel_coords_min_all_new = np.argwhere((diff_min_new >= 0) & (diff_min_new <= max_threshold_min))

            valid_pixel_count_max_new = len(valid_pixel_coords_max_new)
            valid_pixel_count_min_new = len(valid_pixel_coords_min_new)
            valid_pixel_counts_max_new.append(valid_pixel_count_max_new)
            valid_pixel_counts_min_new.append(valid_pixel_count_min_new)       


        valid_pixel_count_mean_min_new = np.mean([value for value in valid_pixel_counts_min_new if value not in [0,'']]) 

        valid_pixel_count_mean_round_min_new = round(valid_pixel_count_mean_min_new) if not np.isnan(valid_pixel_count_mean_min_new) else 0 
        
        valid_pixel_count_mean_max_new = np.mean([value for value in valid_pixel_counts_max_new if value not in [0,'']]) 

        valid_pixel_count_mean_round_max_new = round(valid_pixel_count_mean_max_new) if not np.isnan(valid_pixel_count_mean_max_new) else 0 

        count_zeros_min_new = valid_pixel_counts_min_new.count(0)
        count_zeros_max_new = valid_pixel_counts_max_new.count(0)


        count_all_pixel_in_range_min_new = [valid_pixel_counts_min_new.count(i) for i in range(valid_pixel_count_mean_round_min_new)]
        count_all_pixel_in_range_max_new = [valid_pixel_counts_max_new.count(i) for i in range(valid_pixel_count_mean_round_max_new)]
        
        sum_count_all_pixel_in_range_min_new = sum(count_all_pixel_in_range_min_new, 0)
        sum_count_all_pixel_in_range_max_new = sum(count_all_pixel_in_range_max_new, 0)

        valid_pixel_all_counts_min_new.append(sum_count_all_pixel_in_range_min_new)
        valid_pixel_all_counts_max_new.append(sum_count_all_pixel_in_range_max_new)
        
        valid_pixel_all_counts_min_max_new = max(valid_pixel_all_counts_min_new)
        valid_pixel_all_counts_max_max_new = max(valid_pixel_all_counts_max_new)

        All_pixel_intensities = []  
        valid_pixel_counts = []
        not_cropped_valid_pixel_counts = []
        selected_max_differences = []
        selected_max_differences_location = []
        Analysis_frames = []
        pixel_intensities_movement_list = []
        pixel_intensities_movement_coordinates_list = []



        for i in range(len(gray_frames)):

            max_image_path_pixel_new = top30_max_image_path_3_new
            min_image_path_pixel_new = top30_max_image_path_2_new
            
            max_image_pixel_new = current_rgb_image_list[image_frame_n2]
            min_image_pixel_new = current_rgb_image_list[image_frame_n1]
            max_image_gray_pixel_new = cv2.cvtColor(max_image_pixel_new, cv2.COLOR_BGR2GRAY)
            min_image_gray_pixel_new = cv2.cvtColor(min_image_pixel_new, cv2.COLOR_BGR2GRAY)
            
            cropped_max_image_pixel_new = max_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_min_image_pixel_new = min_image_gray_pixel_new[y1_small_new:y2_small_new, x1_small_new:x2_small_new]
            cropped_image0_pixel_new = image0[y1_small_new:y2_small_new, x1_small_new:x2_small_new]  
            cropped_image1_pixel_new = image1[y1_small_new:y2_small_new, x1_small_new:x2_small_new]  
            
            if cropped_image0_pixel is None or cropped_image1_pixel is None:
                print(f"Error cropping image {i}. Skipping...")
                continue 
            
            # Select the reference candidate frame based on valid-pixel count statistics.
            # The selected candidate determines which images, thresholds, coordinates,
            # and intensity-difference values are used in downstream displacement analysis.
            # KR: valid-pixel count 통계를 기준으로 기준 후보 frame을 선택합니다.
            # KR: 선택된 후보 frame에 따라 이후 displacement 분석에 사용할 이미지, threshold, 좌표, intensity-difference 값이 결정됩니다.
            
            # Apply default or reversed candidate-selection logic depending on the expected contraction direction.
            # KR: 예상되는 수축 방향에 따라 기본 후보 선택 기준 또는 반전된 후보 선택 기준을 적용합니다.
            if REVERSE_CONTRACTION.lower() == 'off':
                if valid_pixel_all_counts_min_max_new >= valid_pixel_all_counts_max_max_new:
                    # Use the candidate relaxation frame as the reference image
                    # KR: 후보 이완 frame을 기준 이미지로 사용합니다.
                    diff = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new)  
                    # Full-frame difference used for visualization
                    # KR: visualization을 위해 full-frame difference를 계산합니다.
                    not_crop_diff = cv2.subtract(max_image_gray_pixel_new, image1) 
                    # Difference between candidate contraction and relaxation frames
                    # KR: 후보 수축 frame과 후보 이완 frame 사이의 차이입니다.
                    crop_diff = cv2.subtract(cropped_max_image_pixel_new, cropped_min_image_pixel_new) 
                    #=================================================================
                    # Store paired candidate contraction/relaxation frame information
                    # KR: 후보 수축/이완 frame 쌍의 정보를 저장합니다.
                    #=================================================================
                    # Reference image: candidate relaxation frame
                    # KR: 기준 이미지: 후보 이완 frame
                    square_image = min_image_pixel_new
                    # Path to the reference relaxation frame
                    # KR: 기준 이완 frame의 경로
                    square_image_path = min_image_path_pixel_new
                    # Paired candidate contraction frame
                    # KR: 기준 frame과 짝이 되는 후보 수축 frame
                    square_image_path_contraction = max_image_path_pixel_new 
                    # Loaded RGB image of the reference relaxation frame
                    # KR: 기준 이완 frame의 RGB 이미지
                    square_image_imread = current_rgb_image_list[image_frame_n1]
                    # Loaded RGB image of the paired contraction frame
                    # KR: 짝이 되는 후보 수축 frame의 RGB 이미지
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n2]
                    # Index of the selected reference frame
                    # KR: 선택된 기준 frame의 index
                    square_image_index = image_frame_n1
                    # Image showing the maximum difference from the reference frame
                    # KR: 기준 frame과 가장 큰 차이를 보이는 이미지
                    square_differences_crop_indices_index_path_image = max_image_pixel_new 
                    # Minimum intensity-difference threshold
                    # KR: 최소 intensity-difference threshold
                    threshold = min_max_top30_differences_crop_except0_2_new 
                    # Maximum intensity-difference threshold
                    # KR: 최대 intensity-difference threshold
                    max_threshold = max_max_top30_differences_crop2_new
                    # Maximum intensity difference
                    # KR: intensity difference 값
                    diff_val = max_diff_crop_top30_max_2_new   
                    # List of maximum intensity differences
                    # KR: 최대 intensity difference 리스트
                    diff2_val = max_top30_differences_crop_2_new
                    # Coordinates of the maximum intensity difference
                    # KR: 최대 intensity difference 좌표
                    loc = max_diff_loc_crop_top30_max_2_new
                    # Coordinate of the maximum intensity difference within the initial analysis window
                    # KR: 초기 분석 구간 안에서 최대 intensity difference가 나타난 좌표
                    loc2 = max_top30_diff_locations_crop_2_new
                    # Coordinate of the maximum intensity difference within the initial analysis window
                    # KR: 초기 분석 구간 안에서 최대 intensity difference가 나타난 좌표
                    loc3 = location_in_Frames_2 
                    # Maximum intensity difference from the reference frame
                    # KR: 기준 frame과 가장 큰 차이를 보이는 intensity difference 값
                    max_values = max_max_top30_differences_crop2_new 
                    ## Index of the frame showing the maximum difference from the reference frame
                    # KR: 기준 frame과 가장 큰 차이를 보이는 frame의 index
                    max_indices = min_max_top30_differences_crop_except0_2_index_new 
                    # List of maximum intensity differences
                    # KR: 최대 intensity difference 리스트
                    adapt = max_top30_differences_crop_2_new 
                    # Maximum intensity-difference value
                    # KR: 최대 intensity-difference 값
                    adapt_max = max_max_top30_differences_crop2_new 
                    # Mean intensity-difference value
                    # KR: 평균 intensity-difference 값
                    adapt_mean = mean_round_max_top30_differences_crop2_new 
                    
                else:
                    diff = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
                    not_crop_diff = cv2.subtract(min_image_gray_pixel_new, image1) 
                    crop_diff = cv2.subtract(cropped_min_image_pixel_new, cropped_max_image_pixel_new) 
                    square_image = max_image_pixel_new 
                    square_image_path = max_image_path_pixel_new 
                    square_image_path_contraction = max_image_path_pixel_new 
                    square_image_imread = current_rgb_image_list[image_frame_n2] 
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n1] 
                    square_image_index = image_frame_n2 
                    square_differences_crop_indices_index_path_image = min_image_pixel_new 
                    threshold = min_max_top30_differences_crop_except0_3_new  
                    max_threshold = max_max_top30_differences_crop3_new 
                    diff_val = max_diff_crop_top30_max_3_new 
                    diff2_val = max_top30_differences_crop_3_new 
                    loc = max_diff_loc_crop_top30_max_3_new 
                    loc2 = max_top30_diff_locations_crop_3_new 
                    loc3 = location_in_Frames_3 
                    max_values = max_max_top30_differences_crop3_new 
                    max_indices = min_max_top30_differences_crop_except0_3_index_new 
                    adapt = max_top30_differences_crop_3_new 
                    adapt_max = max_max_top30_differences_crop3_new 
                    adapt_mean = mean_round_max_top30_differences_crop3_new 
            else: 
                if valid_pixel_all_counts_min_max_new <= valid_pixel_all_counts_max_max_new:
                    diff = cv2.subtract(cropped_min_image_pixel_new, cropped_image1_pixel_new) 
                    not_crop_diff = cv2.subtract(max_image_gray_pixel_new, image1) 
                    crop_diff = cv2.subtract(cropped_max_image_pixel_new, cropped_min_image_pixel_new) 
                    square_image = min_image_pixel_new 
                    square_image_path = min_image_path_pixel_new 
                    square_image_path_contraction = max_image_path_pixel_new 
                    square_image_imread = current_rgb_image_list[image_frame_n1] 
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n2] 
                    square_image_index = image_frame_n1
                    square_differences_crop_indices_index_path_image = max_image_pixel_new 
                    threshold = min_max_top30_differences_crop_except0_2_new 
                    max_threshold = max_max_top30_differences_crop2_new 
                    diff_val = max_diff_crop_top30_max_2_new  
                    diff2_val = max_top30_differences_crop_2_new 
                    loc = max_diff_loc_crop_top30_max_2_new 
                    loc2 = max_top30_diff_locations_crop_2_new 
                    loc3 = location_in_Frames_2 
                    max_values = max_max_top30_differences_crop2_new 
                    max_indices = min_max_top30_differences_crop_except0_2_index_new 
                    adapt = max_top30_differences_crop_2_new 
                    adapt_max = max_max_top30_differences_crop2_new 
                    adapt_mean = mean_round_max_top30_differences_crop2_new 
                    
                else:
                    diff = cv2.subtract(cropped_max_image_pixel_new, cropped_image1_pixel_new) 
                    not_crop_diff = cv2.subtract(min_image_gray_pixel_new, image1) 
                    crop_diff = cv2.subtract(cropped_min_image_pixel_new, cropped_max_image_pixel_new) 
                    square_image = max_image_pixel_new 
                    square_image_path = max_image_path_pixel_new 
                    square_image_path_contraction = max_image_path_pixel_new 
                    square_image_imread = current_rgb_image_list[image_frame_n2] 
                    square_image_imread_contraction = current_rgb_image_list[image_frame_n1] 
                    square_image_index = image_frame_n2 
                    square_differences_crop_indices_index_path_image = min_image_pixel_new 
                    threshold = min_max_top30_differences_crop_except0_3_new  
                    max_threshold = max_max_top30_differences_crop3_new 
                    diff_val = max_diff_crop_top30_max_3_new 
                    diff2_val = max_top30_differences_crop_3_new 
                    loc = max_diff_loc_crop_top30_max_3_new 
                    loc2 = max_top30_diff_locations_crop_3_new  
                    loc3 = location_in_Frames_3 
                    max_values = max_max_top30_differences_crop3_new 
                    max_indices = min_max_top30_differences_crop_except0_3_index_new 
                    adapt = max_top30_differences_crop_3_new 
                    adapt_max = max_max_top30_differences_crop3_new 
                    adapt_mean = mean_round_max_top30_differences_crop3_new 
                

            max_pixel_intensities = int(not_crop_diff.max())
            All_pixel_intensities.append(max_pixel_intensities)
            max_All_pixel_intensities = int(np.max(All_pixel_intensities))
            All_pixel_intensities.append(max_pixel_intensities)
            max_All_pixel_intensities = np.max(All_pixel_intensities)

            selected_image_path = square_image_path  
            
            selected_image = square_image_imread 
            selected_image_contraction = square_image_imread_contraction 
            selected_image_gray = cv2.cvtColor(selected_image, cv2.COLOR_BGR2GRAY) 
            selected_image_contraction_gray = cv2.cvtColor(selected_image_contraction, cv2.COLOR_BGR2GRAY) 
            
            all_image_gray = current_gray_image_list[i] 
            adapt = np.array(adapt) 
            adapt_min = adapt[adapt >= threshold].tolist()
            try:
                sorted_adapt_min = min(adapt_min)
                sorted_adapt_max = max(adapt_min)
                index_of_sorted_adapt_min = int(np.where(adapt == sorted_adapt_min)[0][0])
                index_of_sorted_adapt_max = int(np.where(adapt == sorted_adapt_max)[0][0])
                coordinate_of_sorted_adapt_min = loc2[index_of_sorted_adapt_min]
                coordinate_of_sorted_adapt_max = loc2[index_of_sorted_adapt_max]
            except ValueError:
                break
            
            cropped_selected_image_gray = selected_image_gray[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
            cropped_all_image_gray = all_image_gray[y1_small_new:y2_small_new, x1_small_new:x2_small_new] 
            
            cropped_selected_image_gray_reference = selected_image_gray[y1_small_new_optical:y2_small_new_optical, x1_small_new_optical:x2_small_new_optical] 
            cropped_all_image_gray_reference = all_image_gray[y1_small_new_optical:y2_small_new_optical, x1_small_new_optical:x2_small_new_optical] 
            
            
            diff_selected = cv2.absdiff( cropped_all_image_gray, cropped_selected_image_gray) 
            diff_selected_reference = cv2.subtract( cropped_all_image_gray_reference, cropped_selected_image_gray_reference)

            pixel_intensities_movement = diff_selected.ravel()

            diff_selected_column = np.argwhere(diff_selected >= threshold)

            pixel_intensities_movement_list.append(pixel_intensities_movement)
            pixel_intensities_movement_coordinates_list.append(diff_selected_column)

            max_pixel_intensity = int(diff_selected.max()) if diff_selected.size else 0

            
            reference_selected_image_gray = selected_image_contraction_gray[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference]
            reference_all_image_gray = all_image_gray[y1_small_new_reference:y2_small_new_reference, x1_small_new_reference:x2_small_new_reference] 
            
            not_crop_selected_image_gray = selected_image_contraction_gray 
            not_crop_all_image_gray = all_image_gray

            reference_diff_selected = cv2.absdiff(reference_selected_image_gray, reference_all_image_gray) 
            not_crop_diff_selected = cv2.absdiff(not_crop_selected_image_gray, not_crop_all_image_gray) 
            
            diff_loc = np.argwhere((diff_selected_reference >= threshold) & (diff_selected_reference <= max_threshold))
            not_cropped_diff_loc = np.argwhere((reference_diff_selected >= threshold) & (reference_diff_selected <= max_threshold)) 
            
            threshold_pixel_image = np.zeros_like(diff_selected_reference, dtype = np.uint8)
            threshold_pixel_image_color = cv2.cvtColor(threshold_pixel_image, cv2.COLOR_GRAY2BGR)
            for loc in diff_loc:
                threshold_pixel_image_color[loc[0], loc[1]] = diff_selected_reference[loc[0], loc[1]]
            
            result_folder_diff = os.path.join(folder_path, 'result', 'Analysis Square') 
            if not os.path.exists(result_folder_diff):
                os.makedirs(result_folder_diff, exist_ok = True) 
            x_coord, y_coord = x_selected_new - x1_small_new_optical ,y_selected_new - y1_small_new_optical
        
            height, width = diff_selected.shape[:2]
            not_cropped_height, not_cropped_width = reference_diff_selected.shape[:2]

            new_width = 1080

            new_height = int(height * (new_width / width))
            diff_resized = cv2.resize(threshold_pixel_image_color, (new_width, new_height), interpolation=cv2.INTER_LINEAR) 
            result_image_path = os.path.join(result_folder_diff, f'Analysis_Frame_{i:04d}.{FRAME_EXTENSION}')
            Analysis_frames.append(diff_resized)
            white_pixel_coords = np.argwhere(diff_selected >= adapt_mean)
            
            not_cropped_white_pixel_coords = np.argwhere(reference_diff_selected >= adapt_mean)
            
            white_pixel_count = len(white_pixel_coords)
            
            not_cropped_white_pixel_count = len(not_cropped_white_pixel_coords)
            
            min_threshold = threshold
            max_threshold = max_threshold
            
            valid_pixel_coords = np.argwhere((diff_selected >= min_threshold) & (diff_selected <= max_threshold))
            not_cropped_valid_pixel_coords = np.argwhere((reference_diff_selected >= min_threshold) & (reference_diff_selected <= max_threshold))
            
            valid_pixel_count = len(valid_pixel_coords)
            valid_pixel_counts.append(valid_pixel_count)
            total_pixel_count = resolution_x * resolution_y 
            
            not_cropped_valid_pixel_count = len(not_cropped_valid_pixel_coords)
            not_cropped_valid_pixel_counts.append(not_cropped_valid_pixel_count)   



        valid_pixel_counts_np = np.array(valid_pixel_counts, dtype='float32') 

        valid_values = valid_pixel_counts_np[(valid_pixel_counts_np != 0) & ~np.isnan(valid_pixel_counts_np)]

        valid_pixel_count_mean = np.mean(valid_values) if valid_values.size > 0 else 0
        valid_pixel_count_mean_round = round(valid_pixel_count_mean)

        valid_pixel_count_maximum = np.max(valid_values) if valid_values.size > 0 else 0
        valid_pixel_count_maximum_round = round(valid_pixel_count_maximum)

        min_valid_pixel_counts = np.min(valid_values) if valid_values.size > 0 else 0
    
        if not valid_pixel_counts:
            print("!!!!THIS SAMPLE DOESN'T HAVE ANY MOVEMENT!!!!.")
            classification_dir = os.path.join(BAD_FILE_PATH, 'FINISHED')
            destination_folder = os.path.join(classification_dir, 'BAD')
            folder_path = os.path.join(BAD_FILE_PATH, folder_name)
            video_path = os.path.join(BAD_FILE_PATH, f'{folder_name}.{extension}')  
            os.makedirs(classification_dir, exist_ok=True)
            os.makedirs(destination_folder, exist_ok=True)
            try:
                shutil.move(folder_path, destination_folder)
                shutil.move(video_path, destination_folder)
            except Exception as e:
                pass
            continue

        difference_square_image = cv2.subtract(square_differences_crop_indices_index_path_image, square_image)
        difference_square_image_gray = cv2.cvtColor(difference_square_image, cv2.COLOR_BGR2GRAY)
        


        
        
        # 동영상 파일명과 확장자 설정
        video_name_base = f'{result_folder_diff}\\Analysis_Video'
        video_name = f'{video_name_base}\\{folder_name}_Analysis_Video.avi'
        if not os.path.exists(video_name_base):
            os.makedirs(video_name_base)
        # VideoWriter 객체 생성
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(video_name, fourcc, fps, (new_width, new_height))
        # 프레임 이미지들을 읽어와 동영상에 추가
        for i in range(len(gray_frames)):
            out.write(Analysis_frames[i])

        # VideoWriter 객체 해제
        out.release()
        print("Analysis Video Completed.")
        
        gray_color = '#bbbbbd'
        red_color = '#FF6699'
        blue_color = '#009fd6'
        yellow_color = '#d69f00'
        # 가장 많이 움직인 곳 사각형 표시 #214,159,0 (파란색)




        
        # ============================================
                # ============================================
        # Optical flow-based displacement analysis
        # KR: optical flow 기반 displacement 분석
        # ============================================

        # Prepare candidate contraction and relaxation frames for ROI-based optical flow.
        # KR: ROI 기반 optical flow 분석을 위해 후보 수축/이완 frame을 준비합니다.
        roi_contraction = cv2.cvtColor(
            square_differences_crop_indices_index_path_image, cv2.COLOR_BGR2GRAY
        )
        roi_relaxation = cv2.cvtColor(
            square_image, cv2.COLOR_BGR2GRAY
        )

        roi_contraction = roi_contraction[
            y1_small_new_optical:y2_small_new_optical,
            x1_small_new_optical:x2_small_new_optical
        ]
        roi_relaxation = roi_relaxation[
            y1_small_new_optical:y2_small_new_optical,
            x1_small_new_optical:x2_small_new_optical
        ]

        # Skip optical flow when the ROI is empty.
        # KR: ROI가 비어 있으면 optical flow 분석을 건너뜁니다.
        if roi_contraction.size == 0:
            print("roi_contraction is empty. Optical flow skipped.")
            target_displacements = [0.0]
            maximum_displacement_list = target_displacements
            max_target_displacements = 0.0
            max_displacement_um = 0.0
            min_target_displacements = 0.0

            # Use the selected coordinate as the fallback position.
            # KR: optical flow를 수행할 수 없을 때 선택 좌표를 fallback 위치로 사용합니다.
            a_global = x_selected_new
            b_global = y_selected_new
            c_global = x_selected_new
            d_global = y_selected_new

            global_max_coords_global = (b_global, a_global, d_global, c_global)
            global_max_coords_global1 = (b_global, a_global)
            global_max_coords_global2 = (d_global, c_global)
            global_max_coords_global1_ab = (a_global, b_global)
            global_max_coords_global2_cd = (c_global, d_global)

            tracked_points = [(x_selected_new - x1_small_new_optical,
                            y_selected_new - y1_small_new_optical)]
            best_x_roi = tracked_points[0][0]
            best_y_roi = tracked_points[0][1]
            best_x_roi_min = best_x_roi
            best_y_roi_min = best_y_roi

        else:
            # Set the output path for the optical flow overlay video.
            # KR: optical flow overlay video의 저장 경로를 설정합니다.
            output_video_path = os.path.join(result_folder, f'{folder_name}_Optical_Flow_Overlay.{extension}')

            fps = fps  
            frame_size = (roi_contraction.shape[1], roi_contraction.shape[0])  # Video frame size: width, height

            fourcc = cv2.VideoWriter_fourcc(*'XVID')
            out = cv2.VideoWriter(output_video_path, fourcc, fps, frame_size)

            # ------------------------------------------------
            # 2) Define the baseline tracking point within the ROI
            #    The selected motion center is clamped to remain inside the ROI.
            # KR: ROI 내부에서 baseline tracking point를 정의합니다.
            # KR: 선택된 motion center가 ROI 밖으로 벗어나지 않도록 좌표를 제한합니다.
            # ------------------------------------------------
            H, W = roi_contraction.shape[:2]

            x0 = int(x_selected_new - x1_small_new_optical)
            y0 = int(y_selected_new - y1_small_new_optical)

            half_win = ANALYSIS_SQUARE_SIZE_REFERENCE // 2    # Half-window size for coordinate clamping
            # Clamp the tracking point so that the local window remains inside the ROI.
            # KR: local tracking window가 ROI 내부에 유지되도록 tracking point 좌표를 제한합니다.
            x0 = max(half_win, min(W - 1 - half_win, x0))
            y0 = max(half_win, min(H - 1 - half_win, y0))

            # Use a single baseline tracking point for Lucas-Kanade optical flow.
            # KR: Lucas-Kanade optical flow에서 하나의 baseline tracking point만 사용합니다.
            p0 = np.array([[[x0, y0]]], dtype=np.float32)

            # Store optical-flow tracking results.
            # KR: optical flow tracking 결과를 저장합니다.
            target_displacements = []   # Displacement for each frame in µm
            tracked_points = []         # Tracked ROI coordinates for each frame

            # Smoothing parameters for visualization only.
            # KR: visualization에만 사용하는 smoothing 설정입니다.
            smoothed_x, smoothed_y = None, None
            alpha = 0.5  # EMA smoothing factor

            # ------------------------------------------------
            # 3) Optical flow tracking from the baseline ROI to each frame ROI
            #    A single point p0 is tracked to define optical flow-derived displacement.
            # KR: baseline ROI에서 각 frame ROI로 optical flow tracking을 수행합니다.
            # KR: 하나의 p0 point를 추적하여 optical flow-derived displacement를 정의합니다.
            # ------------------------------------------------
            for frame_idx, frame_full in enumerate(current_gray_image_list):

                cur_gray = frame_full[
                    y1_small_new_optical:y2_small_new_optical,
                    x1_small_new_optical:x2_small_new_optical
                ]

                # Keep the previous displacement value when the current ROI is empty.
                # KR: 현재 ROI가 비어 있으면 이전 displacement 값을 유지합니다.
                if cur_gray.size == 0:
                    last_disp = target_displacements[-1] if target_displacements else 0.0
                    target_displacements.append(float(last_disp))
                    tracked_points.append((float(x0), float(y0)))
                    continue

                p1, st, err = cv2.calcOpticalFlowPyrLK(
                    roi_contraction,
                    cur_gray,
                    p0,
                    None,
                    winSize=(15, 15),
                    maxLevel=2,
                    criteria=(cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, 10, 0.03)
                )

                frame_rgb = cv2.cvtColor(cur_gray, cv2.COLOR_GRAY2BGR)

                # Handle optical-flow tracking failure.
                # KR: optical flow tracking에 실패한 경우를 처리합니다.
                if p1 is None or st is None or st[0, 0] == 0:
                    last_disp = target_displacements[-1] if target_displacements else 0.0
                    target_displacements.append(float(last_disp))
                    tracked_points.append((float(x0), float(y0)))

                    # Write the frame even when optical-flow tracking fails.
                    # KR: optical flow tracking에 실패해도 output video frame은 저장합니다.
                    out.write(frame_rgb)
                    continue

                # New tracked coordinate in the baseline ROI coordinate system.
                # KR: baseline ROI 좌표계에서 새롭게 추적된 좌표입니다.
                new_x = float(p1[0, 0, 0])
                new_y = float(p1[0, 0, 1])

                # Calculate displacement from the baseline point p0.
                # This is not cumulative; the same baseline point is used for all frames.
                # KR: baseline point p0로부터의 displacement를 계산합니다.
                # KR: 누적 변위가 아니라 모든 frame에서 동일한 기준점 p0를 사용합니다.
                dx = new_x - x0
                dy = new_y - y0
                disp_um = np.sqrt(dx * dx + dy * dy) * PIXEL_UM

                target_displacements.append(float(disp_um))
                tracked_points.append((new_x, new_y))

                # Smooth the displayed tracking point for visualization.
                # KR: visualization에서 표시되는 tracking point를 부드럽게 보이도록 smoothing합니다.
                if smoothed_x is None:
                    smoothed_x, smoothed_y = new_x, new_y
                else:
                    smoothed_x = alpha * new_x + (1.0 - alpha) * smoothed_x
                    smoothed_y = alpha * new_y + (1.0 - alpha) * smoothed_y

                # Draw a green arrow from the baseline point to the current tracked position,
                # and mark the smoothed tracked point in red.
                # KR: baseline point에서 현재 추적 위치까지 초록색 화살표를 그리고,
                # KR: smoothing된 tracking point를 빨간색으로 표시합니다.
                cv2.arrowedLine(
                    frame_rgb,
                    (int(x0), int(y0)),
                    (int(new_x), int(new_y)),
                    (0, 255, 0),
                    1
                )
                cv2.circle(
                    frame_rgb,
                    (int(smoothed_x), int(smoothed_y)),
                    2,
                    (0, 0, 255),
                    -1
                )

                out.write(frame_rgb)
            # Finalize the optical flow overlay video.
            # KR: optical flow overlay video 저장을 종료합니다.
            out.release()

            # ------------------------------------------------
            # 4) Displacement time-series post-processing
            #    - target_displacements are stored in micrometers.
            #    - The analysis window is selected based on start_number, NUM_FRAMES, and fps.
            # KR: displacement time-series를 후처리합니다.
            # KR: target_displacements는 micrometer 단위입니다.
            # KR: 분석 구간은 start_number, NUM_FRAMES, fps를 기준으로 선택합니다.
            # ------------------------------------------------
            if len(target_displacements) == 0:
                target_displacements = [0.0]

            # Shift the minimum displacement to zero.
            # KR: 최소 displacement가 0이 되도록 baseline을 보정합니다.
            min_all = min(target_displacements)
            target_displacements = [float(v - min_all) for v in target_displacements]
            
            # Invert the displacement curve for visualization consistency.
            # KR: visualization 방향을 맞추기 위해 displacement curve를 반전합니다.
            max_all = max(target_displacements)
            target_displacements = [float(max_all - v) for v in target_displacements]

            # Select the initial analysis window.
            # KR: 초기 분석 구간을 선택합니다.
            end_idx = int(fps * NUM_FRAMES)
            end_idx = min(end_idx, len(target_displacements))

            if start_number + 1 < end_idx:
                window_vals = target_displacements[start_number + 1:end_idx]
                window_indices = list(range(start_number + 1, end_idx))
            else:
                window_vals = target_displacements
                window_indices = list(range(len(target_displacements)))

            if len(window_vals) == 0:
                window_vals = target_displacements
                window_indices = list(range(len(target_displacements)))

            # Define the final displacement time series and maximum displacement value.
            # KR: 최종 displacement time series와 maximum displacement 값을 정의합니다.
            maximum_displacement_list = target_displacements
            max_target_displacements = round(max(window_vals), 2)
            min_target_displacements = float(min(window_vals))
            max_displacement_um = max_target_displacements

            max_target_displacements_before_all = round(max(target_displacements), 2)

            # Create the time axis for displacement plotting.
            # KR: displacement plot에 사용할 시간축을 생성합니다.
            if fps > 0:
                total_time_seconds = len(maximum_displacement_list) / fps
            else:
                total_time_seconds = float(len(maximum_displacement_list))
            time_axis = np.linspace(0, total_time_seconds, len(maximum_displacement_list))

            # Frame index where the maximum displacement occurs within the analysis window.
            # KR: 분석 구간 안에서 maximum displacement가 발생한 frame index입니다.
            idx_in_window = int(np.argmax(window_vals))
            frame_idx_max = int(window_indices[idx_in_window])

            # Extract ROI coordinates of the tracked point at maximum and baseline displacement.
            # KR: maximum displacement와 baseline displacement 시점의 tracked point ROI 좌표를 가져옵니다.
            if frame_idx_max < len(tracked_points):
                best_x_roi, best_y_roi = tracked_points[frame_idx_max]
            else:
                best_x_roi, best_y_roi = tracked_points[-1]

            if start_number < len(tracked_points):
                best_x_roi_min, best_y_roi_min = tracked_points[start_number]
            else:
                best_x_roi_min, best_y_roi_min = tracked_points[0]

            # Convert ROI coordinates back to full-frame coordinates.
            # KR: ROI 좌표를 full-frame 좌표로 변환합니다.
            a_global = int(best_x_roi + x1_small_new_optical)
            b_global = int(best_y_roi + y1_small_new_optical)
            c_global = int(best_x_roi_min + x1_small_new_optical)
            d_global = int(best_y_roi_min + y1_small_new_optical)

            global_max_coords_global = (b_global, a_global, d_global, c_global)
            global_max_coords_global1 = (b_global, a_global)
            global_max_coords_global2 = (d_global, c_global)
            global_max_coords_global1_ab = (a_global, b_global)
            global_max_coords_global2_cd = (c_global, d_global)
            
    
        # Convert the displacement time series to a comma-separated string for Excel output.
        # KR: Excel output에 저장하기 위해 displacement time series를 comma-separated string으로 변환합니다.
        maximum_displacement_str = ", ".join(f"{x:.2f}" for x in maximum_displacement_list)
                
        y_max_coordinates = int(y_selected_new - y1_small_new_optical)
        x_max_coordinates = int(x_selected_new - x1_small_new_optical)

        x1_max_coordinates = x_max_coordinates  - ANALYSIS_SQUARE_SIZE_REFERENCE
        y1_max_coordinates = y_max_coordinates  - ANALYSIS_SQUARE_SIZE_REFERENCE
        x2_max_coordinates = x_max_coordinates  + ANALYSIS_SQUARE_SIZE_REFERENCE
        y2_max_coordinates = y_max_coordinates  + ANALYSIS_SQUARE_SIZE_REFERENCE
        
        x1_max_coordinates = max( x1_max_coordinates, 0)
        y1_max_coordinates = max( y1_max_coordinates, 0)
        x2_max_coordinates = min( x2_max_coordinates, resolution_x -1)
        y2_max_coordinates = min( y2_max_coordinates, resolution_y -1)

        max_value_displacement = max(maximum_displacement_list) 
        
        if max_target_displacements == 0 or max_target_displacements is None:
            print("No valid displacement was detected.")
            print("This sample could not be tracked by optical flow.")
            print(f"Move to BAD Folder")
            classification_dir = os.path.join(BAD_FILE_PATH, 'FINISHED')
            destination_folder = os.path.join(classification_dir, 'BAD')
            folder_path = os.path.join(BAD_FILE_PATH, folder_name)
            video_path = os.path.join(BAD_FILE_PATH, f'{folder_name}.{extension}')  
            os.makedirs(classification_dir, exist_ok=True)
            os.makedirs(destination_folder, exist_ok=True)
            try:
                shutil.move(folder_path, destination_folder)
                shutil.move(video_path, destination_folder)
                continue
            except Exception as e:
                continue
        else:
            pass



        
        # Compute summary statistics after excluding zero-count frames.
        # KR: pixel count가 0인 frame을 제외하고 moved-region 통계값을 계산합니다.
        valid_pixel_count_mean = np.mean([value for value in valid_pixel_counts if value not in [0, '']])
        valid_pixel_count_mean_round = round(valid_pixel_count_mean) if not np.isnan(valid_pixel_count_mean) else 0
        valid_pixel_count_maximum = max([value for value in valid_pixel_counts if value not in [0, '']], default=0)
        valid_pixel_count_maximum_round = round(valid_pixel_count_maximum) if not np.isnan(valid_pixel_count_maximum) else 0

        # Convert moved-region pixel counts to physical area.
        # KR: moved region의 pixel count를 실제 면적으로 변환합니다.
        valid_pixel_count_mean_area = valid_pixel_count_mean_round * AREA_MOVED
        valid_pixel_count_maximum_area = valid_pixel_count_maximum_round * AREA_MOVED
        valid_pixel_count_mean_area_round = round(valid_pixel_count_mean_area,2)
        valid_pixel_count_maximum_area_round = round(valid_pixel_count_maximum_area, 2)
        
        # Frame-wise moved-region area and length-equivalent values.
        # KR: 각 frame별 moved-region area와 length-equivalent 값을 계산합니다.

        displacement_list = [count * PIXEL_UM for count in valid_pixel_counts]

        # =========================================================
        # Moved-region area time-series visualization
        # =========================================================
        # This plot shows the frame-wise moved-region area estimated from
        # thresholded motion pixels. The moved-region area is calculated by
        # multiplying the number of thresholded pixels by the calibrated pixel area.
        # This output is used as an auxiliary quality-control visualization and is
        # separate from optical-flow displacement.
        #
        # KR:
        # 이 plot은 threshold 조건을 통과한 moved pixel 수를 기반으로
        # frame별 moved-region area 변화를 보여줍니다.
        # moved-region area는 thresholded pixel count에 pixel area 보정값을
        # 곱하여 계산됩니다.
        # 이 결과는 보조적인 QC 시각화이며, optical-flow displacement 값과는
        # 별개의 지표입니다.
        # =========================================================

        fig, ax = plt.subplots(figsize=(8, 8))
        ax.plot(np.linspace(0, total_time_seconds, len(maximum_displacement_list)), 
                maximum_displacement_list, 
                color=blue_color, 
                linewidth=4)

        ax.set_xlabel('Time (s)',  fontsize=40, fontweight='bold', fontname='Arial')
        ax.set_ylabel(f'Displacement ({LENGTH_UNIT})',  fontsize=40,fontweight='bold', fontname='Arial')

        ax.set_ylim(0, Y_LIM)
        
        ax.set_yticks(range(0, Y_LIM + 1, 1))
        
        for spine in ax.spines.values():
            spine.set_linewidth(4) 
            
        ax.set_xlim(0, total_time_seconds)
        
        ax.set_xticks(range(0, int(total_time_seconds) + 1, X_TICKS))

        ax.axhline(y=max_target_displacements, color=red_color, linewidth=3.5, linestyle='--')

        ax.grid(False)
        plt.text(total_time_seconds - (total_time_seconds*0.0125), max_target_displacements + 0.625, f'Max Displacement ({LENGTH_UNIT}) : {max_target_displacements}',
                verticalalignment='top', horizontalalignment='right', color=red_color,fontweight='bold', fontsize=25)

        ax.tick_params(axis='x', direction='in', length=10, width=4, labelsize=40, pad=10)
        ax.tick_params(axis='y', direction='in', length=10, width=4, labelsize=40, pad=20)

        ax.spines['left'].set_position(('data', 0))  
        ax.spines['bottom'].set_position(('data', 0))  


        ax.spines['right'].set_color('black') 
        ax.spines['top'].set_color('black')    


        plt.xticks( fontweight='bold')
        plt.yticks( fontweight='bold')
        
        Savefig(folder_path, 'Optical_Flow_Displacement', 300, 0.2)
        
        
        
        if FRAME_REMOVE.lower() == 'on':
            paths_to_clean = [
                (folder_path, "result"),  
            ]
            
            for base_path, exclude_name in paths_to_clean:
                if os.path.exists(base_path): 
                    for item_name in os.listdir(base_path):
                        item_path = os.path.join(base_path, item_name)
                        if item_name != exclude_name:  
                            if os.path.isfile(item_path):
                                os.remove(item_path)
                            elif os.path.isdir(item_path):
                                shutil.rmtree(item_path)

        seq_raw_np = np.array(maximum_displacement_list, dtype=np.float32)

        if len(seq_raw_np) == 0:
            max_abs_disp = 0.0
            abs_net_change_ratio = 0.0
            sign_changes = 0
            diff_std = 0.0
        else:
            max_abs_disp = float(np.max(np.abs(seq_raw_np)))

            peak_to_peak = float(np.max(seq_raw_np) - np.min(seq_raw_np))

            abs_net_change_ratio = float(
                np.abs(seq_raw_np[-1] - seq_raw_np[0]) / (peak_to_peak + 1e-8)
            )

            diff_seq = np.diff(seq_raw_np)

            if len(diff_seq) >= 2:
                sign_changes = int(np.sum(np.diff(np.sign(diff_seq)) != 0))
            else:
                sign_changes = 0

            diff_std = float(np.std(diff_seq)) if len(diff_seq) > 0 else 0.0
            
    
        band_energy_optical_displacement, max_value_optical_displacement_band_energy, max_optical_amp_at_max_freq, optical_freq_at_max_amp_number  = plot_fft(folder_path, fps, target_displacements, X_LIM_AMPLITUDE, Y_LIM_AMPLITUDE,'Optical_Flow_Displacement_FFT')
        
        
        Excel_result_data_noise = [
        folder_name,
        adapt_max,
        max_target_displacements,
        band_energy_optical_displacement,
        max_value_optical_displacement_band_energy,
        x_selected_new,
        y_selected_new,
        optical_freq_at_max_amp_number,
        maximum_displacement_str,
        max_abs_disp,
        abs_net_change_ratio,
        sign_changes,
        diff_std
        ]
        
        finish_dir = os.path.join(NOISE_FILE_PATH, 'FINISHED')
        if not os.path.exists(finish_dir):
            os.makedirs(finish_dir, exist_ok=True)
        folder_path = os.path.join(NOISE_FILE_PATH, folder_name)
        video_path = os.path.join(NOISE_FILE_PATH, f'{folder_name}.{extension}')
        try:
            shutil.move(folder_path, finish_dir)
            shutil.move(video_path, finish_dir)
        except Exception as e:
            pass 

        folder_end_time = time.time()
        
        folder_execution_time = round((folder_end_time - folder_start_time),2)
        estimated_execution_times.append(folder_execution_time)  

        average_execution_time = round(sum(estimated_execution_times) / len(estimated_execution_times),2)
        estimated_left_time = round((((average_execution_time * number_of_folders) - sum(estimated_execution_times))/60),1)
        print(f"\n[INFO] Folder completed: {folder_name}")
        print(f"[INFO] Processing time: {folder_execution_time} seconds")
        print(f"[INFO] Estimated remaining time: {estimated_left_time} minutes\n")
        
        try:
            noise_ws.append(Excel_result_data_noise)
            noise_wb.save(NOISE_EXCEL)
            print(f'Save The  『 {folder_name} 』 Information to {NOISE_SAMPLE_EXCEL_NAME}')

        except NameError:
            noise_ws.append(Excel_result_data_noise)
            noise_wb.save(NOISE_EXCEL)
            print(f'Save The  『 {folder_name} 』 Information to {NOISE_SAMPLE_EXCEL_NAME}')

        plt.close()
        noise_wb.close()

    try:
        plt.close()
        noise_wb.close()
    except :
        pass
    print('============================ NOISE FILE FINISHED ============================')
    
def load_data(BAD_EXCEL, TARGET_EXCEL, NOISE_EXCEL, columns, extension = 'avi'):
    def parse_displacement_list(cell_value):
        if pd.isna(cell_value):
            return []

        text = str(cell_value).strip()
        if text == "":
            return []

        try:
            if text.startswith("[") and text.endswith("]"):
                parsed = ast.literal_eval(text)
                return [float(x) for x in parsed]

            return [float(x.strip()) for x in text.split(",") if x.strip() != ""]
        except Exception as e:
            print(f"[WARN] Displacement List parse error: {cell_value} | {e}")
            return []
    def make_raw_sequence(seq):
        seq = np.asarray(seq, dtype=np.float32)
        if len(seq) == 0:
            return []

        seq = np.sign(seq) * np.log1p(np.abs(seq))
        return seq.tolist()
    

    def make_shape_sequence(seq):
        seq = np.asarray(seq, dtype=np.float32)
        if len(seq) == 0:
            return []

        max_abs = np.max(np.abs(seq))
        if max_abs < 1e-8:
            return seq.tolist()

        return (seq / max_abs).tolist()


    try:
        bad_sample_df = pd.read_excel(BAD_EXCEL)
        target_sample_df = pd.read_excel(TARGET_EXCEL)
        noise_sample_df = pd.read_excel(NOISE_EXCEL)

        bad_sample_df.columns = bad_sample_df.columns.str.strip()
        target_sample_df.columns = target_sample_df.columns.str.strip()
        noise_sample_df.columns = noise_sample_df.columns.str.strip()

        for df in [bad_sample_df, target_sample_df, noise_sample_df]:
            if 'MAX BAND ENERGY' in df.columns:
                df['MAX BAND ENERGY'] = pd.to_numeric(df['MAX BAND ENERGY'], errors='coerce')
                df['MAX BAND ENERGY'] = df['MAX BAND ENERGY'].fillna(0).astype(np.float32)
                df['MAX BAND ENERGY'] = np.log1p(df['MAX BAND ENERGY'].to_numpy(dtype=np.float32))

            if 'MAX Apeak' in df.columns:
                df['MAX Apeak'] = pd.to_numeric(df['MAX Apeak'], errors='coerce')
                df['MAX Apeak'] = df['MAX Apeak'].fillna(0).astype(np.float32)

        def extract_data(df, class_idx):
            seq_list = []
            scalar_list = []
            label_list = []
            length_list = []
            meta_list = []
            if 'Displacement List' not in df.columns:
                raise ValueError("'Displacement List' column not found in Excel file.")

            derived_cols = {
                'MAX ABS DISP',
                'ABS NET CHANGE RATIO',
                'NUM SIGN CHANGES DIFF',
                'DIFF STD'
            }

            missing_cols = [col for col in columns if col not in df.columns and col not in derived_cols]
            if missing_cols:
                raise ValueError(f"Missing columns in dataframe: {missing_cols}")

            for _, row in df.iterrows():
                seq_raw = parse_displacement_list(row['Displacement List'])

                if len(seq_raw) == 0:
                    continue


                seq_raw_np = np.array(seq_raw, dtype=np.float32)

                max_abs_disp = float(np.max(np.abs(seq_raw_np)))

                peak_to_peak = float(np.max(seq_raw_np) - np.min(seq_raw_np))

                abs_net_change_ratio = float(
                    np.abs(seq_raw_np[-1] - seq_raw_np[0]) / (peak_to_peak + 1e-8)
                )

                diff_seq = np.diff(seq_raw_np)
                if len(diff_seq) >= 2:
                    sign_changes = int(np.sum(np.diff(np.sign(diff_seq)) != 0))
                else:
                    sign_changes = 0
                    
                diff_std = float(np.std(diff_seq)) if len(diff_seq) > 0 else 0.0

 
                seq_raw_processed = make_raw_sequence(seq_raw_np)
                seq_shape_processed = make_shape_sequence(seq_raw_np)

                if len(seq_raw_processed) == 0 or len(seq_shape_processed) == 0:
                    continue

                seq_len = len(seq_raw_processed)
                seq_2ch = np.stack(
                    [
                        np.asarray(seq_raw_processed, dtype=np.float32),
                        np.asarray(seq_shape_processed, dtype=np.float32)
                    ],
                    axis=-1
                )  # [T, 2]

                row_values = row.to_dict()

                row_values['MAX ABS DISP'] = max_abs_disp
                row_values['ABS NET CHANGE RATIO'] = abs_net_change_ratio
                row_values['NUM SIGN CHANGES DIFF'] = sign_changes
                row_values['DIFF STD'] = diff_std

                scalar_values = np.array([row_values[col] for col in columns], dtype=np.float32)
                sample_name = str(row['Folder name']).strip()
                source_path = os.path.join(REFERENCE_FILE_PATH, f"{sample_name}.{extension}")

                seq_list.append(seq_2ch)
                scalar_list.append(scalar_values)
                label_list.append(class_idx)
                length_list.append(seq_len)
                meta_list.append({
                    "sample_name": sample_name,
                    "source_path": source_path
                })

            return seq_list, scalar_list, label_list, length_list, meta_list


        bad_seq, bad_scalar, bad_label, bad_len, bad_meta = extract_data(bad_sample_df, 0)
        target_seq, target_scalar, target_label, target_len, target_meta = extract_data(target_sample_df, 1)
        noise_seq, noise_scalar, noise_label, noise_len, noise_meta = extract_data(noise_sample_df, 2)
        
        all_seq = bad_seq + target_seq + noise_seq
        all_scalar = bad_scalar + target_scalar + noise_scalar
        all_label = bad_label + target_label + noise_label
        all_lengths = bad_len + target_len + noise_len
        all_meta = bad_meta + target_meta + noise_meta

        if len(all_seq) == 0:
            print("[ERROR] No valid sequence data found in 'Displacement List'.")
            return None, None, None, None, None

        x_seq = pad_sequence_list(all_seq, pad_value=0.0)   # [N, T, 2]
        x_scalar = np.array(all_scalar, dtype=np.float32)
        y = np.array(all_label, dtype=np.int64)
        lengths = np.array(all_lengths, dtype=np.int64)

        return x_seq, x_scalar, y, lengths, all_meta

    except Exception as e:
        print(f"[ERROR] load_data failed: {e}")
        return None, None, None, None, None
    
class SequenceScalarDataset(Dataset):
    def __init__(self, x_seq, x_scalar, y, lengths, meta_list):
        self.x_seq = torch.tensor(x_seq, dtype=torch.float32)
        self.x_scalar = torch.tensor(x_scalar, dtype=torch.float32)
        self.y = torch.tensor(y, dtype=torch.long)
        self.lengths = torch.tensor(lengths, dtype=torch.long)
        self.meta_list = meta_list

    def __len__(self):
        return len(self.y)

    def __getitem__(self, idx):
        return self.x_seq[idx], self.x_scalar[idx], self.y[idx], self.lengths[idx], self.meta_list[idx]

def train_one_epoch(model, loader, optimizer, criterion, device):

    model.train()

    running_loss = 0.0
    correct = 0
    total = 0

    pbar = tqdm(loader, desc="Train", leave=False)

    for seq_x, scalar_x, y, lengths, _ in pbar:
        seq_x = seq_x.to(device)
        scalar_x = scalar_x.to(device)
        y = y.to(device)
        lengths = lengths.to(device)

        optimizer.zero_grad()
        logits = model(seq_x, scalar_x, lengths)
        loss = criterion(logits, y)
        loss.backward()
        optimizer.step()

        running_loss += loss.item() * y.size(0)
        preds = torch.argmax(logits, dim=1)
        correct += (preds == y).sum().item()
        total += y.size(0)

        pbar.set_postfix({
            "loss": f"{loss.item():.4f}",
            "acc": f"{correct/total:.3f}"
        })

    epoch_loss = running_loss / total
    epoch_acc = correct / total
    return epoch_loss, epoch_acc


@torch.no_grad()
def evaluate_model(model, loader, criterion, device):
    model.eval()
    running_loss = 0.0
    correct = 0
    total = 0

    all_preds = []
    all_true = []
    all_sample_names = []
    
    for seq_x, scalar_x, y, lengths, meta in loader:
        seq_x = seq_x.to(device)
        scalar_x = scalar_x.to(device)
        y = y.to(device)
        lengths = lengths.to(device)

        logits = model(seq_x, scalar_x, lengths)
        loss = criterion(logits, y)

        running_loss += loss.item() * y.size(0)
        preds = torch.argmax(logits, dim=1)

        correct += (preds == y).sum().item()
        total += y.size(0)

        all_preds.extend(preds.cpu().numpy().tolist())
        all_true.extend(y.cpu().numpy().tolist())
        all_sample_names.extend(meta["sample_name"])

    epoch_loss = running_loss / total
    epoch_acc = correct / total
    return epoch_loss, epoch_acc, np.array(all_preds), np.array(all_true), all_sample_names
    
def classify_and_copy_files(model, loader, device, output_root, move_files=False):
    model.eval()

    os.makedirs(output_root, exist_ok=True)

    active_dir = os.path.join(output_root, "ACTIVE")
    inactive_dir = os.path.join(output_root, "INACTIVE")
    noise_dir = os.path.join(output_root, "NOISE")

    os.makedirs(active_dir, exist_ok=True)
    os.makedirs(inactive_dir, exist_ok=True)
    os.makedirs(noise_dir, exist_ok=True)

    with torch.no_grad():
        for seq_x, scalar_x, y, lengths, meta in loader:
            seq_x = seq_x.to(device)
            scalar_x = scalar_x.to(device)
            lengths = lengths.to(device)

            logits = model(seq_x, scalar_x, lengths)
            preds = torch.argmax(logits, dim=1).cpu().numpy()

            for i in range(len(preds)):
                pred_class = int(preds[i])

                sample_name = meta["sample_name"][i]
                source_path = meta["source_path"][i]
                if pred_class == 1:
                    destination_dir = active_dir
                elif pred_class == 0:
                    destination_dir = inactive_dir
                else:
                    destination_dir = noise_dir

                if not os.path.exists(source_path):
                    print(f"[WARN] Source path not found: {source_path}")
                    continue

                dst_path = os.path.join(destination_dir, os.path.basename(source_path))

                try:
                    if move_files:
                        shutil.move(source_path, dst_path)
                    else:
                        if os.path.isdir(source_path):
                            if os.path.exists(dst_path):
                                shutil.rmtree(dst_path)
                            shutil.copytree(source_path, dst_path)
                        else:
                            shutil.copy2(source_path, dst_path)

                    print(f"[OK] {sample_name} -> pred={pred_class} -> {destination_dir}")

                except Exception as e:
                    print(f"[ERROR] Failed to classify {sample_name}: {e}")
    
def load_latest_checkpoint_torch(model, checkpoint_dir, device):
    checkpoints = [
        f for f in os.listdir(checkpoint_dir)
        if f.endswith('.pth') and f.startswith('epoch_')
    ]

    if checkpoints:
        latest_checkpoint = max(checkpoints, key=lambda x: int(x.split('_')[1]))
        checkpoint_path = os.path.join(checkpoint_dir, latest_checkpoint)
        model.load_state_dict(torch.load(checkpoint_path, map_location=device))
        print(f"Loaded weights from {latest_checkpoint}")
        return int(latest_checkpoint.split('_')[1])
    else:
        print("No checkpoint found, starting training from scratch.")
        return 0
    
class HybridLSTMClassifier(nn.Module):
    def __init__(
        self,
        seq_input_dim=2,
        scalar_input_dim=6,
        num_classes=3,
        lstm_hidden1=64,
        lstm_hidden2=32,
        dropout=0.3
    ):
        super().__init__()

        self.lstm1 = nn.LSTM(
            input_size=seq_input_dim,
            hidden_size=lstm_hidden1,
            batch_first=True,
            bidirectional=True
        )

        self.lstm2 = nn.LSTM(
            input_size=lstm_hidden1 * 2,
            hidden_size=lstm_hidden2,
            batch_first=True,
            bidirectional=True
        )

        self.dropout = nn.Dropout(dropout)

        # attention score
        self.attn_fc = nn.Linear(lstm_hidden2 * 2, 1)

        self.seq_fc = nn.Sequential(
            nn.Linear(lstm_hidden2 * 2, 64),
            nn.ReLU(),
            nn.LayerNorm(64),
            nn.Dropout(dropout),
            nn.Linear(64, 32),
            nn.ReLU()
        )

        self.scalar_fc = nn.Sequential(
            nn.Linear(scalar_input_dim, 32),
            nn.ReLU(),
            nn.LayerNorm(32),
            nn.Dropout(dropout),
            nn.Linear(32, 16),
            nn.ReLU()
        )

        self.classifier = nn.Sequential(
            nn.Linear(32 + 16, 128),
            nn.ReLU(),
            nn.LayerNorm(128),
            nn.Dropout(dropout),
            nn.Linear(128, 64),
            nn.ReLU(),
            nn.LayerNorm(64),
            nn.Dropout(dropout),
            nn.Linear(64, 32),
            nn.ReLU(),
            nn.Linear(32, num_classes)
        )

    def forward(self, seq_x, scalar_x, lengths):
        """
        seq_x   : [B, T, 2]
        scalar_x: [B, F]
        lengths : [B]
        """

        lengths_cpu = lengths.detach().cpu()

        packed = nn.utils.rnn.pack_padded_sequence(
            seq_x,
            lengths_cpu,
            batch_first=True,
            enforce_sorted=False
        )

        packed_out1, _ = self.lstm1(packed)
        packed_out2, _ = self.lstm2(packed_out1)

        lstm_out, _ = nn.utils.rnn.pad_packed_sequence(
            packed_out2,
            batch_first=True
        )  # [B, T_valid, hidden*2]

        lstm_out = self.dropout(lstm_out)

        # -----------------------------
        # Attention pooling with mask
        # -----------------------------
        B, T, C = lstm_out.shape

        mask = torch.arange(T, device=seq_x.device).unsqueeze(0) < lengths.unsqueeze(1)
        # mask shape: [B, T]

        attn_scores = self.attn_fc(lstm_out).squeeze(-1)   # [B, T]
        attn_scores = attn_scores.masked_fill(~mask, float('-inf'))
        attn_weights = torch.softmax(attn_scores, dim=1)   # [B, T]

        seq_feat = torch.sum(lstm_out * attn_weights.unsqueeze(-1), dim=1)  # [B, C]
        seq_feat = self.seq_fc(seq_feat)   # [B, 32]

        scalar_feat = self.scalar_fc(scalar_x)  # [B, 16]

        x = torch.cat([seq_feat, scalar_feat], dim=1)
        logits = self.classifier(x)

        return logits


def get_all_checkpoint_files(checkpoint_dir):
    if not os.path.exists(checkpoint_dir):
        return []

    checkpoint_files = []
    for fname in os.listdir(checkpoint_dir):
        if fname.endswith(".pth") and fname.startswith("epoch_"):
            checkpoint_files.append(os.path.join(checkpoint_dir, fname))

    checkpoint_files = natsort.natsorted(checkpoint_files)
    return checkpoint_files


def pad_sequence_list(seq_list, pad_value=0.0):
    max_len = max(seq.shape[0] for seq in seq_list)
    feat_dim = seq_list[0].shape[1]
    padded = []

    for seq in seq_list:
        seq = np.asarray(seq, dtype=np.float32)   # [T, C]
        pad_len = max_len - seq.shape[0]
        if pad_len > 0:
            pad_block = np.full((pad_len, feat_dim), pad_value, dtype=np.float32)
            seq = np.vstack([seq, pad_block])
        padded.append(seq)

    return np.stack(padded, axis=0)   # [N, T, C]

def evaluate_single_checkpoint(model, checkpoint_file, test_loader, criterion, device):
    print("\n" + "=" * 80)
    print(f"Evaluating checkpoint: {os.path.basename(checkpoint_file)}")
    print("=" * 80)

    model.load_state_dict(torch.load(checkpoint_file, map_location=device))

    test_loss, test_accuracy, y_pred, y_true, sample_names = evaluate_model(
        model, test_loader, criterion, device
    )

    print(f"Test loss    : {test_loss:.6f}")
    print(f"Test accuracy: {test_accuracy:.6f}")

    return {
        "checkpoint_file": checkpoint_file,
        "checkpoint_name": os.path.basename(checkpoint_file),
        "loss": float(test_loss),
        "accuracy": float(test_accuracy),
        "y_pred_classes": y_pred,
        "y_true": y_true
    }
    
    
def plot_confusion_matrix(cm, CLASS_NAMES):
    figure = plt.figure(figsize=(8, 8))
    plt.imshow(cm, interpolation='nearest', cmap=plt.cm.Blues)
    plt.title("Confusion matrix")
    plt.colorbar()
    tick_marks = np.arange(len(CLASS_NAMES))
    plt.xticks(tick_marks, CLASS_NAMES, rotation=45)
    plt.yticks(tick_marks, CLASS_NAMES)

    row_sums = cm.sum(axis=1, keepdims=True)
    row_sums[row_sums == 0] = 1
    cm_norm = np.around(cm.astype('float') / row_sums, decimals=2)

    threshold = cm_norm.max() / 2.
    for i, j in itertools.product(range(cm_norm.shape[0]), range(cm_norm.shape[1])):
        color = "white" if cm_norm[i, j] > threshold else "black"
        plt.text(j, i, cm_norm[i, j], horizontalalignment="center", color=color)

    plt.tight_layout()
    plt.ylabel('True label')
    plt.xlabel('Predicted label')
    return figure


def save_confusion_matrix_image(conf_matrix, class_names, save_path):
    fig = plot_confusion_matrix(conf_matrix, class_names)
    fig.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved confusion matrix: {save_path}")
    
# =========================================================
# PyTorch device setup
# =========================================================
device = torch.device("cuda:0" if (str(GPU_MODE).lower() == "on" and torch.cuda.is_available()) else "cpu")
print("=" * 60)
print("PyTorch device:", device)
print("torch.cuda.is_available():", torch.cuda.is_available())
if torch.cuda.is_available():
    print("GPU count:", torch.cuda.device_count())
    print("GPU name :", torch.cuda.get_device_name(0))
print("=" * 60)


if MODE.lower() == 'train':

    TRAIN_MODE(BAD_FILE_PATH_TRAIN, TARGET_FILE_PATH_TRAIN, NOISE_FILE_PATH_TRAIN,
               bad_sample_excel_path_train, target_sample_excel_path_train, noise_sample_excel_path_train)
    TRAIN_MODE(BAD_FILE_PATH_VAL, TARGET_FILE_PATH_VAL, NOISE_FILE_PATH_VAL,
               bad_sample_excel_path_val, target_sample_excel_path_val, noise_sample_excel_path_val)

    x_train_seq, x_train_scalar, y_train, train_lengths, train_meta = load_data(
        bad_sample_excel_path_train,
        target_sample_excel_path_train,
        noise_sample_excel_path_train,
        COLUMNS
    )

    x_val_seq, x_val_scalar, y_val, val_lengths, val_meta = load_data(
        bad_sample_excel_path_val,
        target_sample_excel_path_val,
        noise_sample_excel_path_val,
        COLUMNS
    )

    if x_train_seq is None or x_train_scalar is None or y_train is None or train_lengths is None:
        print("Failed to load training data.")
    elif x_val_seq is None or x_val_scalar is None or y_val is None or val_lengths is None:
        print("Failed to load validation data.")
    else:
        print("=" * 60)
        print("LSTM input check")
        print("=" * 60)
        print("x_train_seq shape   :", x_train_seq.shape)
        print("x_train_scalar shape:", x_train_scalar.shape)
        print("y_train shape       :", y_train.shape)
        print("example seq shape   :", x_train_seq[0].shape)
        print("example scalar      :", x_train_scalar[0])
        print("example label       :", y_train[0])
        print("=" * 60)

        train_dataset = SequenceScalarDataset(x_train_seq, x_train_scalar, y_train, train_lengths, train_meta)
        val_dataset = SequenceScalarDataset(x_val_seq, x_val_scalar, y_val, val_lengths, val_meta)

        train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True, num_workers=0)
        val_loader = DataLoader(val_dataset, batch_size=BATCH_SIZE, shuffle=False, num_workers=0)
        
        model = HybridLSTMClassifier(
            seq_input_dim=2,
            scalar_input_dim=x_train_scalar.shape[1],
            num_classes=NB_CLASSES
        ).to(device)
        criterion = nn.CrossEntropyLoss()
        optimizer = torch.optim.Adam(model.parameters(), lr=LEARNING_RATE)

        os.makedirs(CHECKPOINT_DIR, exist_ok=True)

        start_epoch = 1
        if KEEP_TRAINING.lower() == 'on':
            loaded_epoch = load_latest_checkpoint_torch(model, CHECKPOINT_DIR, device)
            if loaded_epoch > 0:
                start_epoch = loaded_epoch + 1

        best_val_acc = 0.0

        for epoch in range(start_epoch, EPOCHS + 1):
            train_loss, train_acc = train_one_epoch(model, train_loader, optimizer, criterion, device)
            val_loss, val_acc, val_preds, val_true, val_sample_names = evaluate_model(
                    model, val_loader, criterion, device
                )

            print(f"[Epoch {epoch:03d}] "
                  f"Train Loss={train_loss:.6f} | Train Acc={train_acc:.4f} | "
                  f"Val Loss={val_loss:.6f} | Val Acc={val_acc:.4f}")

            ckpt_path = os.path.join(
                CHECKPOINT_DIR,
                f"epoch_{epoch:04d}_val_acc_{val_acc:.4f}_loss_{val_loss:.4f}.pth"
            )
            torch.save(model.state_dict(), ckpt_path)

            if val_acc > best_val_acc:
                best_val_acc = val_acc
                best_path = os.path.join(CHECKPOINT_DIR, "best_model.pth")
                torch.save(model.state_dict(), best_path)
                print(f"🔥 Best model updated: {best_path}")


elif MODE.lower() in ['test', 'prediction', 'classification']:
    TRAIN_MODE(
        BAD_FILE_PATH_TEST, TARGET_FILE_PATH_TEST, NOISE_FILE_PATH_TEST,
        bad_sample_excel_path_test, target_sample_excel_path_test, noise_sample_excel_path_test
    )

    x_test_seq, x_test_scalar, y_test, test_lengths, test_meta = load_data(
        bad_sample_excel_path_test,
        target_sample_excel_path_test,
        noise_sample_excel_path_test,
        COLUMNS
    )

    if x_test_seq is None or x_test_scalar is None or y_test is None or test_lengths is None:
        print("Failed to load test data. Please check the data loading process.")
    else:
        print("=" * 60)
        print("TEST input check")
        print("=" * 60)
        print("x_test_seq shape   :", x_test_seq.shape)
        print("x_test_scalar shape:", x_test_scalar.shape)
        print("y_test shape       :", y_test.shape)
        print("=" * 60)

        test_dataset = SequenceScalarDataset(x_test_seq, x_test_scalar, y_test, test_lengths, test_meta)
        test_loader = DataLoader(test_dataset, batch_size=BATCH_SIZE, shuffle=False, num_workers=0)

        model = HybridLSTMClassifier(
            seq_input_dim=2,
            scalar_input_dim=x_test_scalar.shape[1],
            num_classes=NB_CLASSES
        ).to(device)

        criterion = nn.CrossEntropyLoss()

        # =========================================================
        # TEST MODE: evaluate all checkpoint files
        # =========================================================
        if MODE.lower() == 'test':
            checkpoint_files = get_all_checkpoint_files(CHECKPOINT_DIR)

            if len(checkpoint_files) == 0:
                print(f"No checkpoint files found in: {CHECKPOINT_DIR}")
            else:
                all_results = []

                for checkpoint_file in checkpoint_files:
                    result = evaluate_single_checkpoint(
                        model,
                        checkpoint_file,
                        test_loader,
                        criterion,
                        device
                    )
                    all_results.append(result)

                all_results = sorted(
                    all_results,
                    key=lambda x: (x["accuracy"], -x["loss"]),
                    reverse=True
                )

                TOP_K = min(10, len(all_results))
                top_results = all_results[:TOP_K]

                print("\n" + "=" * 100)
                print(f"TOP {TOP_K} CHECKPOINT TEST RESULTS (sorted by accuracy)")
                print("=" * 100)

                for idx, result in enumerate(top_results, start=1):
                    print(
                        f"{idx:02d}. {result['checkpoint_name']} | "
                        f"acc={result['accuracy']:.6f} | "
                        f"loss={result['loss']:.6f}"
                    )
                    print(f"    Pred Label : {result['y_pred_classes']}")
                    print(f"    True Label : {result['y_true']}")
                    print("-" * 100)

                top10_df = pd.DataFrame([
                    {
                        "rank": idx,
                        "checkpoint_name": result["checkpoint_name"],
                        "accuracy": result["accuracy"],
                        "loss": result["loss"],
                        "pred_label": " ".join(map(str, result["y_pred_classes"])),
                        "true_label": " ".join(map(str, result["y_true"]))
                    }
                    for idx, result in enumerate(top_results, start=1)
                ])

                top10_csv_path = os.path.join(CHECKPOINT_DIR, "top10_checkpoint_results.csv")
                top10_df.to_csv(top10_csv_path, index=False, encoding="utf-8-sig")
                print(f"\nTop {TOP_K} results saved to: {top10_csv_path}")

                best_result = top_results[0]

                print("\n" + "=" * 100)
                print("BEST CHECKPOINT RESULT")
                print("=" * 100)
                print(f"Best checkpoint : {best_result['checkpoint_name']}")
                print(f"Best accuracy   : {best_result['accuracy']:.6f}")
                print(f"Best loss       : {best_result['loss']:.6f}")
                print(f"Pred Label      : {best_result['y_pred_classes']}")
                print(f"True Label      : {best_result['y_true']}")

                conf_matrix = confusion_matrix(
                    best_result['y_true'],
                    best_result['y_pred_classes']
                )

                conf_save_path = os.path.join(CHECKPOINT_DIR, "best_confusion_matrix.png")
                save_confusion_matrix_image(conf_matrix, CLASS_NAMES, conf_save_path)

        # =========================================================
        # PREDICTION MODE: use one checkpoint only
        # =========================================================
        elif MODE.lower() == 'prediction':
            if not os.path.exists(BEST_CHECKPOINT_FILE):
                print(f"Checkpoint file not found: {BEST_CHECKPOINT_FILE}")
            else:
                model.load_state_dict(torch.load(BEST_CHECKPOINT_FILE, map_location=device))

                
                test_loss, test_accuracy, y_pred_classes, y_true, sample_names = evaluate_model(
                    model, test_loader, criterion, device
                )

                print("\n" + "=" * 100)
                print("PREDICTION MODE")
                print("=" * 100)
                print(f"Checkpoint file : {BEST_CHECKPOINT_FILE}")
                print(f"Pred Label      : {y_pred_classes}")
                print(f"True Label      : {y_true}")
                print(f"Prediction loss     : {test_loss:.6f}")
                print(f"Prediction accuracy : {test_accuracy:.6f}")

                conf_matrix = confusion_matrix(y_true, y_pred_classes)
                conf_save_path = os.path.join(CHECKPOINT_DIR, "prediction_confusion_matrix.png")
                save_confusion_matrix_image(conf_matrix, CLASS_NAMES, conf_save_path)
                
            
        # =========================================================
        # CLASSIFICATION MODE: classify real files
        # =========================================================

        
        elif MODE.lower() == 'classification':
        
            CLASSIFICATION_ROOT = os.path.join(REFERENCE_FILE_PATH, "Classification")

            ACTIVE_DIR = os.path.join(CLASSIFICATION_ROOT, "ACTIVE")
            INACTIVE_DIR = os.path.join(CLASSIFICATION_ROOT, "INACTIVE")
            NOISE_DIR = os.path.join(CLASSIFICATION_ROOT, "NOISE")

            os.makedirs(ACTIVE_DIR, exist_ok=True)
            os.makedirs(INACTIVE_DIR, exist_ok=True)
            os.makedirs(NOISE_DIR, exist_ok=True)

            if not os.path.exists(BEST_CHECKPOINT_FILE):
                print(f"Checkpoint file not found: {BEST_CHECKPOINT_FILE}")

            else:
                model.load_state_dict(torch.load(BEST_CHECKPOINT_FILE, map_location=device))

                print("\n" + "=" * 100)
                print("CLASSIFICATION MODE")
                print("=" * 100)

                model.eval()

                with torch.no_grad():

                    for seq_x, scalar_x, y, lengths, meta in test_loader:

                        seq_x = seq_x.to(device)
                        scalar_x = scalar_x.to(device)
                        lengths = lengths.to(device)

                        logits = model(seq_x, scalar_x, lengths)

                        preds = torch.argmax(logits, dim=1).cpu().numpy()

                        for i in range(len(preds)):

                            pred_class = int(preds[i])

                            sample_name = meta["sample_name"][i]
                            source_path = meta["source_path"][i]
                            if not os.path.exists(source_path):
                                print(f"[WARN] File not found: {source_path}")
                                continue

                            if pred_class == 1:
                                dst_dir = ACTIVE_DIR

                            elif pred_class == 0:
                                dst_dir = INACTIVE_DIR

                            else:
                                dst_dir = NOISE_DIR

                            dst_path = os.path.join(dst_dir, os.path.basename(source_path))

                            try:

                                if os.path.isdir(source_path):

                                    if os.path.exists(dst_path):
                                        shutil.rmtree(dst_path)

                                    shutil.copytree(source_path, dst_path)

                                else:

                                    shutil.copy2(source_path, dst_path)

                                print(f"[OK] {sample_name} -> class {pred_class}")

                            except Exception as e:

                                print(f"[ERROR] {sample_name} : {e}")